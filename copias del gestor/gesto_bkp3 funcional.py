# gestor.py (completo, corregido - SQLite local)
print("🔄 Cargando dependencias...")

try:
    from flask import Flask, render_template, request, redirect, url_for, flash, session, jsonify
    print("✅ Flask importado correctamente")
    
    from datetime import datetime, timedelta
    print("✅ DateTime importado correctamente")
    
    from openpyxl import Workbook, load_workbook
    import openpyxl
    print("✅ OpenPyXL importado correctamente")
    
    import pandas as pd
    print("✅ Pandas importado correctamente")
    
    import os
    import sqlite3
    import traceback
    print("✅ Librerías del sistema importadas correctamente")
    
except ImportError as e:
    print(f"❌ Error de importación: {e}")
    print("💡 Instala las dependencias con: pip install flask pandas openpyxl")
    exit(1)
except Exception as e:
    print(f"❌ Error inesperado en importaciones: {e}")
    exit(1)
try:
    from flask_wtf import CSRFProtect
    from flask_wtf.csrf import generate_csrf
    _HAS_FLASK_WTF = True
    print("✅ Flask-WTF importado correctamente")
except ImportError:
    print("⚠️ Flask-WTF no disponible, usando fallback simple")
    # Flask-WTF not available, provide a minimal fallback
    CSRFProtect = None
    _HAS_FLASK_WTF = False
    def generate_csrf():
        # lazy import to avoid top-level secret import when not needed
        import secrets
        if '_csrf_token' not in session:
            session['_csrf_token'] = secrets.token_urlsafe(16)
        return session['_csrf_token']
except Exception as e:
    print(f"⚠️ Error con Flask-WTF: {e}, usando fallback simple")
    CSRFProtect = None
    _HAS_FLASK_WTF = False
    def generate_csrf():
        import secrets
        if '_csrf_token' not in session:
            session['_csrf_token'] = secrets.token_urlsafe(16)
        return session['_csrf_token']
from decimal import Decimal, InvalidOperation
from werkzeug.security import generate_password_hash, check_password_hash
from functools import wraps
print("✅ Werkzeug y functools importados correctamente")

# --- Configuración de la Aplicación ---
app = Flask(__name__)
app.secret_key = os.getenv('SECRET_KEY', 'supersecretkey')
app.permanent_session_lifetime = timedelta(hours=3)

if _HAS_FLASK_WTF:
    # Use Flask-WTF CSRFProtect when available
    csrf = CSRFProtect(app)
    # Expose generate_csrf to templates as csrf_token()
    app.jinja_env.globals['csrf_token'] = generate_csrf
else:
    # Fallback: expose our simple generate_csrf and validate on POST
    app.jinja_env.globals['csrf_token'] = generate_csrf

    # Endpoints seguros/legacy que omiten la verificación CSRF (AJAX/legacy forms)
    CSRF_WHITELIST = {
        'eliminar_seleccionados',
        'eliminar_todo_historial',
        'eliminar_producto_stock',
        'eliminar_proveedor_manual',
        'agregar_carrito_ajax',
        'agregar_carrito_manual_ajax'
    }
    # Rutas exactas que también pueden omitir CSRF (por si request.endpoint no coincide)
    CSRF_WHITELIST_PATHS = {
        '/eliminar_seleccionados',
        '/eliminar_todo_historial',
        '/eliminar_producto_stock',
        '/eliminar_proveedor_manual',
        '/agregar_carrito_ajax',
        '/agregar_carrito_manual_ajax'
    }

    @app.before_request
    def csrf_protect_fallback():
        # Allow specific endpoints or paths to bypass CSRF fallback (legacy forms or JS not sending token)
        try:
            ep = request.endpoint
        except Exception:
            ep = None
        try:
            path = request.path
        except Exception:
            path = None
        if ep in CSRF_WHITELIST or (path in CSRF_WHITELIST_PATHS):
            return

        if request.method == 'POST':
            token = session.get('_csrf_token')
            if not token:
                import secrets
                session['_csrf_token'] = secrets.token_urlsafe(16)
                token = session['_csrf_token']

            form_token = request.form.get('csrf_token')
            header_token = request.headers.get('X-CSRFToken') or request.headers.get('X-CSRF-Token')
            json_token = None
            if request.is_json:
                try:
                    data = request.get_json(silent=True) or {}
                    json_token = data.get('csrf_token')
                except Exception:
                    json_token = None

            if form_token == token or header_token == token or json_token == token:
                return

            if request.is_json:
                return jsonify({'success': False, 'error': 'CSRF token inválido o ausente.'}), 400
            flash('CSRF token inválido o ausente.', 'danger')
            if 'user_id' in session:
                return redirect(url_for('index'))
            return redirect(url_for('login'))

# --- Rutas de Archivos y Carpetas ---
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_FOLDER = os.path.join(BASE_DIR, 'listas_excel')
MANUAL_PRODUCTS_FILE = os.path.join(EXCEL_FOLDER, 'productos_manual.xlsx')
DATABASE_FILE = os.path.join(BASE_DIR, 'gestor_stock.db')

# --- Configuración de Dueños ---
DUENOS_CONFIG = {
    'ricky': {
        'nombre': 'Ricky',
        'proveedores_excel': ['brementools', 'berger', 'cachan', 'chiesa', 'crossmaster'],
        'puede_excel': True
    },
    'ferreteria_general': {
        'nombre': 'Ferretería General', 
        'proveedores_excel': [],
        'puede_excel': False
    }
}

# --- Configuración de Proveedores Excel ---
PROVEEDOR_CONFIG = {
    'brementools': {
        'fila_encabezado': 5,
        'codigo': ['codigo', 'Código', 'CODIGO'],
        'producto': ['producto', 'Producto', 'PRODUCTO'],
        'precio': ['precio', 'Precio', 'PRECIO']
    },
    'crossmaster': {
        'fila_encabezado': 11,
        'codigo': ['codigo', 'Codigo', 'CODIGO'],
        'producto': ['descripcion', 'Descripcion', 'DESCRIPCION'],
        'precio': ['precio', 'Precio', 'PRECIO']
    },
    'berger': {
        'fila_encabezado': 0,
        'codigo': ['cod', 'COD', 'codigo', 'Codigo'],
        'producto': ['detalle', 'DETALLE', 'producto', 'Producto'],
        'precio': ['P.VENTA', 'precio', 'Precio', 'PRECIO']
    },
    'chiesa': {
        'fila_encabezado': 1,
        'codigo': ['codigo', 'Codigo', 'CODIGO'],
        'producto': ['descripción', 'Descripción', 'descripcion', 'Descripcion'],
        'precio': ['precio', 'Precio', 'PRECIO']
    },
    'cachan': {
        'fila_encabezado': 0,
        'codigo': ['codigo', 'Codigo', 'CODIGO'],
        'producto': ['nombre', 'Nombre', 'NOMBRE'],
        'precio': ['precio', 'Precio', 'PRECIO']
    }
}

# --- Funciones de Utilidad ---
def parse_price(price_str):
    """
    Interpreta precios con distintos formatos:
    - 200.000 -> 200000
    - 200,50  -> 200.50
    - 200.000,75 -> 200000.75
    - 200,000.75 -> 200000.75 (formato inglés)
    - '' (vacío) -> 0.0
    Devuelve (float, texto_error)
    """
    if price_str is None:
        return 0.0, None
    
    # Convertir a string si no lo es
    if isinstance(price_str, (int, float)):
        return float(price_str), None
    
    if not isinstance(price_str, str):
        try:
            return float(str(price_str)), None
        except (ValueError, TypeError):
            return 0.0, f"Cannot convert {price_str} to number"
    
    # String vacío devuelve 0
    price_str = price_str.strip()
    if not price_str:
        return 0.0, None

    cleaned_str = price_str.replace('$', '').replace(' ', '').strip()
    
    # Si después de limpiar queda vacío, devolver 0
    if not cleaned_str:
        return 0.0, None

    # Caso: contiene tanto punto como coma
    if '.' in cleaned_str and ',' in cleaned_str:
        # Si la coma está después del último punto asumimos formato español: 200.000,75
        if cleaned_str.rfind(',') > cleaned_str.rfind('.'):
            cleaned_str = cleaned_str.replace('.', '').replace(',', '.')
        else:
            # Caso inglés: 200,000.75
            cleaned_str = cleaned_str.replace(',', '')
    elif ',' in cleaned_str:
        # Solo coma: decimal
        cleaned_str = cleaned_str.replace(',', '.')
    else:
        # Solo puntos: probablemente separadores de miles
        # Si hay más de un punto o el segmento después del punto tiene 3 dígitos, quitar puntos
        if cleaned_str.count('.') > 1 or (cleaned_str.count('.') == 1 and len(cleaned_str.split('.')[1]) == 3):
            cleaned_str = cleaned_str.replace('.', '')

    try:
        result = float(cleaned_str)
        return result, None
    except (ValueError, TypeError):
        # Si no se puede convertir, devolver 0 y el texto original como "error"
        return 0.0, f"Cannot parse price: {price_str}"

# --- Funciones de Base de Datos ---
def get_db_connection():
    """Crear conexión a la base de datos SQLite local"""
    try:
        conn = sqlite3.connect(DATABASE_FILE)
        conn.row_factory = sqlite3.Row
        return conn
    except sqlite3.Error as e:
        print(f"Error de conexión a SQLite: {e}")
        return None

def db_query(query, params=(), fetch=False):
    """Ejecutar consulta en la base de datos SQLite"""
    conn = get_db_connection()
    if not conn:
        return None
    cursor = conn.cursor()
    result = None
    try:
        cursor.execute(query, params)
        if fetch:
            rows = cursor.fetchall()
            # Convert sqlite3.Row objects to dictionaries
            result = [dict(row) for row in rows]
        else:
            conn.commit()
            result = True
    except sqlite3.Error as e:
        print(f"Error en la consulta a la base de datos: {e}")
        try:
            conn.rollback()
        except Exception:
            pass
        result = False
    finally:
        try:
            cursor.close()
        except Exception:
            pass
        try:
            conn.close()
        except Exception:
            pass
    return result

def init_excel_manual():
    """Inicializar el archivo Excel de productos manuales"""
    if not os.path.exists(EXCEL_FOLDER):
        os.makedirs(EXCEL_FOLDER)
    
    # Estructura esperada y unificada (orden requerido):
    # Codigo | Proveedor | Nombre | Precio | Observaciones | Dueno
    expected_headers = ['Codigo', 'Proveedor', 'Nombre', 'Precio', 'Observaciones', 'Dueno']
    
    if not os.path.exists(MANUAL_PRODUCTS_FILE):
        # Crear nuevo archivo Excel con las columnas requeridas
        wb = Workbook()
        ws = wb.active
        ws.title = "Productos Manuales"
        
        for col, header in enumerate(expected_headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        wb.save(MANUAL_PRODUCTS_FILE)
        print(f"Archivo Excel creado: {MANUAL_PRODUCTS_FILE}")
    else:
        try:
            # Si existe, normalizar encabezados en caso de que tengan acentos u orden distinto
            wb = load_workbook(MANUAL_PRODUCTS_FILE)
            ws = wb.active
            current = [ws.cell(row=1, column=i).value for i in range(1, 7)]
            # Normalizar valores con posibles acentos
            normalize = lambda s: str(s or '').strip().replace('ó', 'o').replace('Ó', 'O').replace('ñ', 'n').replace('Ñ', 'N').title()
            normalized = [normalize(v) for v in current]
            # Si detectamos el antiguo orden (Nombre antes que Proveedor), migrar filas intercambiando columnas B y C
            old_order = ['Codigo', 'Nombre', 'Proveedor', 'Precio', 'Observaciones', 'Dueno']
            if normalized == old_order:
                for r in range(2, ws.max_row + 1):
                    val_b = ws.cell(row=r, column=2).value  # Nombre
                    val_c = ws.cell(row=r, column=3).value  # Proveedor
                    # Nuevo orden requiere Proveedor en B y Nombre en C
                    ws.cell(row=r, column=2, value=val_c)
                    ws.cell(row=r, column=3, value=val_b)
                # Reescribir encabezados al orden esperado
                for idx, header in enumerate(expected_headers, start=1):
                    ws.cell(row=1, column=idx, value=header)
                wb.save(MANUAL_PRODUCTS_FILE)
            elif normalized != expected_headers:
                # Si solo difieren por acentos u otros detalles, reescribimos encabezados al estándar
                for idx, header in enumerate(expected_headers, start=1):
                    ws.cell(row=1, column=idx, value=header)
                wb.save(MANUAL_PRODUCTS_FILE)
        except Exception as e:
            print(f"Advertencia al normalizar/migrar encabezados del Excel manual: {e}")

def agregar_producto_excel_manual(codigo, proveedor, nombre, precio, observaciones, dueno):
    """Agregar producto al Excel de productos manuales"""
    try:
        if not os.path.exists(EXCEL_FOLDER):
            os.makedirs(EXCEL_FOLDER)
        
        expected_headers = ['Codigo', 'Proveedor', 'Nombre', 'Precio', 'Observaciones', 'Dueno']
        
        if not os.path.exists(MANUAL_PRODUCTS_FILE):
            wb = Workbook()
            ws = wb.active
            ws.title = "Productos Manuales"
            # Encabezados estandarizados sin acentos y en orden requerido
            for col, header in enumerate(expected_headers, 1):
                ws.cell(row=1, column=col, value=header)
            wb.save(MANUAL_PRODUCTS_FILE)
        
        wb = load_workbook(MANUAL_PRODUCTS_FILE)
        ws = wb.active
        
        # Verificar y corregir encabezados si es necesario
        try:
            current = [ws.cell(row=1, column=i).value for i in range(1, 7)]
            normalize = lambda s: str(s or '').strip().replace('ó', 'o').replace('Ó', 'O').replace('ñ', 'n').replace('Ñ', 'N').title()
            normalized = [normalize(v) for v in current]
            old_order = ['Codigo', 'Nombre', 'Proveedor', 'Precio', 'Observaciones', 'Dueno']
            if normalized == old_order:
                # Migrar filas intercambiando columnas B y C para adecuar al nuevo orden
                for r in range(2, ws.max_row + 1):
                    val_b = ws.cell(row=r, column=2).value  # Nombre
                    val_c = ws.cell(row=r, column=3).value  # Proveedor
                    ws.cell(row=r, column=2, value=val_c)
                    ws.cell(row=r, column=3, value=val_b)
                # Actualizar encabezados al estándar
                for idx, header in enumerate(expected_headers, start=1):
                    ws.cell(row=1, column=idx, value=header)
            elif normalized != expected_headers:
                for idx, header in enumerate(expected_headers, start=1):
                    ws.cell(row=1, column=idx, value=header)
        except Exception as _:
            pass
        
        # Agregar fila respetando el orden de columnas esperado
        ws.append([codigo, proveedor, nombre, float(precio) if precio is not None else 0.0, observaciones, dueno])
        wb.save(MANUAL_PRODUCTS_FILE)
        
        return True
    except Exception as e:
        print(f"Error al agregar producto manual: {e}")
        return False

def buscar_en_excel_manual(termino_busqueda, dueno_filtro=None):
    """Buscar productos en el Excel de productos manuales"""
    resultados = []
    
    try:
        if not os.path.exists(MANUAL_PRODUCTS_FILE):
            return resultados
        
        wb = load_workbook(MANUAL_PRODUCTS_FILE)
        ws = wb.active
        
        # Asumir que las columnas son (orden unificado):
        # Codigo (A), Proveedor (B), Nombre (C), Precio (D), Observaciones (E), Dueno (F)
        for row in ws.iter_rows(min_row=2, values_only=True):  # Saltar encabezado
            if len(row) < 6:
                continue
            
            codigo = str(row[0]).strip() if row[0] else ''
            proveedor = str(row[1]).strip() if row[1] else ''
            nombre = str(row[2]).strip() if row[2] else ''
            precio = str(row[3]).strip() if row[3] else ''
            observaciones = str(row[4]).strip() if row[4] else ''
            dueno = str(row[5]).strip() if row[5] else ''
            
            # Filtrar por dueño si se especifica
            if dueno_filtro and dueno.lower() != dueno_filtro.lower():
                continue
            
            # Buscar en los campos relevantes
            termino = termino_busqueda.lower()
            if (termino in codigo.lower() or
                termino in proveedor.lower() or
                termino in nombre.lower() or
                termino in precio.lower() or
                termino in observaciones.lower()):
                
                resultados.append({
                    'codigo': codigo,
                    'proveedor': proveedor,
                    'nombre': nombre,
                    'precio': precio,
                    'observaciones': observaciones,
                    'dueno': dueno
                })
    
    except Exception as e:
        print(f"Error al buscar en Excel manual: {e}")
    
    return resultados

def init_db():
    """Inicializar la base de datos SQLite local"""
    if not os.path.exists(EXCEL_FOLDER):
        os.makedirs(EXCEL_FOLDER)
    try:
        conn = get_db_connection()
        if not conn:
            print("ERROR al inicializar la BD: No se pudo establecer la conexión con la base de datos.")
            return

        cursor = conn.cursor()
        
        # Crear tabla stock
        cursor.execute(''' 
            CREATE TABLE IF NOT EXISTS stock ( 
                id INTEGER PRIMARY KEY AUTOINCREMENT, 
                codigo TEXT, 
                nombre TEXT NOT NULL, 
                precio REAL NOT NULL, 
                cantidad INTEGER NOT NULL, 
                fecha_compra TEXT NOT NULL, 
                proveedor TEXT, 
                observaciones TEXT, 
                precio_texto TEXT, 
                avisar_bajo_stock INTEGER DEFAULT 0, 
                min_stock_aviso INTEGER DEFAULT NULL 
            ) 
        ''')
        
        # Crear tabla proveedores_manual
        cursor.execute(''' 
            CREATE TABLE IF NOT EXISTS proveedores_manual ( 
                id INTEGER PRIMARY KEY AUTOINCREMENT, 
                nombre TEXT NOT NULL UNIQUE 
            ) 
        ''')
        
        # Crear tabla productos_manual
        cursor.execute(''' 
            CREATE TABLE IF NOT EXISTS productos_manual ( 
                id INTEGER PRIMARY KEY AUTOINCREMENT, 
                proveedor_id INTEGER, 
                nombre TEXT NOT NULL, 
                codigo TEXT, 
                precio REAL NOT NULL, 
                FOREIGN KEY (proveedor_id) REFERENCES proveedores_manual(id) ON DELETE CASCADE 
            ) 
        ''')
        
        # Crear tabla users
        cursor.execute(''' 
            CREATE TABLE IF NOT EXISTS users ( 
                id INTEGER PRIMARY KEY AUTOINCREMENT, 
                username TEXT NOT NULL UNIQUE, 
                password_hash TEXT NOT NULL 
            ) 
        ''')
        
        # Crear usuario por defecto
        default_username = 'Pauluk'
        default_password = 'Jap2005'
        cursor.execute("SELECT id FROM users WHERE username = ?", (default_username,))
        if cursor.fetchone() is None:
            hashed_password = generate_password_hash(default_password)
            cursor.execute("INSERT INTO users (username, password_hash) VALUES (?, ?)", 
                         (default_username, hashed_password))
        
        # Crear proveedores por defecto
        proveedores_excel = ['BremenTools', 'Crossmaster', 'Berger', 'Chiesa', 'Cachan', 'Otros Proveedores']
        for nombre in proveedores_excel:
            cursor.execute("INSERT OR IGNORE INTO proveedores_manual (nombre) VALUES (?)", (nombre,))
        
        conn.commit()
        cursor.close()
        conn.close()
        print("Base de datos SQLite inicializada/verificada con éxito.")
        
        # Inicializar también el Excel de productos manuales
        init_excel_manual()
        
    except sqlite3.Error as e:
        print(f"\nERROR al inicializar la BD: {e}")
    except Exception as e:
        print(f"\nERROR al inicializar la BD (no-SQLite): {e}")

# --- Decorador de Autenticación ---
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

# --- Rutas de Autenticación ---
@app.route('/login', methods=['GET', 'POST'])
def login():
    if 'user_id' in session:
        return redirect(url_for('index'))
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        user = db_query("SELECT * FROM users WHERE username = ?", (username,), fetch=True)
        if user and check_password_hash(user[0]['password_hash'], password):
            session.clear()
            session.permanent = True
            session['user_id'] = user[0]['id']
            session['username'] = user[0]['username']
            return redirect(url_for('index'))
        else:
            flash('Usuario o contraseña incorrectos.', 'danger')
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    flash('Has cerrado sesión.', 'success')
    return redirect(url_for('login'))

# --- Rutas de la Aplicación Web (Protegidas) ---
@app.route('/')
@login_required
def index():
    return render_template('index.html')

@app.route('/historial')
@login_required
def historial():
    termino = request.args.get('q', '')
    if termino:
        like_pattern = f'%{termino}%'
        productos = db_query( 
            "SELECT id, fecha_compra, codigo, nombre, proveedor, precio, cantidad, observaciones FROM stock WHERE nombre LIKE ? OR codigo LIKE ? OR proveedor LIKE ? OR observaciones LIKE ? ORDER BY fecha_compra DESC, id DESC", 
            (like_pattern, like_pattern, like_pattern, like_pattern), 
            fetch=True 
        )
    else:
        productos = db_query("SELECT id, fecha_compra, codigo, nombre, proveedor, precio, cantidad, observaciones FROM stock ORDER BY fecha_compra DESC, id DESC", fetch=True)
    return render_template('historial.html', productos=productos, termino_busqueda=termino)

@app.route('/stock')
@login_required
def stock():
    termino = request.args.get('q', '')
    if termino:
        like_pattern = f'%{termino}%'
        productos = db_query(
            "SELECT id, codigo, nombre, precio, cantidad, proveedor, observaciones FROM stock WHERE nombre LIKE ? OR codigo LIKE ? OR proveedor LIKE ? ORDER BY nombre",
            (like_pattern, like_pattern, like_pattern),
            fetch=True
        )
    else:
        productos = db_query("SELECT id, codigo, nombre, precio, cantidad, proveedor, observaciones FROM stock ORDER BY nombre", fetch=True)
    return render_template('stock.html', productos=productos, termino_busqueda=termino)

@app.context_processor
def inject_notificacion_emergente():
    return dict(notificacion_emergente=session.pop('notificacion_emergente', None))

@app.route('/notificaciones')
@login_required
def notificaciones():
    notificaciones = session.get('notificaciones', [])
    leidas = session.get('notificaciones_leidas', False)
    return render_template('notificaciones.html', notificaciones=notificaciones, leidas=leidas)

@app.route('/borrar_todas_notificaciones', methods=['POST'])
@login_required
def borrar_todas_notificaciones():
    session['notificaciones'] = []
    session['notificaciones_leidas'] = False
    return redirect(url_for('notificaciones'))

@app.route('/marcar_notificaciones_leidas', methods=['POST'])
@login_required
def marcar_notificaciones_leidas():
    session['notificaciones_leidas'] = True
    return redirect(url_for('notificaciones'))

# --- Rutas principales de la aplicación ---

@app.route('/agregar_producto', methods=['GET', 'POST'])
@login_required
def agregar_producto():
    # Inicializar carrito si no existe
    if 'carrito' not in session:
        session['carrito'] = []
    
    carrito = session['carrito']
    
    # Manejar formulario manual para agregar al Excel (no directamente al stock)
    if request.method == 'POST' and 'agregar_producto_manual_excel' in request.form:
        return agregar_producto_manual_excel()
    
    # Manejar acciones del carrito
    if request.method == 'POST' and 'cargar_carrito' in request.form:
        # Cargar todos los productos del carrito al stock
        productos_agregados = 0
        for item in carrito:
            try:
                # Insertar en la base de datos
                db_query("""
                    INSERT INTO stock (codigo, nombre, precio, cantidad, fecha_compra, 
                                     proveedor, observaciones, precio_texto, avisar_bajo_stock, min_stock_aviso)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    item.get('codigo', ''),
                    item.get('nombre', ''),
                    float(item.get('precio', 0)),
                    int(item.get('cantidad', 1)),
                    item.get('fecha_compra', datetime.now().strftime('%Y-%m-%d')),
                    item.get('proveedor', ''),
                    item.get('observaciones', ''),
                    item.get('precio_texto', ''),
                    int(item.get('avisar_bajo_stock', 0)),
                    int(item.get('min_stock_aviso')) if item.get('min_stock_aviso') else None
                ))
                productos_agregados += 1
            except Exception as e:
                print(f"Error al agregar producto al stock: {e}")
        
        # Limpiar carrito
        session['carrito'] = []
        
        if productos_agregados > 0:
            flash(f'{productos_agregados} producto(s) agregado(s) al stock exitosamente.', 'success')
        else:
            flash('No se pudo agregar ningún producto al stock.', 'danger')
        
        return redirect(url_for('historial'))
    
    # Obtener parámetros de búsqueda en Excel
    termino_excel = request.args.get('busqueda_excel', '')
    proveedor_excel_filtro = request.args.get('proveedor_excel', '')
    filtro_excel = request.args.get('filtro_excel', '')
    resultados_excel = []
    
    # Realizar búsqueda en Excel si hay término
    if termino_excel:
        print(f"🔍 Buscando: '{termino_excel}' con filtro proveedor: '{proveedor_excel_filtro}' y filtro adicional: '{filtro_excel}'")
        resultados_excel = buscar_en_excel(termino_excel, proveedor_excel_filtro, filtro_excel)
        print(f"📊 Resultados encontrados: {len(resultados_excel)}")
    
    # Obtener proveedores manuales para el selector
    proveedores = db_query("SELECT id, nombre FROM proveedores_manual ORDER BY nombre", fetch=True) or []
    # print(f"DEBUG: proveedores = {proveedores}")
    
    # Obtener lista de proveedores Excel disponibles
    proveedores_excel = []
    for key, config in PROVEEDOR_CONFIG.items():
        # Buscar archivos que empiecen con el nombre del proveedor (case insensitive)
        archivos = [f for f in os.listdir(EXCEL_FOLDER) if f.lower().startswith(key.lower()) and f.endswith('.xlsx') and f != 'productos_manual.xlsx']
        if archivos:
            archivo_excel = os.path.join(EXCEL_FOLDER, archivos[0])  # Tomar el primero que coincida
            if os.path.exists(archivo_excel):
                nombre_display = key.title().replace('tools', 'Tools')
                proveedores_excel.append({
                    'key': key,
                    'nombre': nombre_display
                })
    
    # Agregar proveedores manuales al selector
    proveedores_manuales_db = db_query("SELECT id, nombre FROM proveedores_manual ORDER BY nombre", fetch=True) or []
    for p in proveedores_manuales_db:
        proveedores_excel.append({
            'key': f'manual_{p["id"]}',
            'nombre': p['nombre']
        })
    
    # print(f"DEBUG: proveedores_excel = {proveedores_excel}")
    # print(f"DEBUG: EXCEL_FOLDER = {EXCEL_FOLDER}")
    # print(f"DEBUG: Archivos en EXCEL_FOLDER = {os.listdir(EXCEL_FOLDER) if os.path.exists(EXCEL_FOLDER) else 'Carpeta no existe'}")
    
    return render_template('agregar.html', 
                         fecha_actual=datetime.now().strftime('%Y-%m-%d'),
                         proveedores_excel=proveedores_excel,
                         proveedores=proveedores,
                         termino_excel=termino_excel,
                         proveedor_excel=proveedor_excel_filtro,
                         filtro_excel=filtro_excel,
                         resultados_excel=resultados_excel,
                         carrito=session.get('carrito', []))

@app.route('/procesar_producto', methods=['POST'])
@login_required
def procesar_producto():
    try:
        nombre = request.form.get('nombre', '').strip()
        codigo = request.form.get('codigo', '').strip()
        precio_str = request.form.get('precio', '').strip()
        cantidad = int(request.form.get('cantidad', 0))
        proveedor = request.form.get('proveedor', '').strip()
        observaciones = request.form.get('observaciones', '').strip()
        fecha_compra = request.form.get('fecha_compra', datetime.now().strftime('%Y-%m-%d'))
        
        if not nombre:
            flash('El nombre del producto es obligatorio.', 'danger')
            return redirect(url_for('agregar_producto'))
        
        precio, error_precio = parse_price(precio_str)
        if error_precio:
            flash(f'Error en el precio: {error_precio}', 'danger')
            return redirect(url_for('agregar_producto'))
        
        # Insertar producto en stock
        result = db_query(
            "INSERT INTO stock (codigo, nombre, precio, cantidad, fecha_compra, proveedor, observaciones, precio_texto) VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
            (codigo, nombre, precio, cantidad, fecha_compra, proveedor, observaciones, precio_str)
        )
        
        if result:
            flash(f'Producto "{nombre}" agregado exitosamente.', 'success')
        else:
            flash('Error al agregar el producto.', 'danger')
            
    except Exception as e:
        flash(f'Error al procesar el producto: {str(e)}', 'danger')
    
    return redirect(url_for('agregar_producto'))

@app.route('/productos_manual')
@login_required
def productos_manual():
    proveedores = db_query("SELECT id, nombre FROM proveedores_manual ORDER BY nombre", fetch=True)
    return render_template('productos_manual.html', proveedores=proveedores)

@app.route('/carrito')
@login_required
def carrito():
    productos_carrito = session.get('carrito', [])
    total = sum(float(p.get('precio', 0)) * int(p.get('cantidad', 0)) for p in productos_carrito)
    return render_template('carrito.html', productos=productos_carrito, total=total)

@app.route('/limpiar_carrito', methods=['POST'])
@login_required
def limpiar_carrito():
    session['carrito'] = []
    flash('Carrito limpiado.', 'success')
    return redirect(url_for('carrito'))

@app.route('/reportes')
@login_required
def reportes():
    # Obtener estadísticas del stock
    productos_stock = db_query("SELECT COUNT(*) as total FROM stock", fetch=True)
    total_productos = productos_stock[0]['total'] if productos_stock else 0
    
    # Calcular valor total del inventario
    valor_inventario = db_query("SELECT SUM(precio * cantidad) as total_valor FROM stock", fetch=True)
    valor_total = valor_inventario[0]['total_valor'] if valor_inventario and valor_inventario[0]['total_valor'] else 0
    
    # Contar proveedores
    proveedores_count = db_query("SELECT COUNT(*) as total FROM proveedores_manual", fetch=True)
    total_proveedores = proveedores_count[0]['total'] if proveedores_count else 0
    
    return render_template('reportes.html', 
                         total_productos=total_productos,
                         valor_total=valor_total,
                         total_proveedores=total_proveedores)

@app.route('/proveedores')
@login_required
def proveedores():
    # Manejar búsqueda si hay término
    busqueda_proveedor = request.args.get('busqueda_proveedor', '').strip()
    proveedores_encontrados = []
    
    if busqueda_proveedor:
        like_pattern = f'%{busqueda_proveedor}%'
        proveedores_encontrados = db_query(
            "SELECT id, nombre FROM proveedores_manual WHERE nombre LIKE ? ORDER BY nombre",
            (like_pattern,),
            fetch=True
        ) or []
    
    proveedores = db_query("SELECT id, nombre FROM proveedores_manual ORDER BY nombre", fetch=True)
    return render_template('proveedores.html', 
                         proveedores=proveedores,
                         busqueda_proveedor=busqueda_proveedor,
                         proveedores_encontrados=proveedores_encontrados)

@app.route('/agregar_proveedor', methods=['POST'])
@login_required
def agregar_proveedor():
    try:
        nombre = request.form.get('nombre', '').strip()
        dueno = request.form.get('dueno', '').strip()  # Si no se especifica, será 'ferreteria_general' por defecto
        
        if not nombre:
            flash('El nombre del proveedor no puede estar vacío.', 'danger')
        else:
            # Determinar el dueño - si no se especifica, es ferreteria_general
            if dueno == 'ricky':
                dueno_nombre = 'Ricky'
                flash_class = 'success'
            else:
                dueno_nombre = 'Ferretería General'
                flash_class = 'info'
            
            # Insertar en la base de datos (sin campo dueno por ahora, manejado en Excel)
            if db_query("INSERT INTO proveedores_manual (nombre) VALUES (?)", (nombre,)):
                flash(f'Proveedor "{nombre}" agregado exitosamente a {dueno_nombre}.', flash_class)
            else:
                flash('Error al agregar el proveedor (posiblemente ya existe).', 'danger')
    except Exception as e:
        print(f"Error al agregar proveedor: {e}")
        flash(f'Error al agregar el proveedor: {str(e)}', 'danger')
    return redirect(url_for('proveedores'))


@app.route('/agregar_proveedor_manual', methods=['POST'])
@login_required
def agregar_proveedor_manual():
    nombre = request.form.get('nombre', '').strip()
    if nombre:
        result = db_query("INSERT OR IGNORE INTO proveedores_manual (nombre) VALUES (?)", (nombre,))
        if result:
            flash(f'Proveedor "{nombre}" agregado exitosamente.', 'success')
        else:
            flash('Error al agregar el proveedor.', 'danger')
    else:
        flash('El nombre del proveedor es obligatorio.', 'danger')
    return redirect(url_for('proveedores'))


@app.route('/eliminar_proveedor_manual', methods=['POST'])
@login_required
def eliminar_proveedor_manual():
    try:
        proveedor_id = request.form.get('id')
        if not proveedor_id:
            flash('ID de proveedor no válido.', 'danger')
            return redirect(url_for('proveedores'))
        
        # Obtener el nombre del proveedor antes de eliminarlo
        proveedor = db_query("SELECT nombre FROM proveedores_manual WHERE id = ?", (proveedor_id,), fetch=True)
        if not proveedor:
            flash('Proveedor no encontrado.', 'danger')
            return redirect(url_for('proveedores'))
        
        nombre_proveedor = proveedor[0]['nombre']
        
        # Eliminar el proveedor (esto también eliminará los productos asociados por CASCADE)
        if db_query("DELETE FROM proveedores_manual WHERE id = ?", (proveedor_id,)):
            flash(f'Proveedor "{nombre_proveedor}" y todos sus productos asociados han sido eliminados.', 'success')
        else:
            flash('Error al eliminar el proveedor.', 'danger')
    except Exception as e:
        print(f"Error al eliminar proveedor: {e}")
        flash(f'Error al eliminar el proveedor: {str(e)}', 'danger')
    
    return redirect(url_for('proveedores'))

@app.route('/agregar_carrito_ajax', methods=['POST'])
@login_required
def agregar_carrito_ajax():
    try:
        data = request.get_json()
        
        # Si viene con ID, es de productos_manual (base de datos)
        producto_id = data.get('id')
        if producto_id:
            cantidad = int(data.get('cantidad', 1))
            
            producto = db_query("SELECT * FROM productos_manual WHERE id = ?", (producto_id,), fetch=True)
            if not producto:
                return jsonify({'success': False, 'error': 'Producto no encontrado'})
            
            producto = producto[0]
            carrito = session.get('carrito', [])
            
            # Verificar si el producto ya está en el carrito
            encontrado = False
            for item in carrito:
                if item.get('id') == producto_id:
                    item['cantidad'] += cantidad
                    encontrado = True
                    break
            
            if not encontrado:
                carrito.append({
                    'id': producto_id,
                    'nombre': producto['nombre'],
                    'codigo': producto.get('codigo', ''),
                    'precio': float(producto['precio']) if producto['precio'] else 0.0,
                    'cantidad': cantidad,
                    'fecha_compra': datetime.now().strftime('%Y-%m-%d'),
                    'proveedor': data.get('proveedor', ''),
                    'observaciones': data.get('observaciones', ''),
                    'precio_texto': str(producto['precio']) if producto['precio'] else '0',
                    'avisar_bajo_stock': 0,
                    'min_stock_aviso': None
                })
            
            session['carrito'] = carrito
            
            # Renderizar fragment actualizado del carrito
            try:
                html = render_template('carrito_fragment_simple.html', carrito=carrito)
                return jsonify({'success': True, 'msg': f'Producto "{producto["nombre"]}" agregado al carrito', 'html': html})
            except Exception as render_error:
                print(f"Error renderizando carrito_fragment: {render_error}")
            try:
                html = render_template('carrito_fragment_simple.html', carrito=carrito)
                return jsonify({'success': True, 'msg': f'Producto "{producto["nombre"]}" agregado al carrito', 'html': html})
            except Exception as render_error:
                print(f"Error renderizando carrito_fragment: {render_error}")
                return jsonify({'success': True, 'msg': f'Producto "{producto["nombre"]}" agregado al carrito', 'reload': True})
        
        # Si no viene con ID, es de búsqueda Excel
        else:
            nombre = data.get('nombre', '').strip()
            codigo = data.get('codigo', '').strip()
            precio_raw = data.get('precio', '')
            cantidad_raw = data.get('cantidad', '1')
            fecha_compra = data.get('fecha_compra', datetime.now().strftime('%Y-%m-%d'))
            proveedor = data.get('proveedor', '').strip()
            observaciones = data.get('observaciones', '').strip()
            
            # Validar nombre obligatorio
            if not nombre:
                return jsonify({'success': False, 'error': 'El nombre del producto es obligatorio'})
            
            # Procesar cantidad
            try:
                cantidad = int(cantidad_raw) if cantidad_raw else 1
                if cantidad <= 0:
                    cantidad = 1
            except (ValueError, TypeError):
                cantidad = 1
            
            # Procesar precio (puede ser vacío)
            precio = 0.0
            precio_texto = '0'
            if precio_raw:
                if isinstance(precio_raw, (int, float)):
                    precio = float(precio_raw)
                    precio_texto = str(precio_raw)
                else:
                    precio_str = str(precio_raw).strip()
                    if precio_str:
                        precio, precio_error = parse_price(precio_str)
                        precio_texto = precio_str
            
            carrito = session.get('carrito', [])
            carrito.append({
                'id': f'excel_{len(carrito)}_{datetime.now().timestamp()}',
                'nombre': nombre,
                'codigo': codigo,
                'precio': precio,
                'cantidad': cantidad,
                'fecha_compra': fecha_compra,
                'proveedor': proveedor,
                'observaciones': observaciones,
                'precio_texto': precio_texto,
                'avisar_bajo_stock': 0,
                'min_stock_aviso': None
            })
            
            session['carrito'] = carrito
            
            # Renderizar fragment actualizado del carrito
            try:
                html = render_template('carrito_fragment_simple.html', carrito=carrito)
                return jsonify({'success': True, 'msg': f'Producto "{nombre}" agregado al carrito', 'html': html})
            except Exception as render_error:
                print(f"Error renderizando carrito_fragment: {render_error}")
                return jsonify({'success': True, 'msg': f'Producto "{nombre}" agregado al carrito', 'reload': True})
        
    except Exception as e:
        return jsonify({'success': False, 'error': f'Error al procesar: {str(e)}'})

@app.route('/agregar_carrito_manual_ajax', methods=['POST'])
@login_required
def agregar_carrito_manual_ajax():
    try:
        data = request.get_json()
        nombre = data.get('nombre', '').strip()
        codigo = data.get('codigo', '').strip()
        precio_raw = data.get('precio', '')
        cantidad_raw = data.get('cantidad', '1')
        fecha_compra = data.get('fecha_compra', datetime.now().strftime('%Y-%m-%d'))
        proveedor = data.get('proveedor', '').strip()
        observaciones = data.get('observaciones', '').strip()
        
        # Validar nombre obligatorio
        if not nombre:
            return jsonify({'success': False, 'error': 'El nombre del producto es obligatorio'})
        
        # Procesar cantidad
        try:
            cantidad = int(cantidad_raw) if cantidad_raw else 1
            if cantidad <= 0:
                cantidad = 1
        except (ValueError, TypeError):
            cantidad = 1
        
        # Procesar precio (puede ser vacío)
        precio = 0.0
        precio_texto = ''
        if precio_raw:
            if isinstance(precio_raw, (int, float)):
                precio = float(precio_raw)
                precio_texto = str(precio_raw)
            else:
                precio_str = str(precio_raw).strip()
                if precio_str:
                    precio, precio_error = parse_price(precio_str)
                    precio_texto = precio_str
                else:
                    precio = 0.0
                    precio_texto = '0'
        else:
            precio = 0.0
            precio_texto = '0'
        
        carrito = session.get('carrito', [])
        carrito.append({
            'id': f'manual_{len(carrito)}_{datetime.now().timestamp()}',
            'nombre': nombre,
            'codigo': codigo,
            'precio': precio,
            'cantidad': cantidad,
            'fecha_compra': fecha_compra,
            'proveedor': proveedor,
            'observaciones': observaciones,
            'precio_texto': precio_texto,
            'avisar_bajo_stock': 0,
            'min_stock_aviso': None
        })
        
        session['carrito'] = carrito
        return jsonify({'success': True, 'msg': f'Producto "{nombre}" agregado al carrito'})
        
    except Exception as e:
        return jsonify({'success': False, 'error': f'Error al procesar: {str(e)}'})

@app.route('/eliminar_producto_stock', methods=['POST'])
@login_required
def eliminar_producto_stock():
    producto_id = request.form.get('id')
    if producto_id:
        result = db_query("DELETE FROM stock WHERE id = ?", (producto_id,))
        if result:
            flash('Producto eliminado del stock exitosamente.', 'success')
        else:
            flash('Error al eliminar el producto del stock.', 'danger')
    return redirect(url_for('stock'))

@app.route('/eliminar_seleccionados', methods=['POST'])
@login_required
def eliminar_seleccionados():
    try:
        data = request.get_json()
        ids = data.get('ids', [])
        
        if not ids:
            return jsonify({'success': False, 'error': 'No hay elementos seleccionados'})
        
        placeholders = ','.join(['?'] * len(ids))
        result = db_query(f"DELETE FROM stock WHERE id IN ({placeholders})", ids)
        
        if result:
            return jsonify({'success': True, 'message': f'{len(ids)} productos eliminados'})
        else:
            return jsonify({'success': False, 'error': 'Error al eliminar los productos'})
            
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/eliminar_todo_historial', methods=['POST'])
@login_required
def eliminar_todo_historial():
    try:
        result = db_query("DELETE FROM stock")
        if result:
            flash('Todo el historial ha sido eliminado.', 'success')
        else:
            flash('Error al eliminar el historial.', 'danger')
    except Exception as e:
        flash(f'Error: {str(e)}', 'danger')
    
    return redirect(url_for('historial'))

@app.route('/eliminar_manual', methods=['GET', 'POST'])
@login_required
def eliminar_manual():
    """Ruta para gestionar productos manuales"""
    search_term = request.args.get('search_term', '')
    productos = []
    
    if request.method == 'POST':
        search_term = request.form.get('search_term', '')
        if search_term:
            like_pattern = f'%{search_term}%'
            productos = db_query(
                "SELECT pm.nombre, pm.codigo, pm.precio, p.nombre as Proveedor FROM productos_manual pm LEFT JOIN proveedores_manual p ON pm.proveedor_id = p.id WHERE pm.nombre LIKE ? OR pm.codigo LIKE ? OR p.nombre LIKE ?",
                (like_pattern, like_pattern, like_pattern),
                fetch=True
            )
    
    return render_template('eliminar_manual.html', 
                         search_term=search_term, 
                         productos=productos)

@app.route('/obtener_productos_proveedor/<int:proveedor_id>')
@login_required
def obtener_productos_proveedor(proveedor_id):
    """Obtener productos de un proveedor específico (AJAX)"""
    try:
        productos = db_query(
            "SELECT id, nombre, codigo, precio FROM productos_manual WHERE proveedor_id = ? ORDER BY nombre",
            (proveedor_id,),
            fetch=True
        )
        return jsonify({'success': True, 'productos': productos})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/agregar_producto_manual', methods=['POST'])
@login_required
def agregar_producto_manual():
    """Agregar producto manual al Excel de productos manuales"""
    try:
        # Obtener datos del formulario
        nombre = request.form.get('nombre', '').strip()
        codigo = request.form.get('codigo', '').strip()
        precio_str = request.form.get('precio', '').strip()
        observaciones = request.form.get('observaciones', '').strip()
        
        # Determinar proveedor y dueño
        proveedor_id = request.form.get('proveedor_id')
        nuevo_proveedor = request.form.get('nuevo_proveedor', '').strip()
        dueno_nuevo_proveedor = request.form.get('dueno_nuevo_proveedor', '').strip()
        
        if not nombre:
            flash('El nombre del producto es obligatorio.', 'danger')
            return redirect(url_for('agregar_producto'))
        
        # Validar precio
        precio, error_precio = parse_price(precio_str)
        if error_precio and precio_str:  # Solo mostrar error si se ingresó algo
            flash(f'Error en el precio: {error_precio}', 'danger')
            return redirect(url_for('agregar_producto'))
        
        # Determinar proveedor y dueño
        proveedor_nombre = ''
        dueno = ''
        
        if proveedor_id:
            # Proveedor existente
            proveedor_data = db_query("SELECT nombre FROM proveedores_manual WHERE id = ?", (proveedor_id,), fetch=True)
            if proveedor_data:
                proveedor_nombre = proveedor_data[0]['nombre']
                # Determinar dueño basado en el proveedor
                proveedor_lower = proveedor_nombre.lower()
                if any(p in proveedor_lower for p in DUENOS_CONFIG['ricky']['proveedores_excel']):
                    dueno = 'ricky'
                else:
                    dueno = 'ferreteria_general'
        elif nuevo_proveedor and dueno_nuevo_proveedor:
            # Nuevo proveedor
            proveedor_nombre = nuevo_proveedor
            dueno = dueno_nuevo_proveedor
            
            # Agregar el nuevo proveedor a la base de datos
            db_query("INSERT OR IGNORE INTO proveedores_manual (nombre) VALUES (?)", (proveedor_nombre,))
        else:
            flash('Debe seleccionar un proveedor existente o crear uno nuevo con dueño.', 'danger')
            return redirect(url_for('agregar_producto'))
        
        # Agregar al Excel de productos manuales
        result = agregar_producto_excel_manual(codigo, proveedor_nombre, nombre, precio, observaciones, dueno)
        
        if result:
            flash(f'Producto "{nombre}" agregado al catálogo manual de {DUENOS_CONFIG[dueno]["nombre"]}.', 'success')
        else:
            flash('Error al agregar el producto al catálogo manual.', 'danger')
            
    except Exception as e:
        flash(f'Error al procesar el producto: {str(e)}', 'danger')
    
    return redirect(url_for('agregar_producto'))

@app.route('/eliminar_producto_manual', methods=['POST'])
@login_required
def eliminar_producto_manual():
    """Eliminar producto manual"""
    producto_id = request.form.get('id')
    if producto_id:
        result = db_query("DELETE FROM productos_manual WHERE id = ?", (producto_id,))
        if result:
            flash('Producto eliminado exitosamente.', 'success')
        else:
            flash('Error al eliminar el producto.', 'danger')
    return redirect(url_for('productos_manual'))

# Rutas adicionales que los templates esperan
@app.route('/stock.html')
@login_required
def stock_html():
    """Crear template stock.html si no existe"""
    return render_template('stock.html')

@app.route('/productos_manual.html')
@login_required
def productos_manual_html():
    """Crear template productos_manual.html si no existe"""
    return render_template('productos_manual.html')

@app.route('/reportes.html')
@login_required
def reportes_html():
    """Crear template reportes.html si no existe"""
    return render_template('reportes.html')

@app.route('/carrito.html') 
@login_required
def carrito_html():
    """Crear template carrito.html si no existe"""
    return render_template('carrito.html')

# Rutas adicionales que podrían faltar
@app.route('/carrito_accion', methods=['POST'])
@login_required
def carrito_accion():
    """Manejar acciones del carrito (sumar, restar, eliminar, actualizar)"""
    try:
        data = request.get_json()
        accion = data.get('accion')
        idx = int(data.get('idx', 0))
        cantidad = data.get('cantidad')
        
        carrito = session.get('carrito', [])
        
        if idx >= len(carrito):
            return jsonify({'success': False, 'error': 'Índice inválido'})
        
        if accion == 'eliminar':
            carrito.pop(idx)
        elif accion == 'sumar':
            carrito[idx]['cantidad'] += 1
        elif accion == 'restar':
            if carrito[idx]['cantidad'] > 1:
                carrito[idx]['cantidad'] -= 1
        elif accion == 'actualizar' and cantidad:
            carrito[idx]['cantidad'] = max(1, int(cantidad))
        
        session['carrito'] = carrito
        
        # Renderizar fragment del carrito actualizado
        try:
            html = render_template('carrito_fragment_simple.html', carrito=carrito)
            return jsonify({'success': True, 'html': html})
        except Exception as render_error:
            print(f"Error renderizando carrito_fragment: {render_error}")
            return jsonify({'success': True, 'reload': True})
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/actualizar_stock/<int:id>', methods=['POST'])
@login_required
def actualizar_stock(id):
    """Actualizar stock cuando se vende un producto"""
    try:
        cantidad_vendida = int(request.form.get('cantidad_vendida', 1))
        
        # Obtener producto actual
        producto = db_query("SELECT * FROM stock WHERE id = ?", (id,), fetch=True)
        if not producto:
            flash('Producto no encontrado.', 'danger')
            return redirect(url_for('historial'))
        
        producto = producto[0]
        nueva_cantidad = max(0, producto['cantidad'] - cantidad_vendida)
        
        # Actualizar cantidad
        result = db_query("UPDATE stock SET cantidad = ? WHERE id = ?", (nueva_cantidad, id))
        
        if result:
            flash(f'Vendidas {cantidad_vendida} unidades. Nueva cantidad: {nueva_cantidad}', 'success')
        else:
            flash('Error al actualizar el stock.', 'danger')
            
    except Exception as e:
        flash(f'Error: {str(e)}', 'danger')
    
    return redirect(url_for('historial'))

@app.route('/agregar_proveedor_manual_ajax', methods=['POST'])
@login_required
def agregar_proveedor_manual_ajax():
    try:
        data = request.get_json()
        nombre = data.get('nombre', '').strip()
        
        if not nombre:
            return jsonify({'success': False, 'msg': 'El nombre del proveedor no puede estar vacío'})
        
        # Agregar el proveedor
        if db_query("INSERT INTO proveedores_manual (nombre) VALUES (?)", (nombre,)):
            # Obtener lista actualizada de proveedores
            proveedores = db_query("SELECT id, nombre FROM proveedores_manual ORDER BY nombre", fetch=True) or []
            return jsonify({
                'success': True, 
                'msg': f'Proveedor "{nombre}" agregado exitosamente',
                'proveedores': [{'id': p['id'], 'nombre': p['nombre']} for p in proveedores]
            })
        else:
            return jsonify({'success': False, 'msg': 'Error al agregar el proveedor (posiblemente ya existe)'})
    except Exception as e:
        print(f"Error al agregar proveedor AJAX: {e}")
        return jsonify({'success': False, 'msg': f'Error: {str(e)}'})

# Rutas AJAX que faltan para eliminar_manual.html
@app.route('/eliminar_manual_ajax', methods=['POST'])
@login_required
def eliminar_manual_ajax():
    """Eliminar producto manual via AJAX"""
    try:
        data = request.get_json()
        codigo_a_eliminar = data.get('codigo_a_eliminar', '').strip()
        search_term = data.get('search_term', '').strip()
        
        if not codigo_a_eliminar:
            return jsonify({'success': False, 'msg': 'Código del producto es obligatorio.'})
        
        # Para este ejemplo, como no hay Excel, simulo eliminar de productos_manual
        result = db_query("DELETE FROM productos_manual WHERE codigo = ?", (codigo_a_eliminar,))
        
        if result:
            # Buscar productos restantes
            productos = []
            if search_term:
                like_pattern = f'%{search_term}%'
                productos = db_query(
                    "SELECT pm.nombre, pm.codigo, pm.precio, p.nombre as proveedor FROM productos_manual pm LEFT JOIN proveedores_manual p ON pm.proveedor_id = p.id WHERE pm.nombre LIKE ? OR pm.codigo LIKE ? OR p.nombre LIKE ?",
                    (like_pattern, like_pattern, like_pattern),
                    fetch=True
                )
            
            html = render_template('eliminar_manual_resultados.html', productos=productos, search_term=search_term)
            return jsonify({
                'success': True, 
                'msg': f'Producto con código "{codigo_a_eliminar}" eliminado exitosamente.',
                'html': html
            })
        else:
            return jsonify({'success': False, 'msg': 'Error al eliminar el producto o no existe.'})
            
    except Exception as e:
        return jsonify({'success': False, 'msg': f'Error: {str(e)}'})

@app.route('/eliminar_manual_todo_ajax', methods=['POST'])
@login_required
def eliminar_manual_todo_ajax():
    """Eliminar todos los productos manuales via AJAX"""
    try:
        data = request.get_json()
        search_term = data.get('search_term', '').strip()
        
        # Eliminar todos los productos manuales
        result = db_query("DELETE FROM productos_manual")
        
        if result:
            html = render_template('eliminar_manual_resultados.html', productos=[], search_term=search_term)
            return jsonify({
                'success': True, 
                'msg': 'Todos los productos manuales han sido eliminados.',
                'html': html
            })
        else:
            return jsonify({'success': False, 'msg': 'Error al eliminar los productos.'})
            
    except Exception as e:
        return jsonify({'success': False, 'msg': f'Error: {str(e)}'})

@app.route('/eliminar_manual_seleccionados_ajax', methods=['POST'])
@login_required
def eliminar_manual_seleccionados_ajax():
    """Eliminar productos manuales seleccionados via AJAX"""
    try:
        data = request.get_json()
        codigos = data.get('codigos', [])
        search_term = data.get('search_term', '').strip()
        
        if not codigos:
            return jsonify({'success': False, 'msg': 'No hay productos seleccionados.'})
        
        # Eliminar productos seleccionados
        placeholders = ','.join(['?'] * len(codigos))
        result = db_query(f"DELETE FROM productos_manual WHERE codigo IN ({placeholders})", codigos)
        
        if result:
            # Buscar productos restantes
            productos = []
            if search_term:
                like_pattern = f'%{search_term}%'
                productos = db_query(
                    "SELECT pm.nombre, pm.codigo, pm.precio, p.nombre as proveedor FROM productos_manual pm LEFT JOIN proveedores_manual p ON pm.proveedor_id = p.id WHERE pm.nombre LIKE ? OR pm.codigo LIKE ? OR p.nombre LIKE ?",
                    (like_pattern, like_pattern, like_pattern),
                    fetch=True
                )
            
            html = render_template('eliminar_manual_resultados.html', productos=productos, search_term=search_term)
            return jsonify({
                'success': True, 
                'msg': f'{len(codigos)} productos eliminados exitosamente.',
                'html': html
            })
        else:
            return jsonify({'success': False, 'msg': 'Error al eliminar los productos seleccionados.'})
            
    except Exception as e:
        return jsonify({'success': False, 'msg': f'Error: {str(e)}'})

@app.route('/gestionar_manual_buscar_ajax', methods=['POST'])
@login_required
def gestionar_manual_buscar_ajax():
    """Buscar producto manual para editar via AJAX"""
    try:
        data = request.get_json()
        valor = data.get('valor', '').strip()
        
        if not valor:
            return jsonify({'html': '<div class="alert alert-warning">Debe ingresar un valor para buscar.</div>'})
        
        # Buscar producto por código, nombre, proveedor o precio
        like_pattern = f'%{valor}%'
        productos = db_query(
            "SELECT pm.*, p.nombre as proveedor FROM productos_manual pm LEFT JOIN proveedores_manual p ON pm.proveedor_id = p.id WHERE pm.codigo LIKE ? OR pm.nombre LIKE ? OR p.nombre LIKE ? OR CAST(pm.precio AS TEXT) LIKE ? LIMIT 1",
            (like_pattern, like_pattern, like_pattern, like_pattern),
            fetch=True
        )
        
        if productos:
            producto = productos[0]
            # Crear un objeto compatible con el template
            producto_template = {
                'Codigo': producto['codigo'],
                'Nombre': producto['nombre'],
                'Precio': producto['precio'],
                'Proveedor': producto['proveedor']
            }
            html = render_template('editar_manual_formulario.html', producto=producto_template)
        else:
            html = '<div class="alert alert-info">No se encontró ningún producto con ese criterio.</div>'
        
        return jsonify({'html': html})
            
    except Exception as e:
        return jsonify({'html': f'<div class="alert alert-danger">Error: {str(e)}</div>'})

@app.route('/gestionar_manual_actualizar_ajax', methods=['POST'])
@login_required
def gestionar_manual_actualizar_ajax():
    """Actualizar producto manual via AJAX"""
    try:
        data = request.get_json()
        codigo_original = data.get('codigo_original', '').strip()
        nombre = data.get('nombre', '').strip()
        proveedor = data.get('proveedor', '').strip()
        codigo = data.get('codigo', '').strip()
        precio_str = data.get('precio', '').strip()
        
        if not codigo_original or not nombre:
            return jsonify({'success': False, 'msg': 'Datos incompletos.'})
        
        precio, error_precio = parse_price(precio_str)
        if error_precio:
            return jsonify({'success': False, 'msg': f'Error en el precio: {error_precio}'})
        
        # Actualizar producto
        result = db_query(
            "UPDATE productos_manual SET nombre = ?, codigo = ?, precio = ? WHERE codigo = ?",
            (nombre, codigo, precio, codigo_original)
        )
        
        if result:
            return jsonify({
                'success': True, 
                'msg': f'Producto "{nombre}" actualizado exitosamente.',
                'html': '<div class="alert alert-success">Producto actualizado correctamente.</div>'
            })
        else:
            return jsonify({'success': False, 'msg': 'Error al actualizar el producto.'})
            
    except Exception as e:
        return jsonify({'success': False, 'msg': f'Error: {str(e)}'})

def agregar_producto_manual_excel():
    """Agregar producto manual al Excel (no directamente al stock)"""
    try:
        codigo = request.form.get('codigo', '').strip()
        proveedor_id = request.form.get('proveedor_id', '').strip()
        nombre = request.form.get('nombre', '').strip()
        precio_str = request.form.get('precio', '').strip()
        observaciones = request.form.get('observaciones', '').strip()
        dueno = request.form.get('dueno_nuevo_proveedor', '').strip()
        
        # Validaciones
        if not nombre:
            flash('El nombre del producto es obligatorio.', 'danger')
            return redirect(url_for('agregar_producto'))
        
        if not dueno:
            flash('Debe seleccionar un dueño para el producto.', 'danger')
            return redirect(url_for('agregar_producto'))
        
        # Obtener nombre del proveedor si se proporcionó proveedor_id
        proveedor_nombre = ''
        if proveedor_id:
            proveedor_data = db_query("SELECT nombre FROM proveedores_manual WHERE id = ?", (proveedor_id,), fetch=True)
            if proveedor_data:
                proveedor_nombre = proveedor_data[0]['nombre']
        
        # Procesar precio
        precio = 0.0
        if precio_str:
            precio, error_precio = parse_price(precio_str)
            if error_precio:
                flash(f'Error en el precio: {precio_str}', 'warning')
                precio = 0.0
        
        # Agregar al Excel
        if agregar_producto_excel_manual(codigo, proveedor_nombre, nombre, precio, observaciones, dueno):
            dueno_nombre = DUENOS_CONFIG.get(dueno, {}).get('nombre', dueno)
            flash(f'Producto "{nombre}" agregado a la lista manual de {dueno_nombre}. Puede buscarlo en "Buscar en Excel".', 'success')
        else:
            flash('Error al agregar el producto a la lista manual.', 'danger')
            
    except Exception as e:
        flash(f'Error al procesar el producto: {str(e)}', 'danger')
    
    return redirect(url_for('agregar_producto'))

# --- Funciones de Búsqueda en Excel ---
def buscar_en_excel(termino_busqueda, proveedor_filtro=None, filtro_adicional=None):
    """Buscar productos en archivos Excel de proveedores y productos manuales"""
    resultados = []
    
    # 1. Buscar en productos manuales
    if proveedor_filtro and proveedor_filtro.startswith('manual_'):
        # Filtro específico de proveedor manual
        proveedor_id = proveedor_filtro.replace('manual_', '')
        try:
            proveedor_id = int(proveedor_id)
            # Buscar solo en este proveedor específico - permitir todos los proveedores para ambos dueños
            resultados_manuales = buscar_en_excel_manual_por_proveedor(termino_busqueda, proveedor_id)
            resultados.extend(resultados_manuales)
        except (ValueError, TypeError):
            pass
    elif not proveedor_filtro or proveedor_filtro not in PROVEEDOR_CONFIG:
        # Si no hay filtro específico de Excel, incluir todos los manuales
        resultados_manuales = buscar_en_excel_manual(termino_busqueda)
        resultados.extend(resultados_manuales)
    
    # 2. Buscar en archivos Excel de proveedores (solo de Ricky)
    if proveedor_filtro:
        if proveedor_filtro in PROVEEDOR_CONFIG:
            archivos_a_buscar = [proveedor_filtro]
        else:
            archivos_a_buscar = []
    else:
        # Buscar en todos los archivos Excel disponibles
        archivos_a_buscar = []
        for key in PROVEEDOR_CONFIG.keys():
            # Buscar archivos que empiecen con el nombre del proveedor (case insensitive)
            archivos = [f for f in os.listdir(EXCEL_FOLDER) if f.lower().startswith(key.lower()) and f.endswith('.xlsx') and f != 'productos_manual.xlsx']
            if archivos:
                archivos_a_buscar.append(key)
    
    # Procesar cada archivo Excel
    for archivo in archivos_a_buscar:
        if archivo in PROVEEDOR_CONFIG:
            config = PROVEEDOR_CONFIG[archivo]
            resultados_archivo = procesar_archivo_excel(archivo, config, termino_busqueda, filtro_adicional, archivo)
            resultados.extend(resultados_archivo)
    
    return resultados

def buscar_en_excel_manual_por_proveedor(termino_busqueda, proveedor_id):
    """Buscar productos en el Excel de productos manuales por proveedor específico"""
    resultados = []
    
    try:
        if not os.path.exists(MANUAL_PRODUCTS_FILE):
            return resultados
        
        df = pd.read_excel(MANUAL_PRODUCTS_FILE)
        # Normalizar nombres de columnas por si existen acentos
        df.rename(columns={'Código': 'Codigo', 'Dueño': 'Dueno'}, inplace=True)
        
        if df.empty:
            return resultados
        
        # Obtener nombre del proveedor
        proveedor_info = db_query("SELECT nombre FROM proveedores_manual WHERE id = ?", (proveedor_id,), fetch=True)
        if not proveedor_info:
            return resultados
        
        proveedor_nombre = proveedor_info[0]['nombre']
        
        # Filtrar por proveedor específico - sin importar el dueño
        df = df[df['Proveedor'].str.contains(proveedor_nombre, case=False, na=False)]
        
        # Filtrar por término de búsqueda si existe
        if termino_busqueda:
            mask = (
                df['Nombre'].str.contains(termino_busqueda, case=False, na=False) |
                df['Codigo'].str.contains(termino_busqueda, case=False, na=False) |
                df['Proveedor'].str.contains(termino_busqueda, case=False, na=False)
            )
            df = df[mask]
        
        # Convertir a lista de diccionarios
        for _, row in df.iterrows():
            precio_val, precio_error = parse_price(row.get('Precio', ''))
            resultado = {
                'codigo': row.get('Codigo', ''),
                'nombre': row.get('Nombre', ''),
                'precio': precio_val,
                'precio_texto': str(row.get('Precio', '')) if precio_error else None,
                'proveedor': row.get('Proveedor', ''),
                'observaciones': row.get('Observaciones', ''),
                'dueno': row.get('Dueno', ''),
                'es_manual': True
            }
            resultados.append(resultado)
    
    except Exception as e:
        print(f"Error al buscar en Excel manual por proveedor: {e}")
    
    return resultados

def procesar_archivo_excel(archivo, config, termino_busqueda, filtro_adicional, proveedor_key):
    """Procesar un archivo Excel específico"""
    resultados = []
    
    try:
        # Buscar el archivo que coincida
        archivos = [f for f in os.listdir(EXCEL_FOLDER) if f.lower().startswith(archivo.lower()) and f.endswith('.xlsx') and f != 'productos_manual.xlsx']
        if not archivos:
            return resultados
        
        archivo_path = os.path.join(EXCEL_FOLDER, archivos[0])
        if not os.path.exists(archivo_path):
            return resultados
        
        df = pd.read_excel(archivo_path, header=config['fila_encabezado'])
        
        if df.empty:
            return resultados
        
        # Encontrar las columnas correctas
        col_codigo = encontrar_columna(df.columns, config['codigo'])
        col_producto = encontrar_columna(df.columns, config['producto'])
        col_precio = encontrar_columna(df.columns, config['precio'])
        
        if not col_producto:
            return resultados
        
        # Convertir columnas a string para evitar errores con .str accessor
        df[col_producto] = df[col_producto].astype(str)
        if col_codigo:
            df[col_codigo] = df[col_codigo].astype(str)
        
        # Filtrar por término de búsqueda
        if termino_busqueda:
            mask = df[col_producto].str.contains(termino_busqueda, case=False, na=False)
            if col_codigo:
                mask |= df[col_codigo].str.contains(termino_busqueda, case=False, na=False)
            df = df[mask]
        
        # Aplicar filtro adicional si existe
        if filtro_adicional:
            mask_adicional = df[col_producto].str.contains(filtro_adicional, case=False, na=False)
            if col_codigo:
                mask_adicional |= df[col_codigo].str.contains(filtro_adicional, case=False, na=False)
            df = df[mask_adicional]
        
        # Convertir a lista de resultados
        for _, row in df.iterrows():
            codigo = str(row[col_codigo]).strip() if col_codigo and pd.notna(row[col_codigo]) else ''
            nombre = str(row[col_producto]).strip() if pd.notna(row[col_producto]) else ''
            precio_raw = row[col_precio] if col_precio and pd.notna(row[col_precio]) else ''
            
            if not nombre:
                continue
            
            precio_val, precio_error = parse_price(precio_raw)
            
            resultado = {
                'codigo': codigo,
                'nombre': nombre,
                'precio': precio_val,
                'precio_texto': str(precio_raw) if precio_error else None,
                'proveedor': proveedor_key.title(),
                'observaciones': '',
                'dueno': 'ricky',
                'es_manual': False
            }
            resultados.append(resultado)
    
    except Exception as e:
        print(f"Error al procesar archivo {archivo}: {e}")
    
    return resultados

def encontrar_columna(columnas, aliases):
    """Encontrar la columna correcta basada en los aliases"""
    for alias in aliases:
        for col in columnas:
            if str(col).lower().strip() == alias.lower():
                return col
    return None

if __name__ == '__main__':
    print("🚀 Iniciando Gestor de Stock...")
    print("📁 Directorio base:", BASE_DIR)
    print("📊 Carpeta Excel:", EXCEL_FOLDER)
    print("🗄️ Base de datos:", DATABASE_FILE)
    
    try:
        print("🔧 Inicializando base de datos...")
        init_db()
        print("✅ Base de datos inicializada correctamente")
        
        print("🌐 Iniciando servidor Flask...")
        print("📋 Accede a: http://localhost:5000")
        print("👤 Usuario: Pauluk")
        print("🔑 Contraseña: Jap2005")
        print("=" * 50)
        
        app.run(debug=True, host='0.0.0.0', port=5000)
        
    except KeyboardInterrupt:
        print("\n🛑 Servidor detenido por el usuario")
    except Exception as e:
        print(f"❌ Error al iniciar el servidor: {e}")
        import traceback
        traceback.print_exc()