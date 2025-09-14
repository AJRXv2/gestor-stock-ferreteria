import os
import pandas as pd
from openpyxl import load_workbook, Workbook

# Constantes (simuladas para el script)
EXCEL_FOLDER = "listas_excel"
MANUAL_PRODUCTS_FILE = os.path.join(EXCEL_FOLDER, "productos_manual.xlsx")

# Función para agregar un producto de prueba
def agregar_producto_excel_manual(codigo, proveedor, nombre, precio, observaciones, dueno):
    """Agregar producto al Excel de productos manuales"""
    try:
        if not os.path.exists(EXCEL_FOLDER):
            os.makedirs(EXCEL_FOLDER)
        
        expected_headers = ['Codigo', 'Proveedor', 'Nombre', 'Precio', 'Observaciones', 'Dueno']
        
        if not os.path.exists(MANUAL_PRODUCTS_FILE):
            wb = Workbook()
            ws = wb.active
            ws.title = "Productos Manuales"
            # Encabezados estandarizados sin acentos y en orden requerido
            for col, header in enumerate(expected_headers, 1):
                ws.cell(row=1, column=col, value=header)
            wb.save(MANUAL_PRODUCTS_FILE)
        
        wb = load_workbook(MANUAL_PRODUCTS_FILE)
        ws = wb.active
        
        # Verificar y corregir encabezados si es necesario
        try:
            current = [ws.cell(row=1, column=i).value for i in range(1, 7)]
            normalize = lambda s: str(s or '').strip().replace('ó', 'o').replace('Ó', 'O').replace('ñ', 'n').replace('Ñ', 'N').title()
            normalized = [normalize(v) for v in current]
            old_order = ['Codigo', 'Nombre', 'Proveedor', 'Precio', 'Observaciones', 'Dueno']
            if normalized == old_order:
                # Migrar filas intercambiando columnas B y C para adecuar al nuevo orden
                for r in range(2, ws.max_row + 1):
                    val_b = ws.cell(row=r, column=2).value  # Nombre
                    val_c = ws.cell(row=r, column=3).value  # Proveedor
                    ws.cell(row=r, column=2, value=val_c)
                    ws.cell(row=r, column=3, value=val_b)
                # Actualizar encabezados al estándar
                for idx, header in enumerate(expected_headers, start=1):
                    ws.cell(row=1, column=idx, value=header)
            elif normalized != expected_headers:
                for idx, header in enumerate(expected_headers, start=1):
                    ws.cell(row=1, column=idx, value=header)
        except Exception as e:
            print(f"Error al verificar encabezados: {e}")
        
        # Agregar fila respetando el orden de columnas esperado
        ws.append([codigo, proveedor, nombre, float(precio) if precio is not None else 0.0, observaciones, dueno])
        wb.save(MANUAL_PRODUCTS_FILE)
        
        print(f"Producto agregado exitosamente: {codigo} - {nombre} (Proveedor: {proveedor}, Dueño: {dueno})")
        return True
    except Exception as e:
        print(f"Error al agregar producto manual: {e}")
        return False

# Agregar el proveedor JELUZ a la base de datos
import sqlite3

def agregar_proveedor(nombre, dueno):
    """Agrega un proveedor a la base de datos y lo asocia con un dueño"""
    try:
        conn = sqlite3.connect('gestor_stock.db')
        cursor = conn.cursor()
        
        # Verificar si ya existe
        cursor.execute("SELECT id FROM proveedores_manual WHERE nombre = ?", (nombre,))
        existing = cursor.fetchone()
        
        if existing:
            proveedor_id = existing[0]
            print(f"Proveedor '{nombre}' ya existe con ID {proveedor_id}")
        else:
            # Crear proveedor
            cursor.execute("INSERT INTO proveedores_manual (nombre, dueno) VALUES (?, ?)", (nombre, dueno))
            proveedor_id = cursor.lastrowid
            print(f"Proveedor '{nombre}' creado con ID {proveedor_id}")
            
            # Asociar a dueño
            try:
                cursor.execute("INSERT INTO proveedores_duenos (proveedor_id, dueno) VALUES (?, ?)", (proveedor_id, dueno))
                print(f"Proveedor asociado al dueño '{dueno}'")
            except Exception as e:
                print(f"Error al asociar proveedor a dueño: {e}")
        
        conn.commit()
        conn.close()
        return proveedor_id
    except Exception as e:
        print(f"Error al agregar proveedor: {e}")
        return None

# Ejecutar las operaciones
print("========== AGREGANDO PROVEEDOR JELUZ A LA BASE DE DATOS ==========")
proveedor_id = agregar_proveedor("JELUZ", "ferreteria_general")

print("\n========== AGREGANDO PRODUCTO TERM32A AL EXCEL ==========")
agregar_producto_excel_manual(
    codigo="TERM32A",
    proveedor="JELUZ", 
    nombre="TERMICA 32A JELUZ", 
    precio=5000, 
    observaciones="Producto agregado manualmente", 
    dueno="ferreteria_general"
)

print("\n========== VERIFICANDO EL EXCEL DESPUÉS DE AGREGAR EL PRODUCTO ==========")
try:
    if os.path.exists(MANUAL_PRODUCTS_FILE):
        df = pd.read_excel(MANUAL_PRODUCTS_FILE)
        print(f"Total de productos en Excel: {len(df)}")
        print("\nListado de productos:")
        for i, (_, row) in enumerate(df.iterrows(), 1):
            print(f"{i}. Código: {row.get('Codigo', '')}, Nombre: {row.get('Nombre', '')}, Proveedor: {row.get('Proveedor', '')}")
    else:
        print(f"El archivo {MANUAL_PRODUCTS_FILE} no existe")
except Exception as e:
    print(f"Error al verificar el Excel: {e}")

print("\n========== PRUEBA FINALIZADA ==========")