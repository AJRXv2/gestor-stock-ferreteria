print("🔄 Cargando dependencias...")
try:
    from flask import Flask, render_template, request, redirect, url_for, flash, session, jsonify
    from flask import send_file
    from datetime import datetime, timedelta
    from openpyxl import Workbook, load_workbook
    import pandas as pd
    import os, unicodedata, sqlite3, re
    try:
        import psycopg2
        HAS_POSTGRES = True
    except Exception:
        HAS_POSTGRES = False
except ImportError as e:
    print(f"❌ Error de importación: {e}")
    print("💡 Instala las dependencias con: pip install flask pandas openpyxl")
    raise
except Exception as e:
    print(f"❌ Error inesperado en importaciones: {e}")
    raise
try:
    from flask_wtf import CSRFProtect
    from flask_wtf.csrf import generate_csrf
    _HAS_FLASK_WTF = True
    print("✅ Flask-WTF importado correctamente")
except ImportError:
    print("⚠️ Flask-WTF no disponible, usando fallback simple")
    # Flask-WTF not available, provide a minimal fallback
    CSRFProtect = None
    _HAS_FLASK_WTF = False
    def generate_csrf():
        # lazy import to avoid top-level secret import when not needed
        import secrets
        if '_csrf_token' not in session:
            session['_csrf_token'] = secrets.token_urlsafe(16)
        return session['_csrf_token']
except Exception as e:
    print(f"⚠️ Error con Flask-WTF: {e}, usando fallback simple")
    CSRFProtect = None
    _HAS_FLASK_WTF = False
    def generate_csrf():
        import secrets
        if '_csrf_token' not in session:
            session['_csrf_token'] = secrets.token_urlsafe(16)
        return session['_csrf_token']
from decimal import Decimal, InvalidOperation
from werkzeug.security import generate_password_hash, check_password_hash
from functools import wraps
print("✅ Werkzeug y functools importados correctamente")
try:
    import pdfplumber
    _HAS_PDF = True
    print("✅ pdfplumber disponible para lectura de facturas PDF")
except Exception as e:
    _HAS_PDF = False
    import sys, importlib.util
    print(f"⚠️ pdfplumber no disponible ({e}); la importación de facturas PDF estará deshabilitada")
    print("🛠️ Debug pdfplumber:")
    print("  - Python exe:", sys.executable)
    print("  - Version:", sys.version)
    print("  - sys.path entries:")
    for p in sys.path:
        print("     *", p)
    spec = importlib.util.find_spec('pdfplumber')
    print("  - importlib.util.find_spec('pdfplumber') ->", bool(spec))

# --- Utilidades de normalización de nombres de proveedor ---
def _normalizar_nombre_proveedor(nombre: str) -> str:
    try:
        if not nombre:
            return ''
        import unicodedata
        # Quitar espacios repetidos y trimming
        base = ' '.join(str(nombre).strip().split())
        # Normalizar unicode y quitar acentos
        nfkd = unicodedata.normalize('NFKD', base)
        sin_acentos = ''.join(ch for ch in nfkd if not unicodedata.combining(ch))
        return sin_acentos.lower()
    except Exception:
        return nombre or ''

# --- Configuración de la Aplicación ---
app = Flask(__name__)
app.secret_key = os.getenv('SECRET_KEY', 'supersecretkey')
app.permanent_session_lifetime = timedelta(hours=3)

if _HAS_FLASK_WTF:
    # Use Flask-WTF CSRFProtect when available
    csrf = CSRFProtect(app)
    # Expose generate_csrf to templates as csrf_token()
    app.jinja_env.globals['csrf_token'] = generate_csrf
else:
    # Fallback: expose our simple generate_csrf and validate on POST
    app.jinja_env.globals['csrf_token'] = generate_csrf

    # Endpoints seguros/legacy que omiten la verificación CSRF (AJAX/legacy forms)
    CSRF_WHITELIST = {
        'eliminar_seleccionados',
        'eliminar_todo_historial',
        'eliminar_producto_stock',
        'eliminar_proveedor_manual',
        'agregar_carrito_ajax',
        'agregar_carrito_manual_ajax',
        'carrito_accion',
        'manual_listar_ajax',
        'manual_eliminar_seleccionados_ajax',
        'manual_eliminar_por_proveedor_ajax',
        'manual_actualizar_ajax',
        'agregar_producto',
        'carrito_cargar'
    }

    # Endpoint de diagnóstico rápido para revisar estado de pdfplumber y entorno Python
    @app.route('/debug_pdf')
    def debug_pdf():
        import sys, importlib.util, traceback
        info = {}
        info['HAS_FLAG'] = _HAS_PDF
        info['executable'] = sys.executable
        info['version'] = sys.version
        info['cwd'] = os.getcwd()
        info['sys_path'] = sys.path
        spec = importlib.util.find_spec('pdfplumber')
        info['find_spec'] = bool(spec)
        if spec:
            info['spec_origin'] = spec.origin
        # Intentar import en vivo
        try:
            import pdfplumber as pp
            info['live_import'] = True
            info['pdfplumber_file'] = getattr(pp, '__file__', 'desconocido')
        except Exception as e:
            info['live_import'] = False
            info['live_import_error'] = str(e)
            info['trace_tail'] = traceback.format_exc().splitlines()[-5:]
        return jsonify(info)
    # Rutas exactas que también pueden omitir CSRF (por si request.endpoint no coincide)
    CSRF_WHITELIST_PATHS = {
        '/eliminar_seleccionados',
        '/eliminar_todo_historial',
        '/eliminar_producto_stock',
        '/eliminar_proveedor_manual',
        '/agregar_carrito_ajax',
        '/agregar_carrito_manual_ajax',
        '/carrito_accion',
        '/manual_listar_ajax',
        '/manual_eliminar_seleccionados_ajax',
        '/manual_eliminar_por_proveedor_ajax',
        '/manual_actualizar_ajax',
        '/agregar_producto',
        '/carrito_cargar'
    }

    @app.before_request
    def csrf_protect_fallback():
        # Allow specific endpoints or paths to bypass CSRF fallback (legacy forms or JS not sending token)
        try:
            ep = request.endpoint
        except Exception:
            ep = None
        try:
            path = request.path
        except Exception:
            path = None
        if ep in CSRF_WHITELIST or (path in CSRF_WHITELIST_PATHS):
            return

        if request.method == 'POST':
            token = session.get('_csrf_token')
            if not token:
                import secrets
                session['_csrf_token'] = secrets.token_urlsafe(16)
                token = session['_csrf_token']

            form_token = request.form.get('csrf_token')
            header_token = request.headers.get('X-CSRFToken') or request.headers.get('X-CSRF-Token')
            json_token = None
            if request.is_json:
                try:
                    data = request.get_json(silent=True) or {}
                    json_token = data.get('csrf_token')
                except Exception:
                    json_token = None

            if form_token == token or header_token == token or json_token == token:
                return

            if request.is_json:
                return jsonify({'success': False, 'error': 'CSRF token inválido o ausente.'}), 400
            flash('CSRF token inválido o ausente.', 'danger')
            if 'user_id' in session:
                return redirect(url_for('index'))
            return redirect(url_for('login'))

    # Dummy csrf object to allow using @csrf.exempt in routes uniformly
    class _DummyCSRF:
        def exempt(self, f):
            return f
    csrf = _DummyCSRF()

# --- Rutas de Archivos y Carpetas ---
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_FOLDER = os.path.join(BASE_DIR, 'listas_excel')
MANUAL_PRODUCTS_FILE = os.path.join(EXCEL_FOLDER, 'productos_manual.xlsx')
# Permite override en desarrollo: set DATABASE_FILE=/ruta/custom.db
DATABASE_FILE = os.path.join(BASE_DIR, os.getenv('DATABASE_FILE', 'gestor_stock.db'))

# --- Configuración de Dueños ---
DUENOS_CONFIG = {
    'ricky': {
        'nombre': 'Ricky',
        'proveedores_excel': ['brementools', 'berger', 'cachan', 'chiesa', 'crossmaster'],
        'puede_excel': True,
        'carpeta_excel': 'ricky'
    },
    'ferreteria_general': {
        'nombre': 'Ferretería General', 
        # Renombrados: ferreteria_prov1->dewalt, ferreteria_prov2->sica, ferreteria_prov3->sorbalok
        'proveedores_excel': ['dewalt', 'sica', 'sorbalok', 'nortedist'],
        'puede_excel': True,
        'carpeta_excel': 'ferreteria_general'
    }
}

# --- Configuración de Proveedores Excel ---
PROVEEDOR_CONFIG = {
    'brementools': {
        'fila_encabezado': 5,
        'codigo': ['codigo', 'Código', 'CODIGO'],
        'producto': ['producto', 'Producto', 'PRODUCTO'],
        'precio': ['precio', 'Precio', 'PRECIO'],
        'dueno': 'ricky',
        'folder': 'ricky'
    },
    'crossmaster': {
        'fila_encabezado': 11,
        'codigo': ['codigo', 'Codigo', 'CODIGO'],
        'producto': ['descripcion', 'Descripcion', 'DESCRIPCION'],
        'precio': ['precio', 'Precio', 'PRECIO'],
        'dueno': 'ricky',
        'folder': 'ricky'
    },
    'berger': {
        'fila_encabezado': 0,
        'codigo': ['cod', 'COD', 'codigo', 'Codigo'],
        'producto': ['detalle', 'DETALLE', 'producto', 'Producto'],
        'precio': ['P.VENTA', 'precio', 'Precio', 'PRECIO'],
        'dueno': 'ricky',
        'folder': 'ricky'
    },
    'chiesa': {
        'fila_encabezado': 1,
        'codigo': ['codigo', 'Codigo', 'CODIGO'],
        'producto': ['descripción', 'Descripción', 'descripcion', 'Descripcion'],
        'precio': ['precio', 'Precio', 'PRECIO', 'Pr.Unit', 'pr.unit', 'PR.UNIT'],
        'dueno': 'ricky',
        'folder': 'ricky'
    },
    'cachan': {
        'fila_encabezado': 0,
        'codigo': ['codigo', 'Codigo', 'CODIGO'],
        'producto': ['nombre', 'Nombre', 'NOMBRE'],
        'precio': ['precio', 'Precio', 'PRECIO'],
        'dueno': 'ricky',
        'folder': 'ricky'
    },
    # Proveedores de Ferretería General (renombrados)
    'dewalt': {
        'fila_encabezado': 0,
        'codigo': ['codigo', 'Codigo', 'CODIGO', 'cod', 'COD'],
        'producto': ['producto', 'Producto', 'PRODUCTO', 'descripcion', 'Descripcion', 'nombre', 'Nombre'],
        'precio': ['precio', 'Precio', 'PRECIO', 'p.venta', 'P.VENTA'],
        'dueno': 'ferreteria_general',
        'folder': 'ferreteria_general'
    },
    'sica': {
        'fila_encabezado': 0,
        'codigo': ['codigo', 'Codigo', 'CODIGO', 'cod', 'COD'],
        'producto': ['producto', 'Producto', 'PRODUCTO', 'descripcion', 'Descripcion', 'nombre', 'Nombre'],
        'precio': ['precio', 'Precio', 'PRECIO', 'p.venta', 'P.VENTA'],
        'dueno': 'ferreteria_general',
        'folder': 'ferreteria_general'
    },
    'sorbalok': {
        'fila_encabezado': 0,
        'codigo': ['codigo', 'Codigo', 'CODIGO', 'cod', 'COD'],
        'producto': ['producto', 'Producto', 'PRODUCTO', 'descripcion', 'Descripcion', 'nombre', 'Nombre'],
        'precio': ['precio', 'Precio', 'PRECIO', 'p.venta', 'P.VENTA'],
        'dueno': 'ferreteria_general',
        'folder': 'ferreteria_general'
    },
    'nortedist': {
        'fila_encabezado': 0,
        'codigo': ['codigo_barra', 'codigo', 'Codigo', 'CODIGO'],
        'producto': ['descripcion', 'Descripcion', 'DESCRIPCION', 'producto', 'Producto'],
        'precio': ['NORTE SIN IVA', 'NORTE', 'precio', 'Precio'],
        'marca': ['marca', 'Marca', 'MARCA'],
        'usar_colores': True,
        'color_precio': 'FFFF00',  # Amarillo en formato RGB
        'columnas_precio': ['NORTE SIN IVA', 'NORTE'],
        'dueno': 'ferreteria_general',
        'folder': 'ferreteria_general'
    },
    'jeluz': {
        'fila_encabezado': 0,
        'codigo': ['codigo', 'Codigo', 'CODIGO', 'cod', 'COD'],
        'producto': ['producto', 'Producto', 'PRODUCTO', 'descripcion', 'Descripcion', 'nombre', 'Nombre'],
        'precio': ['precio', 'Precio', 'PRECIO', 'p.venta', 'P.VENTA'],
        'dueno': 'ferreteria_general',
        'folder': 'ferreteria_general'
    }
}

# --- Funciones de Utilidad ---
def get_excel_folder_for_dueno(dueno):
    """Obtener la carpeta Excel específica para un dueño"""
    if dueno not in DUENOS_CONFIG:
        return EXCEL_FOLDER
    
    carpeta_dueno = DUENOS_CONFIG[dueno].get('carpeta_excel', dueno)
    carpeta_path = os.path.join(EXCEL_FOLDER, carpeta_dueno)
    
    # Crear la carpeta si no existe
    if not os.path.exists(carpeta_path):
        os.makedirs(carpeta_path)
    
    return carpeta_path

def leer_celda_coloreada(ws, row, col, color_objetivo):
    """Leer el valor de una celda si tiene el color de fondo especificado"""
    try:
        cell = ws.cell(row=row, column=col)
        
        # Verificar si la celda tiene color de fondo
        if cell.fill and cell.fill.start_color:
            color_celda = cell.fill.start_color.rgb
            if color_celda:
                # Convertir RGB object a string si es necesario
                if hasattr(color_celda, 'upper'):
                    color_celda_str = color_celda.upper()
                else:
                    color_celda_str = str(color_celda).upper()
                
                color_objetivo_normalizado = color_objetivo.upper()
                
                # Remover transparencia si existe (FFFFFF00 -> FFFF00)
                if len(color_celda_str) == 8 and color_celda_str.startswith('FF'):
                    color_celda_str = color_celda_str[2:]
                
                if color_celda_str == color_objetivo_normalizado:
                    return cell.value
        return None
    except Exception as e:
        print(f"Error leyendo celda coloreada: {e}")
        return None

def obtener_precio_por_color(ws, row, columnas_precio, color_objetivo):
    """Obtener el precio de la primera celda coloreada encontrada en las columnas especificadas"""
    for col_name in columnas_precio:
        # Buscar la columna por nombre
        for col_idx in range(1, ws.max_column + 1):
            header_cell = ws.cell(row=1, column=col_idx)
            if header_cell.value and str(header_cell.value).strip().upper() == col_name.upper():
                precio = leer_celda_coloreada(ws, row, col_idx, color_objetivo)
                if precio is not None:
                    return precio, col_name
    return None, None

def parse_price(price_str):
    """
    Interpreta precios con distintos formatos:
    - 200.000 -> 200000
    - 200,50  -> 200.50
    - 200.000,75 -> 200000.75
    - 200,000.75 -> 200000.75 (formato inglés)
    - '' (vacío) -> 0.0
    Devuelve (float, texto_error)
    """
    if price_str is None:
        return 0.0, None
    
    # Convertir a string si no lo es
    if isinstance(price_str, (int, float)):
        return float(price_str), None
    
    if not isinstance(price_str, str):
        try:
            return float(str(price_str)), None
        except (ValueError, TypeError):
            return 0.0, f"Cannot convert {price_str} to number"
    
    # String vacío devuelve 0
    price_str = price_str.strip()
    if not price_str:
        return 0.0, None

    cleaned_str = price_str.replace('$', '').replace(' ', '').strip()
    
    # Si después de limpiar queda vacío, devolver 0
    if not cleaned_str:
        return 0.0, None

    # Caso: contiene tanto punto como coma
    if '.' in cleaned_str and ',' in cleaned_str:
        # Si la coma está después del último punto asumimos formato español: 200.000,75
        if cleaned_str.rfind(',') > cleaned_str.rfind('.'):
            cleaned_str = cleaned_str.replace('.', '').replace(',', '.')
        else:
            # Caso inglés: 200,000.75
            cleaned_str = cleaned_str.replace(',', '')
    elif ',' in cleaned_str:
        # Solo coma: decimal
        cleaned_str = cleaned_str.replace(',', '.')
    else:
        # Solo puntos: probablemente separadores de miles
        # Si hay más de un punto o el segmento después del punto tiene 3 dígitos, quitar puntos
        if cleaned_str.count('.') > 1 or (cleaned_str.count('.') == 1 and len(cleaned_str.split('.')[1]) == 3):
            cleaned_str = cleaned_str.replace('.', '')

    try:
        result = float(cleaned_str)
        return result, None
    except (ValueError, TypeError):
        # Si no se puede convertir, devolver 0 y el texto original como "error"
        return 0.0, f"Cannot parse price: {price_str}"

def _is_postgres_configured() -> bool:
    try:
        return all(os.environ.get(k) for k in ['DATABASE_URL']) and HAS_POSTGRES
    except Exception:
        return False

# --- Funciones de Base de Datos ---
def get_db_connection():
    """Crear conexión a la base de datos (PostgreSQL si está configurado; si no, SQLite)."""
    if _is_postgres_configured():
        try:
            # Railway a veces entrega postgres:// en lugar de postgresql://
            dsn = os.environ.get('DATABASE_URL')
            if dsn and dsn.startswith('postgres://'):
                # psycopg2 recomienda reemplazar el esquema
                dsn = dsn.replace('postgres://', 'postgresql://', 1)
            conn = psycopg2.connect(dsn)
            # Usar autocommit para simplificar lógica (commit manual solo en SQLite)
            try:
                conn.autocommit = True
            except Exception:
                pass
            print(f"\n[INIT] 🐘 Motor de base de datos activo: PostgreSQL ({dsn.split('@')[1] if '@' in dsn else 'remoto'})\n")
            return conn
        except Exception as e:
            print(f"Error de conexión a PostgreSQL: {e}")
            return None
    # Fallback: SQLite
    try:
        conn = sqlite3.connect(DATABASE_FILE)
        conn.row_factory = sqlite3.Row
        print(f"[INIT] Motor de base de datos activo: SQLite")
        return conn
    except sqlite3.Error as e:
        print(f"Error de conexión a SQLite: {e}")
        return None

_SQL_INSERT_OR_IGNORE_REGEX = re.compile(r"INSERT\s+OR\s+IGNORE\s+INTO\s+([A-Za-z0-9_\.\"]+)", re.IGNORECASE)

def _adapt_sql_for_postgres(sql: str) -> str:
    """Adaptaciones mínimas de sintaxis para PostgreSQL.
    - INSERT OR IGNORE INTO tabla ...  => INSERT INTO tabla ... ON CONFLICT DO NOTHING
    - AUTOINCREMENT => (ignorado, PostgreSQL usa SERIAL / IDENTITY)
    - INTEGER PRIMARY KEY AUTOINCREMENT => SERIAL PRIMARY KEY
    Nota: No intenta cobertura completa de dialecto; sólo los casos usados.
    """
    original = sql
    # Transformar INSERT OR IGNORE INTO <tabla>
    def repl(m):
        table = m.group(1)
        return f"INSERT INTO {table}"
    sql2 = _SQL_INSERT_OR_IGNORE_REGEX.sub(repl, sql)
    # Si hicimos un cambio de OR IGNORE añadimos ON CONFLICT DO NOTHING al final antes de ; si no existe ya.
    if sql2 != original and 'ON CONFLICT' not in sql2.upper():
        # Insertar justo antes de RETURNING si existiera, de lo contrario al final.
        upper_sql = sql2.upper()
        if 'RETURNING' in upper_sql:
            parts = re.split(r'(?i)RETURNING', sql2, maxsplit=1)
            sql2 = parts[0].rstrip() + ' ON CONFLICT DO NOTHING RETURNING' + parts[1]
        else:
            sql2 = sql2.rstrip().rstrip(';') + ' ON CONFLICT DO NOTHING'
    # Adaptar autoincrement
    sql2 = re.sub(r'INTEGER\s+PRIMARY\s+KEY\s+AUTOINCREMENT', 'SERIAL PRIMARY KEY', sql2, flags=re.IGNORECASE)
    # No reemplazar genérico AUTOINCREMENT dentro de contexto ya transformado, pero limpiar palabra suelta
    sql2 = re.sub(r'AUTOINCREMENT', '', sql2, flags=re.IGNORECASE)
    return sql2

def _convert_placeholders(sql: str) -> str:
    """Convertir placeholders estilo SQLite (?) a estilo psycopg2 (%s). Evita reemplazar dentro de strings simples.
    Aproximación sencilla: contar '?' y reemplazar secuencialmente por %s salvo que ya existan %s.
    """
    if '%s' in sql:
        return sql  # ya adaptado
    # Ignorar signos de pregunta dentro de literales simples '...?...'
    parts = []
    in_str = False
    for ch in sql:
        if ch == "'":
            in_str = not in_str
            parts.append(ch)
        elif ch == '?' and not in_str:
            parts.append('%s')
        else:
            parts.append(ch)
    return ''.join(parts)

def _prepare_sql(query: str, use_postgres: bool) -> str:
    if not use_postgres:
        return query
    q = _adapt_sql_for_postgres(query)
    q = _convert_placeholders(q)
    return q

# Alias para get_db_connection para compatibilidad con diagnósticos
def db_connect():
    """Alias para get_db_connection para compatibilidad con scripts de diagnóstico."""
    return get_db_connection()

# Función para ejecutar consultas directamente con una conexión
def execute_query(conn, query, params=(), fetch=False):
    """Ejecutar una consulta SQL con parámetros opcionales."""
    cursor = conn.cursor()
    try:
        cursor.execute(query, params)
        if fetch:
            if _is_postgres_configured():
                result = [dict(row) for row in cursor.fetchall()]
            else:
                result = [dict(zip([col[0] for col in cursor.description], row)) for row in cursor.fetchall()]
            return result
        else:
            conn.commit()
            return cursor.rowcount
    except Exception as e:
        print(f"Error en execute_query: {e}")
        raise
    finally:
        cursor.close()

def db_query(query, params=(), fetch=False, conn=None):
    """Ejecutar consulta en la base de datos (PostgreSQL o SQLite).

    - Adapta sintaxis y placeholders automáticamente para PostgreSQL.
    - Retorna lista de dicts si fetch=True, True/False según éxito si fetch=False.
    - Opcionalmente acepta una conexión existente a través del parámetro conn.
    - Cierra siempre la conexión (uso simple por operación). Para alto volumen podría añadirse pool.
    """
    close_conn = False
    if conn is None:
        conn = get_db_connection()
        close_conn = True
    if not conn:
        return None
    use_postgres = _is_postgres_configured()
    sql_exec = _prepare_sql(query, use_postgres)
    cursor = None
    try:
        cursor = conn.cursor()
        cursor.execute(sql_exec, tuple(params))
        if fetch:
            rows = cursor.fetchall()
            if use_postgres:
                cols = [d[0] for d in cursor.description]
                return [dict(zip(cols, r)) for r in rows]
            else:
                return [dict(r) for r in rows]
        else:
            if not use_postgres:
                # En PostgreSQL autocommit ya activo
                conn.commit()
            return True
    except Exception as e:
        print(f"Error en la consulta a la base de datos: {e}\nSQL original: {query}\nSQL ejecutado: {sql_exec}\nParams: {params}")
        try:
            if not use_postgres:
                conn.rollback()
        except Exception:
            pass
        return False
    finally:
        try:
            cursor and cursor.close()
        except Exception:
            pass
        if close_conn:
            try:
                conn.close()
            except Exception:
                pass
        try:
            conn.close()
        except Exception:
            pass

def init_excel_manual():
    """Inicializar el archivo Excel de productos manuales"""
    if not os.path.exists(EXCEL_FOLDER):
        os.makedirs(EXCEL_FOLDER)
    
    # Estructura esperada y unificada (orden requerido):
    # Codigo | Proveedor | Nombre | Precio | Observaciones | Dueno
    expected_headers = ['Codigo', 'Proveedor', 'Nombre', 'Precio', 'Observaciones', 'Dueno']
    
    if not os.path.exists(MANUAL_PRODUCTS_FILE):
        # Crear nuevo archivo Excel con las columnas requeridas
        wb = Workbook()
        ws = wb.active
        ws.title = "Productos Manuales"
        
        for col, header in enumerate(expected_headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        wb.save(MANUAL_PRODUCTS_FILE)
        print(f"Archivo Excel creado: {MANUAL_PRODUCTS_FILE}")
    else:
        try:
            # Si existe, normalizar encabezados en caso de que tengan acentos u orden distinto
            wb = load_workbook(MANUAL_PRODUCTS_FILE)
            ws = wb.active
            current = [ws.cell(row=1, column=i).value for i in range(1, 7)]
            # Normalizar valores con posibles acentos
            normalize = lambda s: str(s or '').strip().replace('ó', 'o').replace('Ó', 'O').replace('ñ', 'n').replace('Ñ', 'N').title()
            normalized = [normalize(v) for v in current]
            # Si detectamos el antiguo orden (Nombre antes que Proveedor), migrar filas intercambiando columnas B y C
            old_order = ['Codigo', 'Nombre', 'Proveedor', 'Precio', 'Observaciones', 'Dueno']
            if normalized == old_order:
                for r in range(2, ws.max_row + 1):
                    val_b = ws.cell(row=r, column=2).value  # Nombre
                    val_c = ws.cell(row=r, column=3).value  # Proveedor
                    # Nuevo orden requiere Proveedor en B y Nombre en C
                    ws.cell(row=r, column=2, value=val_c)
                    ws.cell(row=r, column=3, value=val_b)
                # Reescribir encabezados al orden esperado
                for idx, header in enumerate(expected_headers, start=1):
                    ws.cell(row=1, column=idx, value=header)
                wb.save(MANUAL_PRODUCTS_FILE)
            elif normalized != expected_headers:
                # Si solo difieren por acentos u otros detalles, reescribimos encabezados al estándar
                for idx, header in enumerate(expected_headers, start=1):
                    ws.cell(row=1, column=idx, value=header)
                wb.save(MANUAL_PRODUCTS_FILE)
        except Exception as e:
            print(f"Advertencia al normalizar/migrar encabezados del Excel manual: {e}")

def ensure_productos_manual_columns():
    """Auto-reparar tabla productos_manual agregando columnas faltantes (proveedor, observaciones, dueno, created_at).
    Evita fallos en instalaciones antiguas que no corrieron baseline completa.
    """
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        # Detectar motor para sintaxis created_at default
        use_pg = os.environ.get('USE_POSTGRES', '0') == '1'
        # Obtener info de columnas
        columnas = {}
        try:
            if use_pg:
                cur.execute("SELECT column_name FROM information_schema.columns WHERE table_name='productos_manual'")
                for r in cur.fetchall():
                    columnas[r[0]] = True
            else:
                cur.execute("PRAGMA table_info(productos_manual)")
                for r in cur.fetchall():
                    # PRAGMA: (cid, name, type, notnull, dflt_value, pk)
                    columnas[r[1]] = True
        except Exception as e_cols:
            print(f"[ensure_productos_manual_columns] No se pudo leer columnas: {e_cols}")
            return
        alter_stmts = []
        if 'proveedor' not in columnas:
            alter_stmts.append("ALTER TABLE productos_manual ADD COLUMN proveedor TEXT")
        if 'observaciones' not in columnas:
            alter_stmts.append("ALTER TABLE productos_manual ADD COLUMN observaciones TEXT")
        if 'dueno' not in columnas:
            alter_stmts.append("ALTER TABLE productos_manual ADD COLUMN dueno TEXT")
        if 'created_at' not in columnas:
            if use_pg:
                alter_stmts.append("ALTER TABLE productos_manual ADD COLUMN created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP")
            else:
                # SQLite: evitar default no constante en ALTER
                alter_stmts.append("ALTER TABLE productos_manual ADD COLUMN created_at TEXT")
        for sql in alter_stmts:
            try:
                cur.execute(sql)
                print(f"[ensure_productos_manual_columns] Ejecutado: {sql}")
            except Exception as e_alter:
                print(f"[ensure_productos_manual_columns] Falló: {sql} -> {e_alter}")
        if alter_stmts:
            conn.commit()
    except Exception as e:
        print(f"[ensure_productos_manual_columns] Error general: {e}")
    finally:
        try:
            cur.close()
        except Exception:
            pass
        try:
            conn.close()
        except Exception:
            pass

def ensure_proveedores_manual_columns():
    """Auto-reparar tabla proveedores_manual agregando columnas faltantes (dueno, created_at)."""
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        use_pg = os.environ.get('USE_POSTGRES', '0') == '1'
        columnas = {}
        try:
            if use_pg:
                cur.execute("SELECT column_name FROM information_schema.columns WHERE table_name='proveedores_manual'")
                for r in cur.fetchall():
                    columnas[r[0]] = True
            else:
                cur.execute("PRAGMA table_info(proveedores_manual)")
                for r in cur.fetchall():
                    columnas[r[1]] = True
        except Exception as e_cols:
            print(f"[ensure_proveedores_manual_columns] No se pudo leer columnas: {e_cols}")
            return
        alters = []
        if 'dueno' not in columnas:
            alters.append("ALTER TABLE proveedores_manual ADD COLUMN dueno TEXT")
        if 'created_at' not in columnas:
            # Evitar default no constante en SQLite: agregar sin default
            if use_pg:
                alters.append("ALTER TABLE proveedores_manual ADD COLUMN created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP")
            else:
                alters.append("ALTER TABLE proveedores_manual ADD COLUMN created_at TEXT")
        for sql in alters:
            try:
                cur.execute(sql)
                print(f"[ensure_proveedores_manual_columns] Ejecutado: {sql}")
            except Exception as e_alter:
                print(f"[ensure_proveedores_manual_columns] Falló: {sql} -> {e_alter}")
        if alters:
            conn.commit()
    except Exception as e:
        print(f"[ensure_proveedores_manual_columns] Error general: {e}")
    finally:
        try:
            cur.close()
        except Exception:
            pass
        try:
            conn.close()
        except Exception:
            pass

def inferir_dueno_proveedor_si_unico(nombre_proveedor: str):
    """Si el proveedor tiene dueno vacío y todos los productos manuales que lo usan son de un único dueno, asignarle ese dueno.
    Retorna el dueno asignado o None.
    """
    try:
        if not nombre_proveedor:
            return None
        rows = db_query("SELECT id, dueno FROM proveedores_manual WHERE LOWER(nombre)=LOWER(?)", (nombre_proveedor,), fetch=True)
        if not rows:
            return None
        # Filtrar los que no tienen dueno
        vacios = [r for r in rows if not (r.get('dueno') or '').strip()]
        if not vacios:
            return None
        # Buscar duenos distintos en productos_manual que referencien este proveedor por texto
        productos = db_query("SELECT DISTINCT LOWER(dueno) as du FROM productos_manual WHERE LOWER(proveedor)=LOWER(?)", (nombre_proveedor,), fetch=True)
        if not productos:
            return None
        duenos = { (r.get('du') or '').lower() for r in productos if (r.get('du') or '').strip() }
        if len(duenos) == 1:
            unico = list(duenos)[0]
            # Actualizar filas vacías
            db_query("UPDATE proveedores_manual SET dueno = ? WHERE LOWER(nombre)=LOWER(?) AND (dueno IS NULL OR TRIM(dueno)='')", (unico, nombre_proveedor))
            print(f"[inferir_dueno_proveedor] Asignado dueno='{unico}' a proveedor='{nombre_proveedor}'")
            return unico
    except Exception as e:
        print(f"[inferir_dueno_proveedor] Error: {e}")
    return None

def canonicalize_proveedor_name(nombre: str) -> str:
    """Normaliza nombre de proveedor: quita espacios extremos, colapsa espacios internos,
    normaliza unicode (NFKC) y mayúsculas/minúsculas (tal cual, sin forzar upper/lower para no romper matching exacto, solo trim y collapse).
    """
    import unicodedata, re
    if not nombre:
        return ''
    n = unicodedata.normalize('NFKC', nombre)
    n = n.replace('\u00A0', ' ')  # NBSP
    n = n.strip()
    n = re.sub(r'\s+', ' ', n)
    return n

def agregar_producto_db_manual(codigo, proveedor, nombre, precio, observaciones, dueno):
    """Agregar producto a la base de datos de productos manuales"""
    try:
        # Asegurar que las tablas necesarias existen
        ensure_productos_manual_columns()
        
        # Verificar si el producto ya existe (por código y dueño)
        if codigo:
            existing = db_query(
                "SELECT id FROM productos_manual WHERE LOWER(codigo) = LOWER(?) AND dueno = ?", 
                (codigo, dueno), 
                fetch=True
            )
            if existing:
                print(f"Producto con código '{codigo}' ya existe para {dueno}, actualizando...")
                # Actualizar producto existente
                db_query(
                    """UPDATE productos_manual 
                       SET nombre = ?, precio = ?, proveedor = ?, observaciones = ? 
                       WHERE LOWER(codigo) = LOWER(?) AND dueno = ?""",
                    (nombre, precio, proveedor, observaciones, codigo, dueno)
                )
                return True
        
        # Insertar nuevo producto
        db_query(
            """INSERT INTO productos_manual (codigo, nombre, precio, proveedor, observaciones, dueno) 
               VALUES (?, ?, ?, ?, ?, ?)""",
            (codigo, nombre, precio, proveedor, observaciones, dueno)
        )
        
        print(f"Producto '{nombre}' agregado a la base de datos para {dueno}")
        return True
        
    except Exception as e:
        print(f"Error al agregar producto a la base de datos: {e}")
        return False

def agregar_producto_excel_manual(codigo, proveedor, nombre, precio, observaciones, dueno):
    """Agregar producto al Excel de productos manuales"""
    try:
        if not os.path.exists(EXCEL_FOLDER):
            os.makedirs(EXCEL_FOLDER)
        
        expected_headers = ['Codigo', 'Proveedor', 'Nombre', 'Precio', 'Observaciones', 'Dueno']
        
        if not os.path.exists(MANUAL_PRODUCTS_FILE):
            wb = Workbook()
            ws = wb.active
            ws.title = "Productos Manuales"
            # Encabezados estandarizados sin acentos y en orden requerido
            for col, header in enumerate(expected_headers, 1):
                ws.cell(row=1, column=col, value=header)
            wb.save(MANUAL_PRODUCTS_FILE)
        
        wb = load_workbook(MANUAL_PRODUCTS_FILE)
        ws = wb.active
        
        # Verificar y corregir encabezados si es necesario
        try:
            current = [ws.cell(row=1, column=i).value for i in range(1, 7)]
            normalize = lambda s: str(s or '').strip().replace('ó', 'o').replace('Ó', 'O').replace('ñ', 'n').replace('Ñ', 'N').title()
            normalized = [normalize(v) for v in current]
            old_order = ['Codigo', 'Nombre', 'Proveedor', 'Precio', 'Observaciones', 'Dueno']
            if normalized == old_order:
                # Migrar filas intercambiando columnas B y C para adecuar al nuevo orden
                for r in range(2, ws.max_row + 1):
                    val_b = ws.cell(row=r, column=2).value  # Nombre
                    val_c = ws.cell(row=r, column=3).value  # Proveedor
                    ws.cell(row=r, column=2, value=val_c)
                    ws.cell(row=r, column=3, value=val_b)
                # Actualizar encabezados al estándar
                for idx, header in enumerate(expected_headers, start=1):
                    ws.cell(row=1, column=idx, value=header)
            elif normalized != expected_headers:
                for idx, header in enumerate(expected_headers, start=1):
                    ws.cell(row=1, column=idx, value=header)
        except Exception as _:
            pass
        
        # Agregar fila respetando el orden de columnas esperado
        ws.append([codigo, proveedor, nombre, float(precio) if precio is not None else 0.0, observaciones, dueno])
        wb.save(MANUAL_PRODUCTS_FILE)
        
        return True
    except Exception as e:
        print(f"Error al agregar producto manual: {e}")
        return False

def buscar_en_excel_manual(termino_busqueda, dueno_filtro=None):
    """Buscar productos en el Excel de productos manuales"""
    resultados = []
    
    try:
        if not os.path.exists(MANUAL_PRODUCTS_FILE):
            return resultados
        
        wb = load_workbook(MANUAL_PRODUCTS_FILE)
        ws = wb.active
        
        # Asumir que las columnas son (orden unificado):
        # Codigo (A), Proveedor (B), Nombre (C), Precio (D), Observaciones (E), Dueno (F)
        tokens = [t.strip().lower() for t in str(termino_busqueda).split() if t.strip()]
        for row in ws.iter_rows(min_row=2, values_only=True):  # Saltar encabezado
            if len(row) < 6:
                continue
            
            codigo = str(row[0]).strip() if row[0] else ''
            proveedor = str(row[1]).strip() if row[1] else ''
            nombre = str(row[2]).strip() if row[2] else ''
            precio = str(row[3]).strip() if row[3] else ''
            observaciones = str(row[4]).strip() if row[4] else ''
            dueno = str(row[5]).strip() if row[5] else ''
            
            # Filtrar por dueño si se especifica
            if dueno_filtro and dueno.lower() != dueno_filtro.lower():
                continue
            
            # Buscar en los campos relevantes con soporte de combinaciones (AND entre tokens)
            campos = [codigo.lower(), proveedor.lower(), nombre.lower(), precio.lower(), observaciones.lower()]
            coincide = True
            for tok in tokens:
                if not any(tok in campo for campo in campos):
                    coincide = False
                    break
            if coincide:
                
                resultados.append({
                    'codigo': codigo,
                    'proveedor': proveedor,
                    'nombre': nombre,
                    'precio': precio,
                    'observaciones': observaciones,
                    'dueno': dueno
                })
    
    except Exception as e:
        print(f"Error al buscar en Excel manual: {e}")
    
    return resultados

def init_db():
    """Inicializar la base de datos (MySQL si está configurado; si no, SQLite)."""
    if not os.path.exists(EXCEL_FOLDER):
        os.makedirs(EXCEL_FOLDER)
    try:
        conn = get_db_connection()
        if not conn:
            print("ERROR al inicializar la BD: No se pudo establecer la conexión con la base de datos.")
            return

        use_postgres = _is_postgres_configured()
        print("🔧 Inicializando base de datos...")
        print(f"[INIT] Motor de base de datos activo: {'Postgres' if use_postgres else 'SQLite'}")
        cursor = conn.cursor()
        
        # Crear tabla stock
        cursor.execute(_adapt_sql_for_postgres(''' 
            CREATE TABLE IF NOT EXISTS stock ( 
                id SERIAL PRIMARY KEY, 
                codigo TEXT, 
                nombre TEXT NOT NULL, 
                precio REAL NOT NULL, 
                cantidad INTEGER NOT NULL, 
                fecha_compra TEXT NOT NULL, 
                proveedor TEXT, 
                observaciones TEXT, 
                precio_texto TEXT, 
                avisar_bajo_stock INTEGER DEFAULT 0, 
                min_stock_aviso INTEGER DEFAULT NULL,
                dueno TEXT DEFAULT 'ferreteria_general'
            ) 
        '''))
        # Migración created_at: añadir columna si no existe (timestamp ISO completo)
        try:
            if use_postgres:
                cursor.execute("SELECT column_name FROM information_schema.columns WHERE table_name='stock'")
                existing_cols = {r[0].lower() for r in cursor.fetchall()}
                if 'created_at' not in existing_cols:
                    cursor.execute("ALTER TABLE stock ADD COLUMN created_at TEXT")
            else:
                cols = cursor.execute("PRAGMA table_info(stock)").fetchall()
                col_names = [c[1] for c in cols]
                if 'created_at' not in col_names:
                    cursor.execute("ALTER TABLE stock ADD COLUMN created_at TEXT")
            # Rellenar nulos con fecha_compra + ' 00:00:00' si es NULL
            cursor.execute("UPDATE stock SET created_at = fecha_compra || ' 00:00:00' WHERE created_at IS NULL OR TRIM(created_at)='' ")
        except Exception as e:
            print(f"Aviso: no se pudo aplicar migración created_at: {e}")
        # Migración: añadir columna dueno si faltara (solo para SQLite)
        if not use_postgres:
            try:
                cols = cursor.execute("PRAGMA table_info(stock)").fetchall()
                col_names = [c[1] for c in cols]
                if 'dueno' not in col_names:
                    cursor.execute("ALTER TABLE stock ADD COLUMN dueno TEXT DEFAULT 'ferreteria_general'")
            except Exception:
                pass
        # Asegurar columnas críticas (idempotente) por si algún despliegue quedó a medias
        try:
            if use_postgres:
                cursor.execute("SELECT column_name FROM information_schema.columns WHERE table_name='stock'")
                existing_all = {r[0].lower() for r in cursor.fetchall()}
                # Lista completa esperada
                expected = ['codigo','nombre','precio','cantidad','fecha_compra','proveedor','observaciones','precio_texto','avisar_bajo_stock','min_stock_aviso','dueno','created_at']
                # created_at ya se intentó antes, repetimos defensa
                for col in expected:
                    if col not in existing_all:
                        alter_map = {
                            'precio_texto': "ALTER TABLE stock ADD COLUMN precio_texto TEXT",
                            'avisar_bajo_stock': "ALTER TABLE stock ADD COLUMN avisar_bajo_stock INTEGER DEFAULT 0",
                            'min_stock_aviso': "ALTER TABLE stock ADD COLUMN min_stock_aviso INTEGER",
                            'dueno': "ALTER TABLE stock ADD COLUMN dueno TEXT DEFAULT 'ferreteria_general'",
                            'created_at': "ALTER TABLE stock ADD COLUMN created_at TEXT"
                        }
                        stmt = alter_map.get(col)
                        if stmt:
                            try: cursor.execute(stmt)
                            except Exception as ee: print(f"[WARN] Alter PG {col} fallo: {ee}")
            else:
                cols_now = cursor.execute("PRAGMA table_info(stock)").fetchall()
                names = {c[1].lower() for c in cols_now}
                def alter(sql):
                    try: cursor.execute(sql)
                    except Exception as ee: print(f"[WARN] Alter SQLite fallo: {ee}")
                if 'precio_texto' not in names: alter("ALTER TABLE stock ADD COLUMN precio_texto TEXT")
                if 'avisar_bajo_stock' not in names: alter("ALTER TABLE stock ADD COLUMN avisar_bajo_stock INTEGER DEFAULT 0")
                if 'min_stock_aviso' not in names: alter("ALTER TABLE stock ADD COLUMN min_stock_aviso INTEGER")
                if 'dueno' not in names: alter("ALTER TABLE stock ADD COLUMN dueno TEXT DEFAULT 'ferreteria_general'")
                if 'created_at' not in names: alter("ALTER TABLE stock ADD COLUMN created_at TEXT")
        except Exception as e:
            print(f"[WARN] ensure_stock_columns segunda fase: {e}")

        # Crear tabla stock_history (modelo híbrido) si no existe
        try:
            cursor.execute(_adapt_sql_for_postgres('''
                CREATE TABLE IF NOT EXISTS stock_history (
                    id SERIAL PRIMARY KEY,
                    stock_id INTEGER,                -- id de la fila en stock (si existía antes de update)
                    codigo TEXT,
                    nombre TEXT,
                    precio REAL,
                    cantidad INTEGER,
                    fecha_compra TEXT,
                    proveedor TEXT,
                    observaciones TEXT,
                    precio_texto TEXT,
                    dueno TEXT,
                    created_at TEXT,                 -- created_at original de la fila stock (si aplica)
                    fecha_evento TEXT NOT NULL,      -- timestamp del momento del cambio
                    tipo_cambio TEXT NOT NULL,       -- insert | update | manual_edit | import | sync | otro
                    fuente TEXT,                     -- origen más específico (endpoint, script, etc.)
                    usuario TEXT                     -- username si está logueado
                )
            '''))
            # Índices útiles para consultas (ignoramos errores si ya existen)
            try: cursor.execute("CREATE INDEX IF NOT EXISTS idx_stock_history_codigo ON stock_history(codigo)")
            except Exception: pass
            try: cursor.execute("CREATE INDEX IF NOT EXISTS idx_stock_history_fecha_evento ON stock_history(fecha_evento)")
            except Exception: pass
            try: cursor.execute("CREATE INDEX IF NOT EXISTS idx_stock_history_stock_id ON stock_history(stock_id)")
            except Exception: pass
        except Exception as e:
            print(f"[WARN] No se pudo crear stock_history: {e}")
        
        # Crear tabla proveedores_manual
        cursor.execute(_adapt_sql_for_postgres(''' 
            CREATE TABLE IF NOT EXISTS proveedores_manual ( 
                id SERIAL PRIMARY KEY, 
                nombre TEXT NOT NULL UNIQUE 
            ) 
        '''))
        # Tabla de notificaciones persistentes
        try:
            cursor.execute(_adapt_sql_for_postgres('''
                CREATE TABLE IF NOT EXISTS notificaciones (
                    id SERIAL PRIMARY KEY,
                    codigo TEXT,
                    nombre TEXT,
                    proveedor TEXT,
                    mensaje TEXT NOT NULL,
                    ts TEXT NOT NULL,
                    leida INTEGER DEFAULT 0,
                    user_id INTEGER
                )
            '''))
            try: cursor.execute("CREATE INDEX IF NOT EXISTS idx_notif_ts ON notificaciones(ts)")
            except Exception: pass
            # Migración específica SQLite: asegurar que id sea INTEGER PRIMARY KEY AUTOINCREMENT si quedaron filas con id NULL
            if not use_postgres:
                try:
                    cur2 = conn.cursor()
                    cur2.execute("PRAGMA table_info(notificaciones)")
                    cols = cur2.fetchall()
                    # Detectar si la columna id no es INTEGER PRIMARY KEY (cuando SERIAL no se tradujo bien)
                    col_id = next((c for c in cols if c[1].lower()=="id"), None)
                    need_migration = False
                    if col_id:
                        type_decl = (col_id[2] or '').upper()
                        if 'INT' not in type_decl: # tipo inesperado
                            need_migration = True
                    # Verificar si hay ids NULL
                    cur2.execute("SELECT COUNT(1) FROM notificaciones WHERE id IS NULL")
                    null_count = cur2.fetchone()[0]
                    if null_count > 0:
                        need_migration = True
                    if need_migration:
                        print(f"[MIGRA] Reparando tabla notificaciones (ids nulos={null_count})")
                        # Renombrar tabla actual
                        cur2.execute("ALTER TABLE notificaciones RENAME TO notificaciones_old")
                        cur2.execute('''CREATE TABLE notificaciones (
                            id INTEGER PRIMARY KEY AUTOINCREMENT,
                            codigo TEXT,
                            nombre TEXT,
                            proveedor TEXT,
                            mensaje TEXT NOT NULL,
                            ts TEXT NOT NULL,
                            leida INTEGER DEFAULT 0,
                            user_id INTEGER
                        )''')
                        # Copiar datos (generará nuevos ids autoincrement)
                        cur2.execute('''INSERT INTO notificaciones (codigo,nombre,proveedor,mensaje,ts,leida,user_id)
                                         SELECT codigo,nombre,proveedor,mensaje,ts,leida,user_id FROM notificaciones_old''')
                        cur2.execute("DROP TABLE notificaciones_old")
                        try:
                            cur2.execute("CREATE INDEX IF NOT EXISTS idx_notif_ts ON notificaciones(ts)")
                        except Exception:
                            pass
                        conn.commit()
                except Exception as mig_e:
                    print(f"[WARN] Migración notificaciones SQLite falló: {mig_e}")
        except Exception as e:
            print(f"[WARN] No se pudo crear tabla notificaciones: {e}")
        # Tabla meta para asociar uno o más dueños a un mismo proveedor (permitir duplicados por dueño)
        cursor.execute(_adapt_sql_for_postgres('''
            CREATE TABLE IF NOT EXISTS proveedores_meta (
                id SERIAL PRIMARY KEY,
                nombre TEXT NOT NULL,
                dueno TEXT NOT NULL,
                UNIQUE(nombre, dueno)
            )
        '''))
        # Migración simple a esquema con UNIQUE(nombre,dueno) (solo para SQLite)
        if not use_postgres:
            try:
                schema = cursor.execute("SELECT sql FROM sqlite_master WHERE type='table' AND name='proveedores_meta'").fetchone()
                if schema and 'UNIQUE(nombre, dueno)' not in (schema[0] or ''):
                    cursor.execute('ALTER TABLE proveedores_meta RENAME TO proveedores_meta_old')
                    cursor.execute('''
                        CREATE TABLE proveedores_meta (
                            id SERIAL PRIMARY KEY,
                            nombre TEXT NOT NULL,
                            dueno TEXT NOT NULL,
                            UNIQUE(nombre, dueno)
                        )
                    ''')
                    try:
                        cursor.execute('INSERT OR IGNORE INTO proveedores_meta(nombre, dueno) SELECT nombre, dueno FROM proveedores_meta_old')
                    except Exception:
                        pass
                    cursor.execute('DROP TABLE proveedores_meta_old')
            except Exception:
                pass
        # Tabla de ocultamiento lógico de proveedores (no tocar Excels)
        # Ajuste: originalmente nombre era UNIQUE impidiendo ocultar el mismo proveedor para ambos dueños.
        # Migración a UNIQUE(nombre, dueno)
        try:
            if use_postgres:
                # En Postgres no usamos PRAGMA. Creamos tabla si no existe con clave compuesta.
                cursor.execute('''
                    CREATE TABLE IF NOT EXISTS proveedores_ocultos (
                        id SERIAL PRIMARY KEY,
                        nombre TEXT NOT NULL,
                        dueno TEXT,
                        CONSTRAINT uq_prov_ocultos UNIQUE (nombre, dueno)
                    )
                ''')
            else:
                # Ruta SQLite con migración desde esquema antiguo (sólo nombre UNIQUE) a (nombre,dueno)
                cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='proveedores_ocultos'")
                existe = cursor.fetchone()
                if existe:
                    # Intentar detectar si falta UNIQUE compuesto revisando sql
                    schema = cursor.execute("SELECT sql FROM sqlite_master WHERE type='table' AND name='proveedores_ocultos'").fetchone()
                    if schema and 'UNIQUE(nombre, dueno)' not in (schema[0] or ''):
                        cursor.execute("ALTER TABLE proveedores_ocultos RENAME TO proveedores_ocultos_old")
                        cursor.execute('''CREATE TABLE proveedores_ocultos (
                                id INTEGER PRIMARY KEY AUTOINCREMENT,
                                nombre TEXT NOT NULL,
                                dueno TEXT,
                                UNIQUE(nombre, dueno)
                            )''')
                        cursor.execute('''INSERT OR IGNORE INTO proveedores_ocultos(nombre, dueno)
                                           SELECT nombre, dueno FROM proveedores_ocultos_old''')
                        cursor.execute('DROP TABLE proveedores_ocultos_old')
                else:
                    cursor.execute('''CREATE TABLE proveedores_ocultos (
                            id INTEGER PRIMARY KEY AUTOINCREMENT,
                            nombre TEXT NOT NULL,
                            dueno TEXT,
                            UNIQUE(nombre, dueno)
                        )''')
        except Exception as e:
            try:
                conn.rollback()
            except Exception:
                pass
            print(f"[WARN] Migración proveedores_ocultos (ignorando) : {e}")
        # Tabla para registrar proveedores eliminados definitivamente (para seguir excluyéndolos del historial)
        try:
            cursor.execute(_adapt_sql_for_postgres('''
                CREATE TABLE IF NOT EXISTS proveedores_eliminados (
                    id SERIAL PRIMARY KEY,
                    nombre TEXT NOT NULL,
                    dueno TEXT NOT NULL,
                    fecha_eliminacion TEXT NOT NULL,
                    UNIQUE(nombre, dueno)
                )
            '''))
        except Exception as e:
            print(f"[WARN] Creación proveedores_eliminados falló: {e}")
        
        # Crear tabla productos_manual (forma mínima legacy) y luego asegurar columnas adicionales con ALTER dinámico
        cursor.execute(_adapt_sql_for_postgres(''' 
            CREATE TABLE IF NOT EXISTS productos_manual ( 
                id SERIAL PRIMARY KEY, 
                proveedor_id INTEGER, 
                nombre TEXT NOT NULL, 
                codigo TEXT, 
                precio REAL NOT NULL, 
                FOREIGN KEY (proveedor_id) REFERENCES proveedores_manual(id) ON DELETE CASCADE 
            ) 
        '''))
        
        # Crear tabla users
        cursor.execute(_adapt_sql_for_postgres(''' 
            CREATE TABLE IF NOT EXISTS users ( 
                id SERIAL PRIMARY KEY, 
                username TEXT NOT NULL UNIQUE, 
                password_hash TEXT NOT NULL 
            ) 
        '''))
        
        # Crear usuario por defecto
        default_username = 'Pauluk'
        default_password = 'Jap2005'
        if use_postgres:
            cursor.execute("SELECT id FROM users WHERE username = %s", (default_username,))
        else:
            cursor.execute("SELECT id FROM users WHERE username = ?", (default_username,))
        if cursor.fetchone() is None:
            hashed_password = generate_password_hash(default_password)
            if use_postgres:
                cursor.execute("INSERT INTO users (username, password_hash) VALUES (%s, %s)", 
                               (default_username, hashed_password))
            else:
                cursor.execute("INSERT INTO users (username, password_hash) VALUES (?, ?)", 
                               (default_username, hashed_password))
        
        # Crear proveedores por defecto
        proveedores_excel = ['BremenTools', 'Crossmaster', 'Berger', 'Chiesa', 'Cachan', 'Otros Proveedores']
        for nombre in proveedores_excel:
            if use_postgres:
                cursor.execute("INSERT INTO proveedores_manual (nombre) VALUES (%s) ON CONFLICT (nombre) DO NOTHING", (nombre,))
            else:
                cursor.execute("INSERT OR IGNORE INTO proveedores_manual (nombre) VALUES (?)", (nombre,))
        # Consolidar errores comunes de escritura (BremenTools / Crossmaster)
        try:
            # Buscar variantes mal escritas y reemplazar mappings/productos
            variantes_bremen = ['Brementools', 'brementools', 'Bremetools', 'bremetools']
            for var in variantes_bremen:
                if var == 'BremenTools':
                    continue
                # Actualizar productos manuales / stock que refieran a la variante
                if not use_postgres:
                    cursor.execute("UPDATE stock SET proveedor='BremenTools' WHERE LOWER(proveedor)=LOWER(?)", (var,))
                else:
                    cursor.execute("UPDATE stock SET proveedor='BremenTools' WHERE LOWER(proveedor)=LOWER(%s)", (var,))
                # Eliminar proveedor duplicado en proveedores_manual (manteniendo el correcto)
                if not use_postgres:
                    cursor.execute("DELETE FROM proveedores_manual WHERE LOWER(nombre)=LOWER(?) AND nombre <> 'BremenTools'", (var,))
                else:
                    cursor.execute("DELETE FROM proveedores_manual WHERE LOWER(nombre)=LOWER(%s) AND nombre <> 'BremenTools'", (var,))
                # Normalizar meta
                if not use_postgres:
                    cursor.execute("UPDATE OR IGNORE proveedores_meta SET nombre='BremenTools' WHERE LOWER(nombre)=LOWER(?)", (var,))
                else:
                    # En PostgreSQL: ON CONFLICT manual; intentar update simple
                    try:
                        cursor.execute("UPDATE proveedores_meta SET nombre='BremenTools' WHERE LOWER(nombre)=LOWER(%s)", (var,))
                    except Exception:
                        pass
            # Variantes Crossmaster (normalizar a 'Crossmaster')
            variantes_cross = ['crossmaster', 'CrossMaster', 'CROSSMASTER', 'cross master', 'cross-master']
            for var in variantes_cross:
                if var == 'Crossmaster':
                    continue
                if not use_postgres:
                    cursor.execute("UPDATE stock SET proveedor='Crossmaster' WHERE LOWER(proveedor)=LOWER(?)", (var,))
                else:
                    cursor.execute("UPDATE stock SET proveedor='Crossmaster' WHERE LOWER(proveedor)=LOWER(%s)", (var,))
                if not use_postgres:
                    cursor.execute("DELETE FROM proveedores_manual WHERE LOWER(nombre)=LOWER(?) AND nombre <> 'Crossmaster'", (var,))
                else:
                    cursor.execute("DELETE FROM proveedores_manual WHERE LOWER(nombre)=LOWER(%s) AND nombre <> 'Crossmaster'", (var,))
                if not use_postgres:
                    cursor.execute("UPDATE OR IGNORE proveedores_meta SET nombre='Crossmaster' WHERE LOWER(nombre)=LOWER(?)", (var,))
                else:
                    try:
                        cursor.execute("UPDATE proveedores_meta SET nombre='Crossmaster' WHERE LOWER(nombre)=LOWER(%s)", (var,))
                    except Exception:
                        pass
            # Migración de claves antiguas ferreteria_prov1/2/3 a dewalt/sica/sorbalok
            migraciones_fg = [
                ('ferreteria_prov1', 'dewalt'),
                ('ferreteria_prov2', 'sica'),
                ('ferreteria_prov3', 'sorbalok')
            ]
            for viejo, nuevo in migraciones_fg:
                try:
                    if not use_postgres:
                        cursor.execute("UPDATE stock SET proveedor=? WHERE LOWER(proveedor)=LOWER(?)", (nuevo, viejo))
                        cursor.execute("UPDATE OR IGNORE proveedores_meta SET nombre=? WHERE LOWER(nombre)=LOWER(?)", (nuevo, viejo))
                        cursor.execute("UPDATE proveedores_manual SET nombre=? WHERE LOWER(nombre)=LOWER(?)", (nuevo, viejo))
                    else:
                        cursor.execute("UPDATE stock SET proveedor=%s WHERE LOWER(proveedor)=LOWER(%s)", (nuevo, viejo))
                        try:
                            cursor.execute("UPDATE proveedores_meta SET nombre=%s WHERE LOWER(nombre)=LOWER(%s)", (nuevo, viejo))
                        except Exception:
                            pass
                        cursor.execute("UPDATE proveedores_manual SET nombre=%s WHERE LOWER(nombre)=LOWER(%s)", (nuevo, viejo))
                except Exception:
                    pass
        except Exception:
            pass
        # Asegurar dueños meta por defecto
        try:
            ricky_defaults = ['BremenTools', 'Crossmaster', 'Berger', 'Chiesa', 'Cachan']
            for nombre in ricky_defaults:
                if use_postgres:
                    cursor.execute("INSERT INTO proveedores_meta (nombre, dueno) VALUES (%s, 'ricky') ON CONFLICT (nombre, dueno) DO NOTHING", (nombre,))
                else:
                    cursor.execute("INSERT OR REPLACE INTO proveedores_meta (nombre, dueno) VALUES (?, 'ricky')", (nombre,))
            if use_postgres:
                cursor.execute("INSERT INTO proveedores_meta (nombre, dueno) VALUES ('Otros Proveedores', 'ferreteria_general') ON CONFLICT (nombre, dueno) DO NOTHING")
            else:
                cursor.execute("INSERT OR REPLACE INTO proveedores_meta (nombre, dueno) VALUES ('Otros Proveedores', 'ferreteria_general')")
        except Exception:
            pass
        
        # Crear tabla proveedores_duenos para relaciones muchos-a-muchos
        try:
            cursor.execute(_adapt_sql_for_postgres('''
                CREATE TABLE IF NOT EXISTS proveedores_duenos (
                    id SERIAL PRIMARY KEY,
                    proveedor_id INTEGER NOT NULL,
                    dueno TEXT NOT NULL,
                    UNIQUE(proveedor_id, dueno),
                    FOREIGN KEY (proveedor_id) REFERENCES proveedores_manual(id) ON DELETE CASCADE
                )
            '''))
            # Crear índices para mejorar rendimiento
            try:
                cursor.execute("CREATE INDEX IF NOT EXISTS idx_proveedores_duenos_proveedor_id ON proveedores_duenos(proveedor_id)")
                cursor.execute("CREATE INDEX IF NOT EXISTS idx_proveedores_duenos_dueno ON proveedores_duenos(dueno)")
            except Exception:
                pass
            
            # Migrar datos desde proveedores_meta si no existen en proveedores_duenos
            if use_postgres:
                cursor.execute("""
                    INSERT INTO proveedores_duenos (proveedor_id, dueno)
                    SELECT pm.id, meta.dueno 
                    FROM proveedores_meta meta
                    JOIN proveedores_manual pm ON pm.nombre = meta.nombre
                    ON CONFLICT (proveedor_id, dueno) DO NOTHING
                """)
            else:
                cursor.execute("""
                    INSERT OR IGNORE INTO proveedores_duenos (proveedor_id, dueno)
                    SELECT pm.id, meta.dueno 
                    FROM proveedores_meta meta
                    JOIN proveedores_manual pm ON pm.nombre = meta.nombre
                """)
                
        except Exception as e:
            print(f"[WARN] Error creando tabla proveedores_duenos: {e}")
        
        conn.commit()
        # Índices para acelerar filtros de historial (creación idempotente)
        try:
            cursor.execute(_adapt_sql_for_postgres('CREATE INDEX IF NOT EXISTS idx_stock_proveedor ON stock(proveedor)'))
        except Exception:
            pass
        try:
            cursor.execute(_adapt_sql_for_postgres('CREATE INDEX IF NOT EXISTS idx_stock_dueno ON stock(dueno)'))
        except Exception:
            pass
        try:
            cursor.execute(_adapt_sql_for_postgres('CREATE INDEX IF NOT EXISTS idx_stock_codigo ON stock(codigo)'))
        except Exception:
            pass
        try:
            cursor.execute(_adapt_sql_for_postgres('CREATE INDEX IF NOT EXISTS idx_stock_nombre ON stock(nombre)'))
        except Exception:
            pass
        cursor.close()
        conn.close()
        if use_postgres:
            print("Base de datos PostgreSQL inicializada/verificada con éxito.")
            print("✅ PostgreSQL listo para usarse. Modo: PRODUCCIÓN 🚀")
        else:
            print("Base de datos SQLite inicializada/verificada con éxito.")
            print("✅ Base de datos inicializada correctamente")
        
        # Inicializar también el Excel de productos manuales
        init_excel_manual()
        
    except sqlite3.Error as e:
        print(f"\nERROR al inicializar la BD (SQLite): {e}")
    except Exception as e:
        print(f"\nERROR al inicializar la BD: {e}")

# --- Decorador de Autenticación ---
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

# --- Rutas de Autenticación ---
@app.route('/login', methods=['GET', 'POST'])
def login():
    if 'user_id' in session:
        return redirect(url_for('index'))
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        user = db_query("SELECT * FROM users WHERE username = ?", (username,), fetch=True)
        if user and check_password_hash(user[0]['password_hash'], password):
            session.clear()
            session.permanent = True
            session['user_id'] = user[0]['id']
            session['username'] = user[0]['username']
            return redirect(url_for('index'))
        else:
            flash('Usuario o contraseña incorrectos.', 'danger')
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    flash('Has cerrado sesión.', 'success')
    return redirect(url_for('login'))

# --- Rutas de la Aplicación Web (Protegidas) ---
@app.route('/')
@login_required
def index():
    return render_template('index.html')

@app.route('/historial')
@login_required
def historial():
    # Cargar sólo los proveedores distintos y dueños para construir filtros; datos se cargan vía AJAX
    # Como el usuario indicó: cuando el filtro de dueño está en "Todos" no se deben listar proveedores concretos
    # para evitar que parezca que se limita la búsqueda. Por eso inicializamos vacío y sólo se llenará
    # cuando el usuario seleccione un dueño específico (AJAX a /historial_proveedores).
    return render_template('historial.html', proveedores=[])

@app.route('/historial_datos', methods=['POST'])
@login_required
def historial_datos():
    try:
        data = request.get_json() or {}
        proveedor_raw = (data.get('proveedor') or '').strip()
        dueno_raw = (data.get('dueno') or '').strip()
        proveedor = proveedor_raw
        dueno = dueno_raw
        termino = (data.get('termino') or '').strip()
        params = []
        filtros = []
        if dueno:
            filtros.append('LOWER(dueno)=?')
            params.append(dueno.lower())
        if proveedor:
            filtros.append('LOWER(proveedor)=?')
            params.append(proveedor.lower())
        # Si proveedor está oculto para el dueño solicitado, devolver vacío antes de consultar stock
        if dueno and proveedor:
            prov_norm = _normalizar_nombre_proveedor(proveedor)
            # Recuperar sets normalizados de ocultos/eliminados para dueño
            ocultos_rows = db_query("SELECT LOWER(nombre) as nombre FROM proveedores_ocultos WHERE dueno=?", (dueno.lower(),), fetch=True) or []
            eliminados_rows = db_query("SELECT LOWER(nombre) as nombre FROM proveedores_eliminados WHERE dueno=?", (dueno.lower(),), fetch=True) or []
            # Normalizar cada uno (aplicar misma función)
            ocultos_norm = { _normalizar_nombre_proveedor(r['nombre']) for r in ocultos_rows }
            eliminados_norm = { _normalizar_nombre_proveedor(r['nombre']) for r in eliminados_rows }
            oculto = prov_norm in ocultos_norm
            eliminado = prov_norm in eliminados_norm
            if oculto:
                # Oculto siempre bloquea
                html_vacio = render_template('fragmentos/historial_tabla.html', productos=[])
                return jsonify({'success': True, 'html': html_vacio, 'count': 0})
            if eliminado:
                # Permitir reaparición si hay stock real (reactivación implícita)
                existe_stock = db_query("SELECT 1 FROM stock WHERE LOWER(proveedor)=LOWER(?) AND LOWER(dueno)=? LIMIT 1", (proveedor, dueno.lower()), fetch=True)
                if not existe_stock:
                    html_vacio = render_template('fragmentos/historial_tabla.html', productos=[])
                    return jsonify({'success': True, 'html': html_vacio, 'count': 0})
        if termino:
            like = f"%{termino}%"
            filtros.append('(nombre LIKE ? OR codigo LIKE ? OR proveedor LIKE ? OR observaciones LIKE ?)')
            params.extend([like, like, like, like])
        where_clause = ('WHERE ' + ' AND '.join(filtros)) if filtros else ''
        query = f"SELECT id, fecha_compra, codigo, nombre, proveedor, precio, cantidad, observaciones, precio_texto, avisar_bajo_stock, min_stock_aviso, dueno, created_at FROM stock {where_clause} ORDER BY fecha_compra DESC, id DESC LIMIT 800"  # límite preventivo
        rows = db_query(query, tuple(params), fetch=True) or []
        # --- Sanitización de precios ---
        # Se ha detectado que en algunos casos "precio" puede venir como bytes (ej: b'\x00...')
        # o como cadena vacía, lo que termina mostrándose literalmente en el historial.
        # Aquí normalizamos para que siempre sea un float (0.0 si vacío) y ajustamos precio_texto.
        sanitized_rows = []
        for r in rows:
            try:
                # Convertir a dict editable (sqlite3.Row es inmutable)
                item = dict(r)
            except Exception:
                item = r
            raw_precio = item.get('precio')
            precio_float = None
            if isinstance(raw_precio, (bytes, bytearray)):
                # Intentar decodificar y limpiar bytes nulos
                try:
                    decoded = raw_precio.decode('utf-8', 'ignore')
                except Exception:
                    decoded = ''
                decoded = decoded.replace('\x00', '').strip().strip("b'").strip()
                precio_float, _ = parse_price(decoded)
            elif raw_precio is None:
                precio_float = 0.0
            elif isinstance(raw_precio, (int, float)):
                precio_float = float(raw_precio)
            else:
                # Cualquier otro tipo (str probablemente)
                txt = str(raw_precio).strip()
                # Si la representación comienza con b'... tratar de limpiar
                if txt.startswith("b'") and txt.endswith("'"):
                    inner = txt[2:-1]
                    inner = inner.replace('\x00', '').strip()
                    precio_float, _ = parse_price(inner)
                else:
                    precio_float, _ = parse_price(txt)
            if precio_float is None:
                precio_float = 0.0
            item['precio'] = precio_float
            # Normalizar precio_texto
            pt = item.get('precio_texto')
            if not pt or isinstance(pt, (bytes, bytearray)):
                pt = '' if not pt else (pt.decode('utf-8', 'ignore') if isinstance(pt, (bytes, bytearray)) else str(pt))
            if not pt.strip():
                pt = '0' if precio_float == 0 else str(precio_float)
            # Asegurar consistencia: si precio es 0, mostrar '0'
            if precio_float == 0:
                pt = '0'
            item['precio_texto'] = pt
            sanitized_rows.append(item)

        # Render parcial reutilizando un pequeño fragmento HTML con filas sanitizadas
        html = render_template('fragmentos/historial_tabla.html', productos=sanitized_rows)
        return jsonify({'success': True, 'html': html, 'count': len(rows)})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

# --- Utilidades de Respaldo (para entornos sin consola) ---
@app.route('/export_sqlite')
@login_required
def export_sqlite():
    """Descargar el archivo SQLite principal para respaldo.
    IMPORTANTE: Elimina esta ruta una vez migrado a PostgreSQL.
    """
    try:
        # Determinar nombre (prioridad a gestor_stock.sqlite3)
        for fname in ["gestor_stock.sqlite3", "gestor_stock.db", "stock.db"]:
            if os.path.exists(fname):
                return send_file(fname, as_attachment=True, download_name=fname)
        return jsonify({'success': False, 'error': 'No se encontró archivo .sqlite3 / .db en el contenedor.'}), 404
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/export_sqlite_base64')  # compat heredado
@app.route('/export_stock_csv')
@login_required
def export_stock_csv():
    """Exportar datos de stock en CSV estándar.
    Encabezados: FechaCompra,Codigo,Nombre,Proveedor,Precio,Cantidad,AvisarBajoStock,MinStockAviso,Observaciones,Dueno,CreatedAt
    """
    import csv, io
    try:
        rows = db_query("SELECT fecha_compra,codigo,nombre,proveedor,precio,cantidad,avisar_bajo_stock,min_stock_aviso,observaciones,dueno,created_at FROM stock ORDER BY id ASC", fetch=True) or []
        buff = io.StringIO()
        w = csv.writer(buff)
        w.writerow(['FechaCompra','Codigo','Nombre','Proveedor','Precio','Cantidad','AvisarBajoStock','MinStockAviso','Observaciones','Dueno','CreatedAt'])
        for r in rows:
            w.writerow([
                r.get('fecha_compra') or '',
                r.get('codigo') or '',
                r.get('nombre') or '',
                r.get('proveedor') or '',
                r.get('precio') if r.get('precio') is not None else 0,
                r.get('cantidad') if r.get('cantidad') is not None else 0,
                r.get('avisar_bajo_stock', 0) or 0,
                r.get('min_stock_aviso') if r.get('min_stock_aviso') is not None else '',
                (r.get('observaciones') or '').replace('\n',' ').replace('\r',' '),
                r.get('dueno') or '',
                r.get('created_at') or ''
            ])
        buff.seek(0)
        from flask import Response
        return Response(buff.read(), mimetype='text/csv', headers={'Content-Disposition': 'attachment; filename=stock_export.csv'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/export_stock_history_csv')
@login_required
def export_stock_history_csv():
    """Exportar historial completo (tabla stock_history)."""
    import csv, io
    try:
        conn = get_db_connection()
        if not conn:
            return jsonify({'success': False, 'error': 'Sin conexión BD'}), 500
        cur = conn.cursor()
        cur.execute("SELECT id, stock_id, codigo, nombre, precio, cantidad, fecha_compra, proveedor, observaciones, precio_texto, dueno, created_at, fecha_evento, tipo_cambio, fuente, usuario FROM stock_history ORDER BY id ASC")
        rows = cur.fetchall()
        buff = io.StringIO()
        w = csv.writer(buff)
        w.writerow(["ID","StockID","Codigo","Nombre","Precio","Cantidad","FechaCompra","Proveedor","Observaciones","PrecioTexto","Dueno","CreatedAt","FechaEvento","TipoCambio","Fuente","Usuario"])
        for r in rows:
            w.writerow([
                r[0], r[1] or '', r[2] or '', r[3] or '',
                r[4] if r[4] is not None else '',
                r[5] if r[5] is not None else '',
                r[6] or '', r[7] or '', (r[8] or '').replace('\n',' ').replace('\r',' '), r[9] or '',
                r[10] or '', r[11] or '', r[12] or '', r[13] or '', r[14] or '', r[15] or ''
            ])
        buff.seek(0)
        from flask import Response
        return Response(buff.read(), mimetype='text/csv', headers={'Content-Disposition': 'attachment; filename=stock_history_export.csv'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/export_stock_xlsx')
@login_required
def export_stock_xlsx():
    """Exportar datos de stock en XLSX con columnas de avisos y colores.
    Hoja: Stock
    Columnas: FechaCompra, Codigo, Nombre, Proveedor, Precio, Cantidad, AvisarBajoStock, MinStockAviso, Observaciones, Dueno, CreatedAt
    Colores:
      - Amarillo (FFFEF3C7) si AvisarBajoStock=1
      - Rojo suave (FFF8CBAD) si cantidad <= MinStockAviso y AvisarBajoStock=1 y MinStockAviso válido
    """
    try:
        rows = db_query("SELECT fecha_compra,codigo,nombre,proveedor,precio,cantidad,avisar_bajo_stock,min_stock_aviso,observaciones,dueno,created_at FROM stock ORDER BY id ASC", fetch=True) or []
        wb = Workbook()
        ws = wb.active
        ws.title = "Stock"
        headers = ['FechaCompra','Codigo','Nombre','Proveedor','Precio','Cantidad','AvisarBajoStock','MinStockAviso','Observaciones','Dueno','CreatedAt']
        ws.append(headers)
        for r in rows:
            ws.append([
                r.get('fecha_compra') or '',
                r.get('codigo') or '',
                r.get('nombre') or '',
                r.get('proveedor') or '',
                r.get('precio') if r.get('precio') is not None else 0,
                r.get('cantidad') if r.get('cantidad') is not None else 0,
                r.get('avisar_bajo_stock', 0) or 0,
                r.get('min_stock_aviso') if r.get('min_stock_aviso') is not None else '',
                (r.get('observaciones') or '').replace('\n',' ').replace('\r',' '),
                r.get('dueno') or '',
                r.get('created_at') or ''
            ])
        # Aplicar colores: Verde OK, Amarillo bajo (avisar y cant >0 y cant <= min), Rojo sin stock (cant <=0) o cant<=min con avisar y cant==0
        from openpyxl.styles import PatternFill
        fill_verde = PatternFill(start_color='FFD5F5E3', end_color='FFD5F5E3', fill_type='solid')  # verde suave
        fill_amarillo = PatternFill(start_color='FFFEF3C7', end_color='FFFEF3C7', fill_type='solid')
        fill_rojo = PatternFill(start_color='FFF8CBAD', end_color='FFF8CBAD', fill_type='solid')
        # Índices de columnas (1-based): Cantidad=6, Avisar=7, MinStock=8
        for row in ws.iter_rows(min_row=2, min_col=1, max_col=len(headers)):
            try:
                cant_raw = row[5].value
                avisar = row[6].value
                minimo_raw = row[7].value
                cant_v = None
                min_v = None
                try:
                    cant_v = float(cant_raw) if cant_raw not in (None, '') else None
                except Exception:
                    cant_v = None
                try:
                    min_v = float(minimo_raw) if minimo_raw not in (None, '', 0, '0') else None
                except Exception:
                    min_v = None
                # Decidir color
                use_fill = None
                if cant_v is None:
                    # sin dato cantidad -> no colorear
                    use_fill = None
                else:
                    if cant_v <= 0:
                        use_fill = fill_rojo
                    elif avisar in (1,'1',True) and min_v is not None and cant_v <= min_v:
                        use_fill = fill_amarillo
                    else:
                        use_fill = fill_verde
                if use_fill:
                    for c in row:
                        c.fill = use_fill
            except Exception:
                continue
        from io import BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return send_file(output, as_attachment=True, download_name='stock_export.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/export_stock_history_xlsx')
@login_required
def export_stock_history_xlsx():
    """Exportar historial completo (tabla stock_history) en XLSX.
    Hoja: Historial
    Columnas: ID, StockID, Codigo, Nombre, Precio, Cantidad, FechaCompra, Proveedor, Observaciones, PrecioTexto, Dueno, CreatedAt, FechaEvento, TipoCambio, Fuente, Usuario
    """
    try:
        conn = get_db_connection()
        if not conn:
            return jsonify({'success': False, 'error': 'Sin conexión BD'}), 500
        cur = conn.cursor()
        # Traemos el historial junto con flags/umbral actuales (si existen) para colorear estado.
        try:
            cur.execute("""
                SELECT h.id, h.stock_id, h.codigo, h.nombre, h.precio, h.cantidad, h.fecha_compra,
                       h.proveedor, h.observaciones, h.precio_texto, h.dueno, h.created_at,
                       h.fecha_evento, h.tipo_cambio, h.fuente, h.usuario,
                       s.avisar_bajo_stock, s.min_stock_aviso
                FROM stock_history h
                LEFT JOIN stock s ON s.id = h.stock_id
                ORDER BY h.id ASC
            """)
            rows = cur.fetchall()
            has_flags = True
        except Exception:
            # Fallback si no existen columnas/ join falla
            cur.execute("SELECT id, stock_id, codigo, nombre, precio, cantidad, fecha_compra, proveedor, observaciones, precio_texto, dueno, created_at, fecha_evento, tipo_cambio, fuente, usuario FROM stock_history ORDER BY id ASC")
            rows = cur.fetchall()
            has_flags = False
        wb = Workbook()
        ws = wb.active
        ws.title = "Historial"
        headers = ["ID","StockID","Codigo","Nombre","Precio","Cantidad","FechaCompra","Proveedor","Observaciones","PrecioTexto","Dueno","CreatedAt","FechaEvento","TipoCambio","Fuente","Usuario"]
        ws.append(headers)
        from openpyxl.styles import PatternFill
        fill_verde = PatternFill(start_color='FFD5F5E3', end_color='FFD5F5E3', fill_type='solid')
        fill_amarillo = PatternFill(start_color='FFFEF3C7', end_color='FFFEF3C7', fill_type='solid')
        fill_rojo = PatternFill(start_color='FFF8CBAD', end_color='FFF8CBAD', fill_type='solid')

        for r in rows:
            ws.append([
                r[0],
                r[1] or '',
                r[2] or '',
                r[3] or '',
                r[4] if r[4] is not None else '',
                r[5] if r[5] is not None else '',
                r[6] or '',
                r[7] or '',
                (r[8] or '').replace('\n',' ').replace('\r',' '),
                r[9] or '',
                r[10] or '',
                r[11] or '',
                r[12] or '',
                r[13] or '',
                r[14] or '',
                r[15] or ''
            ])
            # Colorear la fila recién añadida usando la cantidad histórica y umbral actual (si existe)
            try:
                cant_hist = r[5]
                avisar_flag = r[16] if has_flags and len(r) > 16 else None
                min_flag = r[17] if has_flags and len(r) > 17 else None
                cant_v = float(cant_hist) if cant_hist not in (None, '') else None
                min_v = float(min_flag) if (min_flag not in (None, '', 0, '0') and avisar_flag in (1,'1',True)) else None
                use_fill = None
                if cant_v is not None:
                    if cant_v <= 0:
                        use_fill = fill_rojo
                    elif avisar_flag in (1,'1',True) and min_v is not None and cant_v <= min_v:
                        use_fill = fill_amarillo
                    else:
                        use_fill = fill_verde
                if use_fill:
                    for c in ws[ws.max_row]:
                        c.fill = use_fill
            except Exception:
                pass
        from io import BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        try:
            cur.close()
            conn.close()
        except Exception:
            pass
        return send_file(output, as_attachment=True, download_name='stock_history_export.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500
@app.route('/import_stock_csv', methods=['POST'])
@login_required
def import_stock_csv():
    """Importar un CSV previamente exportado (stock_export.csv) con opciones:
    Parámetros (form-data):
      - file: archivo CSV.
      - dry_run=1 (opcional) -> No aplica cambios, solo muestra resumen.
      - existing_mode in [update, skip] (default update) -> Qué hacer si existe fila.
    Estrategia base:
      - Espera columnas: FechaCompra,Codigo,Nombre,Proveedor,Precio,Cantidad,Observaciones,Dueno,CreatedAt
      - Clave de coincidencia: (codigo, proveedor, dueno) ignorando mayúsculas.
      - INSERT si no existe; si existe y existing_mode=update -> UPDATE; si skip -> se cuenta skipped_exist.
      - Log a stock_history: import (nuevo), import_update (update efectivo)
    Backup: si no es dry_run genera un CSV de respaldo previo en carpeta 'backups' exportando estado actual.
    Devuelve JSON con métricas.
    """
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'error': 'No se envió archivo'}), 400
        f = request.files['file']
        if not f.filename.lower().endswith('.csv'):
            return jsonify({'success': False, 'error': 'Debe ser .csv'}), 400
        dry_run = request.form.get('dry_run') == '1'
        existing_mode = request.form.get('existing_mode', 'update')
        if existing_mode not in ('update','skip'):
            return jsonify({'success': False, 'error': 'existing_mode inválido'}), 400
        import csv, io
        content = f.read().decode('utf-8', errors='replace')
        reader = csv.reader(io.StringIO(content))
        header = next(reader, None)
        if not header:
            return jsonify({'success': False, 'error': 'CSV vacío'}), 400
        # Columnas posibles (backup completo con avisos) y requeridas mínimas para un backup viejo.
        full_cols = ['FechaCompra','Codigo','Nombre','Proveedor','Precio','Cantidad','AvisarBajoStock','MinStockAviso','Observaciones','Dueno','CreatedAt']
        minimal_required = ['Codigo']  # lo único estrictamente necesario para intentar importar
        col_index = {}
        for i, raw in enumerate(header):
            clean = (raw or '').strip()
            col_index[clean.lower()] = i
        missing_min = [c for c in minimal_required if c.lower() not in col_index]
        if missing_min:
            return jsonify({'success': False, 'error': f'Faltan columnas esenciales: {missing_min}'}), 400
        has_dueno_col = 'dueno' in col_index
        has_avisar_col = 'avisarbajostock' in col_index
        has_min_col = 'minstockaviso' in col_index
        has_created_col = 'createdat' in col_index
        default_dueno = (request.form.get('default_dueno') or '').strip().lower() or 'general'
        conn = get_db_connection()
        if not conn:
            return jsonify({'success': False, 'error': 'Sin conexión BD'}), 500
        cur = conn.cursor()
        use_postgres = _is_postgres_configured()
        # Backup previo si no es dry_run
        backup_filename = None
        if not dry_run:
            try:
                import os, datetime as _dt
                os.makedirs('backups', exist_ok=True)
                ts = _dt.datetime.utcnow().strftime('%Y%m%d_%H%M%S')
                backup_filename = f'backups/stock_pre_import_{ts}.csv'
                bcur = conn.cursor()
                bcur.execute("SELECT fecha_compra,codigo,nombre,proveedor,precio,cantidad,avisar_bajo_stock,min_stock_aviso,observaciones,dueno,created_at FROM stock")
                rows_b = bcur.fetchall()
                with open(backup_filename,'w',encoding='utf-8',newline='') as bf:
                    bw = csv.writer(bf)
                    bw.writerow(full_cols)
                    for rb in rows_b:
                        bw.writerow([rb[0], rb[1], rb[2], rb[3], rb[4], rb[5], rb[6], rb[7], rb[8], rb[9], rb[10]])
            except Exception as e_bk:
                print(f"[WARN] Backup previo falló: {e_bk}")
        inserted = 0
        updated = 0
        skipped = 0
        skipped_exist = 0
        errores = []
        history_batch = []  # acumularemos tuplas para inserción masiva en stock_history
        from flask_login import current_user as _cu
        usuario_actual = None
        try:
            usuario_actual = getattr(_cu, 'username', None) or getattr(_cu, 'id', None) or 'import'
        except Exception:
            usuario_actual = 'import'
        for line_no, row in enumerate(reader, start=2):
            def take(col):
                idx = col_index.get(col.lower())
                if idx is None or idx >= len(row):
                    return ''
                return (row[idx] or '').strip()
            fecha_compra = take('FechaCompra')
            codigo = take('Codigo')
            nombre = take('Nombre') or ''
            proveedor = take('Proveedor')
            precio_raw = take('Precio')
            cantidad_raw = take('Cantidad')
            avisar_raw = take('AvisarBajoStock') if has_avisar_col else ''
            min_raw = take('MinStockAviso') if has_min_col else ''
            observ = take('Observaciones')
            dueno = take('Dueno') if has_dueno_col else default_dueno
            created_at = take('CreatedAt') if has_created_col else ''
            if not codigo and not nombre:
                skipped += 1
                continue
            try:
                precio_val = 0.0
                if precio_raw:
                    try:
                        precio_val = float(str(precio_raw).replace(',', '.'))
                    except Exception:
                        precio_val = 0.0
                cantidad_val = 0
                if cantidad_raw:
                    try:
                        cantidad_val = int(float(cantidad_raw))
                    except Exception:
                        cantidad_val = 0
                # Flags avisos
                avisar_flag = 0
                try:
                    if avisar_raw not in (None, '', '0'):
                        avisar_flag = 1 if str(avisar_raw).strip() in ('1','true','True','YES','yes') else int(float(avisar_raw))
                except Exception:
                    avisar_flag = 0
                min_val_flag = None
                if min_raw not in (None, '', '0'):
                    try:
                        mv = int(float(min_raw))
                        if mv > 0:
                            min_val_flag = mv
                    except Exception:
                        min_val_flag = None
                # Coherencia: si no hay umbral válido, desactivar flag
                if not min_val_flag or min_val_flag <= 0:
                    avisar_flag = 0
                    min_val_flag = None
                # Claves normalizadas
                codigo_l = (codigo or '').lower()
                proveedor_l = (proveedor or '').lower()
                dueno_l = (dueno or '').lower() if dueno else ''
                # Buscar existente
                if use_postgres:
                    sel_sql = "SELECT * FROM stock WHERE LOWER(codigo)=%s AND COALESCE(LOWER(proveedor),'')=%s AND COALESCE(LOWER(dueno),'')=%s ORDER BY id DESC LIMIT 1"
                    cur.execute(sel_sql, (codigo_l, proveedor_l, dueno_l))
                else:
                    sel_sql = "SELECT * FROM stock WHERE LOWER(codigo)=? AND COALESCE(LOWER(proveedor),'')=? AND COALESCE(LOWER(dueno),'')=? ORDER BY id DESC LIMIT 1"
                    cur.execute(sel_sql, (codigo_l, proveedor_l, dueno_l))
                row_exist = cur.fetchone()
                if row_exist:
                    if existing_mode == 'skip':
                        skipped_exist += 1
                    else:
                        if not dry_run:
                            if use_postgres:
                                upd_sql = ("UPDATE stock SET nombre=%s, precio=%s, cantidad=%s, fecha_compra=%s, proveedor=%s, observaciones=%s, precio_texto=%s, dueno=%s, avisar_bajo_stock=%s, min_stock_aviso=%s WHERE id=%s")
                                cur.execute(upd_sql, (nombre, precio_val, cantidad_val, fecha_compra or datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S'), proveedor or None, observ, str(precio_val), dueno_l or None, avisar_flag, min_val_flag, row_exist[0]))
                            else:
                                upd_sql = ("UPDATE stock SET nombre=?, precio=?, cantidad=?, fecha_compra=?, proveedor=?, observaciones=?, precio_texto=?, dueno=?, avisar_bajo_stock=?, min_stock_aviso=? WHERE id=?")
                                cur.execute(upd_sql, (nombre, precio_val, cantidad_val, fecha_compra or datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S'), proveedor or None, observ, str(precio_val), dueno_l or None, avisar_flag, min_val_flag, row_exist[0]))
                        updated += 1
                        if not dry_run:
                            # Preparar registro histórico en batch
                            fecha_evento = datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')
                            history_batch.append((
                                row_exist[0], codigo, nombre, precio_val, cantidad_val, fecha_compra or '', proveedor or '', observ, str(precio_val), dueno_l or '', created_at or '',
                                fecha_evento, 'import_update', 'import_stock_csv', usuario_actual
                            ))
                else:
                    if not dry_run:
                        if use_postgres:
                            ins_sql = ("INSERT INTO stock (codigo,nombre,precio,cantidad,fecha_compra,proveedor,observaciones,precio_texto,dueno,created_at,avisar_bajo_stock,min_stock_aviso) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id")
                            cur.execute(ins_sql, (codigo, nombre, precio_val, cantidad_val, fecha_compra or datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S'), proveedor or None, observ, str(precio_val), dueno_l or None, created_at or datetime.utcnow().isoformat(), avisar_flag, min_val_flag))
                            new_id = cur.fetchone()[0]
                        else:
                            ins_sql = ("INSERT INTO stock (codigo,nombre,precio,cantidad,fecha_compra,proveedor,observaciones,precio_texto,dueno,created_at,avisar_bajo_stock,min_stock_aviso) VALUES (?,?,?,?,?,?,?,?,?,?,?,?)")
                            cur.execute(ins_sql, (codigo, nombre, precio_val, cantidad_val, fecha_compra or datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S'), proveedor or None, observ, str(precio_val), dueno_l or None, created_at or datetime.utcnow().isoformat(), avisar_flag, min_val_flag))
                            new_id = cur.lastrowid
                    else:
                        new_id = -1  # marcador
                    inserted += 1
                    if not dry_run:
                        fecha_evento = datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')
                        history_batch.append((
                            new_id, codigo, nombre, precio_val, cantidad_val, fecha_compra or '', proveedor or '', observ, str(precio_val), dueno_l or '', created_at or '',
                            fecha_evento, 'import', 'import_stock_csv', usuario_actual
                        ))
            except Exception as ex_row:
                errores.append(f'L{line_no}: {ex_row}')
        if dry_run:
            conn.rollback()
        else:
            # Insert batch histórico antes del commit
            if history_batch:
                if use_postgres:
                    cur.executemany("INSERT INTO stock_history (stock_id,codigo,nombre,precio,cantidad,fecha_compra,proveedor,observaciones,precio_texto,dueno,created_at,fecha_evento,tipo_cambio,fuente,usuario) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)", history_batch)
                else:
                    cur.executemany("INSERT INTO stock_history (stock_id,codigo,nombre,precio,cantidad,fecha_compra,proveedor,observaciones,precio_texto,dueno,created_at,fecha_evento,tipo_cambio,fuente,usuario) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)", history_batch)
            conn.commit()
        return jsonify({'success': True, 'inserted': inserted, 'updated': updated, 'skipped': skipped, 'skipped_existentes': skipped_exist, 'dry_run': dry_run, 'existing_mode': existing_mode, 'backup': backup_filename, 'errores': errores[:15], 'errores_total': len(errores)})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/import_stock_xlsx', methods=['POST'])
@login_required
def import_stock_xlsx():
    """Importar un XLSX previamente exportado (stock_export.xlsx) con mismas reglas que CSV.
    Parámetros (form-data):
      - file: archivo XLSX.
      - dry_run=1 opcional.
      - existing_mode in [update, skip].
        Columnas esperadas (headers hoja activa, primera fila):
            FechaCompra, Codigo, Nombre, Proveedor, Precio, Cantidad, AvisarBajoStock, MinStockAviso, Observaciones, Dueno, CreatedAt
    Tolerancia: si faltan Dueno/CreatedAt se aplican defaults como en CSV.
    """
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'error': 'No se envió archivo'}), 400
        f = request.files['file']
        if not f.filename.lower().endswith('.xlsx'):
            return jsonify({'success': False, 'error': 'Debe ser .xlsx'}), 400
        dry_run = request.form.get('dry_run') == '1'
        existing_mode = request.form.get('existing_mode', 'update')
        if existing_mode not in ('update','skip'):
            return jsonify({'success': False, 'error': 'existing_mode inválido'}), 400
        from openpyxl import load_workbook
        from io import BytesIO
        data_bytes = f.read()
        try:
            wb = load_workbook(BytesIO(data_bytes), read_only=True, data_only=True)
        except Exception as ex:
            return jsonify({'success': False, 'error': f'No se pudo leer XLSX: {ex}'}), 400
        ws = wb.active
        header = []
        for cell in ws[1]:
            header.append((cell.value or '').strip() if isinstance(cell.value,str) else (cell.value if cell.value is not None else ''))
        if not header:
            return jsonify({'success': False, 'error': 'XLSX sin encabezado'}), 400
        full_cols = ['FechaCompra','Codigo','Nombre','Proveedor','Precio','Cantidad','AvisarBajoStock','MinStockAviso','Observaciones','Dueno','CreatedAt']
        minimal_required = ['Codigo']
        col_index = {}  # map lower->pos
        for i, raw in enumerate(header):
            clean = (str(raw) if raw is not None else '').strip()
            col_index[clean.lower()] = i
        missing_min = [c for c in minimal_required if c.lower() not in col_index]
        if missing_min:
            return jsonify({'success': False, 'error': f'Faltan columnas esenciales: {missing_min}'}), 400
        has_dueno_col = 'dueno' in col_index
        has_avisar_col = 'avisarbajostock' in col_index
        has_min_col = 'minstockaviso' in col_index
        has_created_col = 'createdat' in col_index
        default_dueno = (request.form.get('default_dueno') or '').strip().lower() or 'general'
        conn = get_db_connection()
        if not conn:
            return jsonify({'success': False, 'error': 'Sin conexión BD'}), 500
        cur = conn.cursor()
        use_postgres = _is_postgres_configured()
        backup_filename = None
        if not dry_run:
            try:
                import os, datetime as _dt, csv
                os.makedirs('backups', exist_ok=True)
                ts = _dt.datetime.utcnow().strftime('%Y%m%d_%H%M%S')
                backup_filename = f'backups/stock_pre_import_{ts}.csv'
                bcur = conn.cursor()
                bcur.execute("SELECT fecha_compra,codigo,nombre,proveedor,precio,cantidad,avisar_bajo_stock,min_stock_aviso,observaciones,dueno,created_at FROM stock")
                rows_b = bcur.fetchall()
                with open(backup_filename,'w',encoding='utf-8',newline='') as bf:
                    bw = csv.writer(bf)
                    bw.writerow(full_cols)
                    for rb in rows_b:
                        bw.writerow([rb[0], rb[1], rb[2], rb[3], rb[4], rb[5], rb[6], rb[7], rb[8], rb[9], rb[10]])
            except Exception as e_bk:
                print(f"[WARN] Backup previo XLSX falló: {e_bk}")
        inserted = 0
        updated = 0
        skipped = 0
        skipped_exist = 0
        errores = []
        history_batch = []
        usuario_actual = 'import'
        try:
            from flask_login import current_user as _cu
            usuario_actual = getattr(_cu, 'username', None) or getattr(_cu, 'id', None) or 'import'
        except Exception:
            pass
        # Iterar filas (read_only: usar ws.iter_rows)
        for line_no, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if row is None:
                continue
            def take(col_name):
                idx = col_index.get(col_name.lower())
                if idx is None:
                    return ''
                try:
                    val = row[idx]
                    return str(val).strip() if val is not None else ''
                except Exception:
                    return ''
            fecha_compra = take('FechaCompra')
            codigo = take('Codigo')
            nombre = take('Nombre') or ''
            proveedor = take('Proveedor')
            precio_raw = take('Precio')
            cantidad_raw = take('Cantidad')
            avisar_raw = take('AvisarBajoStock') if has_avisar_col else ''
            min_raw = take('MinStockAviso') if has_min_col else ''
            observ = take('Observaciones')
            dueno = take('Dueno') if has_dueno_col else default_dueno
            created_at = take('CreatedAt') if has_created_col else ''
            if not codigo and not nombre:
                skipped += 1
                continue
            try:
                # Precio
                precio_val = 0.0
                if precio_raw:
                    try:
                        precio_val = float(str(precio_raw).replace(',', '.'))
                    except Exception:
                        precio_val = 0.0
                cantidad_val = 0
                if cantidad_raw:
                    try:
                        cantidad_val = int(float(cantidad_raw))
                    except Exception:
                        cantidad_val = 0
                # Flags avisos
                avisar_flag = 0
                try:
                    if avisar_raw not in (None, '', '0'):
                        avisar_flag = 1 if str(avisar_raw).strip() in ('1','true','True','YES','yes') else int(float(avisar_raw))
                except Exception:
                    avisar_flag = 0
                min_val_flag = None
                if min_raw not in (None, '', '0'):
                    try:
                        mv = int(float(min_raw))
                        if mv > 0:
                            min_val_flag = mv
                    except Exception:
                        min_val_flag = None
                if not min_val_flag or min_val_flag <= 0:
                    avisar_flag = 0
                    min_val_flag = None
                codigo_l = (codigo or '').lower()
                proveedor_l = (proveedor or '').lower()
                dueno_l = (dueno or '').lower() if dueno else ''
                if use_postgres:
                    sel_sql = "SELECT * FROM stock WHERE LOWER(codigo)=%s AND COALESCE(LOWER(proveedor),'')=%s AND COALESCE(LOWER(dueno),'')=%s ORDER BY id DESC LIMIT 1"
                    cur.execute(sel_sql, (codigo_l, proveedor_l, dueno_l))
                else:
                    sel_sql = "SELECT * FROM stock WHERE LOWER(codigo)=? AND COALESCE(LOWER(proveedor),'')=? AND COALESCE(LOWER(dueno),'')=? ORDER BY id DESC LIMIT 1"
                    cur.execute(sel_sql, (codigo_l, proveedor_l, dueno_l))
                row_exist = cur.fetchone()
                if row_exist:
                    if existing_mode == 'skip':
                        skipped_exist += 1
                    else:
                        if not dry_run:
                            if use_postgres:
                                upd_sql = ("UPDATE stock SET nombre=%s, precio=%s, cantidad=%s, fecha_compra=%s, proveedor=%s, observaciones=%s, precio_texto=%s, dueno=%s, avisar_bajo_stock=%s, min_stock_aviso=%s WHERE id=%s")
                                cur.execute(upd_sql, (nombre, precio_val, cantidad_val, fecha_compra or datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S'), proveedor or None, observ, str(precio_val), dueno_l or None, avisar_flag, min_val_flag, row_exist[0]))
                            else:
                                upd_sql = ("UPDATE stock SET nombre=?, precio=?, cantidad=?, fecha_compra=?, proveedor=?, observaciones=?, precio_texto=?, dueno=?, avisar_bajo_stock=?, min_stock_aviso=? WHERE id=?")
                                cur.execute(upd_sql, (nombre, precio_val, cantidad_val, fecha_compra or datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S'), proveedor or None, observ, str(precio_val), dueno_l or None, avisar_flag, min_val_flag, row_exist[0]))
                        updated += 1
                        if not dry_run:
                            fecha_evento = datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')
                            history_batch.append((
                                row_exist[0], codigo, nombre, precio_val, cantidad_val, fecha_compra or '', proveedor or '', observ, str(precio_val), dueno_l or '', created_at or '',
                                fecha_evento, 'import_update', 'import_stock_xlsx', usuario_actual
                            ))
                else:
                    if not dry_run:
                        if use_postgres:
                            ins_sql = ("INSERT INTO stock (codigo,nombre,precio,cantidad,fecha_compra,proveedor,observaciones,precio_texto,dueno,created_at,avisar_bajo_stock,min_stock_aviso) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id")
                            cur.execute(ins_sql, (codigo, nombre, precio_val, cantidad_val, fecha_compra or datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S'), proveedor or None, observ, str(precio_val), dueno_l or None, created_at or datetime.utcnow().isoformat(), avisar_flag, min_val_flag))
                            new_id = cur.fetchone()[0]
                        else:
                            ins_sql = ("INSERT INTO stock (codigo,nombre,precio,cantidad,fecha_compra,proveedor,observaciones,precio_texto,dueno,created_at,avisar_bajo_stock,min_stock_aviso) VALUES (?,?,?,?,?,?,?,?,?,?,?,?)")
                            cur.execute(ins_sql, (codigo, nombre, precio_val, cantidad_val, fecha_compra or datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S'), proveedor or None, observ, str(precio_val), dueno_l or None, created_at or datetime.utcnow().isoformat(), avisar_flag, min_val_flag))
                            new_id = cur.lastrowid
                    else:
                        new_id = -1
                    inserted += 1
                    if not dry_run:
                        fecha_evento = datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')
                        history_batch.append((
                            new_id, codigo, nombre, precio_val, cantidad_val, fecha_compra or '', proveedor or '', observ, str(precio_val), dueno_l or '', created_at or '',
                            fecha_evento, 'import', 'import_stock_xlsx', usuario_actual
                        ))
            except Exception as ex_row:
                errores.append(f'L{line_no}: {ex_row}')
        if dry_run:
            conn.rollback()
        else:
            if history_batch:
                if use_postgres:
                    cur.executemany("INSERT INTO stock_history (stock_id,codigo,nombre,precio,cantidad,fecha_compra,proveedor,observaciones,precio_texto,dueno,created_at,fecha_evento,tipo_cambio,fuente,usuario) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)", history_batch)
                else:
                    cur.executemany("INSERT INTO stock_history (stock_id,codigo,nombre,precio,cantidad,fecha_compra,proveedor,observaciones,precio_texto,dueno,created_at,fecha_evento,tipo_cambio,fuente,usuario) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)", history_batch)
            conn.commit()
        return jsonify({'success': True, 'inserted': inserted, 'updated': updated, 'skipped': skipped, 'skipped_existentes': skipped_exist, 'dry_run': dry_run, 'existing_mode': existing_mode, 'backup': backup_filename, 'errores': errores[:15], 'errores_total': len(errores)})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/debug_routes')
@login_required
def debug_routes():
    """Listar rutas registradas para diagnosticar en producción."""
    rutas = []
    for rule in app.url_map.iter_rules():
        rutas.append({
            'endpoint': rule.endpoint,
            'methods': sorted(m for m in rule.methods if m not in ('HEAD', 'OPTIONS')),
            'rule': str(rule)
        })
    rutas = sorted(rutas, key=lambda r: r['rule'])
    return jsonify({'success': True, 'routes': rutas, 'count': len(rutas)})

@app.route('/historial_proveedores', methods=['POST'])
@login_required
def historial_proveedores():
    """Devolver lista de proveedores distintos para un dueño (case-insensitive)."""
    try:
        data = request.get_json() or {}
        dueno = (data.get('dueno') or '').strip().lower()
        if not dueno:
            return jsonify({'success': True, 'proveedores': []})

        # Cargar listas base
        ocultos_rows = db_query("SELECT nombre FROM proveedores_ocultos WHERE dueno=?", (dueno,), fetch=True) or []
        eliminados_rows = db_query("SELECT nombre FROM proveedores_eliminados WHERE dueno=?", (dueno,), fetch=True) or []
        ocultos_norm = { _normalizar_nombre_proveedor(r['nombre']) for r in ocultos_rows }
        eliminados_norm = { _normalizar_nombre_proveedor(r['nombre']) for r in eliminados_rows }

        proveedores_raw = set()

        # STOCK (Regla especial: si hay stock real, permitimos que un proveedor "eliminado" reaparezca automáticamente.
        # Sólo se bloquean los que están ocultos explícitamente, no los eliminados, para que nueva actividad lo reactive.)
        q_rows = db_query("SELECT DISTINCT proveedor FROM stock WHERE proveedor IS NOT NULL AND TRIM(proveedor)!='' AND LOWER(dueno)=?", (dueno,), fetch=True) or []
        for r in q_rows:
            prov = (r.get('proveedor') or '').strip()
            if not prov:
                continue
            norm = _normalizar_nombre_proveedor(prov)
            if norm in ocultos_norm:  # oculto sigue siendo bloqueo fuerte
                continue
            # Si está eliminado, lo reactivamos implícitamente: limpiar de eliminados para este render
            if norm in eliminados_norm:
                eliminados_norm.discard(norm)
            proveedores_raw.add(prov)

        # META
        meta_rows = db_query("SELECT DISTINCT nombre FROM proveedores_meta WHERE dueno=?", (dueno,), fetch=True) or []
        for r in meta_rows:
            prov = (r.get('nombre') or '').strip()
            if not prov:
                continue
            if _normalizar_nombre_proveedor(prov) in ocultos_norm or _normalizar_nombre_proveedor(prov) in eliminados_norm:
                continue
            proveedores_raw.add(prov)

        # EXCEL MANUAL
        try:
            if os.path.exists(MANUAL_PRODUCTS_FILE):
                import pandas as pd
                dfp = pd.read_excel(MANUAL_PRODUCTS_FILE)
                dfp.rename(columns={'Código': 'Codigo', 'Dueño': 'Dueno'}, inplace=True)
                if not dfp.empty:
                    dfp = dfp[dfp['Dueno'].astype(str).str.lower() == dueno]
                    for v in dfp['Proveedor'].dropna().unique():
                        s = str(v).strip()
                        if not s:
                            continue
                        if _normalizar_nombre_proveedor(s) in ocultos_norm or _normalizar_nombre_proveedor(s) in eliminados_norm:
                            continue
                        proveedores_raw.add(s)
        except Exception:
            pass

        # CONFIG
        cfg = DUENOS_CONFIG.get(dueno, {})
        for prov_key in cfg.get('proveedores_excel', []) or []:
            if not prov_key:
                continue
            if any(p.lower() == prov_key.lower() for p in proveedores_raw):
                continue
            # Comprobar existencia en alguna fuente para justificar mostrarlo
            existe_stock = db_query("SELECT 1 FROM stock WHERE LOWER(proveedor)=LOWER(?) AND LOWER(dueno)=? LIMIT 1", (prov_key, dueno), fetch=True)
            existe_meta = db_query("SELECT 1 FROM proveedores_meta WHERE LOWER(nombre)=LOWER(?) AND dueno=? LIMIT 1", (prov_key, dueno), fetch=True)
            existe_manual = False
            try:
                if os.path.exists(MANUAL_PRODUCTS_FILE):
                    import pandas as pd
                    dfx = pd.read_excel(MANUAL_PRODUCTS_FILE)
                    dfx.rename(columns={'Código': 'Codigo', 'Dueño': 'Dueno'}, inplace=True)
                    if not dfx.empty:
                        existe_manual = (
                            (dfx['Dueno'].astype(str).str.lower() == dueno) &
                            (dfx['Proveedor'].astype(str).str.lower() == prov_key.lower())
                        ).any()
            except Exception:
                pass
            if (existe_stock or existe_meta or existe_manual):
                if _normalizar_nombre_proveedor(prov_key) in ocultos_norm or _normalizar_nombre_proveedor(prov_key) in eliminados_norm:
                    continue
                proveedores_raw.add(prov_key)

        # Normalizar visualmente (estética) manteniendo exclusión
        normalizados = set()
        for p in proveedores_raw:
            if p.islower() and '_' not in p and p.replace(' ', '') == p:
                normalizados.add(p.title())
            else:
                if '_' in p and p == p.lower():
                    normalizados.add(p.replace('_', ' ').title())
                else:
                    normalizados.add(p)
        proveedores_lista = sorted(normalizados, key=lambda x: x.lower())

        leak_eliminados = [p for p in proveedores_lista if _normalizar_nombre_proveedor(p) in eliminados_norm]
        if leak_eliminados:
            print(f"[DEBUG] leak eliminados previo filtrado historial_proveedores dueno={dueno}: {leak_eliminados}")

        proveedores_filtrados = [p for p in proveedores_lista if _normalizar_nombre_proveedor(p) not in ocultos_norm and _normalizar_nombre_proveedor(p) not in eliminados_norm]
        removed = len(proveedores_lista) - len(proveedores_filtrados)
        if removed:
            print(f"[DEBUG] Filtrado final historial_proveedores removio {removed} (ocultos/eliminados) para dueno={dueno}")
        return jsonify({'success': True, 'proveedores': proveedores_filtrados})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/stock')
@login_required
def stock():
    termino = request.args.get('q', '')
    if termino:
        like_pattern = f'%{termino}%'
        productos = db_query(
            "SELECT id, codigo, nombre, precio, cantidad, proveedor, observaciones FROM stock WHERE nombre LIKE ? OR codigo LIKE ? OR proveedor LIKE ? ORDER BY nombre",
            (like_pattern, like_pattern, like_pattern),
            fetch=True
        )
    else:
        productos = db_query("SELECT id, codigo, nombre, precio, cantidad, proveedor, observaciones FROM stock ORDER BY nombre", fetch=True)
    return render_template('stock.html', productos=productos, termino_busqueda=termino)

@app.context_processor
def inject_notificacion_emergente():
    """Proveer flags para mostrar avisos una sola vez por nueva notificación."""
    mostrar_aviso = False
    try:
        user_id = session.get('user_id')
        if user_id:
            filas = db_query("SELECT COUNT(1) as c FROM notificaciones WHERE (user_id IS NULL OR user_id=?) AND leida=0", (user_id,), fetch=True)
            count_no_leidas = filas[0]['c'] if filas else 0
            if count_no_leidas > 0 and not session.get('notificacion_aviso_mostrado'):
                mostrar_aviso = True
                session['notificacion_aviso_mostrado'] = True
    except Exception as e:
        print(f"[WARN] inject_notificacion_emergente fallo: {e}")
    return dict(notificacion_emergente=None, mostrar_aviso_notificaciones=mostrar_aviso)

@app.route('/notificaciones')
@login_required
def notificaciones():
    user_id = session.get('user_id')
    # Limpiamos las notificaciones de la sesión ya que ahora usamos solo la BD
    # Esto evita duplicación de notificaciones
    if session.get('notificaciones'):
        session['notificaciones'] = []
        session['notificaciones_leidas'] = True
    filas = db_query("SELECT id,codigo,nombre,proveedor,mensaje,ts,leida FROM notificaciones WHERE user_id IS NULL OR user_id=? ORDER BY ts DESC", (user_id,), fetch=True) or []
    # Debug: imprimir primeras filas y detectar si alguna no tiene id
    try:
        print(f"[DEBUG] notificaciones: total={len(filas)} sample={filas[:3]}")
    except Exception:
        pass
    # Filtrar cualquier entrada anómala sin 'id'
    filas_filtradas = [f for f in filas if 'id' in f and f['id'] is not None]
    if len(filas_filtradas) != len(filas):
        print(f"[WARN] Filtrando {len(filas) - len(filas_filtradas)} notificaciones sin id antes de render")
    filas = filas_filtradas
    # Marcar en memoria si hay todas leídas
    leidas = all(f.get('leida') for f in filas)
    return render_template('notificaciones.html', notificaciones=filas, leidas=leidas)

@app.route('/borrar_notificacion/<int:idx>', methods=['POST'])
@login_required
def borrar_notificacion(idx):
    try:
        # idx ahora es id real de la tabla, mantenemos compat previa si viene el index
        # Intentar borrar por id directo
        db_query("DELETE FROM notificaciones WHERE id = ?", (idx,))
    except Exception as e:
        print(f"[WARN] borrar_notificacion: {e}")
    return redirect(url_for('notificaciones'))

@app.route('/borrar_todas_notificaciones', methods=['POST'])
@login_required
def borrar_todas_notificaciones():
    try:
        user_id = session.get('user_id')
        if user_id:
            db_query("DELETE FROM notificaciones WHERE user_id=?", (user_id,))
        else:
            db_query("DELETE FROM notificaciones WHERE user_id IS NULL")
    except Exception as e:
        print(f"[WARN] borrar_todas_notificaciones: {e}")
    return redirect(url_for('notificaciones'))

@app.route('/marcar_notificaciones_leidas', methods=['POST'])
@login_required
def marcar_notificaciones_leidas():
    try:
        user_id = session.get('user_id')
        if user_id:
            db_query("UPDATE notificaciones SET leida=1 WHERE user_id=?", (user_id,))
        else:
            db_query("UPDATE notificaciones SET leida=1 WHERE user_id IS NULL")
    except Exception as e:
        print(f"[WARN] marcar_notificaciones_leidas: {e}")
    return redirect(url_for('notificaciones'))

# --- Rutas principales de la aplicación ---

@app.route('/agregar_producto', methods=['GET', 'POST'])
@login_required
def agregar_producto():
    # Inicializar carrito si no existe
    if 'carrito' not in session:
        session['carrito'] = []
    
    carrito = session['carrito']
    
    # Manejar formulario manual para agregar al Excel (no directamente al stock)
    if request.method == 'POST' and 'agregar_producto_manual_excel' in request.form:
        return agregar_producto_manual_excel()
    
    # Manejar acciones del carrito
    if request.method == 'POST':
        # Soportar JSON (AJAX) y form tradicional
        if request.is_json:
            data = request.get_json(silent=True) or {}
            if data.get('cargar_carrito'):
                productos_agregados = 0
                dueno_dest = data.get('dueno_dest', 'ferreteria_general')
                for item in carrito:
                    try:
                        avisar_bajo_stock = int(item.get('avisar_bajo_stock', 0))
                        min_stock_aviso_val = item.get('min_stock_aviso')
                        try:
                            min_stock_aviso_int = int(min_stock_aviso_val) if min_stock_aviso_val is not None and str(min_stock_aviso_val) != '' else None
                        except (ValueError, TypeError):
                            min_stock_aviso_int = None
                        cantidad_int = int(item.get('cantidad', 1))
                        if avisar_bajo_stock and (min_stock_aviso_int is None or min_stock_aviso_int <= 0):
                            avisar_bajo_stock = 0
                            min_stock_aviso_int = None
                        if avisar_bajo_stock and min_stock_aviso_int and min_stock_aviso_int > cantidad_int:
                            min_stock_aviso_int = cantidad_int
                        db_query(
                            """
                            INSERT INTO stock (codigo, nombre, precio, cantidad, fecha_compra, 
                                             proveedor, observaciones, precio_texto, avisar_bajo_stock, min_stock_aviso, dueno)
                            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                            """,
                            (
                                item.get('codigo', ''),
                                item.get('nombre', ''),
                                float(item.get('precio', 0)),
                                cantidad_int,
                                item.get('fecha_compra', datetime.now().strftime('%Y-%m-%d')),
                                item.get('proveedor', ''),
                                item.get('observaciones', ''),
                                item.get('precio_texto', ''),
                                avisar_bajo_stock,
                                min_stock_aviso_int,
                                dueno_dest,
                            ),
                        )
                        productos_agregados += 1
                    except Exception as e:
                        print(f"Error al agregar producto al stock: {e}")
                # Limpiar carrito y devolver fragmento vacío/actualizado
                session['carrito'] = []
                try:
                    html = render_template('carrito_fragment_simple.html', carrito=session['carrito'])
                except Exception:
                    html = ''
                return jsonify({'success': True, 'agregados': productos_agregados, 'html': html})
        elif 'cargar_carrito' in request.form:
            # Cargar todos los productos del carrito al stock (form tradicional)
            productos_agregados = 0
            dueno_dest = request.form.get('dueno_dest', 'ferreteria_general')
            for item in carrito:
                try:
                    avisar_bajo_stock = int(item.get('avisar_bajo_stock', 0))
                    min_stock_aviso_val = item.get('min_stock_aviso')
                    try:
                        min_stock_aviso_int = int(min_stock_aviso_val) if min_stock_aviso_val is not None and str(min_stock_aviso_val) != '' else None
                    except (ValueError, TypeError):
                        min_stock_aviso_int = None
                    cantidad_int = int(item.get('cantidad', 1))
                    if avisar_bajo_stock and (min_stock_aviso_int is None or min_stock_aviso_int <= 0):
                        avisar_bajo_stock = 0
                        min_stock_aviso_int = None
                    if avisar_bajo_stock and min_stock_aviso_int and min_stock_aviso_int > cantidad_int:
                        min_stock_aviso_int = cantidad_int
                    db_query(
                        """
                        INSERT INTO stock (codigo, nombre, precio, cantidad, fecha_compra, 
                                         proveedor, observaciones, precio_texto, avisar_bajo_stock, min_stock_aviso, dueno)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """,
                        (
                            item.get('codigo', ''),
                            item.get('nombre', ''),
                            float(item.get('precio', 0)),
                            cantidad_int,
                            item.get('fecha_compra', datetime.now().strftime('%Y-%m-%d')),
                            item.get('proveedor', ''),
                            item.get('observaciones', ''),
                            item.get('precio_texto', ''),
                            avisar_bajo_stock,
                            min_stock_aviso_int,
                            dueno_dest,
                        ),
                    )
                    productos_agregados += 1
                except Exception as e:
                    print(f"Error al agregar producto al stock: {e}")
            session['carrito'] = []
            if productos_agregados > 0:
                flash(f'{productos_agregados} producto(s) agregado(s) al stock exitosamente.', 'success')
            else:
                flash('No se pudo agregar ningún producto al stock.', 'danger')
            # Mantenerse en agregar_producto en vez de redirigir a historial
            return redirect(url_for('agregar_producto'))
    
    # Obtener parámetros de búsqueda en Excel
    termino_excel = request.args.get('busqueda_excel', '')
    proveedor_excel_ricky = request.args.get('proveedor_excel_ricky', '')
    proveedor_excel_fg = request.args.get('proveedor_excel_fg', '')
    
    # Usar exclusivamente uno de los selectores como filtro de proveedor
    if proveedor_excel_ricky:
        proveedor_excel_filtro = proveedor_excel_ricky
        solo_ricky = True
        solo_fg = False
    elif proveedor_excel_fg:
        proveedor_excel_filtro = proveedor_excel_fg
        solo_ricky = False
        solo_fg = True
    else:
        proveedor_excel_filtro = ''
        
        # Si no se seleccionaron cajas de verificación, no aplicar filtros exclusivos
        # Esto permite la búsqueda en todos los proveedores de ambos dueños
        solo_ricky = True if request.args.get('solo_ricky') else False
        solo_fg = True if request.args.get('solo_fg') else False
        
        # Si ambas cajas están desmarcadas, buscar en todos los proveedores
        if not solo_ricky and not solo_fg:
            solo_ricky = False
            solo_fg = False
            
    filtro_excel = request.args.get('filtro_excel', '')
    resultados_excel = []
    
    # Realizar búsqueda en Excel si hay término
    if termino_excel:
        print(f"🔍 Buscando: '{termino_excel}' con filtro proveedor: '{proveedor_excel_filtro}' | solo_ricky: {solo_ricky} | solo_fg: {solo_fg}")
        
        # Si el término es un código numérico y tenemos proveedor específico, usar búsqueda precisa
        if termino_excel.isdigit() and proveedor_excel_filtro:
            print(f"🔢 Realizando búsqueda específica de código exacto: {termino_excel} en {proveedor_excel_filtro}")
            # Buscar el código exacto en el proveedor específico
            resultados_exactos = buscar_codigo_exacto_en_proveedor(
                termino_excel, 
                proveedor_excel_filtro, 
                solo_ricky=solo_ricky, 
                solo_fg=solo_fg
            )
            
            if resultados_exactos:
                print(f"📊 Encontrados {len(resultados_exactos)} resultados exactos")
                resultados_excel = resultados_exactos
            else:
                # Hacer la búsqueda normal y luego filtrar estrictamente
                print(f"🔍 No hay coincidencias exactas, buscando alternativas...")
                resultados_excel = buscar_en_excel(termino_excel, proveedor_excel_filtro, filtro_excel, solo_ricky=solo_ricky, solo_fg=solo_fg)
                
                # Aplicar filtro extremadamente estricto - solo el código exacto y proveedor
                resultados_filtrados = []
                for r in resultados_excel:
                    r_codigo = str(r.get('codigo', '')).strip()
                    r_proveedor = str(r.get('proveedor', '')).lower().strip()
                    r_archivo = str(r.get('archivo', '')).lower()
                    
                    # Verificar código exacto
                    if r_codigo != termino_excel:
                        continue
                        
                    # Verificar que el proveedor coincida
                    if r_proveedor != proveedor_excel_filtro.lower():
                        continue
                        
                    # Verificar que el nombre del archivo comience con el proveedor
                    if not r_archivo.startswith(proveedor_excel_filtro.lower()):
                        continue
                        
                    resultados_filtrados.append(r)
                    print(f"✅ Coincidencia estricta: Código {r_codigo} en proveedor {r_proveedor}, archivo {r_archivo}")
                
                if resultados_filtrados:
                    resultados_excel = resultados_filtrados
                    print(f"📊 Filtrado a {len(resultados_excel)} coincidencias exactas de código")
        else:
            # Búsqueda normal para términos de texto o sin proveedor específico
            resultados_excel = buscar_en_excel(termino_excel, proveedor_excel_filtro, filtro_excel, solo_ricky=solo_ricky, solo_fg=solo_fg)
            print(f"📊 Resultados encontrados: {len(resultados_excel)}")
            
            # Filtrar resultados para eliminar entradas inválidas
            resultados_excel = [r for r in resultados_excel if r.get('nombre') and not r.get('nombre').startswith('Fila ')]
            print(f"📊 Resultados después de filtrar entradas inválidas: {len(resultados_excel)}")
            
            # Si hay un proveedor seleccionado, aplicar filtro muy estricto
            if proveedor_excel_filtro:
                resultados_filtrados = []
                for r in resultados_excel:
                    r_proveedor = str(r.get('proveedor', '')).lower().strip()
                    r_archivo = str(r.get('archivo', '')).lower()
                    
                    # Asegurarse que coincida el proveedor 
                    if r_proveedor != proveedor_excel_filtro.lower():
                        continue
                    
                    # Verificar que el nombre del archivo comience con el proveedor
                    if not r_archivo.startswith(proveedor_excel_filtro.lower()):
                        print(f"⚠️ Omitiendo resultado de {r_archivo}, no coincide con proveedor {proveedor_excel_filtro.lower()}")
                        continue
                        
                    resultados_filtrados.append(r)
                
                resultados_excel = resultados_filtrados
                print(f"📊 Resultados estrictamente filtrados por proveedor: {len(resultados_excel)}")
        
        # Aplicar deduplicación para mostrar resultados únicos por producto Y proveedor
        # Agrupar productos por nombre, código Y proveedor para mantener la variedad de proveedores
        productos_unicos = {}
        
        # Criterio de deduplicación: considerar código, nombre del producto Y proveedor
        # Mantener diferentes proveedores para el mismo producto
        for r in resultados_excel:
            codigo = str(r.get('codigo', '')).strip()
            nombre = str(r.get('nombre', '')).lower().strip()
            proveedor = str(r.get('proveedor', '')).lower().strip()
            
            # Clave única basada en código, nombre del producto y proveedor
            clave_producto = f"{codigo}|{nombre}|{proveedor}"
            
            # Almacenar un resultado por cada combinación única de producto/proveedor
            if clave_producto not in productos_unicos:
                productos_unicos[clave_producto] = r
        
        # Convertir el diccionario en lista final de resultados
        resultados_previos = len(resultados_excel)
        resultados_excel = list(productos_unicos.values())
        
        # Log simple sin mostrar datos detallados
        print(f"✅ Deduplicación completada: {len(resultados_excel)} productos únicos (de {resultados_previos} resultados originales)")
    
    # Obtener proveedores manuales para el selector
    proveedores = db_query("SELECT id, nombre FROM proveedores_manual ORDER BY nombre", fetch=True) or []
    # print(f"DEBUG: proveedores = {proveedores}")
    
    # Obtener lista de proveedores disponibles divididos por dueño (para UI ordenada)
    proveedores_excel_ricky = []
    proveedores_excel_fg = []
    
    # 1) Excel por dueño
    for dueno in ['ricky', 'ferreteria_general']:
        if dueno in DUENOS_CONFIG:
            proveedores_dueno = DUENOS_CONFIG[dueno]['proveedores_excel']
            ocultos_excel = db_query("SELECT LOWER(nombre) as nombre FROM proveedores_ocultos WHERE dueno=?", (dueno,), fetch=True) or []
            ocultos_excel_set = {o['nombre'] for o in ocultos_excel}
            
            for key in proveedores_dueno:
                if key in PROVEEDOR_CONFIG and key.lower() not in ocultos_excel_set:
                    # Buscar archivos en la carpeta específica del dueño
                    carpeta_dueno = get_excel_folder_for_dueno(dueno)
                    archivos = [f for f in os.listdir(carpeta_dueno) if f.lower().startswith(key.lower()) and f.endswith('.xlsx') and f != 'productos_manual.xlsx']
                    if archivos:
                        dueno_display = 'Ricky' if dueno == 'ricky' else 'Ferretería General'
                        item = { 'key': key, 'nombre': key.title().replace('tools','Tools') + f' ({dueno_display})' }
                        if dueno == 'ricky':
                            proveedores_excel_ricky.append(item)
                        else:
                            proveedores_excel_fg.append(item)
    # 2) Manuales por dueño (desde mappings activos)
    ocultos_rows = db_query("SELECT LOWER(nombre) as nombre, dueno FROM proveedores_ocultos", fetch=True) or []
    ocultos_pairs = {(o['nombre'], o['dueno']) for o in ocultos_rows}
    
    # Usar proveedores_duenos como fuente principal, fallback a proveedores_meta
    try:
        mappings = db_query("""
            SELECT pm.id, pm.nombre, pd.dueno 
            FROM proveedores_manual pm 
            JOIN proveedores_duenos pd ON pm.id = pd.proveedor_id 
            ORDER BY pm.nombre, pd.dueno
        """, fetch=True) or []
        print(f"[DEBUG] Usando proveedores_duenos: {len(mappings)} mappings encontrados")
    except Exception as e:
        print(f"[DEBUG] proveedores_duenos no disponible, usando proveedores_meta fallback: {e}")
        try:
            mappings = db_query("SELECT pm.id, pm.nombre, m.dueno FROM proveedores_manual pm JOIN proveedores_meta m ON LOWER(m.nombre)=LOWER(pm.nombre) ORDER BY pm.nombre, m.dueno", fetch=True) or []
            print(f"[DEBUG] Usando proveedores_meta fallback: {len(mappings)} mappings encontrados")
        except Exception as e2:
            print(f"[DEBUG] Ambas tablas fallan, usando lista vacía: {e2}")
            mappings = []
    for row in mappings:
        base = (row['nombre'] or '').strip()
        dueno_val = row['dueno']
        if (base.lower(), dueno_val) in ocultos_pairs:
            continue
        # Evitar duplicar Ricky si ya existe como proveedor Excel nativo
        if dueno_val == 'ricky' and base.lower() in PROVEEDOR_CONFIG:
            continue
        dueno_display = 'Ricky' if dueno_val == 'ricky' else 'Ferretería General'
        item = { 'key': f"manual_{row['id']}_{dueno_val}", 'nombre': f"{base} ({dueno_display})" }
        if dueno_val == 'ricky':
            proveedores_excel_ricky.append(item)
        else:
            proveedores_excel_fg.append(item)
    
    # print(f"DEBUG: proveedores_excel = {proveedores_excel}")
    # print(f"DEBUG: EXCEL_FOLDER = {EXCEL_FOLDER}")
    # print(f"DEBUG: Archivos en EXCEL_FOLDER = {os.listdir(EXCEL_FOLDER) if os.path.exists(EXCEL_FOLDER) else 'Carpeta no existe'}")
    
    # Construir listas para select de proveedor en Agregar Producto Manual (mismos mappings por dueño)
    proveedores_select_ricky = []
    proveedores_select_fg = []
    for row in mappings:
        base = (row['nombre'] or '').strip()
        dueno_val = row['dueno']
        dueno_display = 'Ricky' if dueno_val == 'ricky' else 'Ferretería General'
        if (base.lower(), dueno_val) in ocultos_pairs:
            continue
        entry = { 'value': f"{row['id']}|{dueno_val}", 'label': f"{base} ({dueno_display})" }
        if dueno_val == 'ricky':
            proveedores_select_ricky.append(entry)
        else:
            proveedores_select_fg.append(entry)

    # Determinar selección activa por grupo
    selected_ricky = proveedor_excel_filtro if any(p['key'] == proveedor_excel_filtro for p in proveedores_excel_ricky) else ''
    selected_fg = proveedor_excel_filtro if any(p['key'] == proveedor_excel_filtro for p in proveedores_excel_fg) else ''

    return render_template('agregar.html', 
                         fecha_actual=datetime.now().strftime('%Y-%m-%d'),
                         proveedores_excel_ricky=proveedores_excel_ricky,
                         proveedores_excel_fg=proveedores_excel_fg,
                         proveedores=proveedores,
                         proveedores_select_ricky=proveedores_select_ricky,
                         proveedores_select_fg=proveedores_select_fg,
                         proveedor_excel_ricky_selected=selected_ricky,
                         proveedor_excel_fg_selected=selected_fg,
                         termino_excel=termino_excel,
                         proveedor_excel=proveedor_excel_filtro,
                         filtro_excel=filtro_excel,
                         resultados_excel=resultados_excel,
                         solo_ricky=solo_ricky,
                         solo_fg=solo_fg,
                         carrito=session.get('carrito', []))

@app.route('/procesar_producto', methods=['POST'])
@login_required
def procesar_producto():
    try:
        nombre = request.form.get('nombre', '').strip()
        codigo = request.form.get('codigo', '').strip()
        precio_str = request.form.get('precio', '').strip()
        cantidad = int(request.form.get('cantidad', 0))
        proveedor = request.form.get('proveedor', '').strip()
        observaciones = request.form.get('observaciones', '').strip()
        fecha_compra = request.form.get('fecha_compra', datetime.now().strftime('%Y-%m-%d'))
        
        if not nombre:
            flash('El nombre del producto es obligatorio.', 'danger')
            return redirect(url_for('agregar_producto'))
        
        precio, error_precio = parse_price(precio_str)
        if error_precio:
            flash(f'Error en el precio: {error_precio}', 'danger')
            return redirect(url_for('agregar_producto'))
        
        # Insertar producto en stock
        result = db_query(
            "INSERT INTO stock (codigo, nombre, precio, cantidad, fecha_compra, proveedor, observaciones, precio_texto) VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
            (codigo, nombre, precio, cantidad, fecha_compra, proveedor, observaciones, precio_str)
        )
        
        if result:
            flash(f'Producto "{nombre}" agregado exitosamente.', 'success')
        else:
            flash('Error al agregar el producto.', 'danger')
            
    except Exception as e:
        flash(f'Error al procesar el producto: {str(e)}', 'danger')
    
    return redirect(url_for('agregar_producto'))

@app.route('/productos_manual')
@login_required
def productos_manual():
    proveedores = db_query("SELECT id, nombre FROM proveedores_manual ORDER BY nombre", fetch=True)
    return render_template('productos_manual.html', proveedores=proveedores)

@app.route('/carrito')
@login_required
def carrito():
    productos_carrito = session.get('carrito', [])
    total = sum(float(p.get('precio', 0)) * int(p.get('cantidad', 0)) for p in productos_carrito)
    return render_template('carrito.html', productos=productos_carrito, total=total)

@app.route('/limpiar_carrito', methods=['POST'])
@login_required
def limpiar_carrito():
    session['carrito'] = []
    flash('Carrito limpiado.', 'success')
    return redirect(url_for('carrito'))

@app.route('/reportes')
@login_required
def reportes():
    # Obtener estadísticas del stock
    productos_stock = db_query("SELECT COUNT(*) as total FROM stock", fetch=True)
    total_productos = productos_stock[0]['total'] if productos_stock else 0
    
    # Calcular valor total del inventario
    valor_inventario = db_query("SELECT SUM(precio * cantidad) as total_valor FROM stock", fetch=True)
    valor_total = valor_inventario[0]['total_valor'] if valor_inventario and valor_inventario[0]['total_valor'] else 0
    
    # Contar proveedores
    proveedores_count = db_query("SELECT COUNT(*) as total FROM proveedores_manual", fetch=True)
    total_proveedores = proveedores_count[0]['total'] if proveedores_count else 0
    
    return render_template('reportes.html', 
                         total_productos=total_productos,
                         valor_total=valor_total,
                         total_proveedores=total_proveedores)

@app.route('/proveedores')
@login_required
def proveedores():
    # Manejar búsqueda si hay término
    busqueda_proveedor = request.args.get('busqueda_proveedor', '').strip()
    proveedores_encontrados = []
    
    if busqueda_proveedor:
        like_pattern = f'%{busqueda_proveedor}%'
        proveedores_encontrados = db_query(
            """
            SELECT DISTINCT pm.id, pm.nombre, pd.dueno 
            FROM proveedores_manual pm 
            JOIN proveedores_duenos pd ON pd.proveedor_id = pm.id 
            WHERE pm.nombre LIKE ? 
            ORDER BY pm.nombre
            """,
            (like_pattern,),
            fetch=True
        ) or []
    
    # Construir listas por dueño a partir de la tabla de relaciones
    # Obtener todos los proveedores con sus dueños
    mappings = db_query(
        """
        SELECT DISTINCT pm.id, pm.nombre, pd.dueno 
        FROM proveedores_manual pm 
        JOIN proveedores_duenos pd ON pd.proveedor_id = pm.id 
        ORDER BY pm.nombre, pd.dueno
        """, 
        fetch=True
    ) or []
    
    lista_fg, lista_ricky, todos = [], [], []
    for row in mappings:
        entry = { 'id': row['id'], 'nombre': row['nombre'], 'dueno': row['dueno'] }
        todos.append(entry)
        if row['dueno'] == 'ricky':
            # Verificar si ya existe en la lista para evitar duplicados
            if not any(p['id'] == row['id'] for p in lista_ricky):
                lista_ricky.append({'id': row['id'], 'nombre': row['nombre']})
        elif row['dueno'] == 'ferreteria_general':
            # Verificar si ya existe en la lista para evitar duplicados
            if not any(p['id'] == row['id'] for p in lista_fg):
                lista_fg.append({'id': row['id'], 'nombre': row['nombre']})
    todos_proveedores = todos
    return render_template('proveedores.html', 
                         proveedores=proveedores,
                         busqueda_proveedor=busqueda_proveedor,
                         proveedores_encontrados=proveedores_encontrados,
                         proveedores_fg=lista_fg,
                         proveedores_ricky=lista_ricky,
                         todos_proveedores=todos_proveedores)

@app.route('/agregar_proveedor', methods=['POST'])
@login_required
def agregar_proveedor():
    try:
        nombre_raw = request.form.get('nombre', '').strip()
        # Normalizar: quitar espacios dobles, uniformar mayúsculas iniciales
        nombre = ' '.join(nombre_raw.split())
        # Correcciones automáticas de errores comunes / variantes
        # Se normaliza a la forma canónica EXACTA usada internamente para evitar duplicados por casing/espacios
        correcciones = {
            # BremenTools variantes
            'brementools': 'BremenTools',
            'brementool': 'BremenTools',
            'bremetools': 'BremenTools',
            'bremetool': 'BremenTools',
            'bremen tools': 'BremenTools',
            'bremen-tools': 'BremenTools',
            # Crossmaster variantes
            'crossmaster': 'Crossmaster',  # fuerza capitalización
            'cross master': 'Crossmaster',
            'cross-master': 'Crossmaster',
            'cross master.': 'Crossmaster',
            'crossmaster.': 'Crossmaster',
            'crossmaster,': 'Crossmaster',
            'crossmaster tools': 'Crossmaster',
            'cross master tools': 'Crossmaster',
            # Otras capitalizaciones raras
            'crossMaster': 'Crossmaster',
            'crossMaster tools': 'Crossmaster'
        }
        nombre_lower = nombre.lower()
        if nombre_lower in correcciones:
            nombre = correcciones[nombre_lower]
        dueno = request.form.get('dueno', '').strip()  # valores: '', 'ricky', 'ferreteria_general', 'ambos'
        print(f"[DEBUG agregar_proveedor] nombre='{nombre}', dueno='{dueno}'")

        if not nombre:
            flash('El nombre del proveedor no puede estar vacío.', 'danger')
        else:
            # Conjunto de dueños destino
            if dueno == 'ambos':
                duenos_destino = ['ricky', 'ferreteria_general']
            else:
                duenos_destino = ['ricky'] if dueno == 'ricky' else ['ferreteria_general']

            existente = db_query("SELECT id, nombre FROM proveedores_manual WHERE LOWER(nombre)=LOWER(?) LIMIT 1", (nombre,), fetch=True)
            if not existente:
                # Insertar proveedor base sin dueño (el dueño ahora va en proveedores_duenos)
                if not dueno:
                    flash('Se requiere especificar un dueño para el proveedor.', 'danger')
                    return redirect(url_for('proveedores'))
                    
                # Insertar el proveedor sin especificar dueño
                db_query("INSERT INTO proveedores_manual (nombre) VALUES (?)", (nombre,))
                
                # Obtener el ID del proveedor recién insertado
                nuevo_proveedor = db_query("SELECT id, nombre FROM proveedores_manual WHERE LOWER(nombre)=LOWER(?) LIMIT 1", (nombre,), fetch=True)
                if not nuevo_proveedor:
                    flash('Error al crear el proveedor.', 'danger')
                    return redirect(url_for('proveedores'))
                    
                proveedor_id = nuevo_proveedor[0]['id']
                nombre_guardado = nuevo_proveedor[0]['nombre']
                nuevo_flag = True
            else:
                proveedor_id = existente[0]['id']
                nombre_guardado = existente[0]['nombre']
                nuevo_flag = False

            # Asegurar relaciones para cada dueño solicitado en proveedores_duenos
            agregados = []
            print(f"[DEBUG] Agregando proveedor '{nombre_guardado}' (ID: {proveedor_id}) para dueños: {duenos_destino}")
            
            for d in duenos_destino:
                # Insertar en la tabla de relación
                ok = db_query("INSERT OR IGNORE INTO proveedores_duenos (proveedor_id, dueno) VALUES (?, ?)", 
                             (proveedor_id, d))
                if ok:
                    agregados.append(d)
                    print(f"[DEBUG] Proveedor '{nombre_guardado}' asociado a dueño '{d}' en proveedores_duenos")
                else:
                    print(f"[DEBUG] Proveedor '{nombre_guardado}' ya estaba asociado a dueño '{d}' en proveedores_duenos")
                
                # Mantener compatibilidad con proveedores_meta (legacy)
                db_query("INSERT OR IGNORE INTO proveedores_meta (nombre, dueno) VALUES (?, ?)", 
                       (nombre_guardado, d))
                print(f"[DEBUG] Proveedor '{nombre_guardado}' asociado a dueño '{d}' en proveedores_meta (legacy)")
                
                # Borrar ocultamientos previos de ese dueño
                db_query("DELETE FROM proveedores_ocultos WHERE LOWER(nombre)=LOWER(?) AND (dueno IS NULL OR dueno=?)", 
                       (nombre_guardado, d))

            if nuevo_flag:
                if dueno == 'ambos':
                    flash(f'Proveedor "{nombre_guardado}" creado y asociado a ambos dueños.', 'success')
                else:
                    destino_nombre = 'Ricky' if duenos_destino == ['ricky'] else 'Ferretería General'
                    flash(f'Proveedor "{nombre_guardado}" creado para {destino_nombre}.', 'success')
            else:
                if agregados:
                    if dueno == 'ambos':
                        flash(f'Proveedor "{nombre_guardado}" ya existía. Se garantizó asociación a ambos dueños.', 'info')
                    else:
                        destino_nombre = 'Ricky' if duenos_destino == ['ricky'] else 'Ferretería General'
                        flash(f'Proveedor "{nombre_guardado}" ya existía. Se asoció a {destino_nombre}.', 'info')
                else:
                    flash(f'El proveedor "{nombre_guardado}" ya estaba asociado al/los dueño(s) seleccionado(s).', 'warning')
    except Exception as e:
        print(f"Error al agregar proveedor: {e}")
        flash(f'Error al agregar el proveedor: {str(e)}', 'danger')
    return redirect(url_for('proveedores'))


# Helper reutilizable para insertar / asociar proveedor a uno o varios dueños
def sincronizar_proveedores_meta_duenos():
    """
    Sincroniza las tablas proveedores_meta y proveedores_duenos.
    Esta función debe ejecutarse cada vez que se modifican proveedores para mantener consistencia.
    """
    try:
        conn = get_db_connection()
        if not conn:
            return False, "No se pudo conectar a la base de datos"
        
        cursor = conn.cursor()
        use_postgres = _is_postgres_configured()
        
        # Sincronizar desde proveedores_duenos hacia proveedores_meta
        if use_postgres:
            cursor.execute("""
                INSERT INTO proveedores_meta (nombre, dueno)
                SELECT pm.nombre, pd.dueno 
                FROM proveedores_duenos pd
                JOIN proveedores_manual pm ON pm.id = pd.proveedor_id
                ON CONFLICT (nombre, dueno) DO NOTHING
            """)
        else:
            cursor.execute("""
                INSERT OR IGNORE INTO proveedores_meta (nombre, dueno)
                SELECT pm.nombre, pd.dueno 
                FROM proveedores_duenos pd
                JOIN proveedores_manual pm ON pm.id = pd.proveedor_id
            """)
        
        # Sincronizar desde proveedores_meta hacia proveedores_duenos
        if use_postgres:
            cursor.execute("""
                INSERT INTO proveedores_duenos (proveedor_id, dueno)
                SELECT pm.id, meta.dueno 
                FROM proveedores_meta meta
                JOIN proveedores_manual pm ON pm.nombre = meta.nombre
                ON CONFLICT (proveedor_id, dueno) DO NOTHING
            """)
        else:
            cursor.execute("""
                INSERT OR IGNORE INTO proveedores_duenos (proveedor_id, dueno)
                SELECT pm.id, meta.dueno 
                FROM proveedores_meta meta
                JOIN proveedores_manual pm ON pm.nombre = meta.nombre
            """)
        
        conn.commit()
        cursor.close()
        conn.close()
        
        return True, "Sincronización completada exitosamente"
        
    except Exception as e:
        return False, f"Error en sincronización: {str(e)}"

def _upsert_proveedor(nombre: str, dueno_param: str):
    """Crea el proveedor si no existe y asegura mappings según dueno_param ('ricky','ferreteria_general','ambos').
       Devuelve (nombre_guardado, nuevo_flag, agregados(list_dueños_nuevos)).
       Mantiene sincronizadas las tablas proveedores_meta y proveedores_duenos."""
    nombre_guardado = nombre
    existente = db_query("SELECT id, nombre FROM proveedores_manual WHERE LOWER(nombre)=LOWER(?) LIMIT 1", (nombre,), fetch=True)
    if not existente:
        db_query("INSERT INTO proveedores_manual (nombre) VALUES (?)", (nombre_guardado,))
        # Obtener el ID del proveedor recién creado
        proveedor_data = db_query("SELECT id, nombre FROM proveedores_manual WHERE LOWER(nombre)=LOWER(?) LIMIT 1", (nombre_guardado,), fetch=True)
        proveedor_id = proveedor_data[0]['id'] if proveedor_data else None
        nuevo_flag = True
    else:
        nombre_guardado = existente[0]['nombre']
        proveedor_id = existente[0]['id']
        nuevo_flag = False
    
    if dueno_param == 'ambos':
        destinos = ['ricky', 'ferreteria_general']
    else:
        destinos = ['ricky'] if dueno_param == 'ricky' else ['ferreteria_general']
    
    agregados = []
    for d in destinos:
        # Actualizar proveedores_duenos (tabla principal para consultas) - PRIORITARIO
        if proveedor_id:
            ok_duenos = db_query("INSERT OR IGNORE INTO proveedores_duenos (proveedor_id, dueno) VALUES (?, ?)", (proveedor_id, d))
            print(f"[DEBUG] Proveedor '{nombre_guardado}' asociado a dueño '{d}' en proveedores_duenos")
        
        # Actualizar proveedores_meta (solo si la tabla existe)
        try:
            ok_meta = db_query("INSERT OR IGNORE INTO proveedores_meta (nombre, dueno) VALUES (?, ?)", (nombre_guardado, d))
            print(f"[DEBUG] Proveedor '{nombre_guardado}' asociado a dueño '{d}' en proveedores_meta (legacy)")
        except Exception as e:
            print(f"[DEBUG] No se pudo actualizar proveedores_meta (tabla puede no existir): {e}")
            ok_meta = True  # No fallar por esto
        
        # Eliminar de ocultos (si la tabla existe)
        try:
            db_query("DELETE FROM proveedores_ocultos WHERE LOWER(nombre)=LOWER(?) AND (dueno IS NULL OR dueno=?)", (nombre_guardado, d))
        except Exception as e:
            print(f"[DEBUG] No se pudo actualizar proveedores_ocultos: {e}")
        
        # Considerar exitoso si al menos proveedores_duenos funcionó
        if proveedor_id and ok_duenos:
            agregados.append(d)
    
    return nombre_guardado, nuevo_flag, agregados, destinos


@app.route('/agregar_proveedor_manual', methods=['POST'])
@login_required
def agregar_proveedor_manual():
    try:
        nombre_raw = request.form.get('nombre', '').strip()
        nombre = ' '.join(nombre_raw.split())
        dueno = request.form.get('dueno', '').strip()  # '', 'ricky', 'ferreteria_general', 'ambos'
        correcciones = {
            'brementools': 'BremenTools', 'brementool': 'BremenTools', 'bremetools': 'BremenTools', 'bremetool': 'BremenTools', 'bremen tools': 'BremenTools', 'bremen-tools': 'BremenTools',
            'crossmaster': 'Crossmaster', 'cross master': 'Crossmaster', 'cross-master': 'Crossmaster', 'cross master.': 'Crossmaster', 'crossmaster.': 'Crossmaster', 'crossmaster,': 'Crossmaster', 'crossmaster tools': 'Crossmaster', 'cross master tools': 'Crossmaster', 'crossmastertools': 'Crossmaster'
        }
        nombre_lower = nombre.lower()
        if nombre_lower in correcciones:
            nombre = correcciones[nombre_lower]
        if not nombre:
            flash('El nombre del proveedor es obligatorio.', 'danger')
            return redirect(url_for('proveedores'))
        # Default dueño si viene vacío
        if dueno not in ('ricky','ferreteria_general','ambos'):
            dueno = 'ferreteria_general'
        nombre_guardado, nuevo_flag, agregados, destinos = _upsert_proveedor(nombre, dueno)
        if nuevo_flag:
            if dueno == 'ambos':
                flash(f'Proveedor "{nombre_guardado}" creado y asociado a ambos dueños.', 'success')
            else:
                destino_nombre = 'Ricky' if destinos == ['ricky'] else 'Ferretería General'
                flash(f'Proveedor "{nombre_guardado}" creado para {destino_nombre}.', 'success')
        else:
            if agregados:
                if dueno == 'ambos':
                    flash(f'Proveedor "{nombre_guardado}" ya existía. Se garantizó asociación a ambos dueños.', 'info')
                else:
                    destino_nombre = 'Ricky' if destinos == ['ricky'] else 'Ferretería General'
                    flash(f'Proveedor "{nombre_guardado}" ya existía. Se asoció a {destino_nombre}.', 'info')
            else:
                flash(f'El proveedor "{nombre_guardado}" ya estaba asociado al/los dueño(s) seleccionado(s).', 'warning')
    except Exception as e:
        print(f"Error en agregar_proveedor_manual: {e}")
        flash('Error al procesar el proveedor manual.', 'danger')
    return redirect(url_for('proveedores'))


@app.route('/eliminar_proveedor_manual', methods=['POST'])
@login_required
def eliminar_proveedor_manual():
    try:
        proveedor_id = request.form.get('id')
        dueno_form = request.form.get('dueno')
        if not proveedor_id:
            flash('ID de proveedor no válido.', 'danger')
            return redirect(url_for('proveedores'))
        
        # Obtener el nombre del proveedor antes de eliminarlo
        proveedor = db_query("SELECT nombre FROM proveedores_manual WHERE id = ?", (proveedor_id,), fetch=True)
        if not proveedor:
            flash('Proveedor no encontrado.', 'danger')
            return redirect(url_for('proveedores'))
        
        nombre_proveedor = proveedor[0]['nombre']
        
        # Determinar dueño a ocultar
        if not dueno_form:
            # Si no se especificó dueño, intentar obtenerlo de la tabla de relaciones
            relacion = db_query(
                """
                SELECT dueno 
                FROM proveedores_duenos 
                WHERE proveedor_id = ? 
                LIMIT 1
                """, 
                (proveedor_id,), fetch=True
            )
            meta_dueno = relacion[0]['dueno'] if relacion else 'ferreteria_general'
        else:
            meta_dueno = dueno_form
            
        # Registrar ocultamiento lógico solo para ese dueño
        db_query("INSERT OR IGNORE INTO proveedores_ocultos (nombre, dueno) VALUES (?, ?)", (nombre_proveedor, meta_dueno))
        
        # Eliminar la relación en la tabla proveedores_duenos
        ok = db_query("DELETE FROM proveedores_duenos WHERE proveedor_id = ? AND dueno = ?", (proveedor_id, meta_dueno))
        
        # También eliminar de la tabla legacy para mantener compatibilidad
        db_query("DELETE FROM proveedores_meta WHERE nombre = ? AND dueno = ?", (nombre_proveedor, meta_dueno))
        
        if ok:
            flash(f'Proveedor "{nombre_proveedor}" ocultado para {"Ricky" if meta_dueno=="ricky" else "Ferretería General"}.', 'success')
        else:
            flash('Error al eliminar el proveedor.', 'danger')
    except Exception as e:
        print(f"Error al eliminar proveedor: {e}")
        flash(f'Error al eliminar el proveedor: {str(e)}', 'danger')
    
    return redirect(url_for('proveedores'))

# -----------------------------
# Proveedores ocultos (listar / restaurar)
# -----------------------------

@app.route('/proveedores_ocultos', methods=['GET'])
@login_required
def proveedores_ocultos_listado():
    """Devuelve en JSON la lista de proveedores ocultos.
       Parámetros:
         - dueno = ricky | ferreteria_general | todos (default: todos)
         - include_visible = 0|1 -> si 1, muestra también aquellos que tienen mapping activo (para otro u mismo dueño)
           y se añaden flags de estado.
    """
    try:
        dueno = request.args.get('dueno', 'todos').strip().lower()
        include_visible = request.args.get('include_visible', '0').strip() == '1'
        include_matrix = request.args.get('include_matrix', '0').strip() == '1'
        params = []
        # Recuperar todos los registros de ocultos (filtrados por dueno si se pide) para poder construir matriz
        base_all = "SELECT nombre, dueno FROM proveedores_ocultos"
        if dueno in ('ricky','ferreteria_general'):
            base_all += " WHERE dueno = ?"
            params.append(dueno)
        rows_all = db_query(base_all, tuple(params), fetch=True) or []

        # Normalizar por nombre canónico
        grouped = {}
        for r in rows_all:
            canon = db_query("SELECT nombre FROM proveedores_manual WHERE LOWER(nombre)=LOWER(?) LIMIT 1", (r['nombre'],), fetch=True)
            nombre_canon = canon[0]['nombre'] if canon else r['nombre']
            key = nombre_canon.lower()
            if key not in grouped:
                grouped[key] = {'nombre': nombre_canon, 'oculto_ricky': False, 'oculto_fg': False}
            if r['dueno'] == 'ricky':
                grouped[key]['oculto_ricky'] = True
            elif r['dueno'] == 'ferreteria_general':
                grouped[key]['oculto_fg'] = True

        data_list = []
        for g in grouped.values():
            # Determinar visibilidad real consultando mappings (una vez por nombre)
            vis_ricky = db_query("SELECT 1 FROM proveedores_meta WHERE LOWER(nombre)=LOWER(?) AND dueno='ricky' LIMIT 1", (g['nombre'],), fetch=True)
            vis_fg = db_query("SELECT 1 FROM proveedores_meta WHERE LOWER(nombre)=LOWER(?) AND dueno='ferreteria_general' LIMIT 1", (g['nombre'],), fetch=True)
            entry = {
                'nombre': g['nombre'],
                'dueno': None,  # legacy (no aplica cuando devolvemos matriz)
                'oculto_ricky': g['oculto_ricky'],
                'oculto_ferreteria_general': g['oculto_fg'],
                'visible_ricky': bool(vis_ricky),
                'visible_ferreteria_general': bool(vis_fg)
            }
            if not include_visible:
                # En modo sin include_visible mantenemos compat: desglosar por dueño realmente oculto (sin mapping)
                if g['oculto_ricky'] and not vis_ricky:
                    data_list.append({
                        'nombre': g['nombre'],
                        'dueno': 'ricky'
                    })
                if g['oculto_fg'] and not vis_fg:
                    data_list.append({
                        'nombre': g['nombre'],
                        'dueno': 'ferreteria_general'
                    })
            else:
                data_list.append(entry)

        # Si se pide include_matrix entregamos siempre una matriz por nombre (data_list ya sirve cuando include_visible=True)
        if include_matrix and include_visible:
            return jsonify({'success': True, 'proveedores': data_list, 'matrix': True})
        return jsonify({'success': True, 'proveedores': data_list, 'matrix': include_matrix and include_visible})
    except Exception as e:
        print(f"Error listando proveedores ocultos: {e}")
        return jsonify({'success': False, 'error': 'Error interno al listar proveedores ocultos'}), 500


@app.route('/restaurar_proveedor_oculto', methods=['POST'])
@login_required
def restaurar_proveedor_oculto():
    """Restaura (desoculta) un proveedor para un dueño dado.
       Datos esperados: nombre, dueno (ricky|ferreteria_general).
       Si el mapping no existe en proveedores_meta se recrea.
       Devuelve JSON si la petición es AJAX/JSON; en otro caso redirige.
    """
    try:
        # Permitir tanto JSON como form-data
        if request.is_json:
            payload = request.get_json(silent=True) or {}
            nombre = (payload.get('nombre') or '').strip()
            dueno = (payload.get('dueno') or '').strip().lower()
        else:
            nombre = (request.form.get('nombre') or '').strip()
            dueno = (request.form.get('dueno') or '').strip().lower()

        if dueno not in ('ricky', 'ferreteria_general'):
            msg = 'Dueño inválido.'
            if request.is_json:
                return jsonify({'success': False, 'error': msg}), 400
            flash(msg, 'danger')
            return redirect(url_for('proveedores'))
        if not nombre:
            msg = 'Nombre requerido.'
            if request.is_json:
                return jsonify({'success': False, 'error': msg}), 400
            flash(msg, 'danger')
            return redirect(url_for('proveedores'))

        # Quitar de ocultos (case-insensitive)
        db_query("DELETE FROM proveedores_ocultos WHERE LOWER(nombre)=LOWER(?) AND dueno=?", (nombre, dueno))
        # Asegurar proveedor base exista
        existente = db_query("SELECT nombre FROM proveedores_manual WHERE LOWER(nombre)=LOWER(?) LIMIT 1", (nombre,), fetch=True)
        if not existente:
            db_query("INSERT INTO proveedores_manual (nombre) VALUES (?)", (nombre,))
            nombre_guardado = nombre
        else:
            nombre_guardado = existente[0]['nombre']
        # Recrear mapping si falta
        db_query("INSERT OR IGNORE INTO proveedores_meta (nombre, dueno) VALUES (?, ?)", (nombre_guardado, dueno))

        msg_ok = f'Proveedor "{nombre_guardado}" restaurado para {"Ricky" if dueno=="ricky" else "Ferretería General"}.'
        if request.is_json:
            return jsonify({'success': True, 'message': msg_ok})
        flash(msg_ok, 'success')
        return redirect(url_for('proveedores'))
    except Exception as e:
        print(f"Error restaurando proveedor oculto: {e}")
        if request.is_json:
            return jsonify({'success': False, 'error': 'Error interno al restaurar'}), 500
        flash('Error interno al restaurar proveedor.', 'danger')
        return redirect(url_for('proveedores'))


@app.route('/eliminar_proveedor_definitivo', methods=['POST'])
@login_required
def eliminar_proveedor_definitivo():
    """Elimina totalmente un proveedor para un dueño: quita oculto, mapping y si ya no existe mapping para otros dueños tampoco productos manuales asociados.
       Espera: nombre, dueno.
    """
    try:
        if request.is_json:
            payload = request.get_json(silent=True) or {}
            nombre = (payload.get('nombre') or '').strip()
            dueno = (payload.get('dueno') or '').strip().lower()
        else:
            nombre = (request.form.get('nombre') or '').strip()
            dueno = (request.form.get('dueno') or '').strip().lower()
        if dueno not in ('ricky','ferreteria_general'):
            msg = 'Dueño inválido.'
            return (jsonify({'success': False, 'error': msg}), 400) if request.is_json else (flash(msg,'danger'), redirect(url_for('proveedores')))
        if not nombre:
            msg = 'Nombre requerido.'
            return (jsonify({'success': False, 'error': msg}), 400) if request.is_json else (flash(msg,'danger'), redirect(url_for('proveedores')))

        # Obtener el ID del proveedor
        proveedor_row = db_query("SELECT id FROM proveedores_manual WHERE LOWER(nombre)=LOWER(?)", (nombre,), fetch=True)
        
        if not proveedor_row:
            msg = f'Proveedor "{nombre}" no encontrado.'
            return (jsonify({'success': False, 'error': msg}), 404) if request.is_json else (flash(msg,'danger'), redirect(url_for('proveedores')))
        
        proveedor_id = proveedor_row[0]['id']
        
        # Eliminar la relación en la tabla proveedores_duenos
        db_query("DELETE FROM proveedores_duenos WHERE proveedor_id = ? AND dueno = ?", (proveedor_id, dueno))
        
        # Eliminar mapping explícito legacy (si todavía existiera por alguna inconsistencia)
        db_query("DELETE FROM proveedores_meta WHERE LOWER(nombre)=LOWER(?) AND dueno=?", (nombre, dueno))
        
        # Eliminar registro de ocultos
        db_query("DELETE FROM proveedores_ocultos WHERE LOWER(nombre)=LOWER(?) AND dueno=?", (nombre, dueno))
        
        # Registrar eliminación definitiva para excluir del historial
        try:
            db_query("INSERT OR IGNORE INTO proveedores_eliminados (nombre, dueno, fecha_eliminacion) VALUES (?, ?, ?)", (nombre, dueno, datetime.utcnow().isoformat()))
        except Exception as e:
            print(f"[WARN] No se pudo registrar proveedor eliminado: {e}")
        
        # Ver si queda alguna relación para otros dueños
        otros = db_query("SELECT 1 FROM proveedores_duenos WHERE proveedor_id = ? LIMIT 1", (proveedor_id,), fetch=True)
        
        if not otros:
            # No queda asociado a ningún dueño: opcionalmente podríamos dejar el proveedor base, pero si quieres limpieza total podríamos borrar solo si no hay productos manuales.
            productos = db_query("SELECT 1 FROM productos_manual WHERE LOWER(proveedor)=LOWER(?) LIMIT 1", (nombre,), fetch=True)
            if not productos:
                db_query("DELETE FROM proveedores_manual WHERE id = ?", (proveedor_id,))
        msg_ok = f'Proveedor "{nombre}" eliminado definitivamente para {"Ricky" if dueno=="ricky" else "Ferretería General"}.'
        if request.is_json:
            return jsonify({'success': True, 'message': msg_ok})
        flash(msg_ok, 'success')
        return redirect(url_for('proveedores'))
    except Exception as e:
        print(f"Error eliminando proveedor definitivo: {e}")
        if request.is_json:
            return jsonify({'success': False, 'error': 'Error interno al eliminar definitivo'}), 500
        flash('Error interno al eliminar proveedor.', 'danger')
        return redirect(url_for('proveedores'))


@app.route('/debug_proveedor_status')
@login_required
def debug_proveedor_status():
    """Devuelve JSON con el estado de un proveedor (diagnóstico).
       Parámetros: nombre (obligatorio, case-insensitive).
    """
    try:
        nombre = request.args.get('nombre','').strip()
        if not nombre:
            return jsonify({'success': False, 'error': 'nombre requerido'}), 400
            
        # Obtener información básica del proveedor
        rows_manual = db_query("SELECT id, nombre FROM proveedores_manual WHERE LOWER(nombre)=LOWER(?)", (nombre,), fetch=True) or []
        
        # Obtener relaciones de la nueva tabla proveedores_duenos
        rows_duenos = []
        if rows_manual:
            proveedor_id = rows_manual[0]['id']
            rows_duenos = db_query(
                """
                SELECT pd.proveedor_id, pd.dueno 
                FROM proveedores_duenos pd
                WHERE pd.proveedor_id = ?
                """, 
                (proveedor_id,), fetch=True
            ) or []
            
        # Obtener información de tablas legacy para diagnóstico
        rows_meta = db_query("SELECT nombre, dueno FROM proveedores_meta WHERE LOWER(nombre)=LOWER(?)", (nombre,), fetch=True) or []
        rows_ocultos = db_query("SELECT id, nombre, dueno FROM proveedores_ocultos WHERE LOWER(nombre)=LOWER(?)", (nombre,), fetch=True) or []
        rows_eliminados = db_query("SELECT id, nombre, dueno, fecha_eliminacion FROM proveedores_eliminados WHERE LOWER(nombre)=LOWER(?)", (nombre,), fetch=True) or []
        
        return jsonify({
            'success': True,
            'input': nombre,
            'manual': rows_manual,
            'relaciones': rows_duenos,  # Nueva información de la tabla de relaciones
            'meta_legacy': rows_meta,   # Renombrado para claridad
            'ocultos': rows_ocultos,
            'eliminados': rows_eliminados
        })
    except Exception as e:
        print(f"Error debug_proveedor_status: {e}")
        return jsonify({'success': False, 'error': 'error interno'}), 500

@app.route('/forzar_eliminar_proveedor', methods=['POST'])
@login_required
def forzar_eliminar_proveedor():
    """Marca un proveedor como eliminado para un dueño aunque no esté actualmente oculto.
       No borra registros de stock; sólo lo excluye de filtros futuros.
       JSON esperado: { nombre: str, dueno: str }
    """
    try:
        payload = request.get_json(silent=True) or {}
        nombre = _normalizar_nombre_proveedor((payload.get('nombre') or '').strip())
        dueno = (payload.get('dueno') or '').strip().lower()
        if dueno not in ('ricky','ferreteria_general'):
            return jsonify({'success': False, 'error': 'Dueño inválido'}), 400
        if not nombre:
            return jsonify({'success': False, 'error': 'Nombre requerido'}), 400
        # Limpiar mapping y ocultos por consistencia
        db_query("DELETE FROM proveedores_meta WHERE LOWER(nombre)=LOWER(?) AND dueno=?", (nombre, dueno))
        db_query("DELETE FROM proveedores_ocultos WHERE LOWER(nombre)=LOWER(?) AND dueno=?", (nombre, dueno))
        # Registrar eliminado (si ya existía, IGNORE evita duplicado)
        db_query("INSERT OR IGNORE INTO proveedores_eliminados (nombre, dueno, fecha_eliminacion) VALUES (?, ?, ?)", (nombre, dueno, datetime.utcnow().isoformat()))
        return jsonify({'success': True, 'message': f'Proveedor "{nombre}" marcado como eliminado para {dueno}.'})
    except Exception as e:
        print(f"Error forzar_eliminar_proveedor: {e}")
        return jsonify({'success': False, 'error': 'Error interno'}), 500

@app.route('/debug_historial_fuentes')
@login_required
def debug_historial_fuentes():
    """Devuelve un desglose detallado de cómo se arma la lista de proveedores para el historial de un dueño.
       Parámetros: dueno (obligatorio)
       Respuesta: JSON con arrays por fuente y un listado consolidado con flags.
    """
    try:
        dueno = (request.args.get('dueno') or '').strip().lower()
        if dueno not in ('ricky','ferreteria_general'):
            return jsonify({'success': False, 'error': 'dueno inválido'}), 400
        # Cargar conjuntos de ocultos y eliminados normalizados
        ocultos_rows = db_query("SELECT nombre FROM proveedores_ocultos WHERE dueno=?", (dueno,), fetch=True) or []
        eliminados_rows = db_query("SELECT nombre FROM proveedores_eliminados WHERE dueno=?", (dueno,), fetch=True) or []
        ocultos_norm = { _normalizar_nombre_proveedor(r['nombre']) for r in ocultos_rows }
        eliminados_norm = { _normalizar_nombre_proveedor(r['nombre']) for r in eliminados_rows }

        # Fuente stock
        stock_rows = db_query("SELECT DISTINCT proveedor FROM stock WHERE proveedor IS NOT NULL AND TRIM(proveedor)!='' AND LOWER(dueno)=?", (dueno,), fetch=True) or []
        stock_list = [ (r['proveedor'] or '').strip() for r in stock_rows if (r.get('proveedor') or '').strip() ]
        # Fuente meta
        meta_rows = db_query("SELECT DISTINCT nombre FROM proveedores_meta WHERE dueno=?", (dueno,), fetch=True) or []
        meta_list = [ (r['nombre'] or '').strip() for r in meta_rows if (r.get('nombre') or '').strip() ]
        # Fuente Excel manual
        manual_list = []
        try:
            if os.path.exists(MANUAL_PRODUCTS_FILE):
                import pandas as pd
                dfp = pd.read_excel(MANUAL_PRODUCTS_FILE)
                dfp.rename(columns={'Código': 'Codigo', 'Dueño': 'Dueno'}, inplace=True)
                if not dfp.empty:
                    dfp = dfp[dfp['Dueno'].astype(str).str.lower() == dueno]
                    manual_list = [str(v).strip() for v in dfp['Proveedor'].dropna().unique() if str(v).strip()]
        except Exception:
            pass
        # Fuente config
        cfg = DUENOS_CONFIG.get(dueno, {})
        cfg_list = [p for p in cfg.get('proveedores_excel', []) if p]

        # Consolidar todas las apariciones con flags
        all_raw = []
        def add_items(src, items):
            for it in items:
                norm = _normalizar_nombre_proveedor(it)
                all_raw.append({
                    'original': it,
                    'normalizado': norm,
                    'fuente': src,
                    'oculto': norm in ocultos_norm,
                    'eliminado': norm in eliminados_norm
                })
        add_items('stock', stock_list)
        add_items('meta', meta_list)
        add_items('excel_manual', manual_list)
        add_items('config', cfg_list)

        # Agrupar por normalizado
        agregados = {}
        for row in all_raw:
            k = row['normalizado']
            if k not in agregados:
                agregados[k] = {
                    'normalizado': k,
                    'originales': set(),
                    'fuentes': set(),
                    'oculto': False,
                    'eliminado': False
                }
            agregados[k]['originales'].add(row['original'])
            agregados[k]['fuentes'].add(row['fuente'])
            agregados[k]['oculto'] = agregados[k]['oculto'] or row['oculto']
            agregados[k]['eliminado'] = agregados[k]['eliminado'] or row['eliminado']
        agregados_list = []
        for k, v in agregados.items():
            agregados_list.append({
                'normalizado': v['normalizado'],
                'originales': sorted(list(v['originales'])),
                'fuentes': sorted(list(v['fuentes'])),
                'oculto': v['oculto'],
                'eliminado': v['eliminado']
            })
        agregados_list.sort(key=lambda x: x['normalizado'])

        return jsonify({
            'success': True,
            'dueno': dueno,
            'ocultos_norm': sorted(list(ocultos_norm)),
            'eliminados_norm': sorted(list(eliminados_norm)),
            'detalle_fuentes': all_raw,
            'agrupados': agregados_list
        })
    except Exception as e:
        print(f"Error debug_historial_fuentes: {e}")
        return jsonify({'success': False, 'error': 'error interno'}), 500

@app.route('/debug_fuentes_proveedor')
@login_required
def debug_fuentes_proveedor():
    """Diagnóstico: muestra de qué fuentes (stock, meta, excel manual, config) se está intentando componer la lista de historial para un dueño.
       Parámetros: nombre (proveedor), dueno.
    """
    try:
        nombre = (request.args.get('nombre') or '').strip()
        dueno = (request.args.get('dueno') or '').strip().lower()
        if not nombre or not dueno:
            return jsonify({'success': False, 'error': 'nombre y dueno requeridos'}), 400
        nombre_l = nombre.lower()
        oculto = db_query("SELECT 1 FROM proveedores_ocultos WHERE LOWER(nombre)=? AND dueno=? LIMIT 1", (nombre_l, dueno), fetch=True)
        eliminado = db_query("SELECT 1 FROM proveedores_eliminados WHERE LOWER(nombre)=? AND dueno=? LIMIT 1", (nombre_l, dueno), fetch=True)
        en_stock = db_query("SELECT DISTINCT proveedor FROM stock WHERE LOWER(proveedor)=? AND LOWER(dueno)=? LIMIT 5", (nombre_l, dueno), fetch=True)
        en_meta = db_query("SELECT nombre FROM proveedores_meta WHERE LOWER(nombre)=? AND dueno=?", (nombre_l, dueno), fetch=True)
        # Excel manual
        excel_manual = []
        try:
            if os.path.exists(MANUAL_PRODUCTS_FILE):
                import pandas as pd
                dfp = pd.read_excel(MANUAL_PRODUCTS_FILE)
                dfp.rename(columns={'Código': 'Codigo', 'Dueño': 'Dueno'}, inplace=True)
                if not dfp.empty:
                    dfp['Proveedor_l'] = dfp['Proveedor'].astype(str).str.lower()
                    dfp['Dueno_l'] = dfp['Dueno'].astype(str).str.lower()
                    excel_manual = dfp[(dfp['Proveedor_l']==nombre_l) & (dfp['Dueno_l']==dueno)].head(5).to_dict(orient='records')
        except Exception as _e:
            pass
        en_config = False
        cfg = DUENOS_CONFIG.get(dueno, {})
        for p in cfg.get('proveedores_excel', []):
            if p.lower() == nombre_l:
                en_config = True
                break
        return jsonify({
            'success': True,
            'nombre': nombre,
            'dueno': dueno,
            'oculto': bool(oculto),
            'eliminado': bool(eliminado),
            'fuentes': {
                'stock': bool(en_stock),
                'meta': bool(en_meta),
                'excel_manual_match': bool(excel_manual),
                'config_excel': en_config
            },
            'detalles': {
                'stock_raw': en_stock,
                'meta_raw': en_meta,
                'excel_manual_sample': excel_manual
            }
        })
    except Exception as e:
        print(f"Error debug_fuentes_proveedor: {e}")
        return jsonify({'success': False, 'error': 'error interno'}), 500

@app.route('/agregar_carrito_ajax', methods=['POST'])
@login_required
def agregar_carrito_ajax():
    try:
        data = request.get_json()
        print(f"🛒 [AGREGAR_CARRITO_AJAX] Datos recibidos: {data}")
        
        # Si viene con ID, es de productos_manual (base de datos)
        producto_id = data.get('id')
        print(f"🛒 [AGREGAR_CARRITO_AJAX] Producto ID: {producto_id}")
        if producto_id:
            cantidad = int(data.get('cantidad', 1))
            print(f"🛒 [AGREGAR_CARRITO_AJAX] Procesando producto manual - ID: {producto_id}, Cantidad: {cantidad}")
            
            producto = db_query("SELECT * FROM productos_manual WHERE id = ?", (producto_id,), fetch=True)
            if not producto:
                print(f"🛒 [AGREGAR_CARRITO_AJAX] ERROR: Producto no encontrado en BD")
                return jsonify({'success': False, 'error': 'Producto no encontrado'})
            
            producto = producto[0]
            carrito = session.get('carrito', [])
            
            # Verificar si el producto ya está en el carrito
            encontrado = False
            for item in carrito:
                if item.get('id') == producto_id:
                    item['cantidad'] += cantidad
                    encontrado = True
                    break
            
            if not encontrado:
                carrito.append({
                    'id': producto_id,
                    'nombre': producto['nombre'],
                    'codigo': producto.get('codigo', ''),
                    'precio': float(producto['precio']) if producto['precio'] else 0.0,
                    'cantidad': cantidad,
                    'fecha_compra': datetime.now().strftime('%Y-%m-%d'),
                    'proveedor': data.get('proveedor', ''),
                    'observaciones': data.get('observaciones', ''),
                    'precio_texto': str(producto['precio']) if producto['precio'] else '0',
                    'avisar_bajo_stock': 0,
                    'min_stock_aviso': None
                })
            
            session['carrito'] = carrito
            
            # Renderizar fragment actualizado del carrito
            try:
                html = render_template('carrito_fragment_simple.html', carrito=carrito)
                return jsonify({'success': True, 'msg': f'Producto "{producto["nombre"]}" agregado al carrito', 'html': html})
            except Exception as render_error:
                print(f"Error renderizando carrito_fragment: {render_error}")
            try:
                html = render_template('carrito_fragment_simple.html', carrito=carrito)
                return jsonify({'success': True, 'msg': f'Producto "{producto["nombre"]}" agregado al carrito', 'html': html})
            except Exception as render_error:
                print(f"Error renderizando carrito_fragment: {render_error}")
                return jsonify({'success': True, 'msg': f'Producto "{producto["nombre"]}" agregado al carrito', 'reload': True})
        
        # Si no viene con ID, es de búsqueda Excel
        else:
            nombre = data.get('nombre', '').strip()
            codigo = data.get('codigo', '').strip()
            precio_raw = data.get('precio', '')
            cantidad_raw = data.get('cantidad', '1')
            fecha_compra = data.get('fecha_compra', datetime.now().strftime('%Y-%m-%d'))
            proveedor = data.get('proveedor', '').strip()
            observaciones = data.get('observaciones', '').strip()
            
            print(f"🛒 [AGREGAR_CARRITO_AJAX] Procesando producto Excel:")
            print(f"  - Nombre: '{nombre}'")
            print(f"  - Código: '{codigo}'")
            print(f"  - Precio raw: '{precio_raw}' (tipo: {type(precio_raw)})")
            print(f"  - Cantidad: '{cantidad_raw}'")
            print(f"  - Proveedor: '{proveedor}'")
            
            # Validar nombre obligatorio
            if not nombre:
                print(f"🛒 [AGREGAR_CARRITO_AJAX] ERROR: Nombre vacío")
                return jsonify({'success': False, 'error': 'El nombre del producto es obligatorio'})
            
            # Procesar cantidad
            try:
                cantidad = int(cantidad_raw) if cantidad_raw else 1
                if cantidad <= 0:
                    cantidad = 1
            except (ValueError, TypeError):
                cantidad = 1
            
            # Procesar precio (puede ser vacío)
            precio = 0.0
            precio_texto = '0'
            if precio_raw:
                if isinstance(precio_raw, (int, float)):
                    precio = float(precio_raw)
                    precio_texto = str(precio_raw)
                    print(f"🛒 [AGREGAR_CARRITO_AJAX] Precio numérico procesado: {precio}")
                else:
                    precio_str = str(precio_raw).strip()
                    if precio_str:
                        precio, precio_error = parse_price(precio_str)
                        precio_texto = precio_str
                        print(f"🛒 [AGREGAR_CARRITO_AJAX] Precio string procesado: '{precio_str}' -> {precio} (error: {precio_error})")
            else:
                print(f"🛒 [AGREGAR_CARRITO_AJAX] Precio vacío, usando 0.0")
            
            carrito = session.get('carrito', [])
            item_carrito = {
                'id': f'excel_{len(carrito)}_{datetime.now().timestamp()}',
                'nombre': nombre,
                'codigo': codigo,
                'precio': precio,
                'cantidad': cantidad,
                'fecha_compra': fecha_compra,
                'proveedor': proveedor,
                'observaciones': observaciones,
                'precio_texto': precio_texto,
                'avisar_bajo_stock': 0,
                'min_stock_aviso': None
            }
            carrito.append(item_carrito)
            
            print(f"🛒 [AGREGAR_CARRITO_AJAX] Item agregado al carrito: {item_carrito}")
            session['carrito'] = carrito
            print(f"🛒 [AGREGAR_CARRITO_AJAX] Carrito actualizado, total items: {len(carrito)}")
            
            # Renderizar fragment actualizado del carrito
            try:
                html = render_template('carrito_fragment_simple.html', carrito=carrito)
                print(f"🛒 [AGREGAR_CARRITO_AJAX] ✅ ÉXITO: Producto agregado, HTML generado")
                return jsonify({'success': True, 'msg': f'Producto "{nombre}" agregado al carrito', 'html': html})
            except Exception as render_error:
                print(f"🛒 [AGREGAR_CARRITO_AJAX] ERROR renderizando carrito_fragment: {render_error}")
                return jsonify({'success': True, 'msg': f'Producto "{nombre}" agregado al carrito', 'reload': True})
        
    except Exception as e:
        print(f"🛒 [AGREGAR_CARRITO_AJAX] ERROR GENERAL: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': f'Error al procesar: {str(e)}'})

@app.route('/agregar_carrito_manual_ajax', methods=['POST'])
@login_required
def agregar_carrito_manual_ajax():
    try:
        data = request.get_json()
        nombre = data.get('nombre', '').strip()
        codigo = data.get('codigo', '').strip()
        precio_raw = data.get('precio', '')
        cantidad_raw = data.get('cantidad', '1')
        fecha_compra = data.get('fecha_compra', datetime.now().strftime('%Y-%m-%d'))
        proveedor = data.get('proveedor', '').strip()
        observaciones = data.get('observaciones', '').strip()
        avisar_bajo_stock_raw = data.get('avisar_bajo_stock', 0)
        min_stock_aviso_raw = data.get('min_stock_aviso', '')
        
        # Validar nombre obligatorio
        if not nombre:
            return jsonify({'success': False, 'error': 'El nombre del producto es obligatorio'})
        
        # Procesar cantidad
        try:
            cantidad = int(cantidad_raw) if cantidad_raw else 1
            if cantidad <= 0:
                cantidad = 1
        except (ValueError, TypeError):
            cantidad = 1
        
        # Procesar precio (puede ser vacío)
        precio = 0.0
        precio_texto = ''
        if precio_raw:
            if isinstance(precio_raw, (int, float)):
                precio = float(precio_raw)
                precio_texto = str(precio_raw)
            else:
                precio_str = str(precio_raw).strip()
                if precio_str:
                    precio, precio_error = parse_price(precio_str)
                    precio_texto = precio_str
                else:
                    precio = 0.0
                    precio_texto = '0'
        else:
            precio = 0.0
            precio_texto = '0'
        
        # Validar aviso de bajo stock
        try:
            avisar_bajo_stock = 1 if int(avisar_bajo_stock_raw) == 1 else 0
        except Exception:
            avisar_bajo_stock = 0
        min_stock_aviso = None
        if avisar_bajo_stock:
            try:
                min_stock_aviso_val = int(min_stock_aviso_raw) if str(min_stock_aviso_raw).strip() != '' else None
            except (ValueError, TypeError):
                min_stock_aviso_val = None
            if min_stock_aviso_val is None or min_stock_aviso_val <= 0:
                avisar_bajo_stock = 0
                min_stock_aviso = None
            else:
                if min_stock_aviso_val > cantidad:
                    return jsonify({'success': False, 'error': 'La cantidad mínima para aviso no puede ser mayor que la cantidad comprada.'})
                min_stock_aviso = min_stock_aviso_val

        carrito = session.get('carrito', [])
        carrito.append({
            'id': f'manual_{len(carrito)}_{datetime.now().timestamp()}',
            'nombre': nombre,
            'codigo': codigo,
            'precio': precio,
            'cantidad': cantidad,
            'fecha_compra': fecha_compra,
            'proveedor': proveedor,
            'observaciones': observaciones,
            'precio_texto': precio_texto,
            'avisar_bajo_stock': avisar_bajo_stock,
            'min_stock_aviso': min_stock_aviso
        })
        
        session['carrito'] = carrito
        # Devolver fragmento actualizado del carrito para refrescar UI sin recargar
        try:
            html = render_template('carrito_fragment_simple.html', carrito=carrito)
            return jsonify({'success': True, 'msg': f'Producto "{nombre}" agregado al carrito', 'html': html})
        except Exception as render_error:
            print(f"Error renderizando carrito_fragment: {render_error}")
            return jsonify({'success': True, 'msg': f'Producto "{nombre}" agregado al carrito', 'reload': True})
        
    except Exception as e:
        return jsonify({'success': False, 'error': f'Error al procesar: {str(e)}'})

@app.route('/eliminar_producto_stock', methods=['POST'])
@login_required
def eliminar_producto_stock():
    producto_id = request.form.get('id')
    if producto_id:
        result = db_query("DELETE FROM stock WHERE id = ?", (producto_id,))
        if result:
            flash('Producto eliminado del stock exitosamente.', 'success')
        else:
            flash('Error al eliminar el producto del stock.', 'danger')
    return redirect(url_for('stock'))

@app.route('/eliminar_seleccionados', methods=['POST'])
@login_required
def eliminar_seleccionados():
    try:
        ids = []
        is_json = False
        dueno = None
        # Soportar tanto JSON (AJAX) como formulario (POST estándar)
        if request.is_json:
            data = request.get_json(silent=True) or {}
            ids = data.get('ids', [])
            dueno = (data.get('dueno') or '').strip().lower() or None
            is_json = True
        else:
            ids = request.form.getlist('seleccionados')
            dueno = (request.form.get('dueno') or '').strip().lower() or None
        
        if not ids:
            if is_json:
                return jsonify({'success': False, 'error': 'No hay elementos seleccionados'})
            flash('No hay elementos seleccionados.', 'warning')
            return redirect(url_for('historial'))
        
        placeholders = ','.join(['?'] * len(ids))
        if dueno:
            # Borrar solo del dueño especificado
            result = db_query(f"DELETE FROM stock WHERE dueno = ? AND id IN ({placeholders})", tuple([dueno] + ids))
        else:
            result = db_query(f"DELETE FROM stock WHERE id IN ({placeholders})", ids)
        
        if is_json:
            if result:
                return jsonify({'success': True, 'message': f'{len(ids)} productos eliminados'})
            else:
                return jsonify({'success': False, 'error': 'Error al eliminar los productos'})
        else:
            if result:
                flash(f'{len(ids)} productos eliminados.', 'success')
            else:
                flash('Error al eliminar los productos seleccionados.', 'danger')
            return redirect(url_for('historial'))
            
    except Exception as e:
        if request.is_json:
            return jsonify({'success': False, 'error': str(e)})
        flash(f'Error: {str(e)}', 'danger')
        return redirect(url_for('historial'))

@app.route('/eliminar_todo_historial', methods=['POST'])
@login_required
def eliminar_todo_historial():
    try:
        dueno = None
        if request.is_json:
            data = request.get_json(silent=True) or {}
            dueno = (data.get('dueno') or '').strip().lower() or None
        else:
            dueno = (request.form.get('dueno') or '').strip().lower() or None
        if dueno:
            result = db_query("DELETE FROM stock WHERE dueno = ?", (dueno,))
        else:
            result = db_query("DELETE FROM stock")
        if request.is_json:
            if result:
                return jsonify({'success': True, 'message': 'Todo el historial eliminado'})
            return jsonify({'success': False, 'error': 'Error al eliminar el historial'})
        else:
            if result:
                flash('Todo el historial ha sido eliminado.', 'success')
            else:
                flash('Error al eliminar el historial.', 'danger')
    except Exception as e:
        if request.is_json:
            return jsonify({'success': False, 'error': str(e)})
        flash(f'Error: {str(e)}', 'danger')
    
    return redirect(url_for('historial'))

@app.route('/eliminar_manual', methods=['GET', 'POST'])
@login_required
def eliminar_manual():
    """Ruta para gestionar productos manuales"""
    # Preparar listas de proveedores por dueño para la UI
    try:
        mappings = db_query("""
            SELECT pm.id, pm.nombre, pd.dueno 
            FROM proveedores_manual pm 
            JOIN proveedores_duenos pd ON pm.id = pd.proveedor_id 
            ORDER BY pm.nombre
        """, fetch=True) or []
        print(f"[DEBUG] eliminar_manual usando proveedores_duenos: {len(mappings)} mappings")
    except Exception as e:
        print(f"[DEBUG] eliminar_manual fallback a proveedores_meta: {e}")
        try:
            mappings = db_query("SELECT pm.id, pm.nombre, m.dueno FROM proveedores_manual pm JOIN proveedores_meta m ON LOWER(m.nombre)=LOWER(pm.nombre) ORDER BY pm.nombre", fetch=True) or []
            print(f"[DEBUG] eliminar_manual usando proveedores_meta: {len(mappings)} mappings")
        except Exception as e2:
            print(f"[DEBUG] eliminar_manual ambas tablas fallan: {e2}")
            mappings = []
    
    proveedores_ricky = []
    proveedores_fg = []
    for row in mappings:
        entry = { 'id': row['id'], 'nombre': row['nombre'], 'dueno': row['dueno'] }
        if row['dueno'] == 'ricky':
            proveedores_ricky.append(entry)
        else:
            proveedores_fg.append(entry)
    return render_template('eliminar_manual.html', proveedores_ricky=proveedores_ricky, proveedores_fg=proveedores_fg)

@app.route('/obtener_productos_proveedor/<int:proveedor_id>')
@login_required
def obtener_productos_proveedor(proveedor_id):
    """Obtener productos de un proveedor específico (AJAX)"""
    try:
        productos = db_query(
            "SELECT id, nombre, codigo, precio FROM productos_manual WHERE proveedor_id = ? ORDER BY nombre",
            (proveedor_id,),
            fetch=True
        )
        return jsonify({'success': True, 'productos': productos})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/agregar_producto_manual', methods=['POST'])
@login_required
def agregar_producto_manual():
    """Agregar producto manual al Excel de productos manuales"""
    try:
        # Obtener datos del formulario
        nombre = request.form.get('nombre', '').strip()
        codigo = request.form.get('codigo', '').strip()
        precio_str = request.form.get('precio', '').strip()
        observaciones = request.form.get('observaciones', '').strip()
        
        # Determinar proveedor y dueño
        proveedor_id = request.form.get('proveedor_id')
        nuevo_proveedor = request.form.get('nuevo_proveedor', '').strip()
        dueno_nuevo_proveedor = request.form.get('dueno_nuevo_proveedor', '').strip()
        
        if not nombre:
            flash('El nombre del producto es obligatorio.', 'danger')
            return redirect(url_for('agregar_producto'))
        
        # Validar precio
        precio, error_precio = parse_price(precio_str)
        if error_precio and precio_str:  # Solo mostrar error si se ingresó algo
            flash(f'Error en el precio: {error_precio}', 'danger')
            return redirect(url_for('agregar_producto'))
        
        # Determinar proveedor y dueño
        proveedor_nombre = ''
        dueno = ''
        
        if proveedor_id:
            # Proveedor existente (valor viene como "id|dueno")
            try:
                prov_parts = proveedor_id.split('|')
                prov_id_int = int(prov_parts[0])
                dueno_sel = prov_parts[1] if len(prov_parts) > 1 else None
                proveedor_data = db_query("SELECT nombre FROM proveedores_manual WHERE id = ?", (prov_id_int,), fetch=True)
                if proveedor_data:
                    proveedor_nombre = proveedor_data[0]['nombre']
                    dueno = dueno_sel or 'ferreteria_general'
                else:
                    flash('Proveedor no encontrado.', 'danger')
                    return redirect(url_for('agregar_producto'))
            except Exception:
                flash('ID de proveedor inválido.', 'danger')
                return redirect(url_for('agregar_producto'))
        elif nuevo_proveedor and dueno_nuevo_proveedor:
            # Nuevo proveedor
            proveedor_nombre = nuevo_proveedor
            dueno = dueno_nuevo_proveedor
            
            # Agregar el nuevo proveedor a la base de datos con dueño específico
            db_query("INSERT OR IGNORE INTO proveedores_manual (nombre, dueno) VALUES (?, ?)", (proveedor_nombre, dueno))
        else:
            flash('Debe seleccionar un proveedor existente o crear uno nuevo con dueño.', 'danger')
            return redirect(url_for('agregar_producto'))
        
        # Agregar a la base de datos de productos manuales
        result_db = agregar_producto_db_manual(codigo, proveedor_nombre, nombre, precio, observaciones, dueno)
        
        # También agregar al Excel para compatibilidad con código legacy
        result_excel = agregar_producto_excel_manual(codigo, proveedor_nombre, nombre, precio, observaciones, dueno)
        
        if result_db:
            flash(f'Producto "{nombre}" agregado al catálogo manual de {DUENOS_CONFIG[dueno]["nombre"]}.', 'success')
        else:
            flash('Error al agregar el producto al catálogo manual.', 'danger')
            
    except Exception as e:
        flash(f'Error al procesar el producto: {str(e)}', 'danger')
    
    return redirect(url_for('agregar_producto'))

@app.route('/eliminar_producto_manual', methods=['POST'])
@login_required
def eliminar_producto_manual():
    """Eliminar producto manual"""
    producto_id = request.form.get('id')
    if producto_id:
        result = db_query("DELETE FROM productos_manual WHERE id = ?", (producto_id,))
        if result:
            flash('Producto eliminado exitosamente.', 'success')
        else:
            flash('Error al eliminar el producto.', 'danger')
    return redirect(url_for('productos_manual'))

# Rutas adicionales que los templates esperan
@app.route('/stock.html')
@login_required
def stock_html():
    """Crear template stock.html si no existe"""
    return render_template('stock.html')

@app.route('/productos_manual.html')
@login_required
def productos_manual_html():
    """Crear template productos_manual.html si no existe"""
    return render_template('productos_manual.html')

@app.route('/reportes.html')
@login_required
def reportes_html():
    """Crear template reportes.html si no existe"""
    return render_template('reportes.html')

@app.route('/carrito.html') 
@login_required
def carrito_html():
    """Crear template carrito.html si no existe"""
    return render_template('carrito.html')

# Rutas adicionales que podrían faltar
@app.route('/carrito_accion', methods=['POST'])
@login_required
def carrito_accion():
    """Manejar acciones del carrito (sumar, restar, eliminar, actualizar)"""
    try:
        data = request.get_json(silent=True) or {}
        accion = data.get('accion')
        try:
            idx = int(data.get('idx', 0))
        except Exception:
            idx = 0
        cantidad = data.get('cantidad')
        
        carrito = session.get('carrito', [])
        
        if idx >= len(carrito):
            return jsonify({'success': False, 'error': 'Índice inválido'})
        
        if accion == 'eliminar':
            carrito.pop(idx)
        elif accion == 'sumar':
            carrito[idx]['cantidad'] += 1
        elif accion == 'restar':
            if carrito[idx]['cantidad'] > 1:
                carrito[idx]['cantidad'] -= 1
        elif accion == 'actualizar' and cantidad:
            carrito[idx]['cantidad'] = max(1, int(cantidad))
        
        session['carrito'] = carrito
        
        # Renderizar fragment del carrito actualizado
        try:
            html = render_template('carrito_fragment_simple.html', carrito=carrito)
            return jsonify({'success': True, 'html': html})
        except Exception as render_error:
            print(f"Error renderizando carrito_fragment: {render_error}")
            return jsonify({'success': True, 'reload': True})
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/carrito_cargar', methods=['POST'])
@login_required
def carrito_cargar():
    """Cargar todos los productos del carrito al stock (AJAX JSON)."""
    try:
        data = request.get_json(silent=True) or {}
        if not data.get('cargar_carrito'):
            return jsonify({'success': False, 'error': 'Solicitud inválida'})
        dueno_dest = data.get('dueno_dest', 'ferreteria_general')
        carrito = session.get('carrito', [])
        productos_agregados = 0
        for item in carrito:
            try:
                avisar_bajo_stock = int(item.get('avisar_bajo_stock', 0))
                min_stock_aviso_val = item.get('min_stock_aviso')
                try:
                    min_stock_aviso_int = int(min_stock_aviso_val) if min_stock_aviso_val is not None and str(min_stock_aviso_val) != '' else None
                except (ValueError, TypeError):
                    min_stock_aviso_int = None
                cantidad_int = int(item.get('cantidad', 1))
                if avisar_bajo_stock and (min_stock_aviso_int is None or min_stock_aviso_int <= 0):
                    avisar_bajo_stock = 0
                    min_stock_aviso_int = None
                if avisar_bajo_stock and min_stock_aviso_int and min_stock_aviso_int > cantidad_int:
                    min_stock_aviso_int = cantidad_int
                db_query(
                    """
                    INSERT INTO stock (codigo, nombre, precio, cantidad, fecha_compra,
                                       proveedor, observaciones, precio_texto, avisar_bajo_stock, min_stock_aviso, dueno)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        item.get('codigo', ''),
                        item.get('nombre', ''),
                        float(item.get('precio', 0)),
                        cantidad_int,
                        item.get('fecha_compra', datetime.now().strftime('%Y-%m-%d')),
                        item.get('proveedor', ''),
                        item.get('observaciones', ''),
                        item.get('precio_texto', ''),
                        avisar_bajo_stock,
                        min_stock_aviso_int,
                        dueno_dest,
                    ),
                )
                productos_agregados += 1
            except Exception as e:
                print(f"Error al agregar producto al stock: {e}")
        # Limpiar carrito y devolver fragmento actualizado
        session['carrito'] = []
        try:
            html = render_template('carrito_fragment_simple.html', carrito=session['carrito'])
        except Exception:
            html = ''
        return jsonify({'success': True, 'agregados': productos_agregados, 'html': html})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/actualizar_stock/<int:id>', methods=['POST'])
@login_required
def actualizar_stock(id):
    """Actualizar stock cuando se vende un producto"""
    try:
        if request.is_json:
            data = request.get_json(silent=True) or {}
            cantidad_vendida = int(data.get('cantidad_vendida', 1))
        else:
            cantidad_vendida = int(request.form.get('cantidad_vendida', 1))
        
        # Obtener producto actual
        producto = db_query("SELECT * FROM stock WHERE id = ?", (id,), fetch=True)
        if not producto:
            if request.is_json:
                return jsonify({'success': False, 'error': 'Producto no encontrado'})
            flash('Producto no encontrado.', 'danger')
            return redirect(url_for('historial'))
        
        producto = producto[0]
        nueva_cantidad = max(0, producto['cantidad'] - cantidad_vendida)
        
        # Actualizar cantidad
        result = db_query("UPDATE stock SET cantidad = ? WHERE id = ?", (nueva_cantidad, id))
        
        if result:
            if not request.is_json:
                flash(f'Vendidas {cantidad_vendida} unidades. Nueva cantidad: {nueva_cantidad}', 'success')
            # Aviso de bajo stock si aplica
            mostrar_toast = False
            try:
                if producto.get('avisar_bajo_stock') and producto.get('min_stock_aviso') is not None:
                    umbral = int(producto.get('min_stock_aviso'))
                    if nueva_cantidad <= umbral:
                        mensaje = f'Producto "{producto.get("nombre", "")}" bajo en stock (quedan {nueva_cantidad}, umbral {umbral}).'
                        notificaciones = session.get('notificaciones', [])
                        notificaciones.append({
                            'mensaje': mensaje,
                            'ts': datetime.now().isoformat(timespec='seconds'),
                            'nombre': producto.get('nombre', ''),
                            'codigo': producto.get('codigo', ''),
                            'proveedor': producto.get('proveedor', ''),
                            'precio': producto.get('precio', 0),
                            'precio_texto': producto.get('precio_texto', '')
                        })
                        session['notificaciones'] = notificaciones
                        session['notificaciones_leidas'] = False
                        # Permitir que el aviso inferior se muestre una vez por nueva notificación
                        session['notificacion_aviso_mostrado'] = False
                        # Persistir en BD
                        try:
                            user_id = session.get('user_id')
                            db_query(
                                "INSERT INTO notificaciones (codigo,nombre,proveedor,mensaje,ts,leida,user_id) VALUES (?,?,?,?,?,?,?)",
                                (
                                    producto.get('codigo',''),
                                    producto.get('nombre',''),
                                    producto.get('proveedor',''),
                                    mensaje,
                                    datetime.now().isoformat(timespec='seconds'),
                                    0,
                                    user_id
                                )
                            )
                        except Exception as _e:
                            print(f"[WARN] No se pudo persistir notificación: {_e}")
                        mostrar_toast = True
            except Exception as _:
                mostrar_toast = False
            # No mostrar toast si no está habilitado el aviso de bajo stock
            if request.is_json:
                avisar_flag = int(producto.get('avisar_bajo_stock') or 0)
                min_aviso = producto.get('min_stock_aviso')
                try:
                    min_aviso_int = int(min_aviso) if min_aviso not in (None, '', '0') else None
                except Exception:
                    min_aviso_int = None
                # Determinar estado
                if nueva_cantidad <= 0:
                    estado = 'sin'
                elif avisar_flag == 1 and min_aviso_int is not None and nueva_cantidad <= min_aviso_int:
                    estado = 'bajo'
                else:
                    estado = 'ok'
                return jsonify({
                    'success': True,
                    'nueva_cantidad': nueva_cantidad,
                    'avisar_bajo_stock': avisar_flag,
                    'min_stock_aviso': min_aviso_int,
                    'estado': estado,
                    'mostrar_toast': mostrar_toast
                })
        else:
            if request.is_json:
                return jsonify({'success': False, 'error': 'Error al actualizar el stock'})
            flash('Error al actualizar el stock.', 'danger')
            
    except Exception as e:
        if request.is_json:
            return jsonify({'success': False, 'error': str(e)})
        flash(f'Error: {str(e)}', 'danger')
    
    return redirect(url_for('historial'))

@app.route('/stock_row/<int:id>')
@login_required
def stock_row(id):
    """Devuelve el HTML <tr> de un producto del stock para reemplazo dinámico.
    Útil tras operaciones (venta, edición futura) sin recargar toda la tabla.
    """
    try:
        filas = db_query("SELECT * FROM stock WHERE id = ?", (id,), fetch=True)
        if not filas:
            return jsonify({'success': False, 'error': 'No encontrado'}), 404
        p = filas[0]
        # Normalizar tipos/valores esperados por el macro
        from flask import render_template
        html = render_template('fragmentos/fila_historial.html', p=p)
        return jsonify({'success': True, 'html': html, 'id': id})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/agregar_proveedor_manual_ajax', methods=['POST'])
@login_required
def agregar_proveedor_manual_ajax():
    try:
        data = request.get_json()
        nombre_raw = data.get('nombre', '').strip()
        nombre = ' '.join(nombre_raw.split())
        dueno = data.get('dueno', '').strip()  # Obtener el dueño del JSON
        
        correcciones = {
            'brementools': 'BremenTools', 'brementool': 'BremenTools', 'bremetools': 'BremenTools', 'bremetool': 'BremenTools', 'bremen tools': 'BremenTools', 'bremen-tools': 'BremenTools',
            'crossmaster': 'Crossmaster', 'cross master': 'Crossmaster', 'cross-master': 'Crossmaster', 'cross master.': 'Crossmaster', 'crossmaster.': 'Crossmaster', 'crossmaster,': 'Crossmaster', 'crossmaster tools': 'Crossmaster', 'cross master tools': 'Crossmaster', 'crossmastertools': 'Crossmaster'
        }
        nombre_lower = nombre.lower()
        if nombre_lower in correcciones:
            nombre = correcciones[nombre_lower]
        
        if not nombre:
            return jsonify({'success': False, 'msg': 'El nombre del proveedor no puede estar vacío'})
        
        # Validar dueño
        if dueno not in ('ricky', 'ferreteria_general', 'ambos'):
            dueno = 'ferreteria_general'  # Default
        
        # Determinar dueños a asignar
        duenos_a_asignar = ['ricky', 'ferreteria_general'] if dueno == 'ambos' else [dueno]
        
        # Verificar existencia case-insensitive antes de insertar
        existente = db_query("SELECT id, nombre FROM proveedores_manual WHERE LOWER(nombre)=LOWER(?) LIMIT 1", (nombre,), fetch=True)
        
        if existente:
            # El proveedor ya existe, añadir relaciones en proveedores_duenos
            proveedor_id = existente[0]['id']
            proveedor_nombre = existente[0]['nombre']
            
            # Añadir relaciones a proveedores_duenos
            for d in duenos_a_asignar:
                db_query("INSERT OR IGNORE INTO proveedores_duenos (proveedor_id, dueno) VALUES (?, ?)", (proveedor_id, d))
            
            # Obtener lista actualizada de proveedores
            proveedores = db_query("SELECT id, nombre FROM proveedores_manual ORDER BY nombre", fetch=True) or []
            
            if dueno == 'ambos':
                msg = f'El proveedor "{proveedor_nombre}" ahora está disponible para ambos dueños.'
            else:
                msg = f'El proveedor "{proveedor_nombre}" ahora está disponible para {dueno}.'
            
            return jsonify({
                'success': True, 
                'msg': msg,
                'proveedores': [{'id': p['id'], 'nombre': p['nombre']} for p in proveedores]
            })
        else:
            # Insertar nuevo proveedor
            db_query("INSERT INTO proveedores_manual (nombre) VALUES (?)", (nombre,))
            
            # Obtener el ID del proveedor recién insertado
            nuevo_proveedor = db_query("SELECT id FROM proveedores_manual WHERE LOWER(nombre)=LOWER(?) LIMIT 1", (nombre,), fetch=True)
            if not nuevo_proveedor:
                return jsonify({'success': False, 'msg': 'Error al crear el proveedor'})
            
            proveedor_id = nuevo_proveedor[0]['id']
            
            # Añadir relaciones a proveedores_duenos
            for d in duenos_a_asignar:
                db_query("INSERT OR IGNORE INTO proveedores_duenos (proveedor_id, dueno) VALUES (?, ?)", (proveedor_id, d))
            
            # Obtener lista actualizada de proveedores
            proveedores = db_query("SELECT id, nombre FROM proveedores_manual ORDER BY nombre", fetch=True) or []
            
            if dueno == 'ambos':
                msg = f'Proveedor "{nombre}" agregado exitosamente para ambos dueños'
            else:
                msg = f'Proveedor "{nombre}" agregado exitosamente para {dueno}'
            
            return jsonify({
                'success': True, 
                'msg': msg,
                'proveedores': [{'id': p['id'], 'nombre': p['nombre']} for p in proveedores]
            })
    except Exception as e:
        print(f"Error al agregar proveedor AJAX: {e}")
        return jsonify({'success': False, 'msg': f'Error: {str(e)}'})

# Rutas AJAX que faltan para eliminar_manual.html
@app.route('/eliminar_manual_ajax', methods=['POST'])
@login_required
def eliminar_manual_ajax():
    """Eliminar producto manual via AJAX"""
    try:
        data = request.get_json()
        codigo_a_eliminar = data.get('codigo_a_eliminar', '').strip()
        search_term = data.get('search_term', '').strip()
        
        if not codigo_a_eliminar:
            return jsonify({'success': False, 'msg': 'Código del producto es obligatorio.'})
        
        # Para este ejemplo, como no hay Excel, simulo eliminar de productos_manual
        result = db_query("DELETE FROM productos_manual WHERE codigo = ?", (codigo_a_eliminar,))
        
        if result:
            # Buscar productos restantes
            productos = []
            if search_term:
                like_pattern = f'%{search_term}%'
                productos = db_query(
                    "SELECT pm.nombre, pm.codigo, pm.precio, p.nombre as proveedor FROM productos_manual pm LEFT JOIN proveedores_manual p ON pm.proveedor_id = p.id WHERE pm.nombre LIKE ? OR pm.codigo LIKE ? OR p.nombre LIKE ?",
                    (like_pattern, like_pattern, like_pattern),
                    fetch=True
                )
            
            html = render_template('eliminar_manual_resultados.html', productos=productos, search_term=search_term)
            return jsonify({
                'success': True, 
                'msg': f'Producto con código "{codigo_a_eliminar}" eliminado exitosamente.',
                'html': html
            })
        else:
            return jsonify({'success': False, 'msg': 'Error al eliminar el producto o no existe.'})
            
    except Exception as e:
        return jsonify({'success': False, 'msg': f'Error: {str(e)}'})

@app.route('/eliminar_manual_todo_ajax', methods=['POST'])
@login_required
def eliminar_manual_todo_ajax():
    """Eliminar todos los productos manuales via AJAX"""
    try:
        data = request.get_json()
        search_term = data.get('search_term', '').strip()
        
        # Eliminar todos los productos manuales
        result = db_query("DELETE FROM productos_manual")
        
        if result:
            html = render_template('eliminar_manual_resultados.html', productos=[], search_term=search_term)
            return jsonify({
                'success': True, 
                'msg': 'Todos los productos manuales han sido eliminados.',
                'html': html
            })
        else:
            return jsonify({'success': False, 'msg': 'Error al eliminar los productos.'})
            
    except Exception as e:
        return jsonify({'success': False, 'msg': f'Error: {str(e)}'})

@app.route('/eliminar_manual_seleccionados_ajax', methods=['POST'])
@login_required
def eliminar_manual_seleccionados_ajax():
    """Eliminar productos manuales seleccionados via AJAX"""
    try:
        data = request.get_json()
        codigos = data.get('codigos', [])
        search_term = data.get('search_term', '').strip()
        
        if not codigos:
            return jsonify({'success': False, 'msg': 'No hay productos seleccionados.'})
        
        # Eliminar productos seleccionados
        placeholders = ','.join(['?'] * len(codigos))
        result = db_query(f"DELETE FROM productos_manual WHERE codigo IN ({placeholders})", codigos)
        
        if result:
            # Buscar productos restantes
            productos = []
            if search_term:
                like_pattern = f'%{search_term}%'
                productos = db_query(
                    "SELECT pm.nombre, pm.codigo, pm.precio, p.nombre as proveedor FROM productos_manual pm LEFT JOIN proveedores_manual p ON pm.proveedor_id = p.id WHERE pm.nombre LIKE ? OR pm.codigo LIKE ? OR p.nombre LIKE ?",
                    (like_pattern, like_pattern, like_pattern),
                    fetch=True
                )
            
            html = render_template('eliminar_manual_resultados.html', productos=productos, search_term=search_term)
            return jsonify({
                'success': True, 
                'msg': f'{len(codigos)} productos eliminados exitosamente.',
                'html': html
            })
        else:
            return jsonify({'success': False, 'msg': 'Error al eliminar los productos seleccionados.'})
            
    except Exception as e:
        return jsonify({'success': False, 'msg': f'Error: {str(e)}'})

@app.route('/gestionar_manual_buscar_ajax', methods=['POST'])
@login_required
def gestionar_manual_buscar_ajax():
    """Buscar producto manual para editar via AJAX"""
    try:
        data = request.get_json()
        valor = data.get('valor', '').strip()
        
        if not valor:
            return jsonify({'html': '<div class="alert alert-warning">Debe ingresar un valor para buscar.</div>'})
        
        # Buscar producto por código, nombre, proveedor o precio
        like_pattern = f'%{valor}%'
        productos = db_query(
            "SELECT pm.*, p.nombre as proveedor FROM productos_manual pm LEFT JOIN proveedores_manual p ON pm.proveedor_id = p.id WHERE pm.codigo LIKE ? OR pm.nombre LIKE ? OR p.nombre LIKE ? OR CAST(pm.precio AS TEXT) LIKE ? LIMIT 1",
            (like_pattern, like_pattern, like_pattern, like_pattern),
            fetch=True
        )
        
        if productos:
            producto = productos[0]
            # Crear un objeto compatible con el template
            producto_template = {
                'Codigo': producto['codigo'],
                'Nombre': producto['nombre'],
                'Precio': producto['precio'],
                'Proveedor': producto['proveedor']
            }
            html = render_template('editar_manual_formulario.html', producto=producto_template)
        else:
            html = '<div class="alert alert-info">No se encontró ningún producto con ese criterio.</div>'
        
        return jsonify({'html': html})
            
    except Exception as e:
        return jsonify({'html': f'<div class="alert alert-danger">Error: {str(e)}</div>'})

# --- Nueva gestión de productos manuales por proveedor/dueño ---
@app.route('/obtener_proveedores_por_dueno', methods=['POST'])
@login_required
def obtener_proveedores_por_dueno():
    try:
        data = request.get_json()
        dueno = data.get('dueno', '').strip()
        
        if not dueno:
            return jsonify({'success': False, 'msg': 'Dueño no especificado'})
        
        print(f"[DEBUG] obtener_proveedores_por_dueno llamado con dueño: '{dueno}'")
        
        # Detectar si es PostgreSQL y usar la sintaxis correcta
        use_postgres = _is_postgres_configured()
        print(f"[DEBUG] Usando PostgreSQL: {use_postgres}")
        
        if use_postgres:
            # Usar sintaxis PostgreSQL explícitamente
            proveedores = db_query(
                """
                SELECT DISTINCT p.nombre 
                FROM proveedores_manual p
                JOIN proveedores_duenos pd ON p.id = pd.proveedor_id
                WHERE pd.dueno = %s
                ORDER BY p.nombre
                """, 
                (dueno,), fetch=True
            )
        else:
            # Usar sintaxis SQLite
            proveedores = db_query(
                """
                SELECT DISTINCT p.nombre 
                FROM proveedores_manual p
                JOIN proveedores_duenos pd ON p.id = pd.proveedor_id
                WHERE pd.dueno = ?
                ORDER BY p.nombre
                """, 
                (dueno,), fetch=True
            )
        
        print(f"[DEBUG] Consulta ejecutada. Resultados: {proveedores}")
        
        resultado = [p['nombre'] for p in proveedores] if proveedores else []
        print(f"[DEBUG] obtener_proveedores_por_dueno - proveedores encontrados: {resultado}")
        
        return jsonify({
            'success': True, 
            'proveedores': resultado
        })
        
    except Exception as e:
        print(f"Error en obtener_proveedores_por_dueno: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'msg': f'Error: {str(e)}'})

# Endpoint de debug para verificar estado de las tablas
@app.route('/debug_proveedores_duenos/<string:dueno>')
def debug_proveedores_duenos_publico(dueno):
    """Endpoint público para debuggear proveedores_duenos sin autenticación"""
    try:
        print(f"[DEBUG_PUBLIC] Debugging proveedores para dueño: '{dueno}'")
        
        # Verificar si la tabla existe
        try:
            count_duenos = db_query("SELECT COUNT(*) as count FROM proveedores_duenos", fetch=True)
            print(f"[DEBUG_PUBLIC] Total registros en proveedores_duenos: {count_duenos[0]['count'] if count_duenos else 0}")
        except Exception as e:
            print(f"[DEBUG_PUBLIC] Error accediendo proveedores_duenos: {e}")
            return jsonify({'error': f'Tabla proveedores_duenos no existe: {e}'})
        
        # Verificar registros para el dueño específico
        try:
            use_postgres = _is_postgres_configured()
            print(f"[DEBUG_PUBLIC] Usando PostgreSQL: {use_postgres}")
            
            if use_postgres:
                registros_dueno = db_query(
                    "SELECT pd.proveedor_id, pd.dueno, pm.nombre FROM proveedores_duenos pd JOIN proveedores_manual pm ON pm.id = pd.proveedor_id WHERE pd.dueno = %s", 
                    (dueno,), fetch=True
                )
            else:
                registros_dueno = db_query(
                    "SELECT pd.proveedor_id, pd.dueno, pm.nombre FROM proveedores_duenos pd JOIN proveedores_manual pm ON pm.id = pd.proveedor_id WHERE pd.dueno = ?", 
                    (dueno,), fetch=True
                )
            
            print(f"[DEBUG_PUBLIC] Registros para {dueno}: {registros_dueno}")
            
            return jsonify({
                'dueno': dueno,
                'postgres': use_postgres,
                'total_registros_duenos': count_duenos[0]['count'] if count_duenos else 0,
                'registros_para_dueno': registros_dueno,
                'proveedores_nombres': [r['nombre'] for r in registros_dueno] if registros_dueno else []
            })
            
        except Exception as e:
            print(f"[DEBUG_PUBLIC] Error en consulta específica: {e}")
            import traceback
            traceback.print_exc()
            return jsonify({'error': f'Error en consulta: {e}'})
            
    except Exception as e:
        print(f"[DEBUG_PUBLIC] Error general: {e}")
        return jsonify({'error': f'Error general: {e}'})

@app.route('/obtener_proveedores_por_dueno_test', methods=['POST'])
@csrf.exempt  # Eximir este endpoint de protección CSRF para pruebas
def obtener_proveedores_por_dueno_test():
    """Endpoint de prueba sin requerir login o CSRF para probar la API"""
    try:
        data = request.get_json()
        dueno = data.get('dueno', '').strip()
        
        if not dueno:
            return jsonify({'success': False, 'msg': 'Dueño no especificado'})
        
        print(f"[DEBUG] obtener_proveedores_por_dueno_test llamado con dueño: '{dueno}'")
        
        # Usar la nueva tabla proveedores_duenos para obtener proveedores
        proveedores = db_query(
            """
            SELECT DISTINCT p.nombre 
            FROM proveedores_manual p
            JOIN proveedores_duenos pd ON p.id = pd.proveedor_id
            WHERE pd.dueno = ?
            ORDER BY p.nombre
            """, 
            (dueno,), fetch=True
        )
        
        resultado = [p['nombre'] for p in proveedores]
        print(f"[DEBUG] obtener_proveedores_por_dueno_test - proveedores encontrados: {resultado}")
        
        return jsonify({
            'success': True,
            'proveedores': resultado
        })
    except Exception as e:
        print(f"Error en obtener_proveedores_por_dueno_test: {e}")
        return jsonify({'success': False, 'msg': f'Error: {str(e)}'})

@app.route('/manual_listar_ajax', methods=['POST'])
@login_required
def manual_listar_ajax():
    try:
        data = request.get_json() or {}
        proveedor_id = data.get('proveedor_id')
        dueno = (data.get('dueno') or '').strip().lower()
        termino = (data.get('termino') or '').strip()
        if not proveedor_id and not dueno:
            return jsonify({'success': False, 'html': '<div class="alert alert-warning">Seleccione un proveedor o "Todos" para un dueño.</div>'})
        # 1) Intentar leer desde Excel (fuente primaria del catálogo manual)
        try:
            if os.path.exists(MANUAL_PRODUCTS_FILE):
                df = pd.read_excel(MANUAL_PRODUCTS_FILE)
                df.rename(columns={'Código': 'Codigo', 'Dueño': 'Dueno'}, inplace=True)
                if proveedor_id:
                    prov = db_query('SELECT nombre FROM proveedores_manual WHERE id = ?', (proveedor_id,), fetch=True)
                    if prov:
                        nombre_prov = prov[0]['nombre']
                        mask = df['Proveedor'].astype(str).str.contains(nombre_prov, case=False, na=False)
                        if dueno:
                            mask &= df['Dueno'].astype(str).str.lower().eq(dueno)
                        # Aplicar filtro de término de búsqueda también para proveedor específico
                        if termino:
                            tokens = [t.strip() for t in termino.split() if t.strip()]
                            for tok in tokens:
                                tok_mask = (
                                    df['Nombre'].astype(str).str.contains(tok, case=False, na=False) |
                                    df['Codigo'].astype(str).str.contains(tok, case=False, na=False)
                                )
                                mask &= tok_mask
                        # Aplicar la máscara y devolver resultados
                        df2 = df[mask]
                        if not df2.empty:
                            rows = []
                            for _, r in df2.iterrows():
                                rows.append({
                                    'codigo': str(r.get('Codigo', '') or ''),
                                    'nombre': str(r.get('Nombre', '') or ''),
                                    'precio': str(r.get('Precio', '') or ''),
                                    'proveedor': str(r.get('Proveedor', '') or ''),
                                    'dueno': str(r.get('Dueno', '') or '')
                                })
                            return jsonify({'success': True, 'productos': rows})
                    else:
                        mask = pd.Series(False, index=df.index)
                else:
                    # Todos los proveedores del dueño
                    mask = pd.Series(True, index=df.index)
                    if dueno:
                        mask &= df['Dueno'].astype(str).str.lower().eq(dueno)
                    if termino:
                        tokens = [t.strip() for t in termino.split() if t.strip()]
                        for tok in tokens:
                            tok_mask = (
                                df['Nombre'].astype(str).str.contains(tok, case=False, na=False) |
                                df['Codigo'].astype(str).str.contains(tok, case=False, na=False)
                            )
                            mask &= tok_mask
                    df2 = df[mask]
                    if not df2.empty:
                        rows = []
                        for _, r in df2.iterrows():
                            rows.append({
                                'codigo': str(r.get('Codigo', '') or ''),
                                'nombre': str(r.get('Nombre', '') or ''),
                                'precio': str(r.get('Precio', '') or ''),
                                'proveedor': str(r.get('Proveedor', '') or ''),
                                'dueno': str(r.get('Dueno', '') or '')
                            })
                        return jsonify({'success': True, 'productos': rows})
        except Exception as _e:
            # si Excel falla, se continúa con la base local
            pass
        # 2) Fallback: Si Excel falló, intentar leer desde Excel de nuevo con manejo de errores más robusto
        try:
            if os.path.exists(MANUAL_PRODUCTS_FILE):
                df = pd.read_excel(MANUAL_PRODUCTS_FILE)
                df.rename(columns={'Código': 'Codigo', 'Dueño': 'Dueno'}, inplace=True)
                
                # Aplicar filtros
                mask = pd.Series(True, index=df.index)
                
                if proveedor_id:
                    prov = db_query('SELECT nombre FROM proveedores_manual WHERE id = ?', (proveedor_id,), fetch=True)
                    if prov:
                        nombre_prov = prov[0]['nombre']
                        mask &= df['Proveedor'].astype(str).str.contains(nombre_prov, case=False, na=False)
                    else:
                        mask = pd.Series(False, index=df.index)
                
                if dueno:
                    mask &= df['Dueno'].astype(str).str.lower().eq(dueno)
                
                if termino:
                    tokens = [t.strip() for t in termino.split() if t.strip()]
                    for tok in tokens:
                        tok_mask = (
                            df['Nombre'].astype(str).str.contains(tok, case=False, na=False) |
                            df['Codigo'].astype(str).str.contains(tok, case=False, na=False)
                        )
                        mask &= tok_mask
                
                df2 = df[mask]
                if not df2.empty:
                    rows = []
                    for _, r in df2.iterrows():
                        rows.append({
                            'codigo': str(r.get('Codigo', '') or ''),
                            'nombre': str(r.get('Nombre', '') or ''),
                            'precio': str(r.get('Precio', '') or ''),
                            'proveedor': str(r.get('Proveedor', '') or ''),
                            'dueno': str(r.get('Dueno', '') or '')
                        })
                    return jsonify({'success': True, 'productos': rows})
        except Exception as e:
            print(f"Error en fallback Excel: {e}")
        
        # 3) Si todo falla, devolver lista vacía
        return jsonify({'success': True, 'productos': []})
    except Exception as e:
        return jsonify({'success': False, 'html': f'<div class="alert alert-danger">Error: {str(e)}</div>'})

@app.route('/manual_eliminar_seleccionados_ajax', methods=['POST'])
@login_required
def manual_eliminar_seleccionados_ajax():
    try:
        data = request.get_json() or {}
        codigos = data.get('codigos') or []
        dueno = (data.get('dueno') or '').strip().lower()
        proveedor_id = data.get('proveedor_id')
        if not codigos:
            return jsonify({'success': False, 'msg': 'No se enviaron códigos para eliminar.'})
        # Eliminar de la tabla productos_manual (si existe)
        eliminados = 0
        for c in codigos:
            c2 = str(c).strip()
            if not c2:
                continue
            try:
                res = db_query("DELETE FROM productos_manual WHERE codigo = ?", (c2,))
                if res:
                    eliminados += 1
            except Exception as e_del:
                print(f"[manual_eliminar_seleccionados_ajax] Error eliminando {c2}: {e_del}")
        # Actualizar Excel productos_manual.xlsx
        if os.path.exists(MANUAL_PRODUCTS_FILE):
            try:
                dfm = pd.read_excel(MANUAL_PRODUCTS_FILE)
                dfm.rename(columns={'Código': 'Codigo'}, inplace=True)
                before = len(dfm)
                dfm = dfm[~dfm['Codigo'].astype(str).isin([str(c).strip() for c in codigos])]  # filtrar
                if len(dfm) != before:
                    with pd.ExcelWriter(MANUAL_PRODUCTS_FILE, engine='openpyxl', mode='w') as writer:
                        dfm.to_excel(writer, index=False)
            except Exception as e_x:
                print(f"[manual_eliminar_seleccionados_ajax] Error actualizando Excel: {e_x}")
        return jsonify({'success': True, 'msg': f'Eliminados {eliminados} código(s).', 'eliminados': eliminados})
    except Exception as e:
        return jsonify({'success': False, 'msg': f'Error: {str(e)}'})
        # Validar proveedor ANTES de actualizar Excel

@app.route('/manual_eliminar_por_proveedor_ajax', methods=['POST'])
@login_required
def manual_eliminar_por_proveedor_ajax():
    """Eliminar TODOS los productos manuales de un proveedor específico (por proveedor_id + dueño).
    Request JSON: { proveedor_id: int, dueno: 'ricky'|'ferreteria_general' }
    - Borra de tabla productos_manual donde proveedor coincide (texto) y opcionalmente dueno.
    - Actualiza productos_manual.xlsx filtrando esas filas.
    """
    try:
        data = request.get_json(silent=True) or {}
        proveedor_id = data.get('proveedor_id')
        dueno = (data.get('dueno') or '').strip().lower() or None
        if not proveedor_id:
            return jsonify({'success': False, 'msg': 'Falta proveedor_id.'})

        # Obtener nombre del proveedor desde proveedores_manual
        prov_rows = db_query("SELECT nombre, dueno FROM proveedores_manual WHERE id = ?", (proveedor_id,), fetch=True)
        if not prov_rows:
            return jsonify({'success': False, 'msg': 'Proveedor no encontrado.'})
        proveedor_nombre = prov_rows[0]['nombre']
        dueno_prov = (prov_rows[0].get('dueno') or '').lower() if isinstance(prov_rows[0], dict) else dueno

        # Criterios de eliminación en BD
        condiciones_sql = ["proveedor = ?"]
        params = [proveedor_nombre]
        if dueno and dueno_prov and dueno_prov == dueno:
            condiciones_sql.append("(dueno IS NULL OR LOWER(dueno)=?)")
            params.append(dueno)
        deleted_count = 0
        try:
            # Obtener códigos para sincronizar Excel luego
            cod_rows = db_query(f"SELECT codigo FROM productos_manual WHERE {' AND '.join(condiciones_sql)}", tuple(params), fetch=True) or []
            codigos_eliminar = [r['codigo'] for r in cod_rows if r.get('codigo')]
            res = db_query(f"DELETE FROM productos_manual WHERE {' AND '.join(condiciones_sql)}", tuple(params))
            if res:
                deleted_count = len(codigos_eliminar)
        except Exception as e_del:
            print(f"[manual_eliminar_por_proveedor_ajax] Error eliminando en BD: {e_del}")

        # Actualizar Excel
        excel_deleted = 0
        if os.path.exists(MANUAL_PRODUCTS_FILE) and deleted_count:
            try:
                dfm = pd.read_excel(MANUAL_PRODUCTS_FILE)
                dfm.rename(columns={'Código': 'Codigo', 'Dueño': 'Dueno'}, inplace=True)
                before = len(dfm)
                # Filtrar filas cuyo Proveedor coincida (case-insensitive)
                mask = dfm['Proveedor'].astype(str).str.lower() == str(proveedor_nombre).lower()
                if dueno:
                    # Si queremos restringir por dueño, solo quitar esas
                    mask_dueno = dfm['Dueno'].astype(str).str.lower() == dueno
                    mask = mask & mask_dueno
                dfm = dfm[~mask]
                excel_deleted = before - len(dfm)
                if excel_deleted:
                    with pd.ExcelWriter(MANUAL_PRODUCTS_FILE, engine='openpyxl', mode='w') as writer:
                        dfm.to_excel(writer, index=False)
            except Exception as e_x:
                print(f"[manual_eliminar_por_proveedor_ajax] Error actualizando Excel: {e_x}")

        return jsonify({'success': True, 'msg': f'Eliminados {deleted_count} producto(s) (Excel removió {excel_deleted}).', 'eliminados': deleted_count, 'excel_eliminados': excel_deleted})
    except Exception as e:
        return jsonify({'success': False, 'msg': f'Error: {str(e)}'})
    except Exception as e:
        return jsonify({'success': False, 'msg': f'Error: {str(e)}'})

@app.route('/agregar_productos_masivo_ajax', methods=['POST'])
@login_required
def agregar_productos_masivo_ajax():
    """Agregar múltiples productos manuales de una vez"""
    try:
        data = request.get_json() or {}
        productos = data.get('productos', [])
        
        if not productos:
            return jsonify({'success': False, 'msg': 'No hay productos para agregar.'})
        
        productos_agregados = 0
        productos_con_error = []
        
        for producto in productos:
            try:
                nombre = producto.get('nombre', '').strip()
                codigo = producto.get('codigo', '').strip()
                precio_str = producto.get('precio', '').strip()
                proveedor_nombre = producto.get('proveedor', '').strip()
                dueno = producto.get('dueno', '').strip()
                observaciones = producto.get('observaciones', '').strip()
                
                if not nombre or not precio_str or not proveedor_nombre or not dueno:
                    productos_con_error.append(f"{nombre or 'Sin nombre'}: Datos incompletos")
                    continue
                
                # Validar precio
                precio, error_precio = parse_price(precio_str)
                if error_precio:
                    productos_con_error.append(f"{nombre}: Error en precio - {error_precio}")
                    continue
                
                # Validar que el proveedor existe para el dueño
                meta = db_query('SELECT 1 FROM proveedores_meta WHERE LOWER(nombre)=LOWER(?) AND dueno=?', (proveedor_nombre, dueno), fetch=True)
                if not meta:
                    dueno_nombre = DUENOS_CONFIG.get(dueno, {}).get('nombre', dueno)
                    productos_con_error.append(f"{nombre}: Proveedor '{proveedor_nombre}' no existe para {dueno_nombre}")
                    continue
                
                # Agregar al Excel
                if agregar_producto_excel_manual(codigo, proveedor_nombre, nombre, precio, observaciones, dueno):
                    productos_agregados += 1
                else:
                    productos_con_error.append(f"{nombre}: Error al agregar al Excel")
                    
            except Exception as e:
                productos_con_error.append(f"{producto.get('nombre', 'Sin nombre')}: {str(e)}")
        
        # Crear mensaje de resultado
        if productos_agregados == len(productos):
            mensaje = f"✅ Todos los productos agregados exitosamente ({productos_agregados} producto(s))."
        elif productos_agregados > 0:
            mensaje = f"⚠️ {productos_agregados} de {len(productos)} productos agregados exitosamente."
            if productos_con_error:
                mensaje += f" Errores: {', '.join(productos_con_error[:3])}"
                if len(productos_con_error) > 3:
                    mensaje += f" y {len(productos_con_error) - 3} más."
        else:
            mensaje = f"❌ No se pudo agregar ningún producto. Errores: {', '.join(productos_con_error[:3])}"
            if len(productos_con_error) > 3:
                mensaje += f" y {len(productos_con_error) - 3} más."
        
        return jsonify({
            'success': productos_agregados > 0,
            'msg': mensaje,
            'agregados': productos_agregados,
            'total': len(productos),
            'errores': productos_con_error
        })
        
    except Exception as e:
        return jsonify({'success': False, 'msg': f'Error general: {str(e)}'})

@app.route('/gestionar_manual_actualizar_ajax', methods=['POST'])
@login_required
def gestionar_manual_actualizar_ajax():
    """Actualizar producto manual via AJAX"""
    try:
        # Asegurar columnas necesarias (auto-reparación en instalaciones antiguas)
        ensure_productos_manual_columns()
        ensure_proveedores_manual_columns()
        data = request.get_json()
        codigo_original = data.get('codigo_original', '').strip()
        nombre = data.get('nombre', '').strip()
        # Aceptar ambas claves desde distintos templates: proveedor / proveedor_nombre
        proveedor_raw = (data.get('proveedor') or data.get('proveedor_nombre') or '')
        proveedor = proveedor_raw.strip()
        codigo = data.get('codigo', '').strip()
        precio_str = data.get('precio', '').strip()
        dueno_input = (data.get('dueno') or '').strip().lower()
        print(f"[DEBUG editar_manual_early] codigo_original={codigo_original} codigo_nuevo={codigo} proveedor_input='{proveedor}' dueno_input='{dueno_input}' nombre='{nombre}'")
        if dueno_input not in ('', 'ricky', 'ferreteria_general'):
            return jsonify({'success': False, 'msg': 'Dueño inválido.'})

        # (Validación de proveedor se hará más adelante, luego de determinar dueno_final, para respetar pertenencia)
        
        if not codigo_original or not nombre:
            return jsonify({'success': False, 'msg': 'Datos incompletos.'})
        
        precio, error_precio = parse_price(precio_str)
        if error_precio:
            return jsonify({'success': False, 'msg': f'Error en el precio: {error_precio}'})

        # Recuperar fila original (case-insensitive). Si no existe, intentaremos crearlo.
        fila_original_rows = db_query("SELECT * FROM productos_manual WHERE LOWER(codigo)=LOWER(?) LIMIT 1", (codigo_original,), fetch=True)
        creado_nuevo = False
        if not fila_original_rows:
            # Intentar inferir proveedor y dueno desde Excel si existe
            infer_dueno = dueno_input or ''
            infer_proveedor = proveedor or ''
            print(f"[DEBUG crear] No se encontró fila original para '{codigo_original}'. Intentando inferir datos. dueno_input={dueno_input} proveedor_input={proveedor}")
            if os.path.exists(MANUAL_PRODUCTS_FILE):
                try:
                    df_inf = pd.read_excel(MANUAL_PRODUCTS_FILE)
                    df_inf.rename(columns={'Código': 'Codigo', 'Dueño': 'Dueno'}, inplace=True)
                    if 'Codigo' in df_inf.columns:
                        mask_inf = df_inf['Codigo'].astype(str).str.strip().str.lower() == codigo_original.lower()
                        if mask_inf.any():
                            row_inf = df_inf[mask_inf].iloc[0]
                            infer_proveedor_excel = str(row_inf.get('Proveedor', '')).strip()
                            infer_dueno_excel = str(row_inf.get('Dueno', '')).strip().lower()
                            if not infer_proveedor:
                                infer_proveedor = infer_proveedor_excel
                            if not infer_dueno:
                                infer_dueno = infer_dueno_excel
                            print(f"[DEBUG crear] Inferido desde Excel -> proveedor={infer_proveedor} dueno={infer_dueno}")
                        else:
                            print("[DEBUG crear] Codigo no encontrado en Excel para inferir.")
                except Exception as _e_inf:
                    print(f"[WARN] No se pudo inferir datos desde Excel para creación: {_e_inf}")
            # Insertar registro base
            try:
                print(f"[DEBUG crear] INSERT base productos_manual nombre={nombre} codigo={codigo_original} precio={precio} proveedor={infer_proveedor} dueno={infer_dueno}")
                db_query(
                    "INSERT INTO productos_manual (nombre, codigo, precio, proveedor, dueno) VALUES (?, ?, ?, ?, ?)",
                    (nombre, codigo_original, precio, infer_proveedor, infer_dueno)
                )
                fila_original_rows = db_query("SELECT * FROM productos_manual WHERE LOWER(codigo)=LOWER(?) LIMIT 1", (codigo_original,), fetch=True)
                print(f"[DEBUG crear] SELECT post-insert (por codigo_original) devolvió {len(fila_original_rows) if fila_original_rows else 0} filas")
                creado_nuevo = True
            except Exception as _e_crea:
                print(f"[ERROR crear] Falló INSERT base: {_e_crea}")
                return jsonify({'success': False, 'msg': f'No se pudo crear registro base: {_e_crea}'})
        if not fila_original_rows:
            # Fallback: si usuario ya envió un nuevo codigo distinto, intentar buscar por el nuevo
            if codigo and codigo.lower() != codigo_original.lower():
                print(f"[DEBUG crear] Fallback SELECT por nuevo codigo '{codigo}' tras fallar búsqueda por original '{codigo_original}'")
                fila_original_rows = db_query("SELECT * FROM productos_manual WHERE LOWER(codigo)=LOWER(?) LIMIT 1", (codigo,), fetch=True)
            if not fila_original_rows:
                # Diagnóstico adicional: contar coincidencias parciales
                try:
                    proximos = db_query("SELECT codigo, dueno, proveedor FROM productos_manual WHERE codigo LIKE ? ORDER BY id DESC LIMIT 5", (f"%{codigo_original[:6]}%",), fetch=True)
                except Exception:
                    proximos = []
                return jsonify({'success': False, 'msg': 'No se pudo recuperar el producto tras crearlo.', 'diagnostico': {
                    'codigo_original': codigo_original,
                    'codigo_nuevo': codigo,
                    'fallback_busqueda': bool(codigo and codigo.lower() != codigo_original.lower()),
                    'coincidencias_parciales': proximos
                }})
        fila_original = fila_original_rows[0]
        dueno_original = (fila_original.get('dueno') if isinstance(fila_original, dict) else None) or ''
        proveedor_original = (fila_original.get('proveedor') if isinstance(fila_original, dict) else None) or ''

        # Si no se envía nuevo código, mantener el original
        if not codigo:
            codigo = codigo_original

        # Determinar dueno final: si se envía dueno_input usarlo, sino conservar
        dueno_final = dueno_input or dueno_original
        if dueno_final and dueno_final not in ('ricky', 'ferreteria_general'):
            dueno_final = dueno_original  # fallback silencioso
        print(f"[DEBUG editar_manual_duenos] dueno_original='{dueno_original}' dueno_input='{dueno_input}' dueno_final='{dueno_final}' proveedor='{proveedor}'")

        # Validar proveedor (si viene) y pertenencia al dueno_final.
        proveedor_normalizado = canonicalize_proveedor_name(proveedor)
        if proveedor_normalizado:
            # Primero, verificar si el proveedor existe
            prov_rows_all = db_query("SELECT id, nombre FROM proveedores_manual WHERE LOWER(nombre)=LOWER(?)", (proveedor_normalizado.lower(),), fetch=True)
            # Si no encuentra, intentar variantes trim / collapse manuales para diagnosticar
            if not prov_rows_all:
                variantes = db_query("SELECT id, nombre FROM proveedores_manual WHERE REPLACE(TRIM(LOWER(nombre)), '  ', ' ') = REPLACE(TRIM(LOWER(?)), '  ', ' ')", (proveedor_normalizado.lower(),), fetch=True)
            else:
                variantes = []
            def _row_debug(r):
                nm = r.get('nombre') if isinstance(r, dict) else r['nombre']
                return {
                    'id': r.get('id'),
                    'nombre': nm,
                    'len': len(nm or ''),
                    'hex': (nm or '').encode('utf-8').hex()
                }
            dbg_rows = [_row_debug(r) for r in (prov_rows_all or [])]
            dbg_var = [_row_debug(r) for r in (variantes or [])]
            print(f"[DEBUG proveedor+] raw='{proveedor_raw}' canon='{proveedor_normalizado}' lower='{proveedor_normalizado.lower()}' filas={dbg_rows} variantes={dbg_var}")
            
            if not prov_rows_all:
                return jsonify({'success': False, 'msg': f'Proveedor "{proveedor_normalizado}" no existe. Agregalo primero en la sección de gestionar proveedores para el dueño {dueno_final}.', 'diagnostico': {'canon': proveedor_normalizado, 'raw': proveedor_raw, 'variantes': dbg_var}})
            
            # Obtener el ID del proveedor
            proveedor_id = prov_rows_all[0]['id']
            
            # Ahora usar la nueva tabla proveedores_duenos para verificar la relación con el dueño
            relaciones = db_query(
                """
                SELECT pd.dueno 
                FROM proveedores_duenos pd
                WHERE pd.proveedor_id = ?
                """, 
                (proveedor_id,), fetch=True
            ) or []
            
            duenos_relacionados = [r['dueno'] for r in relaciones]
            print(f"[DEBUG validacion] Proveedor '{proveedor_normalizado}' (ID: {proveedor_id}) tiene dueños: {duenos_relacionados}, buscando: '{dueno_final}'")
            
            # VALIDACIÓN ESTRICTA: Solo permitir si el proveedor está explícitamente agregado para el dueño específico
            if dueno_final in duenos_relacionados:
                print(f"[DEBUG proveedor] Proveedor explícitamente agregado para dueño '{dueno_final}', permitiendo uso")
                match = True
            else:
                match = False
                print(f"[DEBUG proveedor] ERROR: Proveedor no agregado para '{dueno_final}', dueños disponibles: {duenos_relacionados}")
                return jsonify({'success': False, 'msg': f'Proveedor "{proveedor_normalizado}" no está agregado para "{dueno_final}". Agregalo primero en la sección de gestionar proveedores. Dueños existentes: {", ".join(duenos_relacionados)}', 'dueños_existentes': duenos_relacionados, 'diagnostico': {'filas': dbg_rows}})
            print("[DEBUG proveedor] Validación proveedor estricta OK.")
        
        # 1) Actualizar producto en tabla productos_manual (incluyendo proveedor y dueno)
        result = db_query(
            "UPDATE productos_manual SET nombre = ?, codigo = ?, precio = ?, proveedor = ?, dueno = ? WHERE LOWER(codigo) = LOWER(?)",
            (nombre, codigo, precio, proveedor, dueno_final, codigo_original)
        )
        # Verificación post-update: verificar que el proveedor esté explícitamente agregado para el dueño
        if proveedor_normalizado and dueno_original.lower() != dueno_final.lower():
            # Obtener ID del proveedor
            proveedor_row = db_query("SELECT id FROM proveedores_manual WHERE LOWER(nombre)=LOWER(?)", 
                                   (proveedor_normalizado,), fetch=True)
            
            if proveedor_row:
                proveedor_id = proveedor_row[0]['id']
                
                # Usar la tabla de relaciones para verificar
                relaciones = db_query(
                    """
                    SELECT dueno 
                    FROM proveedores_duenos
                    WHERE proveedor_id = ?
                    """, 
                    (proveedor_id,), fetch=True
                ) or []
                
                duenos_relacionados = [r['dueno'] for r in relaciones]
                
                # Solo revertir si no hay compatibilidad (validación estricta)
                if dueno_final not in duenos_relacionados:
                    # Revertir dueño en producto para evitar traslado inconsistente
                    db_query("UPDATE productos_manual SET dueno = ? WHERE LOWER(codigo)=LOWER(?)", (dueno_original, codigo.lower()))
                    print(f"[WARN traslado] Revertido cambio de dueño '{dueno_final}' -> '{dueno_original}' para codigo={codigo} por proveedor no agregado para este dueño")
                    return jsonify({'success': False, 'msg': f'No se puede trasladar el producto porque el proveedor "{proveedor_normalizado}" no está agregado para {dueno_final}. Agregalo primero en la sección de gestionar proveedores. Se mantuvo en {dueno_original}.'})

        # 2) Reflejar cambios en productos_manual.xlsx (si existe y contiene el código)
        excel_actualizado = False
        if os.path.exists(MANUAL_PRODUCTS_FILE):
            try:
                df = pd.read_excel(MANUAL_PRODUCTS_FILE)
                df.rename(columns={'Código': 'Codigo', 'Dueño': 'Dueno'}, inplace=True)
                if not df.empty and 'Codigo' in df.columns:
                    mask = df['Codigo'].astype(str).str.strip().str.lower() == codigo_original.lower()
                    if mask.any():
                        idx = df[mask].index
                        # Actualizar columnas si existen
                        if 'Nombre' in df.columns:
                            df.loc[idx, 'Nombre'] = nombre
                        df.loc[idx, 'Codigo'] = codigo
                        if 'Precio' in df.columns:
                            df.loc[idx, 'Precio'] = precio
                        if 'Proveedor' in df.columns and proveedor:
                            df.loc[idx, 'Proveedor'] = proveedor
                        if 'Dueno' in df.columns and dueno_final:
                            df.loc[idx, 'Dueno'] = dueno_final
                        # Guardar Excel
                        with pd.ExcelWriter(MANUAL_PRODUCTS_FILE, engine='openpyxl', mode='w') as writer:
                            df.to_excel(writer, index=False)
                        excel_actualizado = True
            except Exception as e_excel_upd:
                print(f"[WARN] No se pudo actualizar productos_manual.xlsx: {e_excel_upd}")

        # Propagar cambios al historial (tabla stock) para coherencia visual:
        # Criterio: actualizar filas cuyo codigo coincida con codigo_original y (opcional) proveedor si viene informado.
        # Si no existe fila afectada, se inserta nueva con cantidad=0. Incluimos dueno si se puede derivar de proveedores_meta.
        if result:
            try:
                if precio is None:
                    precio = 0.0
                precio_texto = '0' if float(precio) == 0.0 else str(precio)
                # Priorizar dueno_final explícito; si no hay, intentar derivar por proveedor
                dueno_val = dueno_final if dueno_final in ('ricky','ferreteria_general') else None
                if not dueno_val and proveedor:
                    dueno_row = db_query("SELECT dueno FROM proveedores_meta WHERE LOWER(nombre)=LOWER(?) LIMIT 1", (proveedor,), fetch=True)
                    if dueno_row:
                        dueno_val = dueno_row[0]['dueno']

                if proveedor and dueno_val:
                    update_sql = ("UPDATE stock SET nombre = ?, codigo = ?, precio = ?, precio_texto = ?, dueno = ? "
                                  "WHERE LOWER(codigo) = ? AND LOWER(proveedor)=LOWER(?)")
                    params = [nombre, codigo, precio, precio_texto, dueno_val, codigo_original.lower(), proveedor]
                elif proveedor and not dueno_val:
                    update_sql = ("UPDATE stock SET nombre = ?, codigo = ?, precio = ?, precio_texto = ? "
                                  "WHERE LOWER(codigo) = ? AND LOWER(proveedor)=LOWER(?)")
                    params = [nombre, codigo, precio, precio_texto, codigo_original.lower(), proveedor]
                elif not proveedor and dueno_val:
                    update_sql = ("UPDATE stock SET nombre = ?, codigo = ?, precio = ?, precio_texto = ?, dueno = ? "
                                  "WHERE LOWER(codigo) = ?")
                    params = [nombre, codigo, precio, precio_texto, dueno_val, codigo_original.lower()]
                else:
                    update_sql = ("UPDATE stock SET nombre = ?, codigo = ?, precio = ?, precio_texto = ? "
                                  "WHERE LOWER(codigo) = ?")
                    params = [nombre, codigo, precio, precio_texto, codigo_original.lower()]
                # Ejecutar y contar filas afectadas manualmente (db_query no retorna rowcount si usa SELECT wrapper)
                conn = get_db_connection()
                cur = conn.cursor()
                cur.execute(update_sql, tuple(params))
                afectados = cur.rowcount
                conn.commit()
                if afectados > 0:
                    try:
                        if proveedor:
                            row_q = db_query("SELECT * FROM stock WHERE LOWER(codigo)=LOWER(?) AND LOWER(proveedor)=LOWER(?) ORDER BY id DESC LIMIT 1", (codigo, proveedor), fetch=True)
                        else:
                            row_q = db_query("SELECT * FROM stock WHERE LOWER(codigo)=LOWER(?) ORDER BY id DESC LIMIT 1", (codigo,), fetch=True)
                        if row_q:
                            log_stock_history('manual_edit', fuente='gestionar_manual_actualizar_ajax_update', stock_row=row_q[0])
                    except Exception as e_log_gest_upd:
                        print(f"[WARN] log history update gestionar_manual_actualizar_ajax: {e_log_gest_upd}")
                if afectados == 0:
                    # Insertar nueva fila para que aparezca
                    prov_ins = proveedor or ''
                    ahora = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    try:
                        if dueno_val:
                            cur.execute("""
                                INSERT INTO stock (codigo, nombre, precio, cantidad, fecha_compra, proveedor, observaciones, precio_texto, dueno, created_at)
                                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                            """, (codigo, nombre, precio, 0, ahora, prov_ins, 'Añadido por edición manual', str(precio), dueno_val, ahora))
                        else:
                            cur.execute("""
                                INSERT INTO stock (codigo, nombre, precio, cantidad, fecha_compra, proveedor, observaciones, precio_texto, created_at)
                                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                            """, (codigo, nombre, precio, 0, ahora, prov_ins, 'Añadido por edición manual', str(precio), ahora))
                        conn.commit()
                        afectados = 1
                        try:
                            row_new = db_query("SELECT * FROM stock WHERE id = (SELECT MAX(id) FROM stock)", fetch=True)
                            if row_new:
                                log_stock_history('insert', fuente='gestionar_manual_actualizar_ajax_insert', stock_row=row_new[0])
                        except Exception as e_log_gest_ins:
                            print(f"[WARN] log history insert gestionar_manual_actualizar_ajax: {e_log_gest_ins}")
                    except Exception as e_ins:
                        print(f"Error insertando fila stock tras edición manual: {e_ins}")
            except Exception as e_upd:
                # No interrumpir flujo si falla la propagación; informar en mensaje.
                return jsonify({
                    'success': True,
                    'msg': f'Producto "{nombre}" actualizado (advertencia al sincronizar historial: {str(e_upd)})',
                    'html': '<div class="alert alert-warning">Actualizado, pero no se pudo sincronizar completamente el historial.</div>'
                })

            # Construir detalle de cambios
            cambios = []
            if codigo_original.lower() != codigo.lower():
                cambios.append('código')
            if nombre != (fila_original.get('nombre') if isinstance(fila_original, dict) else ''):
                cambios.append('nombre')
            if proveedor and proveedor.lower() != proveedor_original.lower():
                cambios.append('proveedor')
            if dueno_final and dueno_final != dueno_original:
                cambios.append('dueño')
            if precio != (fila_original.get('precio') if isinstance(fila_original, dict) else None):
                cambios.append('precio')
            if creado_nuevo:
                cambios.append('creado')
            cambios_txt = (" Cambios: " + ", ".join(cambios) + ".") if cambios else ''
            msg_extra = (' (Excel actualizado)' if excel_actualizado else ' (Excel sin cambios)') + cambios_txt
            return jsonify({
                'success': True,
                'msg': f'Producto "{nombre}" actualizado. Historial sincronizado en {afectados} fila(s).'+msg_extra,
                'html': '<div class="alert alert-success">Producto actualizado y sincronizado en historial.</div>',
                'excel_actualizado': excel_actualizado
            })
        else:
            return jsonify({'success': False, 'msg': 'Error al actualizar el producto.'})
            
    except Exception as e:
        return jsonify({'success': False, 'msg': f'Error: {str(e)}'})

# Alias legacy para compatibilidad con templates que aún llaman manual_actualizar_ajax
@app.route('/manual_actualizar_ajax', methods=['POST'])
@login_required
def manual_actualizar_ajax():
    """Alias que delega en gestionar_manual_actualizar_ajax (mantiene compatibilidad)."""
    return gestionar_manual_actualizar_ajax()

@app.route('/debug_proveedores')
@login_required
def debug_proveedores():
    """Diagnóstico de proveedores: muestra las filas exactas y los productos_manual que los usan.
    Uso: /debug_proveedores?nombre=DEWALT  (nombre opcional -> si falta lista primeros 50)
    """
    try:
        nombre = (request.args.get('nombre') or '').strip()
        params = []
        where = ''
        if nombre:
            where = 'WHERE LOWER(nombre)=LOWER(?)'
            params.append(nombre)
        provs = db_query(f"SELECT id, nombre, dueno FROM proveedores_manual {where} ORDER BY nombre, dueno LIMIT 50", tuple(params), fetch=True) or []
        productos = []
        if nombre:
            productos = db_query("SELECT codigo, nombre, dueno, proveedor FROM productos_manual WHERE LOWER(proveedor)=LOWER(?) ORDER BY id DESC LIMIT 50", (nombre,), fetch=True) or []
        return jsonify({
            'success': True,
            'proveedores': provs,
            'productos_manual': productos,
            'filtro': nombre
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/debug_proveedores_ui')
@login_required
def debug_proveedores_ui():
    return render_template('debug_proveedores.html')

@app.route('/debug_obtener_proveedores_por_dueno/<string:dueno>')
@login_required
def debug_obtener_proveedores_por_dueno_endpoint(dueno):
    try:
        print(f"[DEBUG] debug_obtener_proveedores_por_dueno llamado con dueño: '{dueno}'")
        
        # Usar la nueva tabla proveedores_duenos para obtener proveedores
        proveedores = db_query(
            """
            SELECT DISTINCT p.nombre 
            FROM proveedores_manual p
            JOIN proveedores_duenos pd ON p.id = pd.proveedor_id
            WHERE pd.dueno = ?
            ORDER BY p.nombre
            """, 
            (dueno,), fetch=True
        )
        
        resultado = [p['nombre'] for p in proveedores]
        print(f"[DEBUG] debug_obtener_proveedores_por_dueno - proveedores encontrados: {resultado}")
        
        return jsonify({
            'success': True, 
            'proveedores': resultado
        })
    except Exception as e:
        print(f"Error en debug_obtener_proveedores_por_dueno: {e}")
        return jsonify({'success': False, 'msg': f'Error: {str(e)}'})

@app.route('/debug_stock_item')
@login_required
def debug_stock_item():
    """Endpoint de diagnóstico: muestra filas de stock filtradas por codigo/proveedor/dueno.
    Uso: /debug_stock_item?codigo=XXX&proveedor=YYY&dueno=ZZZ (todos opcionales)
    """
    try:
        codigo = (request.args.get('codigo') or '').strip()
        proveedor = (request.args.get('proveedor') or '').strip()
        dueno = (request.args.get('dueno') or '').strip().lower()
        condiciones = []
        params = []
        if codigo:
            condiciones.append('LOWER(codigo)=LOWER(?)')
            params.append(codigo)
        if proveedor:
            condiciones.append('LOWER(proveedor)=LOWER(?)')
            params.append(proveedor)
        if dueno:
            condiciones.append('LOWER(dueno)=?')
            params.append(dueno)
        where = ('WHERE ' + ' AND '.join(condiciones)) if condiciones else ''
        rows = db_query(f"SELECT id, codigo, nombre, precio, cantidad, proveedor, dueno, fecha_compra, precio_texto FROM stock {where} ORDER BY id DESC LIMIT 50", tuple(params), fetch=True) or []
        return jsonify({
            'success': True,
            'count': len(rows),
            'rows': [dict(r) for r in rows],
            'filtro': {'codigo': codigo, 'proveedor': proveedor, 'dueno': dueno}
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

# --- Funciones de Búsqueda en Excel ---
def buscar_en_excel(termino_busqueda, proveedor_filtro=None, filtro_adicional=None, solo_ricky=False, solo_fg=False):
    """Buscar productos en archivos Excel de proveedores y productos manuales"""
    resultados = []
    print(f"🔍 [BUSCAR_EXCEL] Iniciando búsqueda: '{termino_busqueda}' | Proveedor: '{proveedor_filtro}' | Filtro: '{filtro_adicional}' | Solo Ricky: {solo_ricky} | Solo FG: {solo_fg}")
    # Normalizar el proveedor_filtro a minúsculas para comparaciones case-insensitive
    proveedor_filtro = proveedor_filtro.lower() if proveedor_filtro else None
    
    # 1. Buscar en productos manuales
    if proveedor_filtro and proveedor_filtro.startswith('manual_'):
        # Filtro específico de proveedor manual (incluye dueño)
        try:
            _, rest = proveedor_filtro.split('manual_', 1)
            parts = rest.split('_', 1)
            proveedor_id = int(parts[0])
            dueno_sel = parts[1] if len(parts) > 1 else None
            resultados_manuales = buscar_en_excel_manual_por_proveedor(termino_busqueda, proveedor_id, dueno_sel)
            resultados.extend(resultados_manuales)
        except (ValueError, TypeError):
            pass
    elif proveedor_filtro and proveedor_filtro in [k.lower() for k in PROVEEDOR_CONFIG.keys()]:
        # Obtener la clave original del diccionario (preservando mayúsculas)
        proveedor_key = next((k for k in PROVEEDOR_CONFIG.keys() if k.lower() == proveedor_filtro), proveedor_filtro)
        # Nuevo: también incluir productos manuales que pertenezcan a ese proveedor Excel
        # Determinar alcance de dueños según flags
        if solo_ricky and not solo_fg:
            duenos_manual = ['ricky']
        elif solo_fg and not solo_ricky:
            duenos_manual = ['ferreteria_general']
        else:
            duenos_manual = ['ricky', 'ferreteria_general']
        
        # Buscar tanto por el nombre del proveedor en Excel como por posibles variaciones en mayúsculas/minúsculas
        proveedor_nombre_original = proveedor_filtro
        
        # Intentar primero con el nombre exacto del proveedor
        for d in duenos_manual:
            resultados_manuales = buscar_en_excel_manual_por_nombre_proveedor(termino_busqueda, proveedor_nombre_original, dueno_filtro=d)
            if resultados_manuales:
                resultados.extend(resultados_manuales)
                
        # Si no hay resultados, probar con variaciones de mayúsculas/minúsculas
        if not resultados:
            print(f"[EXCEL DEBUG] No se encontraron resultados con el nombre exacto '{proveedor_nombre_original}', probando con versión en mayúsculas")
            proveedor_upper = proveedor_nombre_original.upper()
            if proveedor_upper != proveedor_nombre_original:
                for d in duenos_manual:
                    resultados_manuales = buscar_en_excel_manual_por_nombre_proveedor(termino_busqueda, proveedor_upper, dueno_filtro=d)
                    if resultados_manuales:
                        resultados.extend(resultados_manuales)
    elif not proveedor_filtro or proveedor_filtro not in PROVEEDOR_CONFIG:
        # Si no hay filtro específico de Excel, incluir todos los manuales
        # Aplicar alcance por dueño si corresponde
        if solo_ricky and not solo_fg:
            resultados_manuales = buscar_en_excel_manual(termino_busqueda, dueno_filtro='ricky')
            print(f"Búsqueda manual solo para dueño 'ricky': {len(resultados_manuales)} resultados")
        elif solo_fg and not solo_ricky:
            resultados_manuales = buscar_en_excel_manual(termino_busqueda, dueno_filtro='ferreteria_general')
            print(f"Búsqueda manual solo para dueño 'ferreteria_general': {len(resultados_manuales)} resultados")
        else:
            resultados_manuales = buscar_en_excel_manual(termino_busqueda)
            print(f"Búsqueda manual para todos los dueños: {len(resultados_manuales)} resultados")
        
        if resultados_manuales:
            print(f"Agregando {len(resultados_manuales)} resultados manuales")
            resultados.extend(resultados_manuales)
        else:
            print("No se encontraron resultados manuales")
        
    # 2. Buscar en Excel de proveedores
    if not proveedor_filtro or proveedor_filtro in [k.lower() for k in PROVEEDOR_CONFIG.keys()]:
        # Verificar si necesitamos buscar en todos o un proveedor específico
        if proveedor_filtro:
            # Verificar si el proveedor está habilitado para el filtro de dueño
            proveedor_key = next((k for k in PROVEEDOR_CONFIG.keys() if k.lower() == proveedor_filtro), proveedor_filtro)
            proveedor_config = PROVEEDOR_CONFIG.get(proveedor_key, {})
            proveedor_dueno = proveedor_config.get('dueno', None)
            
            # Saltar si no coincide con los filtros de dueño
            if (solo_ricky and proveedor_dueno != 'ricky') or (solo_fg and proveedor_dueno != 'ferreteria_general'):
                pass
            else:
                proveedores_a_buscar = [proveedor_filtro]
        else:
            # Determinar qué proveedores buscar según filtro de dueño
            if solo_ricky and not solo_fg:
                proveedores_a_buscar = [p for p, cfg in PROVEEDOR_CONFIG.items() if cfg.get('dueno') == 'ricky']
            elif solo_fg and not solo_ricky:
                proveedores_a_buscar = [p for p, cfg in PROVEEDOR_CONFIG.items() if cfg.get('dueno') == 'ferreteria_general']
            else:
                proveedores_a_buscar = list(PROVEEDOR_CONFIG.keys())
                
        # Realizar búsquedas en los proveedores seleccionados
        for proveedor in proveedores_a_buscar:
            resultados_proveedor = buscar_en_excel_proveedor(
                termino_busqueda, 
                proveedor,
                filtro_adicional
            )
            resultados.extend(resultados_proveedor)
    
    # Primera deduplicación básica - eliminar duplicados exactos
    # Esta deduplicación es más ligera y se hace a nivel de función
    resultados_sin_duplicados = []
    claves_vistas = set()
    
    for r in resultados:
        codigo = str(r.get('codigo', '')).strip()
        proveedor = str(r.get('proveedor', '')).lower().strip()
        nombre = str(r.get('nombre', '')).lower().strip()
        
        # Crear una clave para eliminar duplicados exactos
        clave = f"{codigo}|{proveedor}|{nombre}"
        
        if clave not in claves_vistas:
            claves_vistas.add(clave)
            resultados_sin_duplicados.append(r)
    
    # Log cantidad total de resultados
    print(f"🔍 [BUSCAR_EXCEL] Total de resultados sin duplicados exactos: {len(resultados_sin_duplicados)} de {len(resultados)} originales")
    
    return resultados_sin_duplicados

def agregar_producto_manual_excel():
    """Agregar producto manual al Excel (no directamente al stock)"""
    try:
        codigo = request.form.get('codigo', '').strip()
        proveedor_id = request.form.get('proveedor_id', '').strip()
        nombre = request.form.get('nombre', '').strip()
        precio_str = request.form.get('precio', '').strip()
        observaciones = request.form.get('observaciones', '').strip()
        dueno = request.form.get('dueno_nuevo_proveedor', '').strip()
        
        # Validaciones
        if not nombre:
            flash('El nombre del producto es obligatorio.', 'danger')
            return redirect(url_for('agregar_producto'))
        
        if not dueno:
            flash('Debe seleccionar un dueño para el producto.', 'danger')
            return redirect(url_for('agregar_producto'))
        
        # Obtener nombre del proveedor si se proporcionó proveedor_id
        proveedor_nombre = ''
        if proveedor_id:
            proveedor_data = db_query("SELECT nombre FROM proveedores_manual WHERE id = ?", (proveedor_id,), fetch=True)
            if proveedor_data:
                proveedor_nombre = proveedor_data[0]['nombre']
        
        # Procesar precio
        precio = 0.0
        if precio_str:
            precio, error_precio = parse_price(precio_str)
            if error_precio:
                flash(f'Error en el precio: {precio_str}', 'warning')
                precio = 0.0
        
        # Agregar al Excel
        if agregar_producto_excel_manual(codigo, proveedor_nombre, nombre, precio, observaciones, dueno):
            dueno_nombre = DUENOS_CONFIG.get(dueno, {}).get('nombre', dueno)
            flash(f'Producto "{nombre}" agregado a la lista manual de {dueno_nombre}. Puede buscarlo en "Buscar en Excel".', 'success')
        else:
            flash('Error al agregar el producto a la lista manual.', 'danger')
            
    except Exception as e:
        flash(f'Error al procesar el producto: {str(e)}', 'danger')
    
    return redirect(url_for('agregar_producto'))
    
    # 2. Buscar en archivos Excel de proveedores por dueño
    if proveedor_filtro:
        if proveedor_filtro in PROVEEDOR_CONFIG:
            # Determinar el dueño del proveedor
            dueno_proveedor = None
            for dueno, config in DUENOS_CONFIG.items():
                if proveedor_filtro in config.get('proveedores_excel', []):
                    dueno_proveedor = dueno
                    break
            
            if dueno_proveedor:
                archivos_a_buscar = [(proveedor_filtro, dueno_proveedor)]
                print(f"🔍 Proveedor específico: {proveedor_filtro} -> Dueño: {dueno_proveedor}")
            else:
                archivos_a_buscar = []
                print(f"❌ Proveedor no encontrado en configuración: {proveedor_filtro}")
        else:
            archivos_a_buscar = []
    else:
        # Buscar en todos los archivos Excel disponibles por dueño
        archivos_a_buscar = []
        
        # Determinar qué dueños buscar
        duenos_a_buscar = []
        if solo_ricky and not solo_fg:
            duenos_a_buscar = ['ricky']
        elif solo_fg and not solo_ricky:
            duenos_a_buscar = ['ferreteria_general']
        else:
            duenos_a_buscar = ['ricky', 'ferreteria_general']
        
        for dueno in duenos_a_buscar:
            if dueno in DUENOS_CONFIG:
                proveedores_dueno = DUENOS_CONFIG[dueno]['proveedores_excel']
                for key in proveedores_dueno:
                    if key in PROVEEDOR_CONFIG:
                        # Buscar archivos en la carpeta específica del dueño
                        carpeta_dueno = get_excel_folder_for_dueno(dueno)
                        archivos = [f for f in os.listdir(carpeta_dueno) if f.lower().startswith(key.lower()) and f.endswith('.xlsx') and f != 'productos_manual.xlsx']
                        if archivos:
                            archivos_a_buscar.append((key, dueno))
    
    # Procesar cada archivo Excel
    for archivo_info in archivos_a_buscar:
        if isinstance(archivo_info, tuple):
            archivo, dueno = archivo_info
        else:
            archivo = archivo_info
            dueno = 'ricky'  # Por defecto para compatibilidad
        
        if archivo in PROVEEDOR_CONFIG:
            config = PROVEEDOR_CONFIG[archivo]
            
            # Excluir proveedores ocultos
            ocultos = db_query("SELECT LOWER(nombre) as nombre FROM proveedores_ocultos WHERE dueno = ?", (dueno,), fetch=True) or []
            ocultos_set = {o['nombre'] for o in ocultos}
            if archivo.lower() in ocultos_set:
                continue
            
            # Aplicar filtros de dueño
            if solo_fg and not solo_ricky and dueno != 'ferreteria_general':
                continue
            if solo_ricky and not solo_fg and dueno != 'ricky':
                continue
            
            resultados_archivo = procesar_archivo_excel(archivo, config, termino_busqueda, filtro_adicional, archivo, dueno)
            resultados.extend(resultados_archivo)
    
    print(f"🔍 [BUSCAR_EXCEL] ✅ Búsqueda completada. Total resultados: {len(resultados)}")
    for i, resultado in enumerate(resultados):
        print(f"  {i+1}. {resultado.get('codigo', 'N/A')} - {resultado.get('nombre', 'N/A')} - ${resultado.get('precio', 0)}")
    
    return resultados

# ------------------------------------------------------
# Helper para registrar historial híbrido
# ------------------------------------------------------
def log_stock_history(tipo_cambio: str, fuente: str = None, stock_row: dict = None, extra: dict = None):
    """Inserta una fila en stock_history.
        tipo_cambio: 'insert' | 'update' | 'manual_edit' | 'import' | 'sync' | 'otro'
        fuente: endpoint o script
        stock_row: dict representando la fila actual (de stock) o futura
        extra: campos adicionales (ignorados si col no existe)
    """
    try:
        conn = get_db_connection()
        if not conn:
            return
        cur = conn.cursor()
        fecha_evento = datetime.utcnow().isoformat()
        data = {
            'stock_id': stock_row.get('id') if stock_row else None,
            'codigo': stock_row.get('codigo') if stock_row else None,
            'nombre': stock_row.get('nombre') if stock_row else None,
            'precio': stock_row.get('precio') if stock_row else None,
            'cantidad': stock_row.get('cantidad') if stock_row else None,
            'fecha_compra': stock_row.get('fecha_compra') if stock_row else None,
            'proveedor': stock_row.get('proveedor') if stock_row else None,
            'observaciones': stock_row.get('observaciones') if stock_row else None,
            'precio_texto': stock_row.get('precio_texto') if stock_row else None,
            'dueno': stock_row.get('dueno') if stock_row else None,
            'created_at': stock_row.get('created_at') if stock_row else None,
            'fecha_evento': fecha_evento,
            'tipo_cambio': tipo_cambio,
            'fuente': fuente,
            'usuario': session.get('username') if 'username' in session else None
        }
        if extra:
            data.update(extra)
        cols = ",".join(data.keys())
        placeholders = ",".join(['?'] * len(data)) if not _is_postgres_configured() else ",".join(['%s'] * len(data))
        sql = f"INSERT INTO stock_history ({cols}) VALUES ({placeholders})"
        cur.execute(sql, tuple(data.values()))
        conn.commit()
    except Exception as e:
        print(f"[WARN] log_stock_history fallo: {e}")

def buscar_en_excel_manual_por_proveedor(termino_busqueda, proveedor_id, dueno_filtro=None):
    """Buscar productos en el Excel de productos manuales por proveedor específico"""
    resultados = []
    
    try:
        print(f"[EXCEL DEBUG] Iniciando búsqueda en Excel manual para proveedor_id={proveedor_id} y dueno_filtro={dueno_filtro}")
        
        if not os.path.exists(MANUAL_PRODUCTS_FILE):
            print(f"[EXCEL ERROR] No se encontró el archivo de productos manuales: {MANUAL_PRODUCTS_FILE}")
            return resultados
        
        print(f"[EXCEL DEBUG] Leyendo archivo: {MANUAL_PRODUCTS_FILE}")
        df = pd.read_excel(MANUAL_PRODUCTS_FILE)
        # Normalizar nombres de columnas por si existen acentos
        df.rename(columns={'Código': 'Codigo', 'Dueño': 'Dueno'}, inplace=True)
        
        print(f"[EXCEL DEBUG] DataFrame inicial: {len(df)} filas")
        if df.empty:
            print("[EXCEL ERROR] El DataFrame está vacío")
            return resultados
        
        # DIAGNÓSTICO TOTAL: Mostrar todas las filas del Excel al inicio
        print("[EXCEL DIAGNÓSTICO TOTAL] ===== CONTENIDO COMPLETO DEL EXCEL =====")
        try:
            for idx, row in df.iterrows():
                print(f"[EXCEL DIAGNÓSTICO] Fila {idx}:")
                for col in df.columns:
                    print(f"[EXCEL DIAGNÓSTICO]   {col}: {row[col]}")
                print("[EXCEL DIAGNÓSTICO] ---")
        except Exception as e:
            print(f"[EXCEL ERROR] Error al mostrar diagnóstico total: {e}")
        print("[EXCEL DIAGNÓSTICO TOTAL] ===== FIN CONTENIDO COMPLETO =====")
        
        # Obtener nombre del proveedor
        proveedor_info = db_query("SELECT nombre FROM proveedores_manual WHERE id = ?", (proveedor_id,), fetch=True)
        if not proveedor_info:
            print(f"[EXCEL ERROR] No se encontró información para el proveedor con ID {proveedor_id}")
            # CAMBIO: Si no encontramos el proveedor, continuamos sin filtrar por proveedor
            all_results = buscar_en_excel_manual(termino_busqueda, dueno_filtro)
            print(f"[EXCEL DEBUG] Como no se encontró el proveedor, se devuelven todos los resultados: {len(all_results)}")
            return all_results
        
        proveedor_nombre = proveedor_info[0]['nombre']
        print(f"[EXCEL DEBUG] Nombre del proveedor ID {proveedor_id}: '{proveedor_nombre}'")
        
        # Imprimir todos los proveedores disponibles en el Excel para diagnóstico
        print(f"[EXCEL DEBUG] Proveedores disponibles en Excel: {df['Proveedor'].unique().tolist()}")
        
        # Añadir diagnóstico para buscar específicamente productos con el código buscado
        if termino_busqueda:
            print(f"[EXCEL DEBUG] Buscando productos con código '{termino_busqueda}' en todo el Excel:")
            for idx, row in df.iterrows():
                codigo = str(row.get('Codigo', '')).lower()
                if termino_busqueda.lower() in codigo:
                    print(f"[EXCEL DEBUG] ¡ENCONTRADO! Fila {idx}, Código: {row.get('Codigo', '')}, Proveedor: {row.get('Proveedor', '')}, Dueño: {row.get('Dueno', '')}")
        
        # Verificar si está en mayúsculas o minúsculas
        print(f"[EXCEL DEBUG] Búsqueda exacta de '{proveedor_nombre}': {(df['Proveedor'] == proveedor_nombre).sum()} coincidencias")
        print(f"[EXCEL DEBUG] Búsqueda exacta de '{proveedor_nombre.upper()}': {(df['Proveedor'] == proveedor_nombre.upper()).sum()} coincidencias")
        print(f"[EXCEL DEBUG] Búsqueda exacta de '{proveedor_nombre.lower()}': {(df['Proveedor'] == proveedor_nombre.lower()).sum()} coincidencias")
        
        # Búsqueda específica: Si buscamos un producto exacto por código, lo buscamos en todo el Excel
        if termino_busqueda and len(termino_busqueda.strip()) >= 3 and not termino_busqueda.strip().isdigit():
            exact_matches = df[df['Codigo'].astype(str).str.lower() == termino_busqueda.lower()]
            if len(exact_matches) > 0:
                print(f"[EXCEL DEBUG] ¡ENCONTRAMOS CÓDIGO EXACTO! '{termino_busqueda}' - {len(exact_matches)} coincidencias")
                for _, row in exact_matches.iterrows():
                    precio_val, precio_error = parse_price(row.get('Precio', ''))
                    resultado = {
                        'codigo': row.get('Codigo', ''),
                        'nombre': row.get('Nombre', ''),
                        'precio': precio_val,
                        'precio_texto': str(row.get('Precio', '')) if precio_error else None,
                        'proveedor': row.get('Proveedor', ''),
                        'observaciones': row.get('Observaciones', ''),
                        'dueno': row.get('Dueno', ''),
                        'es_manual': True,
                        'matchExacto': True
                    }
                    resultados.append(resultado)
        
        # Usar una búsqueda más flexible que incluya coincidencias parciales y comparaciones insensibles a mayúsculas/minúsculas
        # Esta es una búsqueda más agresiva que la original
        filtered_df = df[df['Proveedor'].astype(str).str.lower().str.contains(proveedor_nombre.lower(), na=False)]
        print(f"[EXCEL DEBUG] Después de filtrar por proveedor '{proveedor_nombre}' (búsqueda flexible): {len(filtered_df)} filas")
        
        # Si no hay resultados después de filtrar por proveedor pero hay un filtro de dueño,
        # ignoramos el filtro de proveedor y solo filtramos por dueño
        if len(filtered_df) == 0:
            print(f"[EXCEL DEBUG] No se encontraron productos para el proveedor '{proveedor_nombre}' - ignorando filtro de proveedor")
            # CAMBIO: Usamos el DataFrame original sin filtrar por proveedor siempre que no hay resultados
            df_for_dueno = df
            # Si hay término de búsqueda, buscamos específicamente ese término
            if termino_busqueda:
                print(f"[EXCEL DEBUG] Buscando específicamente el término '{termino_busqueda}' en todo el Excel:")
                for idx, row in df.iterrows():
                    codigo = str(row.get('Codigo', '')).lower()
                    nombre = str(row.get('Nombre', '')).lower()
                    if termino_busqueda.lower() in codigo or termino_busqueda.lower() in nombre:
                        print(f"[EXCEL DEBUG] ¡ENCONTRADO SIN FILTRO DE PROVEEDOR! Fila {idx}, Código: {row.get('Codigo', '')}, Nombre: {row.get('Nombre', '')}")
        else:
            # Si encontramos productos con el proveedor, continuamos normalmente
            df_for_dueno = filtered_df
        
        # Aplicamos el filtro de dueño si existe
        if dueno_filtro:
            print(f"[EXCEL DEBUG] Filtrando por dueño: {dueno_filtro}")
            df = df_for_dueno[df_for_dueno['Dueno'].astype(str).str.lower() == str(dueno_filtro).lower()]
            print(f"[EXCEL DEBUG] Después de filtrar por dueño: {len(df)} filas")
        else:
            df = df_for_dueno
        
        # Filtrar por término de búsqueda si existe (soporta combinaciones "palabra1 palabra2")
        if termino_busqueda and len(resultados) == 0:  # Solo filtrar si no tenemos resultados exactos
            print(f"[EXCEL DEBUG] Filtrando por término de búsqueda: {termino_busqueda}")
            tokens = [t.strip() for t in str(termino_busqueda).split() if t.strip()]
            if tokens:
                mask_all = pd.Series(True, index=df.index)
                for tok in tokens:
                    mask_tok = (
                        df['Nombre'].astype(str).str.contains(tok, case=False, na=False) |
                        df['Codigo'].astype(str).str.contains(tok, case=False, na=False) |
                        df['Proveedor'].astype(str).str.contains(tok, case=False, na=False)
                    )
                    mask_all &= mask_tok
                df = df[mask_all]
                print(f"[EXCEL DEBUG] Después de filtrar por tokens de búsqueda: {len(df)} filas")
        
        # Si después de todo esto no hay resultados, pero hay un término de búsqueda,
        # volvemos a buscar en todo el Excel sin filtro de proveedor
        if len(df) == 0 and len(resultados) == 0 and termino_busqueda:
            print(f"[EXCEL DEBUG] No se encontraron resultados - buscando en todo el Excel sin filtros")
            all_df = pd.read_excel(MANUAL_PRODUCTS_FILE)
            # Normalizar nombres de columnas
            all_df.rename(columns={'Código': 'Codigo', 'Dueño': 'Dueno'}, inplace=True)
            
            # Filtrar solo por término de búsqueda
            tokens = [t.strip() for t in str(termino_busqueda).split() if t.strip()]
            if tokens:
                mask_all = pd.Series(True, index=all_df.index)
                for tok in tokens:
                    mask_tok = (
                        all_df['Nombre'].astype(str).str.contains(tok, case=False, na=False) |
                        all_df['Codigo'].astype(str).str.contains(tok, case=False, na=False) |
                        all_df['Proveedor'].astype(str).str.contains(tok, case=False, na=False)
                    )
                    mask_all &= mask_tok
                all_df = all_df[mask_all]
                print(f"[EXCEL DEBUG] Resultados sin filtro de proveedor: {len(all_df)} filas")
                
                # Filtrar por dueño si existe
                if dueno_filtro:
                    all_df = all_df[all_df['Dueno'].astype(str).str.lower() == str(dueno_filtro).lower()]
                    print(f"[EXCEL DEBUG] Después de filtrar por dueño: {len(all_df)} filas")
                
                # Usar estos resultados
                df = all_df
        
        print(f"[EXCEL DEBUG] Resultados finales: {len(df)} filas")
        
        # Convertir a lista de diccionarios
        for _, row in df.iterrows():
            precio_val, precio_error = parse_price(row.get('Precio', ''))
            resultado = {
                'codigo': row.get('Codigo', ''),
                'nombre': row.get('Nombre', ''),
                'precio': precio_val,
                'precio_texto': str(row.get('Precio', '')) if precio_error else None,
                'proveedor': row.get('Proveedor', ''),
                'observaciones': row.get('Observaciones', ''),
                'dueno': row.get('Dueno', ''),
                'es_manual': True
            }
            resultados.append(resultado)
    
    except Exception as e:
        print(f"Error al buscar en Excel manual por proveedor: {e}")
    
    return resultados

def buscar_en_excel_manual_por_nombre_proveedor(termino_busqueda, nombre_proveedor, dueno_filtro=None):
    """Buscar productos manuales filtrando por el nombre (clave) del proveedor Excel seleccionado.
    Esto permite que al filtrar por un proveedor Excel también se muestren los productos agregados manualmente
    asociados a ese proveedor (si el campo 'Proveedor' en productos_manual.xlsx contiene ese nombre).
    """
    resultados = []
    try:
        print(f"[EXCEL DEBUG] Iniciando búsqueda en Excel manual por nombre de proveedor: '{nombre_proveedor}', dueño: {dueno_filtro}")
        
        if not os.path.exists(MANUAL_PRODUCTS_FILE):
            print(f"[EXCEL ERROR] No se encontró el archivo: {MANUAL_PRODUCTS_FILE}")
            return resultados
            
        df = pd.read_excel(MANUAL_PRODUCTS_FILE)
        print(f"[EXCEL DEBUG] DataFrame inicial: {len(df)} filas")
        
        df.rename(columns={'Código': 'Codigo', 'Dueño': 'Dueno'}, inplace=True)
        if df.empty:
            print("[EXCEL ERROR] El DataFrame está vacío")
            return resultados
            
        # Filtrar por nombre de proveedor (coincidencia parcial / case-insensitive)
        print(f"[EXCEL DEBUG] Filtrando por nombre de proveedor: '{nombre_proveedor}'")
        df = df[df['Proveedor'].astype(str).str.contains(str(nombre_proveedor), case=False, na=False)]
        print(f"[EXCEL DEBUG] Después de filtrar por proveedor: {len(df)} filas")
        
        if dueno_filtro:
            print(f"[EXCEL DEBUG] Filtrando por dueño: {dueno_filtro}")
            df = df[df['Dueno'].astype(str).str.lower() == str(dueno_filtro).lower()]
            print(f"[EXCEL DEBUG] Después de filtrar por dueño: {len(df)} filas")
            
        if df.empty:
            print("[EXCEL DEBUG] No hay resultados después de filtrar por proveedor/dueño")
            return resultados
            
        # Aplicar término de búsqueda (tokens AND)
        if termino_busqueda:
            print(f"[EXCEL DEBUG] Filtrando por término de búsqueda: {termino_busqueda}")
            tokens = [t.strip() for t in str(termino_busqueda).split() if t.strip()]
            if tokens:
                mask_all = pd.Series(True, index=df.index)
                for tok in tokens:
                    tok_mask = (
                        df['Nombre'].astype(str).str.contains(tok, case=False, na=False) |
                        df['Codigo'].astype(str).str.contains(tok, case=False, na=False) |
                        df['Proveedor'].astype(str).str.contains(tok, case=False, na=False)
                    )
                    mask_all &= tok_mask
                df = df[mask_all]
                print(f"[EXCEL DEBUG] Después de filtrar por términos de búsqueda: {len(df)} filas")
        for _, row in df.iterrows():
            precio_val, precio_error = parse_price(row.get('Precio', ''))
            resultados.append({
                'codigo': row.get('Codigo', ''),
                'nombre': row.get('Nombre', ''),
                'precio': precio_val,
                'precio_texto': str(row.get('Precio', '')) if precio_error else None,
                'proveedor': row.get('Proveedor', ''),
                'observaciones': row.get('Observaciones', ''),
                'dueno': row.get('Dueno', ''),
                'es_manual': True
            })
    except Exception as e:
        print(f"[EXCEL ERROR] Error en buscar_en_excel_manual_por_nombre_proveedor: {e}")
        import traceback
        print(traceback.format_exc())
    return resultados

def buscar_en_excel_manual(termino_busqueda, dueno_filtro=None):
    """Buscar en productos_manual.xlsx sin proveedor específico. Permite filtrar por dueño."""
    resultados = []
    try:
        print(f"[EXCEL DEBUG] Iniciando búsqueda en Excel manual. Término: '{termino_busqueda}', Dueño: {dueno_filtro}")
        
        if not os.path.exists(MANUAL_PRODUCTS_FILE):
            print(f"[EXCEL ERROR] Archivo de productos manuales no encontrado: {MANUAL_PRODUCTS_FILE}")
            return resultados
            
        print(f"[EXCEL DEBUG] Leyendo archivo: {MANUAL_PRODUCTS_FILE}")
        df = pd.read_excel(MANUAL_PRODUCTS_FILE)
        df.rename(columns={'Código': 'Codigo', 'Dueño': 'Dueno'}, inplace=True)
        
        print(f"[EXCEL DEBUG] DataFrame inicial: {len(df)} filas")
        if df.empty:
            print("[EXCEL ERROR] El archivo de productos manuales está vacío")
            return resultados
            
        if dueno_filtro:
            print(f"[EXCEL DEBUG] Filtrando por dueño: {dueno_filtro}")
            df = df[df['Dueno'].astype(str).str.lower() == str(dueno_filtro).lower()]
            print(f"[EXCEL DEBUG] Después de filtrar por dueño: {len(df)} filas")
            
        if termino_busqueda:
            print(f"[EXCEL DEBUG] Filtrando por término de búsqueda: {termino_busqueda}")
            tokens = [t.strip() for t in str(termino_busqueda).split() if t.strip()]
            if tokens:
                mask_all = pd.Series(True, index=df.index)
                for tok in tokens:
                    tok_mask = (
                        df['Nombre'].astype(str).str.contains(tok, case=False, na=False) |
                        df['Codigo'].astype(str).str.contains(tok, case=False, na=False) |
                        df['Proveedor'].astype(str).str.contains(tok, case=False, na=False)
                    )
                    mask_all &= tok_mask
                
                filtered_by_term = df[mask_all]
                print(f"[EXCEL DEBUG] Después de filtrar por tokens de búsqueda: {len(filtered_by_term)} filas")
                
                # Si no hay resultados después de filtrar por término, mostramos todos los productos disponibles
                if len(filtered_by_term) == 0:
                    print("[EXCEL DEBUG] ======= DIAGNÓSTICO: PRODUCTOS DISPONIBLES =======")
                    print(f"[EXCEL DEBUG] No se encontraron productos que coincidan con '{termino_busqueda}'")
                    print("[EXCEL DEBUG] Mostrando TODOS los productos disponibles después de filtrar por dueño:")
                    try:
                        # Mostrar cada fila individualmente para garantizar que se vea en los logs
                        for idx, row in df.iterrows():
                            print(f"[EXCEL DEBUG] Fila {idx}:")
                            for col in df.columns:
                                print(f"[EXCEL DEBUG]   {col}: {row[col]}")
                            print("[EXCEL DEBUG] ---")
                            
                            # Intentar una búsqueda manual por si hay problemas de formato
                            if termino_busqueda and isinstance(termino_busqueda, str):
                                codigo_str = str(row.get('Codigo', '')).lower()
                                nombre_str = str(row.get('Nombre', '')).lower()
                                term_lower = termino_busqueda.lower()
                                print(f"[EXCEL DEBUG] Prueba manual - Código: '{codigo_str}', Término: '{term_lower}', ¿Coincide?: {term_lower in codigo_str}")
                                print(f"[EXCEL DEBUG] Prueba manual - Nombre: '{nombre_str}', Término: '{term_lower}', ¿Coincide?: {term_lower in nombre_str}")
                                
                    except Exception as e:
                        print(f"[EXCEL DEBUG] Error al mostrar filas: {e}")
                    print("[EXCEL DEBUG] ======= FIN DIAGNÓSTICO =======")
                    
            # Alternativa más agresiva: mostrar SIEMPRE los productos disponibles
        print("[EXCEL DEBUG] Mostrando productos disponibles aunque no coincidan con el término de búsqueda")
        print("[EXCEL DEBUG] Productos disponibles para mostrar en la UI:")
        
        # Guardar el DataFrame filtrado original
        filtered_empty = filtered_by_term.empty
        
        # Si no hay resultados de búsqueda pero hay productos para este dueño, mostrarlos todos
        if filtered_empty and len(df) > 0:
            print(f"[EXCEL DEBUG] No se encontraron coincidencias exactas para '{termino_busqueda}', mostrando todos los productos disponibles")
            
            # DESACTIVAR ESTA LÍNEA PARA VOLVER AL COMPORTAMIENTO NORMAL
            # Mostrar todos los productos disponibles para este dueño, ignorando el filtro de término
            df_result = df  # Usar todos los productos disponibles
            
            # Crear resultados para cada producto
            for _, row in df_result.iterrows():
                precio_val, precio_error = parse_price(row.get('Precio', ''))
                resultado = {
                    'codigo': row.get('Codigo', ''),
                    'nombre': row.get('Nombre', '') + " [SIN FILTRO]",  # Marcar que no coincide con el filtro
                    'precio': precio_val,
                    'precio_texto': str(row.get('Precio', '')) if precio_error else None,
                    'proveedor': row.get('Proveedor', ''),
                    'observaciones': row.get('Observaciones', ''),
                    'dueno': row.get('Dueno', ''),
                    'es_manual': True
                }
                resultados.append(resultado)
                print(f"[EXCEL DEBUG] Producto añadido: {resultado['codigo']} - {resultado['nombre']}")
            
            # Devolver resultados directamente sin seguir procesando
            print(f"[EXCEL DEBUG] Total resultados (mostrados sin filtro): {len(resultados)}")
            return resultados
        else:
            # Continuar con el comportamiento normal si hay resultados o no hay productos disponibles
            df = filtered_by_term
            
        print(f"[EXCEL DEBUG] Resultados finales: {len(df)} filas")
        
        for _, row in df.iterrows():
            precio_val, precio_error = parse_price(row.get('Precio', ''))
            resultados.append({
                'codigo': row.get('Codigo', ''),
                'nombre': row.get('Nombre', ''),
                'precio': precio_val,
                'precio_texto': str(row.get('Precio', '')) if precio_error else None,
                'proveedor': row.get('Proveedor', ''),
                'observaciones': row.get('Observaciones', ''),
                'dueno': row.get('Dueno', ''),
                'es_manual': True
            })
    except Exception as e:
        print(f"[EXCEL ERROR] Error en buscar_en_excel_manual: {e}")
        import traceback
        print(traceback.format_exc())
    return resultados
    
def buscar_en_excel_proveedor(termino_busqueda, proveedor, filtro_adicional=None):
    """Buscar productos en archivos Excel del proveedor especificado"""
    resultados = []
    claves_vistas = set()  # Conjunto para controlar duplicados
    
    try:
        print(f"[EXCEL DEBUG] Iniciando búsqueda para proveedor '{proveedor}'")
        
        # Verificar que exista la configuración del proveedor (case-insensitive)
        proveedor_lower = proveedor.lower() if proveedor else ''
        
        # Verificar primero con la clave exacta
        if proveedor in PROVEEDOR_CONFIG:
            proveedor_key = proveedor
        # Si no existe, buscar de forma case-insensitive
        else:
            proveedor_key = next((k for k in PROVEEDOR_CONFIG.keys() if k.lower() == proveedor_lower), None)
            
        if not proveedor_key:
            print(f"[EXCEL] Error: Proveedor '{proveedor}' no configurado")
            return []
        
        # Obtener configuración del proveedor
        config = PROVEEDOR_CONFIG[proveedor_key]
        print(f"[EXCEL DEBUG] Configuración del proveedor: {config}")
        
        excel_folder = config.get('folder', proveedor)
        dueno = config.get('dueno', 'ferreteria_general')
        print(f"[EXCEL DEBUG] Carpeta Excel: {excel_folder}, Dueño: {dueno}")
        
        # Preparar ruta al directorio de archivos
        directorio_base = os.path.join('listas_excel', excel_folder)
        print(f"[EXCEL DEBUG] Directorio base completo: {directorio_base}")
        directorio_abs = os.path.abspath(directorio_base)
        print(f"[EXCEL DEBUG] Directorio absoluto: {directorio_abs}")
        
        # Si no existe el directorio, salir
        if not os.path.exists(directorio_base):
            print(f"[EXCEL] Error: Directorio no existe '{directorio_base}'")
            # Listar directorios disponibles para ayudar en el diagnóstico
            print(f"[EXCEL DEBUG] Contenido de 'listas_excel': {os.listdir('listas_excel') if os.path.exists('listas_excel') else 'No existe'}")
            return []
        
        # Listar todos los archivos Excel en el directorio
        archivos_excel = []
        try:
            for root, _, files in os.walk(directorio_base):
                print(f"[EXCEL DEBUG] Explorando directorio: {root}")
                for file in files:
                    if file.endswith('.xlsx') or file.endswith('.xls'):
                        ruta_completa = os.path.join(root, file)
                        archivos_excel.append(ruta_completa)
                        print(f"[EXCEL DEBUG] Archivo encontrado: {ruta_completa}")
        except Exception as e:
            print(f"[EXCEL DEBUG] Error al listar archivos: {str(e)}")
            return []
        
        # Verificar si hay archivos
        if not archivos_excel:
            print(f"[EXCEL] No se encontraron archivos Excel en '{directorio_base}'")
            return []
            
        print(f"[EXCEL] Buscando '{termino_busqueda}' en {len(archivos_excel)} archivos de '{proveedor}'")
        
        # Buscar en cada archivo
        termino_busqueda = termino_busqueda.lower()
        for archivo in archivos_excel:
            nombre_archivo = os.path.basename(archivo)
            print(f"[EXCEL DEBUG] Procesando archivo: {nombre_archivo}")
            try:
                # Cargar el libro Excel
                print(f"[EXCEL DEBUG] Intentando cargar archivo: {archivo}")
                try:
                    from openpyxl import load_workbook
                    wb = load_workbook(archivo, read_only=True)
                    print(f"[EXCEL DEBUG] Archivo cargado exitosamente: {nombre_archivo}")
                except Exception as excel_e:
                    print(f"[EXCEL ERROR] Error al cargar archivo {nombre_archivo}: {str(excel_e)}")
                    continue
                
                # Procesar cada hoja
                for ws_name in wb.sheetnames:
                    ws = wb[ws_name]
                    
                    # Determinar filas y columnas a procesar
                    filas = list(ws.rows)
                    if not filas:
                        continue
                    
                    # Buscar en cada fila
                    for row_idx, row in enumerate(filas, 1):
                        # Guardamos valores originales y convertimos a minúsculas solo para búsqueda
                        row_values_orig = [str(cell.value) if cell.value is not None else '' for cell in row]
                        row_values_lower = [val.lower() for val in row_values_orig]
                        row_text_lower = ' '.join(row_values_lower)
                        
                        # Verificar si el término está específicamente en código o nombre
                        codigo_lower = row_values_lower[0] if len(row_values_lower) > 0 else ''
                        nombre_lower = row_values_lower[1] if len(row_values_lower) > 1 else ''
                        
                        # Hacer un filtrado más estricto
                        match_found = False
                        # Si el término parece un código (solo números), buscar coincidencia exacta en código
                        if termino_busqueda.isdigit():
                            # Coincidencia exacta con el código
                            if termino_busqueda == codigo_lower:
                                match_found = True
                                print(f"[EXCEL] ✅ Coincidencia exacta de código: {termino_busqueda} == {codigo_lower}")
                            # Si no coincide exactamente, buscar como parte del código (solo si tiene espacios)
                            elif codigo_lower and termino_busqueda in codigo_lower.split():
                                match_found = True
                                print(f"[EXCEL] ✅ Coincidencia en parte del código: {termino_busqueda} en {codigo_lower}")
                        # Si no es un código, buscar en cualquier parte de la fila
                        else:
                            if termino_busqueda in row_text_lower:
                                match_found = True
                                print(f"[EXCEL] ✅ Coincidencia de texto: {termino_busqueda} en fila")
                        
                        if not match_found:
                            continue
                            
                        # Aplicar filtro adicional si existe
                        if filtro_adicional and filtro_adicional.lower() not in row_text_lower:
                            continue
                            
                        # Extraer datos de la fila
                        codigo = row_values_orig[0] if len(row_values_orig) > 0 else ''
                        nombre = row_values_orig[1] if len(row_values_orig) > 1 else ''
                        precio = 0.0
                        
                        # Intentar extraer precio
                        if len(row_values_orig) > 2:
                            try:
                                precio_text = row_values_orig[2].replace('.', '').replace(',', '.')
                                precio = float(precio_text) if precio_text else 0.0
                            except (ValueError, TypeError):
                                precio = 0.0
                            
                            # Crear clave única para evitar duplicados
                            clave = f"{codigo}_{proveedor.lower()}_{nombre_archivo}_{ws_name}_{row_idx}"
                            
                            # Verificar si ya hemos procesado este resultado exacto
                            if clave in claves_vistas:
                                print(f"[EXCEL] ⚠️ Omitiendo duplicado para código {codigo} en {nombre_archivo}")
                                continue
                                
                            claves_vistas.add(clave)
                            
                            # Crear resultado
                            resultado = {
                                'codigo': codigo,
                                'nombre': nombre if nombre else f"Fila {row_idx}",
                                'precio': precio,
                                'proveedor': proveedor.title(),
                                'archivo': nombre_archivo,
                                'hoja': ws_name,
                                'fila': row_idx,
                                'tipo': 'excel',
                                'dueno': dueno,
                                'row_text': row_text_lower
                            }
                            resultados.append(resultado)
                
            except Exception as e:
                print(f"[EXCEL] Error al procesar archivo '{archivo}': {e}")
                continue
                
    except Exception as e:
        print(f"[EXCEL] Error general en buscar_en_excel_proveedor: {e}")
    
    print(f"[EXCEL] Búsqueda en proveedor '{proveedor}' completada: {len(resultados)} resultados (de {len(claves_vistas)} coincidencias)")
    return resultados

def buscar_codigo_exacto_en_proveedor(codigo, proveedor, solo_ricky=False, solo_fg=False):
    """Busca un código específico en un proveedor específico con criterios muy estrictos"""
    resultados = []
    claves_vistas = set()  # Conjunto para controlar duplicados
    
    try:
        print(f"🎯 [EXCEL] Iniciando búsqueda exacta de código '{codigo}' en proveedor '{proveedor}'")
        if not codigo.isdigit() or not proveedor:
            print(f"⚠️ [EXCEL] Código no numérico o proveedor vacío")
            return []
        
        # Verificar que exista la configuración del proveedor
        if proveedor not in PROVEEDOR_CONFIG:
            print(f"⚠️ [EXCEL] Error: Proveedor '{proveedor}' no configurado")
            return []
            
        # Obtener configuración del proveedor
        config = PROVEEDOR_CONFIG[proveedor]
        dueno = config.get('dueno', 'ferreteria_general')
        print(f"🔍 [EXCEL] Dueño del proveedor: {dueno}")
        
        # Verificar si el proveedor está habilitado para el filtro de dueño
        if (solo_ricky and dueno != 'ricky') or (solo_fg and dueno != 'ferreteria_general'):
            print(f"⚠️ [EXCEL] Proveedor no coincide con filtro de dueño")
            return []
        
        # Directorio base para archivos del proveedor
        directorio_base = os.path.join('listas_excel', config.get('folder', proveedor))
        print(f"📁 [EXCEL] Buscando en directorio: {directorio_base}")
        
        if not os.path.exists(directorio_base):
            print(f"⚠️ [EXCEL] Directorio no existe: {directorio_base}")
            return []
            
        # Buscar en todos los archivos Excel disponibles
        archivos_excel = []
        for root, _, files in os.walk(directorio_base):
            for file in files:
                if file.endswith('.xlsx') or file.endswith('.xls'):
                    ruta_completa = os.path.join(root, file)
                    archivos_excel.append(ruta_completa)
                    
        print(f"📊 [EXCEL] Encontrados {len(archivos_excel)} archivos para analizar")
        
        # Buscar el código exacto en cada archivo
        for archivo in archivos_excel:
            try:
                from openpyxl import load_workbook
                wb = load_workbook(archivo, read_only=True, data_only=True)
                
                for ws_name in wb.sheetnames:
                    ws = wb[ws_name]
                    filas = list(ws.rows)
                    
                    for row_idx, row in enumerate(filas, 1):
                        # Obtener valores de celda
                        row_values = [str(cell.value) if cell.value is not None else '' for cell in row]
                        if len(row_values) == 0:
                            continue
                        
                        # Obtener código y verificar coincidencia exacta
                        fila_codigo = row_values[0].strip() if len(row_values) > 0 else ''
                        if fila_codigo == codigo:
                            print(f"✅ [EXCEL] Coincidencia exacta encontrada en {os.path.basename(archivo)}, hoja {ws_name}, fila {row_idx}")
                            
                            # Extraer datos
                            nombre = row_values[1] if len(row_values) > 1 else f"Fila {row_idx}"
                            precio = 0.0
                            
                            if len(row_values) > 2:
                                try:
                                    precio_text = row_values[2].replace('.', '').replace(',', '.')
                                    precio = float(precio_text) if precio_text else 0.0
                                except (ValueError, TypeError):
                                    precio = 0.0
                                    
                            # Verificar que el archivo corresponda realmente al proveedor buscado
                            archivo_base = os.path.basename(archivo).lower()
                            if not archivo_base.startswith(proveedor.lower()):
                                print(f"⚠️ [EXCEL] Omitiendo resultado de {archivo_base}, no coincide con proveedor {proveedor.lower()}")
                                continue
                                
                            # Crear clave única para evitar duplicados
                            clave = f"{codigo}_{proveedor.lower()}_{archivo_base}_{ws_name}_{row_idx}"
                            
                            # Verificar si ya hemos procesado este resultado exacto
                            if clave in claves_vistas:
                                print(f"⚠️ [EXCEL] Omitiendo duplicado para código {codigo} en {archivo_base}")
                                continue
                                
                            claves_vistas.add(clave)
                            
                            # Crear resultado
                            resultado = {
                                'codigo': codigo,
                                'nombre': nombre,
                                'precio': precio,
                                'proveedor': proveedor.title(),
                                'archivo': os.path.basename(archivo),
                                'hoja': ws_name,
                                'fila': row_idx,
                                'tipo': 'excel',
                                'dueno': dueno,
                                'coincidencia_exacta': True
                            }
                            resultados.append(resultado)
            except Exception as e:
                print(f"⚠️ [EXCEL] Error al procesar archivo {archivo}: {e}")
                
        print(f"🎯 [EXCEL] Búsqueda exacta completada: {len(resultados)} resultados (de {len(claves_vistas)} coincidencias)")
        return resultados
        
    except Exception as e:
        print(f"❌ [EXCEL] Error en buscar_codigo_exacto_en_proveedor: {e}")
        return []

def procesar_archivo_excel(archivo, config, termino_busqueda, filtro_adicional, proveedor_key, dueno='ricky'):
    """Procesar un archivo Excel específico"""
    resultados = []
    print(f"📊 [PROCESAR_EXCEL] Procesando archivo: '{archivo}' | Proveedor: '{proveedor_key}' | Dueño: '{dueno}' | Término: '{termino_busqueda}'")
    
    try:
        # Obtener la carpeta específica del dueño
        carpeta_dueno = get_excel_folder_for_dueno(dueno)
        
        # Buscar el archivo que coincida en la carpeta del dueño
        archivos = [f for f in os.listdir(carpeta_dueno) if f.lower().startswith(archivo.lower()) and f.endswith('.xlsx') and f != 'productos_manual.xlsx']
        if not archivos:
            # Intentar con patrón Nombre-*.xlsx por si hay sufijos de fecha
            archivos = [f for f in os.listdir(carpeta_dueno) if f.lower().startswith(f"{archivo.lower()}-") and f.endswith('.xlsx') and f != 'productos_manual.xlsx']
            if not archivos:
                return resultados
        
        archivo_path = os.path.join(carpeta_dueno, archivos[0])
        if not os.path.exists(archivo_path):
            return resultados
        
        # Intentar leer con la fila de encabezado configurada; si falla, probar varias alternativas
        df = None
        header_candidates = [config.get('fila_encabezado', 0)]
        # Probar también primeras 12 filas como posibles encabezados si no coincide
        header_candidates += [i for i in range(0, 12) if i not in header_candidates]
        for header_idx in header_candidates:
            try:
                df = pd.read_excel(archivo_path, header=header_idx)
                if df is not None and not df.empty:
                    break
            except Exception:
                continue
        if df is None or df.empty:
            return resultados
        
        if df.empty:
            return resultados
        
        # Encontrar las columnas correctas
        col_codigo = encontrar_columna(df.columns, config['codigo'])
        col_producto = encontrar_columna(df.columns, config['producto'])
        col_precio = encontrar_columna(df.columns, config['precio'])
        
        print(f"📊 [PROCESAR_EXCEL] Columnas encontradas - Código: '{col_codigo}' | Producto: '{col_producto}' | Precio: '{col_precio}'")
        print(f"📊 [PROCESAR_EXCEL] Configuración de precio: {config['precio']}")
        print(f"📊 [PROCESAR_EXCEL] Columnas disponibles: {list(df.columns)}")
        
        if not col_producto:
            print(f"📊 [PROCESAR_EXCEL] ERROR: No se encontró columna de producto")
            return resultados
        
        # Convertir columnas a string para evitar errores con .str accessor
        df[col_producto] = df[col_producto].astype(str)
        if col_codigo:
            df[col_codigo] = df[col_codigo].astype(str)
        
        # Filtrar por término de búsqueda (soporta combinaciones con AND entre tokens)
        if termino_busqueda:
            tokens = [t.strip() for t in str(termino_busqueda).split() if t.strip()]
            if tokens:
                mask_all = pd.Series([True] * len(df))
                for tok in tokens:
                    mask_tok = df[col_producto].str.contains(tok, case=False, na=False)
                    if col_codigo:
                        mask_tok |= df[col_codigo].str.contains(tok, case=False, na=False)
                    mask_all &= mask_tok
                df = df[mask_all]
        
        # Aplicar filtro adicional si existe
        if filtro_adicional:
            mask_adicional = df[col_producto].str.contains(filtro_adicional, case=False, na=False)
            if col_codigo:
                mask_adicional |= df[col_codigo].str.contains(filtro_adicional, case=False, na=False)
            df = df[mask_adicional]
        
        # Verificar si este proveedor usa colores
        usar_colores = config.get('usar_colores', False)
        color_precio = config.get('color_precio', 'FFFF00')
        columnas_precio = config.get('columnas_precio', [])
        
        if usar_colores and columnas_precio:
            # Procesar con colores usando openpyxl directamente
            from openpyxl import load_workbook
            print(f"🎨 Procesando archivo con colores: {archivo_path}")
            wb = load_workbook(archivo_path)
            ws = wb.active
            print(f"📊 Archivo cargado: {ws.max_row} filas, {ws.max_column} columnas")
            
            # Encontrar índices de columnas
            col_codigo_idx = None
            col_producto_idx = None
            col_marca_idx = None
            
            for col_idx in range(1, ws.max_column + 1):
                header_cell = ws.cell(row=1, column=col_idx)
                if header_cell.value:
                    header_val = str(header_cell.value).strip().upper()
                    if col_codigo and any(alias.upper() == header_val for alias in config['codigo']):
                        col_codigo_idx = col_idx
                    elif any(alias.upper() == header_val for alias in config['producto']):
                        col_producto_idx = col_idx
                    elif 'marca' in config and any(alias.upper() == header_val for alias in config['marca']):
                        col_marca_idx = col_idx
            
            # Procesar fila por fila
            for row_idx in range(2, ws.max_row + 1):  # Saltar encabezado
                # Obtener datos básicos
                codigo = ws.cell(row=row_idx, column=col_codigo_idx).value if col_codigo_idx else ''
                nombre = ws.cell(row=row_idx, column=col_producto_idx).value if col_producto_idx else ''
                marca = ws.cell(row=row_idx, column=col_marca_idx).value if col_marca_idx else ''
                
                if not nombre:
                    continue
                
                # Aplicar filtros de búsqueda
                if termino_busqueda:
                    tokens = [t.strip() for t in str(termino_busqueda).split() if t.strip()]
                    coincide = True
                    for tok in tokens:
                        if not (tok.lower() in str(nombre).lower() or 
                               (codigo and tok.lower() in str(codigo).lower()) or
                               (marca and tok.lower() in str(marca).lower())):
                            coincide = False
                            break
                    if not coincide:
                        continue
                
                # Obtener precio de celda coloreada
                precio_raw, col_precio_usada = obtener_precio_por_color(ws, row_idx, columnas_precio, color_precio)
                
                if precio_raw is None:
                    continue  # Solo incluir productos con precio coloreado
                
                print(f"✅ Producto encontrado: {nombre} - Precio: {precio_raw} (Columna: {col_precio_usada})")
                
                precio_val, precio_error = parse_price(precio_raw)
                
                resultado = {
                    'codigo': str(codigo).strip() if codigo else '',
                    'nombre': str(nombre).strip(),
                    'precio': precio_val,
                    'precio_texto': str(precio_raw) if precio_error else None,
                    'proveedor': proveedor_key.title(),
                    'observaciones': f'Marca: {marca}' if marca else '',
                    'dueno': dueno,
                    'es_manual': False,
                    'columna_precio': col_precio_usada
                }
                resultados.append(resultado)
        else:
            # Procesamiento normal sin colores
            for _, row in df.iterrows():
                codigo = str(row[col_codigo]).strip() if col_codigo and pd.notna(row[col_codigo]) else ''
                nombre = str(row[col_producto]).strip() if pd.notna(row[col_producto]) else ''
                precio_raw = row[col_precio] if col_precio and pd.notna(row[col_precio]) else ''
                
                if not nombre:
                    continue
                
                precio_val, precio_error = parse_price(precio_raw)
                
                resultado = {
                    'codigo': codigo,
                    'nombre': nombre,
                    'precio': precio_val,
                    'precio_texto': str(precio_raw) if precio_error else None,
                    'proveedor': proveedor_key.title(),
                    'observaciones': '',
                    'dueno': dueno,
                    'es_manual': False
                }
                resultados.append(resultado)
                print(f"📊 [PROCESAR_EXCEL] ✅ Producto agregado: {codigo} - {nombre} - Precio: {precio_val}")
    
    except Exception as e:
        print(f"Error al procesar archivo {archivo}: {e}")
    
    return resultados

def _normalize_text(value):
    try:
        text = str(value or '')
        text = unicodedata.normalize('NFKD', text)
        text = ''.join([c for c in text if not unicodedata.combining(c)])
        return text.strip().lower()
    except Exception:
        return str(value or '').strip().lower()

def encontrar_columna(columnas, aliases):
    """Encontrar la columna correcta basada en los aliases (acentos/espacios/caso-insensible)"""
    normalized_aliases = [_normalize_text(a) for a in aliases]
    for col in columnas:
        col_norm = _normalize_text(col)
        if col_norm in normalized_aliases:
            return col
    # Fallback: coincidencia por contiene
    for col in columnas:
        col_norm = _normalize_text(col)
        for a in normalized_aliases:
            if a in col_norm:
                return col
    return None

# ===================== VERIFICACIÓN DE CÓDIGOS POR PROVEEDOR (PDF Review Helper) =====================
@app.route('/verificar_codigos_proveedor', methods=['POST'])
@login_required
def verificar_codigos_proveedor():
    """Verifica códigos contra el Excel del proveedor y (opcional) productos manuales.
    Request JSON: { proveedor: str, codigos: [..], incluir_manual: bool }
    Respuesta:
      success, detalles (por código ORIGINAL), meta (info extra), debug (diagnóstico)
    Normalización aplicada: quita espacios, NBSP, .0 finales y ceros a la izquierda; genera variantes.
    """
    try:
        import re as _re  # asegurar disponibilidad aun si import global cambia
        data = request.get_json(silent=True) or {}
        proveedor = (data.get('proveedor') or '').strip()
        codigos_input = data.get('codigos') or []
        incluir_manual = bool(data.get('incluir_manual', True))
        if not proveedor:
            return jsonify({'success': False, 'error': 'Proveedor no informado'}), 400
        if not isinstance(codigos_input, list) or not codigos_input:
            return jsonify({'success': False, 'error': 'Lista de códigos vacía'}), 400
        if proveedor not in PROVEEDOR_CONFIG:
            return jsonify({'success': False, 'error': f'Proveedor "{proveedor}" no configurado'}), 404

        def _norm_code_base(c):
            s = str(c).strip()
            if not s:
                return ''
            s = s.replace('\u00A0', '').replace(' ', '')
            if _re.fullmatch(r"\d+\.0+", s):
                s = s.split('.')[0]
            m = _re.fullmatch(r"(\d+)\.(\d+)", s)
            if m and set(m.group(2)) == {'0'}:
                s = m.group(1)
            return s

        def _norm_variants(c):
            base = _norm_code_base(c)
            variants = {base.lower()}
            if base.endswith('.0'):
                variants.add(base[:-2].lower())
            if _re.fullmatch(r"0+\d+", base):
                variants.add(base.lstrip('0').lower() or '0')
            if _re.fullmatch(r"\d+", base):
                variants.add(str(int(base)).lower())
            return variants

        codigos_originales = []
        codigos_norm = []
        codigos_norm_variants = []
        for c in codigos_input:
            orig = str(c).strip()
            if not orig:
                continue
            base = _norm_code_base(orig)
            if not base:
                continue
            codigos_originales.append(orig)
            codigos_norm.append(base)
            codigos_norm_variants.append(_norm_variants(base))

        config = PROVEEDOR_CONFIG[proveedor]
        dueno_proveedor = None
        for d, cfg in DUENOS_CONFIG.items():
            if proveedor in cfg.get('proveedores_excel', []):
                dueno_proveedor = d
                break
        if not dueno_proveedor:
            dueno_proveedor = 'ricky'

        carpeta_dueno = get_excel_folder_for_dueno(dueno_proveedor)
        if not os.path.isdir(carpeta_dueno):
            return jsonify({'success': False, 'error': f'Carpeta Excel para dueño {dueno_proveedor} no existe'}), 500

        archivo_candidates = [f for f in os.listdir(carpeta_dueno) if f.lower().startswith(proveedor.lower()) and f.endswith('.xlsx') and f != 'productos_manual.xlsx']
        if not archivo_candidates:
            archivo_candidates = [f for f in os.listdir(carpeta_dueno) if f.lower().startswith(f"{proveedor.lower()}-") and f.endswith('.xlsx') and f != 'productos_manual.xlsx']
        if not archivo_candidates:
            return jsonify({'success': False, 'error': 'Archivo Excel del proveedor no encontrado'}), 404
        excel_path = os.path.join(carpeta_dueno, archivo_candidates[0])

        header_candidates = [config.get('fila_encabezado', 0)] + [i for i in range(0, 12) if i != config.get('fila_encabezado', 0)]
        df = None
        header_usado = None
        for h in header_candidates:
            try:
                tmp = pd.read_excel(excel_path, header=h)
                if tmp is not None and not tmp.empty:
                    df = tmp
                    header_usado = h
                    break
            except Exception:
                continue
        if df is None or df.empty:
            return jsonify({'success': False, 'error': 'No se pudo leer el Excel o está vacío'}), 500

        col_codigo = encontrar_columna(df.columns, config.get('codigo', [])) if 'codigo' in config else None
        if not col_codigo:
            for c in df.columns:
                if 'codi' in _normalize_text(c):
                    col_codigo = c
                    break
        if not col_codigo:
            return jsonify({'success': False, 'error': 'Columna de código no encontrada en el Excel'}), 500

        serie_codigos = df[col_codigo].fillna('')
        indice_excel = {}
        sample_excel_norm = []
        for val in serie_codigos:
            raw = str(val)
            base = _norm_code_base(raw)
            if not base:
                continue
            variants = _norm_variants(base)
            for v in variants:
                indice_excel.setdefault(v, set()).add(raw)
            if len(sample_excel_norm) < 120:
                sample_excel_norm.append(base)

        encontrados = {}
        meta = {}
        mapping_norm = {}
        for i, base in enumerate(codigos_norm):
            original = codigos_originales[i]
            variants = codigos_norm_variants[i]
            match_raw = None
            match_variant = None
            for v in variants:
                if v in indice_excel:
                    match_variant = v
                    match_raw = sorted(list(indice_excel[v]))[0]
                    break
            if match_variant:
                encontrados[original] = True
                meta[original] = {'variant_usada': match_variant, 'excel_raw': match_raw, 'origen': 'excel'}
            else:
                encontrados[original] = False
                meta[original] = {'motivo': 'no_en_excel', 'variants_generadas': list(variants)}
            mapping_norm[base] = original

        if incluir_manual and os.path.exists(MANUAL_PRODUCTS_FILE):
            try:
                dfm = pd.read_excel(MANUAL_PRODUCTS_FILE)
                dfm.rename(columns={'Código': 'Codigo'}, inplace=True)
                if 'Codigo' in dfm.columns:
                    for val in dfm['Codigo'].fillna(''):
                        raw = str(val)
                        base_m = _norm_code_base(raw)
                        if not base_m:
                            continue
                        variants_m = _norm_variants(base_m)
                        for i, base_user in enumerate(codigos_norm):
                            original_user = codigos_originales[i]
                            if encontrados.get(original_user):
                                continue
                            variants_user = codigos_norm_variants[i]
                            if variants_user & variants_m:
                                encontrados[original_user] = True
                                meta[original_user] = {
                                    'variant_usada': list(variants_user & variants_m)[0],
                                    'excel_raw': raw,
                                    'origen': 'manual'
                                }
            except Exception as e_manu:
                meta['manual_error'] = str(e_manu)

        for original in codigos_originales:
            if not encontrados.get(original):
                motivo = meta.get(original, {}).get('motivo') or 'no_en_excel_ni_manual'
                meta[original]['motivo'] = motivo

        total_ok = sum(1 for v in encontrados.values() if v)
        return jsonify({
            'success': True,
            'proveedor': proveedor,
            'total': len(codigos_originales),
            'encontrados': total_ok,
            'detalles': encontrados,
            'meta': meta,
            'debug': {
                'header_usado': header_usado,
                'col_codigo': col_codigo,
                'archivo_excel': os.path.basename(excel_path),
                'muestra_codigos_excel_normalizados': sample_excel_norm[:50],
                'mapping_norm_to_original': mapping_norm
            }
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

# ===================== FACTURAS PDF =====================
import re

PDF_ITEM_REGEXES = [
    # Formato: cant codigo descripcion precio_unit subtotal  (miles con . y decimales con , o .)
    re.compile(r"^(?P<cant>\d+(?:[.,]\d+)?)\s+(?P<codigo>[A-Za-z0-9\-]{3,})\s+(?P<desc>.+?)\s+(?P<unit>\d[\d\.,]*\d)\s+(?P<subtotal>\d[\d\.,]*\d)$"),
    # Formato sin subtotal (cant codigo desc precio)
    re.compile(r"^(?P<cant>\d+(?:[.,]\d+)?)\s+(?P<codigo>[A-Za-z0-9\-]{3,})\s+(?P<desc>.+?)\s+(?P<unit>\d[\d\.,]*\d)$"),
    # Formato donde precio y subtotal pueden ir con signo $ opcional
    re.compile(r"^(?P<cant>\d+(?:[.,]\d+)?)\s+(?P<codigo>[A-Za-z0-9\-]{3,})\s+(?P<desc>.+?)\s+\$?(?P<unit>\d[\d\.,]*\d)\s+\$?(?P<subtotal>\d[\d\.,]*\d)$"),
]

def _parse_decimal(num_str: str) -> float:
    if not num_str:
        return 0.0
    s = num_str.strip()
    # Normalizar distintos formatos de miles/decimales.
    # Casos:
    #  - Español: 1.234,56
    #  - US: 24,267.03
    #  - Simple: 1234.56 / 1234,56
    # Estrategia: identificar qué separador aparece a la derecha.
    if ',' in s and '.' in s:
        last_comma = s.rfind(',')
        last_dot = s.rfind('.')
        if last_dot > last_comma:
            # Punto es decimal (formato US) -> quitar comas
            s = s.replace(',', '')
        else:
            # Coma es decimal (formato ES) -> quitar puntos y cambiar coma por punto
            s = s.replace('.', '').replace(',', '.')
    elif ',' in s and '.' not in s:
        # Asumir coma decimal
        s = s.replace(',', '.')
    elif '.' in s and s.count('.') > 1:
        # Múltiples puntos: probablemente puntos de miles + decimal final
        parts = s.split('.')
        # Unir todos menos el último como miles
        decimal_part = parts[-1]
        s = ''.join(parts[:-1]) + '.' + decimal_part
    try:
        return float(s)
    except Exception:
        return 0.0

def extraer_items_factura_pdf(file_stream, debug=False):
    """Extrae items (codigo, descripcion, cantidad, precio_unitario) desde un PDF de factura.
    1. Intenta parse línea a línea con regex flexibles.
    2. Fallback heurístico: identifica patrones cant / codigo / precio final.
    3. Devuelve (items, lines) si debug=True, sino solo items.
    """
    if not _HAS_PDF:
        raise RuntimeError("Librería pdfplumber no instalada")
    try:
        with pdfplumber.open(file_stream) as pdf:
            texto_paginas = []
            for page in pdf.pages:
                try:
                    t = page.extract_text() or ''
                    texto_paginas.append(t)
                except Exception:
                    continue
            contenido = '\n'.join(texto_paginas)
    except Exception as e:
        raise RuntimeError(f"Error leyendo PDF: {e}")

    raw_lines = [l.rstrip() for l in contenido.splitlines()]
    # Normalizar espacios múltiples para facilitar regex
    norm_lines = []
    for l in raw_lines:
        l2 = re.sub(r"\s+", " ", l.strip())
        norm_lines.append(l2)

    items = []
    vistos = set()

    def add_item(codigo, desc, cant, unit):
        if not codigo:
            codigo = 'SIN-COD'
        key = (codigo, desc)
        if key in vistos:
            # sumar cantidades
            for it in items:
                if it['codigo'] == codigo and it['nombre'] == desc:
                    it['cantidad'] += cant
                    return
        else:
            vistos.add(key)
            items.append({
                'codigo': codigo[:40],
                'nombre': desc[:120].title(),
                'cantidad': int(round(cant)),
                'precio': unit,
                'precio_texto': str(unit) if unit else '0',
            })

    # Preprocesamiento especial: unir patrones de dígitos/letras separados por espacios simples en líneas largas.
    # Ejemplo: "8 0 4 2 3 . 0 0" -> "80423.00" y "P U M P K I N" -> "PUMPKIN"
    def _compact_fragment(fr):
        # Compactar secuencias de letras o dígitos separadas por un espacio: "P U M P K I N" -> "PUMPKIN", "p i n z a" -> "pinza"
        fr2 = re.sub(r'(?:(?:[A-Za-zÁÉÍÓÚÜÑ0-9]\s){2,}[A-Za-zÁÉÍÓÚÜÑ0-9])', lambda m: m.group(0).replace(' ', ''), fr)
        # Compactar números con puntos/comas dispersos: "7 1 5 8 2 . 0 0" -> "71582.00"
        fr2 = re.sub(r'(?:(?:\d\s){2,}\d)(?:\s[.,]\s(?:\d\s){1,}\d)?', lambda m: m.group(0).replace(' ', ''), fr2)
        return fr2
    norm_lines = [_compact_fragment(l) for l in norm_lines]

    # Paso 1: Regex directos
    for line in norm_lines:
        if not line:
            continue
        low = line.lower()
        if any(h in low for h in ['cliente', 'iva', 'total ', 'total:', 'factura', 'fecha', 'subtotal']):
            continue
        matched = False
        for rgx in PDF_ITEM_REGEXES:
            m = rgx.match(line)
            if m:
                cant = _parse_decimal(m.group('cant'))
                if cant <= 0:
                    matched = True
                    break
                desc = m.group('desc').strip()
                unit = _parse_decimal(m.group('unit')) if m.groupdict().get('unit') else 0.0
                codigo = m.group('codigo').strip()
                if len(desc) >= 2:
                    add_item(codigo, desc, cant, unit)
                matched = True
                break
        if matched:
            continue

    # Paso 2: Heurística especializada para layout COD Cantidad Descripción ... PRELISTA %Bon1 %Bon2 %Bon3 Importe
    if len(items) < 1:
        money_pattern = r"\d[\d\.,]*\d"
        for line in norm_lines:
            if not line or len(line) < 20:
                continue
            low = line.lower()
            if 'descripcion' in low and '%bon1' in low:
                continue
            tokens = line.split()
            if len(tokens) < 5:
                continue
            # Hallar cluster monetario final consecutivo
            i = len(tokens) - 1
            cluster_indices = []
            while i >= 0 and re.fullmatch(money_pattern, tokens[i]):
                cluster_indices.append(i)
                i -= 1
            cluster_indices.reverse()
            if len(cluster_indices) < 2:
                continue  # necesita al menos prelista + importe
            cluster_start = cluster_indices[0]
            importe_pos = cluster_indices[-1]
            codigo_token = tokens[0]
            # Determinar si segundo token es cantidad
            cantidad_val = 1.0
            desc_start = 1
            if len(tokens) > 1 and re.fullmatch(r"\d+(?:[.,]\d+)?", tokens[1]):
                # Tratar segundo token como cantidad SOLO si resto tiene sentido (suficientes tokens antes del cluster)
                posible_cant = _parse_decimal(tokens[1])
                if posible_cant > 0 and cluster_start > 2:
                    cantidad_val = posible_cant
                    desc_start = 2
            # Validar código (permitir numérico/alfanumérico con punto)
            if not re.fullmatch(r"[A-Za-z0-9\-\.]{3,}", codigo_token):
                continue
            if cluster_start - desc_start < 1:
                continue
            desc_tokens = tokens[desc_start:cluster_start]
            desc = ' '.join(desc_tokens)
            prelista_val = _parse_decimal(tokens[cluster_start])
            descuentos_vals = []
            # Tokens de descuento: los intermedios entre prelista y importe (excluyendo ambos extremos)
            if len(cluster_indices) > 2:
                for mid_idx in cluster_indices[1:-1]:
                    val = _parse_decimal(tokens[mid_idx])
                    descuentos_vals.append(val)
            unit = prelista_val
            for dval in descuentos_vals:
                if 0 < dval < 100:
                    unit *= (1 - dval/100.0)
            importe_val = _parse_decimal(tokens[importe_pos])
            if cantidad_val > 0 and importe_val > 0:
                deriv_unit = importe_val / cantidad_val
                if unit == 0 or abs(deriv_unit - unit)/max(unit,1e-6) > 0.05:
                    unit = deriv_unit
            add_item(codigo_token, desc, cantidad_val, unit)
        # Si se llenó, saltar la heurística genérica
    # Paso 3: Heurística genérica básica si aún no hay items
    if len(items) < 1:
        for line in norm_lines:
            if not line or len(line) < 10:
                continue
            tokens = line.split(' ')
            if len(tokens) < 5:
                continue
            # Intentar detectar patrón: COD  CANT  ...  PRELISTA  BON1  BON2  BON3  IMPORTE
            # Buscamos al menos 4 números (prelista + 3 bonos) y un importe final.
            money_pattern = r"\d[\d\.,]*\d"
            numeric_indices = [i for i,t in enumerate(tokens) if re.fullmatch(money_pattern, t)]
            if len(numeric_indices) >= 4:
                # Considerar últimos 4 como (prelista, bon1, bon2, bon3) o (bonos + importe)
                # Mejor: tomar desde el final: importe -> último token con dinero
                importe_idx = numeric_indices[-1]
                if importe_idx <= 4:
                    continue
                # Buscar descuentos y prelista hacia la izquierda (típicamente 4 tokens antes del importe)
                # Layout esperado: ... PRELISTA %Bon1 %Bon2 %Bon3 IMPORTE
                bon3_idx = importe_idx - 1
                bon2_idx = importe_idx - 2
                bon1_idx = importe_idx - 3
                prelista_idx = importe_idx - 4
                if prelista_idx < 2:
                    continue
                # Primer token código, segundo cantidad
                cod_token = tokens[0]
                cant_token = tokens[1]
                if not re.fullmatch(r"\d{2,}", cod_token):
                    continue
                if not re.fullmatch(r"\d+(?:[.,]\d+)?", cant_token):
                    continue
                try:
                    cantidad_val = _parse_decimal(cant_token)
                except Exception:
                    continue
                if cantidad_val <= 0:
                    continue
                # Descripción = desde tokens[2] hasta token antes de prelista_idx
                desc_tokens = tokens[2:prelista_idx]
                if len(desc_tokens) < 1:
                    continue
                desc = ' '.join(desc_tokens)
                prelista_val = _parse_decimal(tokens[prelista_idx])
                bon1_val = _parse_decimal(tokens[bon1_idx]) if bon1_idx > prelista_idx else 0.0
                bon2_val = _parse_decimal(tokens[bon2_idx]) if bon2_idx > bon1_idx else 0.0
                bon3_val = _parse_decimal(tokens[bon3_idx]) if bon3_idx > bon2_idx else 0.0
                # Calcular precio unitario neto aplicando bonificaciones sucesivas
                unit = prelista_val
                for b in (bon1_val, bon2_val, bon3_val):
                    if 0 < b < 100:
                        unit = unit * (1 - b/100.0)
                # Salvaguarda: si importe / cantidad es razonable y difiere mucho, ajustar
                importe_val = _parse_decimal(tokens[importe_idx])
                if cantidad_val > 0 and importe_val > 0:
                    deriv_unit = importe_val / cantidad_val
                    # Si difiere más de 5% usar derivado
                    if unit == 0 or abs(deriv_unit - unit)/max(unit, 1e-6) > 0.05:
                        unit = deriv_unit
                add_item(cod_token, desc, cantidad_val, unit)
                continue
            # Heurística mínima alternativa: primer token cant, segundo código
            first = tokens[0]
            second = tokens[1]
            if re.fullmatch(r"\d+(?:[.,]\d+)?", first) and re.fullmatch(r"[A-Za-z0-9\-]{3,}", second):
                cantidad_val = _parse_decimal(first)
                if cantidad_val <= 0:
                    continue
                # Precio asumido = último token dinero
                money_tokens = [t for t in tokens[2:] if re.fullmatch(money_pattern, t)]
                unit = 0.0
                if money_tokens:
                    unit = _parse_decimal(money_tokens[-1])
                desc = ' '.join(tokens[2:-1]) if len(tokens) > 3 else second
                add_item(second, desc, cantidad_val, unit)

    if debug:
        return items, norm_lines
    return items

@app.route('/importar_factura_pdf', methods=['GET','POST'])
@login_required
def importar_factura_pdf():
    if request.method == 'GET':
        try:
            proveedores_keys = sorted(list(PROVEEDOR_CONFIG.keys())) if isinstance(PROVEEDOR_CONFIG, dict) else []
        except Exception:
            proveedores_keys = []
        return render_template('importar_factura_pdf.html', soporte_pdf=_HAS_PDF, proveedores=proveedores_keys)
    # POST
    if not _HAS_PDF:
        return jsonify({'success': False, 'error': 'Dependencia pdfplumber no instalada en el servidor'}), 500

    # Log de Content-Type, args y form para depurar por qué debug no llega
    try:
        ct = request.headers.get('Content-Type')
        args_snapshot = {k: request.args.get(k) for k in request.args.keys()}
        form_snapshot = {k: request.form.get(k) for k in request.form.keys()}
        print(f"[PDF_IMPORT] CT={ct} args={args_snapshot} form={form_snapshot}")
    except Exception as e_log:
        print(f"[PDF_IMPORT] WARN logging request: {e_log}")

    if 'file' not in request.files:
        return jsonify({'success': False, 'error': 'No se envió archivo (campo file faltante)'}), 400
    f = request.files['file']
    if not f.filename.lower().endswith('.pdf'):
        return jsonify({'success': False, 'error': 'Debe subir un archivo .pdf'}), 400

    # Función helper para interpretar flags
    def _as_bool(val):
        if val is None:
            return False
        if isinstance(val, bool):
            return val
        s = str(val).strip().lower()
        return s in ('1','true','yes','on','y','si','sí')
    json_body = request.get_json(silent=True) if request.is_json else {}
    debug_mode = _as_bool(request.args.get('debug')) or _as_bool(request.form.get('debug')) or _as_bool(json_body.get('debug') if isinstance(json_body, dict) else None)
    print(f"[PDF_IMPORT] debug_mode={debug_mode} filename={f.filename}")

    items = []
    norm_lines = []
    try:
        # Reposicionar stream
        try:
            f.stream.seek(0)
        except Exception:
            pass
        # Siempre ejecutar en modo debug interno para capturar norm_lines
        try:
            result = extraer_items_factura_pdf(f, debug=True)
            if isinstance(result, tuple) and len(result) == 2:
                items, norm_lines = result
                print(f"[PDF_IMPORT] Extraction (forced debug) -> items={len(items)} lines={len(norm_lines)}")
            else:
                items = result or []
                norm_lines = []
                print(f"[PDF_IMPORT] Extraction (forced debug no tuple) -> items={len(items)}")
        except Exception as e_ext:
            return jsonify({'success': False, 'error': f'Error extrayendo items: {e_ext}'}), 500
    except Exception as e:
        return jsonify({'success': False, 'error': f'Excepción leyendo PDF: {e}'}), 500

    if not items:
        resp = {'success': False, 'error': 'No se detectaron productos en la factura'}
        sample = [l for l in norm_lines if l.strip()][:200]
        candidatos = [l for l in sample if len(re.findall(r"\d[\d\.,]*", l)) >= 4 and len(l.split()) >= 5][:40]
        resp['debug'] = {
            'requested_debug': debug_mode,
            'auto_debug': True if not debug_mode else False,
            'total_lines': len(norm_lines),
            'lines_sample': sample[:120],
            'candidatos_items': candidatos,
            'line_length_stats': None
        }
        if sample:
            largos = [len(l) for l in sample]
            resp['debug']['line_length_stats'] = {
                'min': min(largos),
                'max': max(largos),
                'avg': round(sum(largos)/len(largos),2)
            }
        print(f"[PDF_IMPORT] Sin items. total_lines={len(norm_lines)} candidatos={len(candidatos)} (auto_debug={'yes' if not debug_mode else 'requested'})")
        return jsonify(resp), 200

    out = {'success': True, 'items_detectados': len(items), 'items': items}
    if debug_mode:
        out['debug_total_lines'] = len(norm_lines)
    return jsonify(out)

@app.route('/importar_factura_pdf_confirm', methods=['POST'])
@login_required
def importar_factura_pdf_confirm():
    """Recibe lista de items (JSON) editados por el usuario y los agrega al carrito."""
    try:
        data = request.get_json(silent=True) or {}
        items = data.get('items') or []
        proveedor = data.get('proveedor','')
        if not isinstance(items, list) or not items:
            return jsonify({'success': False, 'error': 'Sin items para agregar'})
        carrito = session.get('carrito', [])
        fecha_compra = datetime.now().strftime('%Y-%m-%d')
        agregados = 0
        for it in items:
            try:
                nombre = (it.get('nombre') or '').strip()
                if not nombre:
                    continue
                codigo = (it.get('codigo') or '').strip()
                try:
                    cantidad = int(it.get('cantidad') or 0)
                except Exception:
                    cantidad = 0
                if cantidad <= 0:
                    cantidad = 1
                try:
                    precio_val = float(str(it.get('precio') or 0).replace(',','.'))
                except Exception:
                    precio_val = 0.0
                carrito.append({
                    'id': f'pdf_{len(carrito)}_{datetime.now().timestamp()}',
                    'nombre': nombre,
                    'codigo': codigo,
                    'precio': precio_val,
                    'cantidad': cantidad,
                    'fecha_compra': fecha_compra,
                    'proveedor': it.get('proveedor') or proveedor,
                    'observaciones': (it.get('observaciones') or 'Importado PDF').strip(),
                    'precio_texto': str(precio_val) if precio_val else '0',
                    'avisar_bajo_stock': 1 if (str(it.get('avisar_bajo_stock')).strip() in ('1','true','True')) else 0,
                    'min_stock_aviso': (int(it.get('min_stock_aviso')) if str(it.get('min_stock_aviso')).strip().isdigit() else None)
                })
                agregados += 1
            except Exception:
                continue
        session['carrito'] = carrito
        try:
            html = render_template('carrito_fragment_simple.html', carrito=carrito)
        except Exception:
            html = None
        return jsonify({'success': True, 'items_agregados': agregados, 'html': html})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

# Rutas para el escáner de códigos de barras
@app.route('/escanear', methods=['GET'])
@login_required
def escanear():
    # Obtener el historial de escaneos recientes (últimos 10)
    if _is_postgres_configured():
        # PostgreSQL usa una sintaxis diferente para operaciones de fecha
        historial = db_query("""
            SELECT s.id, s.codigo, s.nombre, h.cantidad, h.fecha_compra, s.cantidad as stock_actual
            FROM historial h
            JOIN stock s ON h.codigo = s.codigo
            WHERE h.fecha_compra >= NOW() - INTERVAL '1 day'
            ORDER BY h.fecha_compra DESC
            LIMIT 10
        """, fetch=True) or []
    else:
        # Consulta para SQLite
        historial = db_query("""
            SELECT s.id, s.codigo, s.nombre, h.cantidad, h.fecha_compra, s.cantidad as stock_actual
            FROM historial h
            JOIN stock s ON h.codigo = s.codigo
            WHERE h.fecha_compra >= datetime('now', '-1 day')
            ORDER BY h.fecha_compra DESC
            LIMIT 10
        """, fetch=True) or []
    
    return render_template('escanear.html', historial=historial, resultado=session.pop('resultado_escaneo', None))

@app.route('/procesar_escaneo', methods=['POST'])
@login_required
def procesar_escaneo():
    codigo_barras = request.form.get('codigo_barras', '').strip()
    cantidad = int(request.form.get('cantidad', 1))
    
    if not codigo_barras:
        flash('Debe proporcionar un código de barras', 'danger')
        return redirect(url_for('escanear'))
    
    # Buscar el producto por código de barras
    producto = db_query("SELECT * FROM stock WHERE codigo = ?", (codigo_barras,), fetch=True)
    
    if not producto:
        session['resultado_escaneo'] = {
            'tipo': 'danger',
            'titulo': 'Producto no encontrado',
            'mensaje': f'No se encontró ningún producto con el código {codigo_barras}'
        }
        return redirect(url_for('escanear'))
    
    producto = producto[0]
    
    # Actualizar el stock
    nueva_cantidad = max(0, producto['cantidad'] - cantidad)
    result = db_query("UPDATE stock SET cantidad = ? WHERE id = ?", (nueva_cantidad, producto['id']))
    
    if result:
        # Registrar en el historial
        db_query(
            "INSERT INTO historial (codigo, nombre, precio, cantidad, fecha_compra, proveedor, observaciones, precio_texto, dueno) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)",
            (
                producto['codigo'],
                producto['nombre'],
                producto['precio'],
                cantidad,
                datetime.now().isoformat(timespec='seconds'),
                producto['proveedor'],
                f'Venta por escáner de código de barras',
                str(producto['precio']),
                producto['dueno']
            )
        )
        
        # Generar mensaje de resultado
        session['resultado_escaneo'] = {
            'tipo': 'success',
            'titulo': 'Stock actualizado',
            'mensaje': f'Se descontaron {cantidad} unidades de "{producto["nombre"]}". Stock actual: {nueva_cantidad}'
        }
        
        # Verificar si el stock está bajo el umbral
        if producto.get('avisar_bajo_stock') and producto.get('min_stock_aviso') is not None:
            umbral = int(producto.get('min_stock_aviso'))
            if nueva_cantidad <= umbral:
                mensaje = f'Producto "{producto.get("nombre", "")}" bajo en stock (quedan {nueva_cantidad}, umbral {umbral}).'
                # Persistir en BD
                try:
                    user_id = session.get('user_id')
                    db_query(
                        "INSERT INTO notificaciones (codigo,nombre,proveedor,mensaje,ts,leida,user_id) VALUES (?,?,?,?,?,?,?)",
                        (
                            producto.get('codigo',''),
                            producto.get('nombre',''),
                            producto.get('proveedor',''),
                            mensaje,
                            datetime.now().isoformat(timespec='seconds'),
                            0,
                            user_id
                        )
                    )
                except Exception as _e:
                    print(f"[WARN] No se pudo persistir notificación: {_e}")
    else:
        session['resultado_escaneo'] = {
            'tipo': 'danger',
            'titulo': 'Error',
            'mensaje': 'No se pudo actualizar el stock del producto'
        }
    
    return redirect(url_for('escanear'))

# Importar el blueprint de diagnóstico Railway
try:
    from diagnostico_railway import diagnostico_railway_bp
    app.register_blueprint(diagnostico_railway_bp)
    print("[INFO] Blueprint de diagnóstico Railway registrado correctamente")
except ImportError as e:
    print(f"[WARN] No se pudo importar el blueprint de diagnóstico Railway: {e}")

# Importar el blueprint de diagnóstico de búsqueda
try:
    from diagnostico_busqueda import diagnostico_busqueda_bp
    app.register_blueprint(diagnostico_busqueda_bp)
    print("[INFO] Blueprint de diagnóstico de búsqueda registrado correctamente")
except ImportError as e:
    print(f"[WARN] No se pudo importar el blueprint de diagnóstico de búsqueda: {e}")

@app.route('/api/clean_railway_db', methods=['POST'])
@login_required
def api_clean_railway_db():
    """Endpoint para limpiar la base de datos PostgreSQL en Railway.
    
    Requiere un token de seguridad para evitar ejecuciones no autorizadas.
    El token se configura mediante la variable de entorno MIGRATION_TOKEN.
    """
    # Verificar si es entorno PostgreSQL
    if not _is_postgres_configured():
        return jsonify({
            "success": False,
            "mensaje": "Este endpoint solo funciona con PostgreSQL en Railway."
        })
    
    # Verificar token de seguridad
    migration_token = os.environ.get('MIGRATION_TOKEN', 'default_migration_token')
    
    # Obtener token de la solicitud (header o form data)
    token = request.headers.get('X-Migration-Token')
    if not token:
        token = request.form.get('token')
    
    # Verificar token (excepto en desarrollo local)
    if token != migration_token and 'DATABASE_URL' in os.environ:
        return jsonify({
            "success": False,
            "mensaje": "Token de migración inválido"
        })
    
    # Importar y ejecutar la función de limpieza
    try:
        from clean_railway_db import clean_railway_db
        resultado = clean_railway_db()
        return jsonify(resultado)
    except ImportError:
        return jsonify({
            "success": False,
            "mensaje": "No se encontró el módulo clean_railway_db.py"
        })
    except Exception as e:
        return jsonify({
            "success": False,
            "error": str(e),
            "mensaje": "Error al ejecutar la limpieza de la base de datos."
        })

# Endpoint público para limpiar la base de datos sin necesidad de consola
# Este endpoint usa un código secreto en la URL para mayor seguridad
@app.route('/limpiar_base_datos_railway/<string:codigo_secreto>', methods=['GET'])
def limpiar_base_datos_railway(codigo_secreto):
    """Endpoint público que limpia la base de datos PostgreSQL en Railway.
    Se accede directamente desde el navegador con un código secreto en la URL.
    """
    # Código secreto fijo para facilitar su uso
    CODIGO_SECRETO_LIMPIEZA = "CleanRailwayDB2025"
    
    # Verificar código secreto
    if codigo_secreto != CODIGO_SECRETO_LIMPIEZA:
        return jsonify({
            "success": False,
            "mensaje": "Código secreto inválido"
        })
    
    # Verificar si es entorno PostgreSQL
    if not _is_postgres_configured():
        return jsonify({
            "success": False,
            "mensaje": "Este endpoint solo funciona con PostgreSQL en Railway."
        })
    
    # Importar y ejecutar la función de limpieza
    try:
        from clean_railway_db import clean_railway_db
        resultado = clean_railway_db()
        
        # Crear una respuesta HTML para mostrar en el navegador
        html_response = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <title>Base de datos limpiada</title>
            <style>
                body {{ font-family: Arial, sans-serif; margin: 20px; line-height: 1.6; }}
                .success {{ color: green; }}
                .error {{ color: red; }}
                .container {{ max-width: 800px; margin: 0 auto; }}
                table {{ border-collapse: collapse; width: 100%; margin-top: 20px; }}
                th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
                th {{ background-color: #f2f2f2; }}
            </style>
        </head>
        <body>
            <div class="container">
                <h1 class="{'success' if resultado.get('success') else 'error'}">
                    {resultado.get('mensaje', 'Operación completada')}
                </h1>
                
                {'<h2>Tablas limpiadas:</h2>' if resultado.get('tablas_limpiadas') else ''}
                {'<ul>' + ''.join([f'<li>{tabla}</li>' for tabla in resultado.get('tablas_limpiadas', [])]) + '</ul>' if resultado.get('tablas_limpiadas') else ''}
                
                {'<h2>Registros eliminados:</h2>' if resultado.get('registros_eliminados') else ''}
                {'<table><tr><th>Tabla</th><th>Estado</th><th>Antes</th><th>Después</th><th>Eliminados</th></tr>' + 
                ''.join([
                    f'<tr><td>{tabla}</td><td>{"Limpiada" if info.get("exists", True) else "No existe"}</td><td>{info.get("before", 0)}</td><td>{info.get("after", 0)}</td><td>{info.get("deleted", 0)}</td></tr>' 
                    if "error" not in info else 
                    f'<tr><td>{tabla}</td><td style="color:red;">Error: {info.get("error", "Desconocido")}</td><td colspan="3">-</td></tr>'
                    for tabla, info in resultado.get('registros_eliminados', {}).items()
                ]) + '</table>' if resultado.get('registros_eliminados') else ''}
                
                <p><a href="/">Volver a la página principal</a></p>
            </div>
        </body>
        </html>
        """
        
        # Devolver respuesta HTML
        return html_response
        
    except ImportError:
        return "Error: No se encontró el módulo clean_railway_db.py"
    except Exception as e:
        return f"Error al ejecutar la limpieza de la base de datos: {str(e)}"

# Endpoint para realizar diagnóstico de productos persistentes
@app.route('/diagnostico_productos/<string:codigo_secreto>', methods=['GET'])
def diagnostico_productos(codigo_secreto):
    """Endpoint público que ejecuta el diagnóstico de productos persistentes.
    Se accede directamente desde el navegador con un código secreto en la URL.
    Puede recibir el parámetro 'clean=true' para eliminar los productos problemáticos.
    """
    # Código secreto fijo para facilitar su uso
    CODIGO_SECRETO_DIAGNOSTICO = "DiagRailwayDB2025"
    
    # Verificar código secreto
    if codigo_secreto != CODIGO_SECRETO_DIAGNOSTICO:
        return jsonify({
            "success": False,
            "mensaje": "Código secreto inválido"
        })
    
    # Verificar si se solicitó limpieza
    clean_option = request.args.get('clean', 'false').lower() == 'true'
    
    # Ejecutar diagnóstico
    try:
        import io
        import sys
        from contextlib import redirect_stdout
        
        # Capturar la salida del diagnóstico
        buffer = io.StringIO()
        with redirect_stdout(buffer):
            from diagnostico_productos import run_diagnostics
            run_diagnostics(clean_products=clean_option)
        
        output = buffer.getvalue()
        
        # Verificar si hay un archivo de resultados JSON
        import os
        import json
        json_results = {}
        if os.path.exists('diagnostico_productos_resultado.json'):
            try:
                with open('diagnostico_productos_resultado.json', 'r', encoding='utf-8') as f:
                    json_results = json.load(f)
            except Exception as e:
                output += f"\nError al leer el archivo JSON de resultados: {e}"
        
        # Crear una respuesta HTML para mostrar en el navegador
        html_response = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <title>Diagnóstico de productos persistentes</title>
            <style>
                body {{ font-family: Arial, sans-serif; margin: 20px; line-height: 1.6; }}
                .success {{ color: green; }}
                .error {{ color: red; }}
                .container {{ max-width: 900px; margin: 0 auto; }}
                pre {{ background-color: #f5f5f5; padding: 15px; overflow-x: auto; white-space: pre-wrap; }}
                table {{ border-collapse: collapse; width: 100%; margin-top: 20px; }}
                th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
                th {{ background-color: #f2f2f2; }}
                .actions {{ margin-top: 20px; }}
                .actions a {{ display: inline-block; padding: 10px 15px; background-color: #4CAF50; color: white; 
                              text-decoration: none; margin-right: 10px; border-radius: 4px; }}
                .actions a.danger {{ background-color: #f44336; }}
            </style>
        </head>
        <body>
            <div class="container">
                <h1>Diagnóstico de productos persistentes</h1>
                
                <div class="actions">
                    <a href="/">Volver a la página principal</a>
                    <a href="/diagnostico_productos/{CODIGO_SECRETO_DIAGNOSTICO}?clean=true" class="danger">
                        Ejecutar con eliminación de productos
                    </a>
                </div>
                
                <h2>Resultados del diagnóstico</h2>
                <pre>{output}</pre>
                
                {'<h2>Resumen del diagnóstico</h2>' if json_results else ''}
                {'<table>' +
                '<tr><th>Categoría</th><th>Productos encontrados</th></tr>' +
                f'<tr><td>Base de datos</td><td>{"Sí" if json_results.get("resultados_db") else "No"}</td></tr>' +
                f'<tr><td>Archivos Excel</td><td>{"Sí" if json_results.get("resultados_excel") else "No"}</td></tr>' +
                f'<tr><td>Productos eliminados</td><td>{"Sí" if json_results.get("productos_eliminados") else "No"}</td></tr>' +
                '</table>' if json_results else ''}
                
                <div class="actions">
                    <a href="/limpiar_base_datos_railway/CleanRailwayDB2025">
                        Ejecutar limpieza completa de la base de datos
                    </a>
                </div>
            </div>
        </body>
        </html>
        """
        
        return html_response
        
    except ImportError as e:
        return f"Error: No se encontró el módulo necesario: {e}"
    except Exception as e:
        import traceback
        return f"Error al ejecutar el diagnóstico: {str(e)}<br><pre>{traceback.format_exc()}</pre>"

@app.route('/api/fix_railway_db', methods=['POST'])
def fix_railway_db():
    """Endpoint para ejecutar la migración de la base de datos en Railway.
    
    Requiere un token de seguridad para evitar ejecuciones no autorizadas.
    El token se configura mediante la variable de entorno MIGRATION_TOKEN.
    """
    try:
        # Verificar si estamos en Railway con PostgreSQL
        if not _is_postgres_configured():
            return jsonify({
                'success': False,
                'message': 'Este endpoint solo funciona en entornos con PostgreSQL (Railway)'
            }), 400
        
        # Verificar token de seguridad
        expected_token = os.environ.get('MIGRATION_TOKEN')
        if not expected_token:
            return jsonify({
                'success': False, 
                'message': 'No se ha configurado MIGRATION_TOKEN en las variables de entorno'
            }), 500
        
        provided_token = request.headers.get('X-Migration-Token') or request.form.get('token')
        if not provided_token or provided_token != expected_token:
            return jsonify({
                'success': False,
                'message': 'Token de migración inválido o no proporcionado'
            }), 403
        
        # Ejecutar la migración
        from fix_railway_pg import execute_migration
        result = execute_migration()
        
        if result:
            return jsonify({
                'success': True,
                'message': 'Migración aplicada correctamente'
            })
        else:
            return jsonify({
                'success': False,
                'message': 'Error al aplicar la migración. Revise los logs del servidor.'
            }), 500
    
    except Exception as e:
        print(f"[ERROR] Error en el endpoint de migración: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'message': f'Error inesperado: {str(e)}'
        }), 500

@app.route('/api/fix_railway_proveedores_case', methods=['POST'])
def fix_railway_proveedores_case():
    """Endpoint para normalizar nombres de proveedores en la base de datos de Railway.
    
    Requiere un token de seguridad para evitar ejecuciones no autorizadas.
    El token se configura mediante la variable de entorno MIGRATION_TOKEN.
    """
    try:
        # Verificar si estamos en Railway con PostgreSQL
        if not _is_postgres_configured():
            return jsonify({
                'success': False,
                'message': 'Este endpoint solo funciona en entornos con PostgreSQL (Railway)'
            }), 400
        
        # Verificar token de seguridad
        expected_token = os.environ.get('MIGRATION_TOKEN')
        if not expected_token:
            return jsonify({
                'success': False, 
                'message': 'No se ha configurado MIGRATION_TOKEN en las variables de entorno'
            }), 500
        
        provided_token = request.headers.get('X-Migration-Token') or request.form.get('token')
        if not provided_token or provided_token != expected_token:
            return jsonify({
                'success': False,
                'message': 'Token de migración inválido o no proporcionado'
            }), 403
        
        # Ejecutar la normalización de proveedores
        from fix_railway_proveedores_case import normalizar_proveedores
        result = normalizar_proveedores()
        
        if result:
            return jsonify({
                'success': True,
                'message': 'Normalización de proveedores aplicada correctamente'
            })
        else:
            return jsonify({
                'success': False,
                'message': 'Error al normalizar proveedores. Revise los logs del servidor.'
            }), 500
    
    except Exception as e:
        print(f"[ERROR] Error en el endpoint de normalización de proveedores: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'message': f'Error inesperado: {str(e)}'
        }), 500

@app.route('/api/sincronizar_proveedores', methods=['POST'])
@login_required
def api_sincronizar_proveedores():
    """Endpoint para sincronizar las tablas proveedores_meta y proveedores_duenos.
    
    Soluciona el problema de proveedores que no aparecen en el formulario
    de agregar productos en Railway.
    """
    try:
        print("[DEBUG] Iniciando sincronización de proveedores...")
        
        # Ejecutar la sincronización
        success, message = sincronizar_proveedores_meta_duenos()
        
        if success:
            print(f"[DEBUG] Sincronización exitosa: {message}")
            return jsonify({
                'success': True,
                'message': message
            })
        else:
            print(f"[DEBUG] Sincronización falló: {message}")
            return jsonify({
                'success': False,
                'message': message
            }), 500
    
    except Exception as e:
        print(f"[ERROR] Error en sincronización de proveedores: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'message': f'Error inesperado: {str(e)}'
        }), 500

# Endpoint alternativo sin autenticación para Railway (con código secreto)
@app.route('/api/sincronizar_proveedores_railway/<string:codigo_secreto>', methods=['GET', 'POST'])
def api_sincronizar_proveedores_railway(codigo_secreto):
    """Endpoint público para sincronizar proveedores en Railway usando código secreto."""
    try:
        # Verificar código secreto
        codigo_esperado = os.environ.get('RAILWAY_SECRET_CODE', 'railway_fix_2024')
        if codigo_secreto != codigo_esperado:
            return jsonify({
                'success': False,
                'message': 'Código secreto inválido'
            }), 403
        
        print("[DEBUG] Iniciando sincronización Railway...")
        
        # Ejecutar la sincronización
        success, message = sincronizar_proveedores_meta_duenos()
        
        if success:
            print(f"[DEBUG] Sincronización Railway exitosa: {message}")
            return jsonify({
                'success': True,
                'message': message
            })
        else:
            print(f"[DEBUG] Sincronización Railway falló: {message}")
            return jsonify({
                'success': False,
                'message': message
            }), 500
    
    except Exception as e:
        print(f"[ERROR] Error en sincronización Railway: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'message': f'Error inesperado: {str(e)}'
        }), 500

@app.route('/fix_railway_simple/<string:codigo_secreto>', methods=['GET'])
def fix_railway_simple(codigo_secreto):
    """Endpoint público simple para crear solo la tabla proveedores_duenos sin migrar datos."""
    try:
        # Verificar código secreto
        codigo_esperado = os.environ.get('RAILWAY_SECRET_CODE', 'railway_fix_2024')
        if codigo_secreto != codigo_esperado:
            return jsonify({
                'success': False,
                'message': 'Código secreto inválido'
            }), 403
        
        print("[DEBUG] Ejecutando fix simple para Railway...")
        
        if not _is_postgres_configured():
            return jsonify({
                'success': False,
                'message': 'Este endpoint solo funciona con PostgreSQL (Railway)'
            }), 400
        
        # Solo crear la tabla, SIN migración de datos
        create_table_sql = """
        CREATE TABLE IF NOT EXISTS proveedores_duenos (
            id SERIAL PRIMARY KEY,
            proveedor_id INTEGER NOT NULL,
            dueno TEXT NOT NULL,
            CONSTRAINT proveedores_duenos_unique UNIQUE (proveedor_id, dueno),
            CONSTRAINT fk_proveedor FOREIGN KEY (proveedor_id) REFERENCES proveedores_manual(id) ON DELETE CASCADE
        )
        """
        
        success = db_query(create_table_sql)
        if not success:
            return jsonify({
                'success': False,
                'message': 'Error creando tabla proveedores_duenos'
            }), 500
        
        # Crear índices
        indices = [
            "CREATE INDEX IF NOT EXISTS idx_proveedores_duenos_proveedor_id ON proveedores_duenos(proveedor_id)",
            "CREATE INDEX IF NOT EXISTS idx_proveedores_duenos_dueno ON proveedores_duenos(dueno)"
        ]
        
        for indice in indices:
            db_query(indice)
        
        return jsonify({
            'success': True,
            'message': 'Tabla proveedores_duenos creada exitosamente. Los nuevos proveedores manuales ahora funcionarán correctamente.'
        })
        
    except Exception as e:
        print(f"[ERROR] Error en fix simple: {e}")
        return jsonify({
            'success': False,
            'message': f'Error inesperado: {str(e)}'
        }), 500

@app.route('/init_railway_db/<string:codigo_secreto>', methods=['GET'])
def init_railway_db(codigo_secreto):
    """Endpoint para forzar la inicialización completa de la base de datos en Railway."""
    try:
        # Verificar código secreto
        codigo_esperado = os.environ.get('RAILWAY_SECRET_CODE', 'railway_fix_2024')
        if codigo_secreto != codigo_esperado:
            return jsonify({
                'success': False,
                'message': 'Código secreto inválido'
            }), 403
        
        print("[DEBUG] Forzando inicialización completa de Railway DB...")
        
        if not _is_postgres_configured():
            return jsonify({
                'success': False,
                'message': 'Este endpoint solo funciona con PostgreSQL (Railway)'
            }), 400
        
        # Ejecutar init_db() completo
        try:
            init_db()
            print("[DEBUG] init_db() ejecutado exitosamente")
        except Exception as e:
            print(f"[ERROR] Error en init_db(): {e}")
            return jsonify({
                'success': False,
                'message': f'Error en init_db(): {str(e)}'
            }), 500
        
        # Verificar que las tablas críticas existan
        tablas_criticas = ['proveedores_manual', 'proveedores_meta', 'proveedores_duenos', 'notificaciones']
        tablas_creadas = []
        tablas_faltantes = []
        
        for tabla in tablas_criticas:
            try:
                result = db_query(f"SELECT COUNT(*) FROM {tabla}", fetch=True)
                if result is not None:
                    count = result[0]['count'] if result else 0
                    tablas_creadas.append(f"{tabla} ({count} registros)")
                else:
                    tablas_faltantes.append(tabla)
            except Exception as e:
                print(f"[ERROR] Error verificando tabla {tabla}: {e}")
                tablas_faltantes.append(f"{tabla} (error: {str(e)})")
        
        # Verificar la consulta específica que falla
        try:
            test_proveedores = db_query("""
                SELECT DISTINCT p.nombre 
                FROM proveedores_manual p
                JOIN proveedores_duenos pd ON p.id = pd.proveedor_id
                WHERE pd.dueno = %s
                ORDER BY p.nombre
            """, ('ferreteria_general',), fetch=True)
            
            consulta_funciona = True
            proveedores_encontrados = [p['nombre'] for p in test_proveedores] if test_proveedores else []
        except Exception as e:
            consulta_funciona = False
            proveedores_encontrados = []
            print(f"[ERROR] Error en consulta de proveedores: {e}")
        
        return jsonify({
            'success': len(tablas_faltantes) == 0,
            'message': 'Inicialización de Railway DB completada',
            'tablas_creadas': tablas_creadas,
            'tablas_faltantes': tablas_faltantes,
            'consulta_proveedores_funciona': consulta_funciona,
            'proveedores_ferreteria_general': proveedores_encontrados
        })
        
    except Exception as e:
        print(f"[ERROR] Error en inicialización Railway: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'message': f'Error inesperado: {str(e)}'
        }), 500

@app.route('/api/diagnostico_proveedores', methods=['GET'])
@login_required
def api_diagnostico_proveedores():
    """Endpoint para diagnosticar el estado de las tablas de proveedores."""
    try:
        conn = get_db_connection()
        if not conn:
            return jsonify({
                'success': False,
                'message': 'No se pudo conectar a la base de datos'
            }), 500
        
        cursor = conn.cursor()
        
        # Contar proveedores en cada tabla
        diagnostico = {}
        
        # proveedores_manual
        cursor.execute("SELECT COUNT(*) FROM proveedores_manual")
        diagnostico['proveedores_manual_count'] = cursor.fetchone()[0]
        
        # proveedores_meta
        cursor.execute("SELECT COUNT(*) FROM proveedores_meta")
        diagnostico['proveedores_meta_count'] = cursor.fetchone()[0]
        
        # proveedores_duenos
        try:
            cursor.execute("SELECT COUNT(*) FROM proveedores_duenos")
            diagnostico['proveedores_duenos_count'] = cursor.fetchone()[0]
            diagnostico['proveedores_duenos_exists'] = True
        except Exception:
            diagnostico['proveedores_duenos_count'] = 0
            diagnostico['proveedores_duenos_exists'] = False
        
        # Obtener distribución por dueño en cada tabla
        try:
            cursor.execute("SELECT dueno, COUNT(*) FROM proveedores_meta GROUP BY dueno")
            diagnostico['meta_por_dueno'] = dict(cursor.fetchall())
        except Exception:
            diagnostico['meta_por_dueno'] = {}
        
        try:
            cursor.execute("""
                SELECT pd.dueno, COUNT(*) 
                FROM proveedores_duenos pd
                GROUP BY pd.dueno
            """)
            diagnostico['duenos_por_dueno'] = dict(cursor.fetchall())
        except Exception:
            diagnostico['duenos_por_dueno'] = {}
        
        cursor.close()
        conn.close()
        
        return jsonify({
            'success': True,
            'diagnostico': diagnostico
        })
    
    except Exception as e:
        print(f"[ERROR] Error en diagnóstico: {e}")
        return jsonify({
            'success': False,
            'message': f'Error: {str(e)}'
        }), 500

@app.route('/admin/proveedores')
@login_required
def admin_proveedores():
    """Página de administración para diagnosticar y corregir problemas con proveedores."""
    return render_template('admin_proveedores.html')

# Ruta para corregir problemas de visibilidad de proveedores en Railway
@app.route('/admin/corregir_proveedores', methods=['GET', 'POST'])
@login_required
def admin_corregir_proveedores():
    """Ruta de administración para diagnosticar y corregir problemas con los proveedores
    en la versión Railway (PostgreSQL). Específicamente soluciona el problema de proveedores
    que no aparecen en la sección de agregar productos ni en gestionar productos."""
    
    # Inicializar resultados
    diagnostico = {}
    correccion = {}
    indices = {}
    
    # Si es una solicitud POST, realizar correcciones
    if request.method == 'POST':
        accion = request.form.get('accion', '')
        
        if accion == 'diagnosticar':
            diagnostico = diagnosticar_proveedores_railway()
        
        elif accion == 'corregir':
            correccion = corregir_proveedores_railway()
            diagnostico = diagnosticar_proveedores_railway()  # Actualizar diagnóstico después de corrección
        
        elif accion == 'indices':
            indices = verificar_indices_railway()
    
    # Para GET, mostrar solo diagnóstico inicial
    else:
        diagnostico = diagnosticar_proveedores_railway()
    
    return render_template(
        'admin_corregir_proveedores.html',
        diagnostico=diagnostico,
        correccion=correccion,
        indices=indices,
        is_postgres=_is_postgres_configured()
    )

def diagnosticar_proveedores_railway():
    """Diagnostica el estado de los proveedores y sus relaciones en la base de datos.
    
    Verifica si hay proveedores sin dueños asociados, lo cual causa que no aparezcan
    en las secciones de agregar productos y gestionar productos.
    
    Returns:
        dict: Un diccionario con los resultados del diagnóstico
    """
    resultados = {
        'proveedores_total': 0,
        'proveedores_sin_dueno': [],
        'duenos': []
    }
    
    conn = get_db_connection()
    if not conn:
        return {"error": "No se pudo conectar a la base de datos"}
    
    try:
        # Verificar proveedores sin dueños asociados
        query_proveedores_sin_dueno = """
        SELECT p.id, p.nombre
        FROM proveedores_manual p
        LEFT JOIN proveedores_duenos pd ON p.id = pd.proveedor_id
        WHERE pd.proveedor_id IS NULL
        ORDER BY p.nombre
        """
        
        sin_dueno = db_query(query_proveedores_sin_dueno, fetch=True, conn=conn)
        
        # Obtener todos los proveedores para estadísticas
        query_todos_proveedores = "SELECT COUNT(*) as total FROM proveedores_manual"
        total = db_query(query_todos_proveedores, fetch=True, conn=conn)
        
        # Obtener dueños existentes
        query_duenos = "SELECT DISTINCT dueno FROM proveedores_duenos"
        duenos = db_query(query_duenos, fetch=True, conn=conn)
        
        resultados['proveedores_total'] = total[0]['total'] if total else 0
        resultados['proveedores_sin_dueno'] = sin_dueno if sin_dueno else []
        resultados['duenos'] = [d['dueno'] for d in duenos] if duenos else []
        
        # Verificar tabla proveedores_duenos
        if _is_postgres_configured():
            # Verificar índices en PostgreSQL
            indices_query = """
            SELECT indexname FROM pg_indexes 
            WHERE tablename = 'proveedores_duenos'
            """
            indices = db_query(indices_query, fetch=True, conn=conn)
            resultados['indices'] = [idx['indexname'] for idx in indices] if indices else []
        
    except Exception as e:
        resultados['error'] = f"Error durante el diagnóstico: {str(e)}"
    finally:
        try:
            conn.close()
        except:
            pass
    
    return resultados

def corregir_proveedores_railway():
    """Corrige el problema de proveedores sin dueños asociados.
    
    Asocia automáticamente los proveedores sin dueño a ambos dueños
    ('ricky' y 'ferreteria_general') para garantizar su visibilidad.
    
    Returns:
        dict: Un diccionario con los resultados de la corrección
    """
    resultados = {
        'corregidos': 0,
        'errores': [],
        'proveedores': []
    }
    
    # Primero diagnosticar
    diagnostico = diagnosticar_proveedores_railway()
    if 'error' in diagnostico:
        return {"error": diagnostico['error']}
    
    if not diagnostico.get('proveedores_sin_dueno'):
        return {"mensaje": "No hay proveedores sin dueño que corregir", "corregidos": 0}
    
    conn = get_db_connection()
    if not conn:
        return {"error": "No se pudo conectar a la base de datos"}
    
    try:
        for p in diagnostico['proveedores_sin_dueno']:
            proveedor_id = p['id']
            nombre = p['nombre']
            
            # Asociar a ambos dueños
            duenos = ['ricky', 'ferreteria_general']
            proveedor_corregido = True
            
            for d in duenos:
                # Insertar en proveedores_duenos
                ok_duenos = db_query(
                    "INSERT OR IGNORE INTO proveedores_duenos (proveedor_id, dueno) VALUES (?, ?)",
                    (proveedor_id, d),
                    conn=conn
                )
                
                # Insertar en proveedores_meta (legacy)
                ok_meta = db_query(
                    "INSERT OR IGNORE INTO proveedores_meta (nombre, dueno) VALUES (?, ?)",
                    (nombre, d),
                    conn=conn
                )
                
                if not (ok_duenos and ok_meta):
                    proveedor_corregido = False
                    resultados['errores'].append(f"Error al asociar proveedor '{nombre}' a dueño '{d}'")
            
            if proveedor_corregido:
                resultados['corregidos'] += 1
                resultados['proveedores'].append(nombre)
    
    except Exception as e:
        resultados['error'] = f"Error durante la corrección: {str(e)}"
    finally:
        try:
            conn.close()
        except:
            pass
    
    return resultados

def verificar_indices_railway():
    """Verifica y crea índices necesarios en PostgreSQL para mejorar el rendimiento.
    
    Esta función solo hace algo en entorno PostgreSQL (Railway).
    
    Returns:
        dict: Un diccionario con los resultados de la verificación
    """
    if not _is_postgres_configured():
        return {"mensaje": "No es necesario verificar índices en SQLite"}
    
    resultados = {
        'indices_creados': 0,
        'errores': []
    }
    
    conn = get_db_connection()
    if not conn:
        return {"error": "No se pudo conectar a la base de datos"}
    
    try:
        # Índices para proveedores_duenos
        indices = [
            "CREATE INDEX IF NOT EXISTS idx_proveedores_duenos_proveedor_id ON proveedores_duenos(proveedor_id)",
            "CREATE INDEX IF NOT EXISTS idx_proveedores_duenos_dueno ON proveedores_duenos(dueno)"
        ]
        
        for idx_query in indices:
            ok = db_query(idx_query, conn=conn)
            if ok:
                resultados['indices_creados'] += 1
            else:
                resultados['errores'].append(f"Error al crear índice: {idx_query}")
    
    except Exception as e:
        resultados['error'] = f"Error durante la verificación de índices: {str(e)}"
    finally:
        try:
            conn.close()
        except:
            pass
    
    return resultados

if __name__ == '__main__':
    print("🚀 Iniciando Gestor de Stock...")
    print("📁 Directorio base:", BASE_DIR)
    print("📊 Carpeta Excel:", EXCEL_FOLDER)
    print("🗄️ Base de datos:", DATABASE_FILE)
    
    try:
        print("🔧 Inicializando base de datos...")
        if os.environ.get('AUTO_INIT_DB') == '1':
            try:
                from init_db import main as _auto_init_main
                print('[AUTO_INIT_DB] Variable activa -> ejecutando init_db.main()')
                _auto_init_main()
            except Exception as _e_ai:
                print(f'[AUTO_INIT_DB] Advertencia: fallo en auto init: {_e_ai}')
        init_db()
        
        # Añadir información del tipo de entorno
        if _is_postgres_configured():
            print("\n🚀 MODO PRODUCCIÓN: PostgreSQL activado 🐘")
            print("🔗 Conectado a: " + os.environ.get('DATABASE_URL', 'DATABASE_URL configurado'))
        
        print("🌐 Iniciando servidor Flask...")
        print("📋 Accede a: http://localhost:5000")
        print("👤 Usuario: Pauluk")
        print("🔑 Contraseña: Jap2005")
        print("=" * 50)
        
        # Configurar para producción
        port = int(os.environ.get('PORT', 5000))
        debug = os.environ.get('FLASK_ENV') == 'development'
        
        app.run(debug=debug, host='0.0.0.0', port=port)
        
    except KeyboardInterrupt:
        print("\n🛑 Servidor detenido por el usuario")
    except Exception as e:
        print(f"❌ Error al iniciar el servidor: {e}")
        import traceback
        traceback.print_exc()