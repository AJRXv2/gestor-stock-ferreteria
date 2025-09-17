#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
DEBUG ESPECÍFICO BREMEN - Análisis de todas las filas que contienen '7044'
"""

import sys
import os

# Agregar el directorio actual al path para importar gestor
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

print("🔄 Cargando dependencias...")

try:
    from werkzeug.datastructures import MultiDict
    from functools import wraps
    print("✅ Werkzeug y functools importados correctamente")
except ImportError as e:
    print(f"❌ Error importando Werkzeug: {e}")
    sys.exit(1)

try:
    from flask_wtf import FlaskForm
    print("✅ Flask-WTF importado correctamente")
except ImportError:
    print("⚠️ Flask-WTF no disponible, usando fallback simple")

# Importar openpyxl para procesar Excel
try:
    from openpyxl import load_workbook
    print("✅ openpyxl importado correctamente")
except ImportError:
    print("❌ openpyxl no disponible")
    sys.exit(1)

print("🔍 DEBUG BREMEN COMPLETO - TODAS LAS FILAS CON '7044'")
print("=" * 60)

# Cargar archivo Bremen
archivo_bremen = r"listas_excel\ricky\BremenTools-092025.xlsx"

try:
    print(f"📁 Cargando archivo: {archivo_bremen}")
    wb = load_workbook(archivo_bremen, data_only=True)
    ws = wb.active
    print(f"📊 Archivo cargado: {ws.max_row} filas, {ws.max_column} columnas")
    
    # Obtener todas las filas como lista
    filas = list(ws.rows)
    
    termino_busqueda = '7044'
    print(f"\n🔍 Buscando todas las filas que contienen: '{termino_busqueda}'")
    
    coincidencias = []
    
    for row_idx, row in enumerate(filas, 1):
        # Crear versión de texto para buscar
        row_values_original = [str(cell.value) if cell.value is not None else '' for cell in row]
        row_values_busqueda = [str(cell.value).lower() if cell.value is not None else '' for cell in row]
        row_text = ' '.join(row_values_busqueda)
        
        # Verificar si contiene el término
        if termino_busqueda.lower() in row_text:
            coincidencias.append((row_idx, row, row_values_original))
            
    print(f"\n✅ Encontradas {len(coincidencias)} filas que contienen '{termino_busqueda}':")
    
    for i, (row_idx, row, row_values_original) in enumerate(coincidencias, 1):
        print(f"\n--- FILA {i} (Índice {row_idx}) ---")
        print(f"📍 Posición: Fila {row_idx}")
        
        # Mostrar primeras 12 columnas (A-L)
        for col_idx, valor in enumerate(row_values_original[:12], 0):
            letra_col = chr(65 + col_idx)  # A, B, C, etc.
            print(f"   {letra_col}: '{valor}'")
            
        # Análisis específico Bremen
        print(f"\n🔍 ANÁLISIS BREMEN:")
        
        # Extraer código (columna A o donde esté)
        codigo = row_values_original[0] if len(row_values_original) > 0 else ''
        print(f"   📋 Código (A): '{codigo}'")
        
        # Extraer nombre (columna B o donde esté)  
        nombre = row_values_original[1] if len(row_values_original) > 1 else ''
        print(f"   📝 Nombre (B): '{nombre}'")
        
        # Extraer precio (columna J - índice 9)
        precio_j = None
        if len(row) > 9:
            precio_j = row[9].value
            print(f"   💰 Precio (J): {precio_j} (tipo: {type(precio_j)})")
        else:
            print(f"   💰 Precio (J): No disponible (fila solo tiene {len(row)} columnas)")
            
        # Verificar si es numérico
        es_numerico = precio_j is not None and isinstance(precio_j, (int, float))
        print(f"   ✅ Es numérico: {es_numerico}")
        
        # Buscar proveedor en columnas adicionales
        proveedor_detectado = None
        if len(row_values_original) > 3:
            for col_idx in range(3, min(len(row_values_original), 8)):
                valor_col = str(row_values_original[col_idx]).strip()
                if valor_col and len(valor_col) > 2 and valor_col.lower() not in ['', 'nan', 'none', '0']:
                    try:
                        float(valor_col.replace(',', '.').replace('.', ''))
                    except (ValueError, TypeError):
                        if not any(char.isdigit() for char in valor_col):
                            proveedor_detectado = valor_col
                            print(f"   🏭 Proveedor detectado (col {chr(65+col_idx)}): '{proveedor_detectado}'")
                            break
        
        if not proveedor_detectado:
            print(f"   🏭 Proveedor: No detectado, usará 'brementools'")
            
        print(f"   🎯 ¿Generaría resultado? {es_numerico or not es_numerico}")
        
except Exception as e:
    print(f"❌ Error: {e}")
    import traceback
    traceback.print_exc()